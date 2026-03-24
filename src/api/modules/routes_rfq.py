# RFQ + Quote Routes
# 9 routes, 484 lines
# Loaded by dashboard.py via load_module()

# ── Explicit imports (S11 refactor: no longer relying solely on injection) ──
from flask import request, jsonify
from src.api.shared import bp, auth_required
import logging
log = logging.getLogger("reytech")
from src.core.error_handler import safe_route
from flask import redirect, flash
from src.core.paths import DATA_DIR, UPLOAD_DIR, OUTPUT_DIR
from src.core.db import get_db
from src.api.render import render_page
from datetime import datetime, timezone, timedelta
import re as _re_mod


def _validate_rid(rid: str):
    """Validate rfq_id to prevent path traversal. Returns None if valid,
    or a (response, status_code) tuple if invalid."""
    if not rid or not _re_mod.match(r'^[a-zA-Z0-9_-]+$', rid):
        return jsonify({"ok": False, "error": "Invalid RFQ ID"}), 400
    return None


# ═══════════════════════════════════════════════════════════════════════
# Pricing Intelligence — catalog + price history integration
# ═══════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════
# F11: Margin Guardrails — configurable pricing rules
# ═══════════════════════════════════════════════════════════════════════

MARGIN_RULES = {
    "min_margin_pct": 15,       # Warn if margin below this
    "critical_margin_pct": 5,   # Fail if margin below this
    "max_over_scprs_pct": 10,   # Warn if bid > SCPRS by this much
    "max_under_scprs_pct": 15,  # Warn if bid < SCPRS by this much (leaving money)
    "require_cost_source": True, # Warn if cost has no backing URL/SCPRS
}


def _check_guardrails(items):
    """F11: Check margin guardrails on all items. Returns list of warnings."""
    warnings = []
    for i, item in enumerate(items):
        bid = item.get("price_per_unit") or 0
        cost = item.get("supplier_cost") or 0
        scprs = item.get("scprs_last_price") or 0
        desc = (item.get("description", "") or "")[:40]
        if not bid or bid <= 0:
            continue

        # Margin check
        if cost > 0:
            margin = (bid - cost) / bid * 100
            if margin < MARGIN_RULES["critical_margin_pct"]:
                warnings.append({
                    "idx": i, "desc": desc, "level": "critical",
                    "msg": f"Margin {margin:.1f}% is below {MARGIN_RULES['critical_margin_pct']}% minimum"
                })
            elif margin < MARGIN_RULES["min_margin_pct"]:
                warnings.append({
                    "idx": i, "desc": desc, "level": "warn",
                    "msg": f"Margin {margin:.1f}% is below {MARGIN_RULES['min_margin_pct']}% target"
                })

        # SCPRS comparison
        if scprs > 0 and bid > 0:
            diff_pct = (bid - scprs) / scprs * 100
            if diff_pct > MARGIN_RULES["max_over_scprs_pct"]:
                warnings.append({
                    "idx": i, "desc": desc, "level": "warn",
                    "msg": f"Bid is {diff_pct:.0f}% above SCPRS — may lose"
                })

        # Cost without source
        if cost > 0 and MARGIN_RULES["require_cost_source"]:
            if not item.get("item_link") and not scprs:
                warnings.append({
                    "idx": i, "desc": desc, "level": "info",
                    "msg": "Cost has no backing source (no URL or SCPRS)"
                })

    return warnings


def _recommend_price(item):
    """Lightweight price recommendation (mirrors routes_analytics logic).
    Returns {recommended, aggressive, safe} tiers or None."""
    scprs = item.get("scprs_last_price") or 0
    amazon = item.get("amazon_price") or 0
    cost = item.get("supplier_cost") or 0

    base = 0
    reason = ""

    if scprs > 0:
        base = scprs
        reason = f"SCPRS ${scprs:.2f}"
    elif amazon > 0:
        base = amazon * 1.15
        reason = f"Amazon ${amazon:.2f}+15%"
    elif cost > 0:
        base = cost * 1.25
        reason = f"Cost ${cost:.2f}+25%"
    else:
        return None

    # Ensure minimum margin over cost
    if cost > 0 and base < cost * 1.10:
        base = cost * 1.10

    return {
        "recommended": round(base * 0.98, 2),  # Undercut by 2%
        "aggressive": round(base * 0.93, 2),    # Undercut by 7%
        "safe": round(base * 1.05, 2),           # 5% above
        "reason": reason,
    }


def _enrich_items_with_intel(items, rfq_number="", agency=""):
    """Enrich line items with catalog matches and price history.
    Called on RFQ detail load to surface pricing intelligence."""
    for item in items:
        desc = item.get("description", "")
        pn = item.get("item_number", "") or ""
        if not desc and not pn:
            continue

        # 1. Catalog match
        if not item.get("catalog_match"):
            try:
                from src.core.catalog import search_catalog
                matches = search_catalog(pn or desc[:40], limit=1)
                if matches:
                    m = matches[0]
                    item["catalog_match"] = {
                        "sku": m.get("sku", ""),
                        "name": m.get("name", ""),
                        "typical_cost": m.get("typical_cost", 0),
                        "list_price": m.get("list_price", 0),
                        "category": m.get("category", ""),
                    }
            except Exception:
                pass

        # 2. Price history (last 5 observations)
        if not item.get("price_intel"):
            try:
                from src.core.db import get_price_history_db
                history = get_price_history_db(
                    description=desc[:60] if not pn else "",
                    part_number=pn,
                    limit=5
                )
                if history:
                    prices = [h["unit_price"] for h in history if h.get("unit_price")]
                    item["price_intel"] = {
                        "history_count": len(history),
                        "avg_price": round(sum(prices) / len(prices), 2) if prices else 0,
                        "min_price": round(min(prices), 2) if prices else 0,
                        "max_price": round(max(prices), 2) if prices else 0,
                        "last_price": prices[0] if prices else 0,
                        "last_source": history[0].get("source", "") if history else "",
                        "last_date": history[0].get("found_at", "")[:10] if history else "",
                        "last_quote": history[0].get("quote_number", "") if history else "",
                    }
            except Exception:
                pass


def _record_rfq_prices(rfq_data, source="rfq_save"):
    """Record all priced items to price_history + auto-ingest to catalog."""
    sol = rfq_data.get("solicitation_number", "")
    agency = rfq_data.get("agency", "")
    for item in rfq_data.get("line_items", []):
        desc = item.get("description", "")
        pn = item.get("item_number", "") or ""
        if not desc:
            continue

        # Record supplier cost
        cost = item.get("supplier_cost") or 0
        if cost and cost > 0:
            try:
                from src.core.db import record_price
                record_price(
                    description=desc, unit_price=cost, source=source,
                    part_number=pn, agency=agency, quote_number=sol,
                    source_url=item.get("item_link", ""),
                    notes=f"Supplier cost from RFQ {sol}"
                )
            except Exception:
                pass

        # Record bid price
        bid = item.get("price_per_unit") or 0
        if bid and bid > 0:
            try:
                from src.core.db import record_price
                record_price(
                    description=desc, unit_price=bid, source=f"{source}_bid",
                    part_number=pn, agency=agency, quote_number=sol,
                    notes=f"Bid price from RFQ {sol}"
                )
            except Exception:
                pass

        # Record SCPRS price
        scprs = item.get("scprs_last_price") or 0
        if scprs and scprs > 0:
            try:
                from src.core.db import record_price
                record_price(
                    description=desc, unit_price=scprs, source="scprs",
                    part_number=pn, agency=agency, quote_number=sol,
                )
            except Exception:
                pass

        # Record Amazon price
        amz = item.get("amazon_price") or 0
        if amz and amz > 0:
            try:
                from src.core.db import record_price
                record_price(
                    description=desc, unit_price=amz, source="amazon",
                    part_number=pn, source_url=item.get("item_link", ""),
                )
            except Exception:
                pass

        # Auto-ingest to product_catalog (same table PC uses + auto-price reads)
        if cost > 0 or bid > 0:
            try:
                from src.agents.product_catalog import add_to_catalog, init_catalog_db
                init_catalog_db()
                add_to_catalog(
                    description=desc,
                    part_number=pn,
                    cost=float(cost) if cost else 0,
                    sell_price=float(bid) if bid else 0,
                    source=f"rfq_{sol}",
                    supplier_name=item.get("item_supplier", ""),
                    supplier_url=item.get("item_link", ""),
                )
            except Exception:
                pass

@bp.route("/health")
def health_check():
    """Health check endpoint for Railway/load balancers. No auth required."""
    checks = {"status": "ok", "timestamp": datetime.now().isoformat()}
    # Check SQLite
    try:
        from src.core.db import get_db
        with get_db() as conn:
            conn.execute("SELECT 1")
        checks["db"] = "ok"
    except Exception as e:
        checks["db"] = f"error: {e}"
        checks["status"] = "degraded"
    # Check data dir writable
    try:
        test_path = os.path.join(DATA_DIR, ".health_check")
        with open(test_path, "w") as f: f.write("ok")
        os.remove(test_path)
        checks["disk"] = "ok"
    except Exception as e:
        checks["disk"] = f"error: {e}"
        checks["status"] = "degraded"
    # Check validation module loads
    try:
        from src.core.validation import validate_price
        v, err = validate_price("12.50")
        checks["validation"] = "ok" if v == 12.5 and err is None else "fail"
    except Exception as e:
        checks["validation"] = f"error: {e}"
    # Active RFQ/PC counts (non-critical)
    try:
        active_rfqs = {k: v for k, v in load_rfqs().items() if v.get("status") not in ("dismissed", "sent", "duplicate")}
        checks["active_rfqs"] = len(active_rfqs)
    except Exception:
        checks["active_rfqs"] = -1
    try:
        from src.api.dashboard import load_price_checks
        pcs = load_price_checks()
        active_pcs = {k: v for k, v in pcs.items() if v.get("status") not in ("dismissed", "duplicate", "archived")}
        checks["active_pcs"] = len(active_pcs)
    except Exception:
        checks["active_pcs"] = -1
    code = 200 if checks["status"] == "ok" else 503
    return jsonify(checks), code

@bp.route("/")
@auth_required
def home():
    import time as _ht
    _t0 = _ht.time()
    log.info("HOME: request started")
    try:
        all_pcs = _load_price_checks()
    except Exception:
        all_pcs = {}
    log.info("HOME: PCs loaded (%d) in %.0fms", len(all_pcs), (_ht.time()-_t0)*1000)
    # Recovery runs at boot (dashboard.py), not on every request
    from src.api.dashboard import _is_user_facing_pc
    user_pcs = {k: v for k, v in all_pcs.items() if _is_user_facing_pc(v)}
    # Additional cleanup: filter PCs with no solicitation AND 0 items
    user_pcs = {k: v for k, v in user_pcs.items()
                if len(v.get("items", [])) > 0
                or v.get("status") in ("sent", "won", "lost", "generated", "ready", "priced")
                or (v.get("solicitation_number") or v.get("pc_number", "")) not in ("", "unknown")}
    # Split: active queue vs sent/completed
    _pc_actionable = {"new", "draft", "parsed", "parse_error", "priced", "ready", "auto_drafted", "quoted", "generated"}
    active_pcs = {k: v for k, v in user_pcs.items() if v.get("status", "") in _pc_actionable}
    sent_pcs = {k: v for k, v in user_pcs.items() if v.get("status", "") in ("sent", "pending_award", "won", "lost")}
    # PST "today" for California-based due date comparisons
    _pst = timezone(timedelta(hours=-8))
    _today = datetime.now(_pst).replace(tzinfo=None)
    # Sort by URGENCY: overdue first, then soonest due date, then newest
    def _pc_sort_key(item):
        pc = item[1]
        due = pc.get("due_date", "") or ""
        status = pc.get("status", "")
        # Terminal statuses go to bottom
        if status in ("won", "lost", "dismissed", "archived", "expired"):
            return (3, "9999-99-99", "")
        # Parse due date and compute urgency
        urgency = 1  # default: normal
        try:
            for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
                try:
                    d = datetime.strptime(due.strip(), fmt)
                    days_left = (d - _today).days
                    if days_left < 0:
                        urgency = 0  # OVERDUE — top of queue
                    elif days_left <= 2:
                        urgency = 0  # Due within 48h — also top
                    due_sort = d.strftime("%Y-%m-%d")
                    return (urgency, due_sort, pc.get("created_at", ""))
                except ValueError:
                    continue
        except Exception:
            pass
        # No parseable due date — sort by creation
        return (2, "", pc.get("created_at", ""))
    sorted_pcs = dict(sorted(active_pcs.items(), key=_pc_sort_key))
    
    # Also compute urgency metadata for template
    for pid, pc in sorted_pcs.items():
        due = pc.get("due_date", "") or ""
        pc["_days_left"] = None
        pc["_urgency"] = "normal"
        try:
            for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
                try:
                    d = datetime.strptime(due.strip(), fmt)
                    days = (d - _today).days
                    pc["_days_left"] = days
                    if days < 0: pc["_urgency"] = "overdue"
                    elif days <= 1: pc["_urgency"] = "critical"
                    elif days <= 3: pc["_urgency"] = "soon"
                    break
                except ValueError:
                    continue
        except Exception:
            pass

    # Same for RFQs — split active from sent/completed
    _actionable_rfq = {"new", "draft", "ready", "generated", "parsed"}
    all_rfqs = load_rfqs()
    active_rfqs = {k: v for k, v in all_rfqs.items() if v.get("status", "") in _actionable_rfq}
    # Filter ghost RFQs: 0 items + no real solicitation
    active_rfqs = {k: v for k, v in active_rfqs.items()
                   if len(v.get("line_items", v.get("items", []))) > 0
                   or (v.get("solicitation_number") or v.get("rfq_number", "")) not in ("", "unknown")}
    sent_rfqs = {k: v for k, v in all_rfqs.items() if v.get("status", "") in ("sent", "won", "lost")}
    for rid, r in active_rfqs.items():
        due = r.get("due_date", "") or ""
        r["_days_left"] = None
        r["_urgency"] = "normal"
        try:
            for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
                try:
                    d = datetime.strptime(due.strip(), fmt)
                    days = (d - _today).days
                    r["_days_left"] = days
                    if days < 0: r["_urgency"] = "overdue"
                    elif days <= 1: r["_urgency"] = "critical"
                    elif days <= 3: r["_urgency"] = "soon"
                    break
                except ValueError:
                    continue
        except Exception:
            pass
    # Sort RFQs by urgency too
    active_rfqs = dict(sorted(active_rfqs.items(), key=lambda x: (
        3 if x[1].get("status") in ("sent","generated") else 0 if x[1].get("_urgency") in ("overdue","critical") else 1,
        x[1].get("due_date", "9999"),
    )))
    log.info("HOME: rendering template, %d PCs + %d RFQs, total %.0fms", 
             len(sorted_pcs), len(active_rfqs), (_ht.time()-_t0)*1000)
    return render_page("home.html", active_page="Home", rfqs=active_rfqs, price_checks=sorted_pcs, sent_rfqs=sent_rfqs, sent_pcs=sent_pcs)

@bp.route("/growth")
@auth_required
@safe_route
def growth_redirect():
    """Growth page — redirects to pipeline."""
    return redirect("/pipeline")


@bp.route("/awards")
@auth_required
@safe_route
def awards_page():
    """Pending PO Award Review page."""
    from src.api.dashboard import _load_pending_pos
    pending = _load_pending_pos()
    return render_page("awards.html", active_page="Awards", pending=pending)


@bp.route("/api/award/<int:idx>/approve", methods=["POST"])
@auth_required
def api_award_approve(idx):
    """Approve a pending PO — creates order and marks RFQ/quote as won."""
    from src.api.dashboard import _load_pending_pos, _save_pending_pos, _pending_po_reviews, _create_order_from_po_email
    pending = _load_pending_pos()
    if idx < 0 or idx >= len(pending):
        return jsonify({"ok": False, "error": "Invalid index"})

    po = pending[idx]
    if po.get("review_status") != "pending":
        return jsonify({"ok": False, "error": "Already processed"})

    # Create the order
    try:
        order = _create_order_from_po_email(po)
        po["review_status"] = "approved"
        po["approved_at"] = datetime.now().isoformat()
        po["order_id"] = order.get("id", "") or order.get("order_id", "")
        _save_pending_pos()

        # Also update RFQ status to won
        sol = po.get("sol_number", "")
        if sol:
            rfqs = load_rfqs()
            for rid, r in rfqs.items():
                if r.get("solicitation_number") == sol or r.get("rfq_number") == sol:
                    r["status"] = "won"
                    r["outcome"] = "won"
                    r["outcome_date"] = datetime.now().isoformat()
                    r["po_number"] = po.get("po_number", "")
                    break
            save_rfqs(rfqs)

        return jsonify({
            "ok": True,
            "order_id": po["order_id"],
            "redirect": f"/order/{po['order_id']}",
        })
    except Exception as e:
        log.error("Award approve failed: %s", e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/award/<int:idx>/dismiss", methods=["POST"])
@auth_required
def api_award_dismiss(idx):
    """Dismiss a pending PO (not a real award)."""
    from src.api.dashboard import _load_pending_pos, _save_pending_pos, _pending_po_reviews
    pending = _load_pending_pos()
    if idx < 0 or idx >= len(pending):
        return jsonify({"ok": False, "error": "Invalid index"})

    data = request.get_json(force=True, silent=True) or {}
    pending[idx]["review_status"] = "dismissed"
    pending[idx]["dismiss_reason"] = data.get("reason", "")
    _save_pending_pos()
    return jsonify({"ok": True})


@bp.route("/api/awards/pending")
@auth_required
def api_awards_pending():
    """Get count of pending PO reviews (for home page banner)."""
    from src.api.dashboard import _load_pending_pos
    pending = [p for p in _load_pending_pos() if p.get("review_status") == "pending"]
    return jsonify({"ok": True, "count": len(pending), "pending": pending})


@bp.route("/api/rfq/create-manual", methods=["POST"])
@auth_required
def api_rfq_create_manual():
    """Create an RFQ manually from the dashboard."""
    data = request.get_json(force=True, silent=True) or {}
    sol = data.get("solicitation_number", "").strip()
    if not sol:
        return jsonify({"ok": False, "error": "solicitation_number required"})

    import uuid
    rid = uuid.uuid4().hex[:8]
    agency_key = data.get("agency", "")
    agency_name = ""
    try:
        from src.core.agency_config import AGENCY_CONFIGS
        if agency_key in AGENCY_CONFIGS:
            agency_name = AGENCY_CONFIGS[agency_key].get("name", agency_key)
    except Exception:
        agency_name = agency_key

    rfq = {
        "id": rid,
        "solicitation_number": sol,
        "rfq_number": sol,
        "agency": agency_key,
        "agency_name": agency_name,
        "requestor_name": data.get("requestor_name", ""),
        "requestor_email": data.get("requestor_email", ""),
        "due_date": data.get("due_date", ""),
        "delivery_location": data.get("delivery_location", ""),
        "status": "new",
        "source": "manual",
        "created_at": datetime.now().isoformat(),
        "received_at": datetime.now().isoformat(),
        "line_items": [],
        "notes": data.get("notes", ""),
    }

    rfqs = load_rfqs()
    rfqs[rid] = rfq
    save_rfqs(rfqs)

    try:
        from src.core.dal import log_lifecycle_event
        log_lifecycle_event("rfq", rid, "manual_create",
            f"RFQ #{sol} created manually — {agency_name} / {data.get('requestor_name', '')}",
            actor="user")
    except Exception:
        pass

    return jsonify({"ok": True, "rfq_id": rid, "sol": sol})


@bp.route("/api/rfq/<rid>/upload-parse-doc", methods=["POST"])
@auth_required
def api_rfq_upload_parse_doc(rid):
    """Upload any document (PDF, image, screenshot) -> parse items -> populate RFQ.

    Tries parsers in order:
    1. AMS 704 (if PDF looks like a 704)
    2. Generic RFQ parser (XFA + text extraction)
    3. Vision parser (Claude vision for scanned/image docs)
    """
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "No file uploaded"})

    filename = f.filename.lower()
    is_pdf = filename.endswith(".pdf")
    is_image = any(filename.endswith(ext) for ext in [".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff"])

    if not is_pdf and not is_image:
        return jsonify({"ok": False, "error": "Upload a PDF or image file"})

    # Save uploaded file
    upload_dir = os.path.join(DATA_DIR, "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    save_name = f"doc_{rid}_{f.filename}"
    save_path = os.path.join(upload_dir, save_name)
    f.save(save_path)

    try:
        items = []
        parser_used = ""
        header = {}
        vision_error = None

        if is_pdf:
            # Try 1: AMS 704
            try:
                from src.forms.price_check import parse_ams704
                parsed = parse_ams704(save_path)
                if not parsed.get("error") and parsed.get("line_items"):
                    items = parsed["line_items"]
                    header = parsed.get("header", {})
                    parser_used = "AMS 704"
                    log.info("Upload parse: AMS 704 found %d items", len(items))
            except Exception as e:
                log.debug("704 parse failed: %s", e)

            # Try 2: Generic RFQ parser
            if not items:
                try:
                    from src.forms.generic_rfq_parser import parse_generic_rfq
                    parsed = parse_generic_rfq([save_path],
                        subject=r.get("email_subject", ""),
                        sender_email=r.get("requestor_email", ""),
                        body=r.get("body_text", ""))
                    if parsed.get("items"):
                        items = parsed["items"]
                        header = parsed.get("header", {})
                        parser_used = "Generic RFQ"
                        log.info("Upload parse: Generic found %d items", len(items))
                except Exception as e:
                    log.debug("Generic parse failed: %s", e)

            # Try 3: Vision parser (PDF -> image -> Claude)
            if not items:
                try:
                    from src.forms.vision_parser import parse_with_vision
                    parsed = parse_with_vision(save_path)
                    _vitems = parsed.get("line_items") or parsed.get("items") if parsed else None
                    if _vitems:
                        items = _vitems
                        header = parsed.get("header", {})
                        parser_used = "Vision AI"
                        log.info("Upload parse: Vision found %d items", len(items))
                except Exception as e:
                    log.debug("Vision parse failed: %s", e)

        elif is_image:
            # Image: go straight to vision parser
            vision_error = None
            try:
                from src.forms.vision_parser import parse_with_vision, is_available
                if not is_available():
                    vision_error = "Vision AI unavailable (API key not set — check ANTHROPIC_API_KEY or AGENT_ITEM_ID_KEY env var)"
                    log.warning("Upload parse: %s", vision_error)
                else:
                    parsed = parse_with_vision(save_path)
                    _vitems = parsed.get("line_items") or parsed.get("items") if parsed else None
                    if _vitems:
                        items = _vitems
                        header = parsed.get("header", {})
                        parser_used = "Vision AI"
                        log.info("Upload parse: Vision (image) found %d items", len(items))
                    else:
                        vision_error = "Vision AI returned no items from image"
                        log.warning("Upload parse: %s", vision_error)
            except Exception as e:
                vision_error = f"Vision AI error: {e}"
                log.warning("Vision image parse failed: %s", e)

            # Fallback: try OCR -> text -> generic parser
            if not items:
                try:
                    import subprocess
                    ocr_result = subprocess.run(["tesseract", save_path, "stdout"],
                        capture_output=True, text=True, timeout=30)
                    if ocr_result.returncode == 0 and ocr_result.stdout.strip():
                        from src.forms.generic_rfq_parser import parse_line_items_from_text
                        items = parse_line_items_from_text(ocr_result.stdout)
                        if items:
                            parser_used = "OCR + Text"
                            log.info("Upload parse: OCR found %d items", len(items))
                except Exception as e:
                    log.debug("OCR parse failed: %s", e)

        if not items:
            err_detail = vision_error or "No items could be extracted"
            return jsonify({
                "ok": False,
                "error": f"Could not extract items. {err_detail}",
                "parser_tried": ["AMS 704", "Generic RFQ", "Vision AI"] if is_pdf else ["Vision AI", "OCR"],
            })

        # Merge header info into RFQ if available
        if header:
            if header.get("institution") and not r.get("agency_name"):
                r["agency_name"] = header["institution"]
            if header.get("ship_to_address") and not r.get("delivery_location"):
                r["delivery_location"] = header["ship_to_address"]
            if header.get("price_check_number") and not r.get("linked_pc_id"):
                r["linked_pc_number"] = header["price_check_number"]

        # Build RFQ line items
        existing_items = r.get("line_items", r.get("items", []))
        added = 0

        for it in items:
            desc = it.get("description", "") or it.get("name", "") or ""
            if not desc or len(desc.strip()) < 3:
                continue

            qty = it.get("qty", 1) or it.get("quantity", 1) or 1
            try:
                qty = int(float(qty))
            except (ValueError, TypeError):
                qty = 1

            uom = it.get("uom", "EA") or it.get("unit_of_measure", "EA") or "EA"
            # item_number from vision parser is sequential (1,2,3) — NOT a part number
            part = it.get("mfg_number", "") or it.get("part_number", "") or ""
            if not part:
                raw_inum = str(it.get("item_number", "")).strip()
                if raw_inum and not raw_inum.isdigit():
                    part = raw_inum
            cost = it.get("price", 0) or it.get("unit_price", 0) or it.get("cost", 0) or 0
            try:
                cost = float(cost)
            except (ValueError, TypeError):
                cost = 0

            rfq_item = {
                "description": desc,
                "qty": qty,
                "uom": uom,
                "part_number": part,
                "item_number": part,
            }
            if cost > 0:
                rfq_item["supplier_cost"] = cost
                rfq_item["cost_source"] = f"Uploaded ({parser_used})"

            existing_items.append(rfq_item)
            added += 1

        r["line_items"] = existing_items
        save_rfqs(rfqs)

        try:
            from src.core.dal import log_lifecycle_event
            log_lifecycle_event("rfq", rid, "doc_uploaded",
                f"Uploaded {f.filename}: {added} items parsed via {parser_used}",
                actor="user", detail={"filename": f.filename, "parser": parser_used, "items": added})
        except Exception:
            pass

        return jsonify({
            "ok": True,
            "items_found": len(items),
            "items_added": added,
            "parser": parser_used,
            "header": header,
        })
    finally:
        # Clean up uploaded file after parsing
        try:
            os.remove(save_path)
        except OSError:
            pass


@bp.route("/upload", methods=["POST"])
@auth_required
def upload():
    files = request.files.getlist("files")
    if not files:
        flash("No files uploaded", "error"); return redirect("/")
    
    rfq_id = str(uuid.uuid4())[:8]
    rfq_dir = os.path.join(UPLOAD_DIR, rfq_id)
    os.makedirs(rfq_dir, exist_ok=True)
    
    saved = []
    for f in files:
        safe_fn = _safe_filename(f.filename)
        if safe_fn and safe_fn.lower().endswith(".pdf"):
            p = os.path.join(rfq_dir, safe_fn)
            f.save(p); saved.append(p)
    
    if not saved:
        flash("No PDFs found", "error"); return redirect("/")
    
    log.info("Upload: %d PDFs saved to %s", len(saved), rfq_id)
    
    # Check if this is a Price Check (AMS 704) instead of an RFQ
    if PRICE_CHECK_AVAILABLE and len(saved) == 1:
        if _is_price_check(saved[0]):
            return _handle_price_check_upload(saved[0], rfq_id)

    templates = identify_attachments(saved)
    if "704b" not in templates:
        flash("Could not identify 704B", "error"); return redirect("/")
    
    rfq = parse_rfq_attachments(templates)
    rfq["id"] = rfq_id
    rfq["source"] = "upload"

    # Filter out junk items (legal text, instructions, boilerplate)
    from src.forms.price_check import _filter_junk_items
    rfq["line_items"] = _filter_junk_items(rfq.get("line_items", []))

    # Auto SCPRS lookup
    rfq["line_items"] = bulk_lookup(rfq.get("line_items", []))

    # ── Dedup check: reject if same solicitation + agency + due_date exists ──
    sol = rfq.get("solicitation_number", "")
    agency = rfq.get("agency", "")
    due = rfq.get("due_date", "")
    if sol and agency:
        existing = load_rfqs()
        for eid, er in existing.items():
            if (er.get("solicitation_number") == sol
                    and er.get("agency") == agency
                    and er.get("due_date") == due):
                log.warning("Duplicate RFQ upload blocked: %s for %s (existing ID: %s)", sol, agency, eid)
                flash(f"Duplicate RFQ: {sol} for {agency} already exists (ID: {eid})", "error")
                return redirect(f"/rfq/{eid}")

    # Carry SCPRS/Amazon cost to supplier_cost so YOUR COST column displays it
    for _item in rfq.get("line_items", []):
        _sp = _item.get("scprs_last_price") or 0
        _ap = _item.get("amazon_price") or 0
        _best_cost = _sp or _ap
        if _best_cost and not _item.get("supplier_cost"):
            try:
                _item["supplier_cost"] = float(_best_cost)
                _item["cost_source"] = "SCPRS" if _sp else "Amazon"
            except (ValueError, TypeError):
                pass

    # Store lookup results summary
    items = rfq.get("line_items", [])
    priced_count = sum(1 for i in items if i.get("price_per_unit") or i.get("scprs_last_price"))
    rfq["auto_lookup_results"] = {
        "scprs_found": sum(1 for i in items if i.get("scprs_last_price")),
        "amazon_found": sum(1 for i in items if i.get("amazon_price")),
        "catalog_found": 0,
        "priced": priced_count,
        "total": len(items),
        "ran_at": datetime.now().isoformat(),
    }
    
    # Set status based on whether prices were actually found
    if priced_count > 0:
        _transition_status(rfq, "priced", actor="system", notes=f"Parsed + {priced_count}/{len(items)} items priced")
    else:
        _transition_status(rfq, "draft", actor="system", notes="Parsed from upload — no prices found yet")
    
    rfqs = load_rfqs()
    rfqs[rfq_id] = rfq
    save_rfqs(rfqs)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rfq_id, rfq.get("status", "draft"))
    except Exception:
        pass

    scprs_found = sum(1 for i in rfq["line_items"] if i.get("scprs_last_price"))
    msg = f"RFQ #{rfq['solicitation_number']} parsed — {len(rfq['line_items'])} items"
    if scprs_found:
        msg += f", {scprs_found} SCPRS prices found"
    flash(msg, "success")
    return redirect(f"/rfq/{rfq_id}")


def _is_price_check(pdf_path):
    """Detect if a PDF is an AMS 704 Price Check (NOT 704B quote worksheet).
    
    Uses filename first (fast, reliable), falls back to PDF content parsing.
    """
    basename = os.path.basename(pdf_path).lower()
    
    # ── Filename-based detection (fast path) ──
    # Exclude 704B / 703B / bid package by filename
    if any(x in basename for x in ["704b", "703b", "bid package", "bid_package", "quote worksheet"]):
        return False
    
    # Positive filename match: "AMS 704" or "ams704" in filename (but NOT 704B)
    if "704" in basename and "ams" in basename:
        return True
    # Also match "Quote - [Name] - [Date]" pattern (Valentina's format)
    # These always carry a single AMS 704 attachment
    if basename.startswith("quote") and basename.endswith(".pdf") and "704b" not in basename:
        # Only if filename looks like a price check attachment, not a generated quote
        if any(x in basename for x in ["ams", "704", "price"]):
            return True
    
    # ── PDF content fallback ──
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        text = reader.pages[0].extract_text() or ""
        text_lower = text.lower()
        
        # Exclude 704B forms
        if any(marker in text_lower for marker in ["704b", "quote worksheet", "acquisition quote"]):
            return False
        
        if "price check" in text_lower and ("ams 704" in text_lower or "worksheet" in text_lower):
            return True
        # Check form fields for AMS 704 patterns
        fields = reader.get_fields()
        if fields:
            field_names = set(fields.keys())
            ams704_markers = {"COMPANY NAME", "Requestor", "PRICE PER UNITRow1", "EXTENSIONRow1"}
            if len(ams704_markers & field_names) >= 3:
                return True
    except Exception as e:
        log.debug("PDF parse fallback failed for %s: %s", basename, e)
    return False


# ═══════════════════════════════════════════════════════════════════════
# Status Lifecycle — tracks every transition for PCs and RFQs
# ═══════════════════════════════════════════════════════════════════════

# PC lifecycle: parsed → priced → completed → won/lost/expired
# RFQ lifecycle: new → pending → ready → generated → sent → won/lost
PC_LIFECYCLE = ["parsed", "priced", "completed", "won", "lost", "expired"]
RFQ_LIFECYCLE = ["new", "pending", "ready", "generated", "sent", "won", "lost"]


def _transition_status(record, new_status, actor="system", notes=""):
    """Record a status transition with full history.

    Mutates record in place. Returns the record for chaining.
    """
    old_status = record.get("status", "")

    # Validate transition
    try:
        from src.core.quote_validator import validate_transition
        check = validate_transition(old_status, new_status)
        if not check["ok"]:
            log.warning("BLOCKED transition: %s -> %s (%s)",
                       old_status, new_status, check["error"])
            try:
                from flask import flash as _flash
                _flash(f"Unusual status change: {old_status} -> {new_status}", "warning")
            except Exception:
                pass
    except Exception:
        pass

    record["status"] = new_status
    now = datetime.now().isoformat()
    record["status_updated"] = now

    # Build status_history (create if missing for legacy records)
    history = record.get("status_history", [])
    entry = {"from": old_status, "to": new_status, "timestamp": now, "actor": actor}
    if notes:
        entry["notes"] = notes
    history.append(entry)
    record["status_history"] = history

    # Speed clock tracking
    try:
        from src.core.pricing_oracle_v2 import record_speed_event
        record_id = record.get("id") or record.get("pc_id") or ""
        record_type = "pc" if "pc_data" in record or "pc_number" in record else "quote"
        speed_map = {"parsed": "received", "draft": "received", "new": "received",
                     "priced": "priced", "ready": "priced", "generated": "generated",
                     "sent": "sent", "submitted": "sent"}
        event = speed_map.get(new_status)
        if event and record_id:
            record_speed_event(record_type, record_id, event)
    except Exception:
        pass

    # On win: confirm item mappings + lock costs
    if new_status in ("won", "awarded"):
        try:
            from src.core.pricing_oracle_v2 import confirm_item_mapping, lock_cost
            items = record.get("line_items", record.get("items", []))
            if isinstance(items, str):
                import json as _json
                items = _json.loads(items)
            for item in (items or []):
                desc = item.get("description", "")
                cost = item.get("supplier_cost") or item.get("unit_cost") or item.get("cost")
                sell = item.get("unit_price") or item.get("sell_price") or item.get("price")
                if desc and sell:
                    confirm_item_mapping(
                        original_description=desc, canonical_description=desc,
                        item_number=item.get("item_number", ""),
                        supplier=item.get("item_supplier", ""),
                        cost=float(str(cost or 0).replace("$", "").replace(",", "")) if cost else None,
                    )
                if desc and cost:
                    try:
                        lock_cost(desc, float(str(cost).replace("$", "").replace(",", "")),
                                  supplier=item.get("item_supplier", ""),
                                  source="won_quote", expires_days=60,
                                  item_number=item.get("item_number", ""))
                    except Exception:
                        pass
        except Exception:
            pass

    # Post-send pipeline: schedule follow-ups and tracking
    if new_status in ("sent", "submitted"):
        try:
            from src.agents.post_send_pipeline import on_quote_sent
            record_type = "pc" if "pc_data" in record or "pc_number" in record else "rfq"
            on_quote_sent(record_type,
                         record.get("id", record.get("pc_id", "")),
                         record)
        except Exception as _e:
            log.warning("Post-send pipeline: %s", _e)

    return record


def _handle_price_check_upload(pdf_path, pc_id, from_email=False):
    """Process an uploaded Price Check PDF.
    
    Full pipeline:
    1. Parse PDF → extract header + line items
    2. Dedup check
    3. Catalog matching → pull costs, MFG#, UOM from known products
    4. Save with status 'new' (ready for work in queue)
    5. Return/redirect to PC detail page
    
    Args:
        from_email: If True, returns dict instead of redirect (email pipeline call)
    """
    # Save to data dir for persistence
    pc_file = os.path.join(DATA_DIR, f"pc_upload_{os.path.basename(pdf_path)}")
    shutil.copy2(pdf_path, pc_file)

    # Parse
    parsed = parse_ams704(pc_file)
    parse_error = parsed.get("error")
    now = datetime.now().isoformat()
    source = "email_auto" if from_email else "manual_upload"
    
    if parse_error:
        if from_email:
            # Still create a minimal PC so the email isn't lost
            log.warning("PC parse failed for %s: %s — creating minimal PC with PDF attached",
                        os.path.basename(pdf_path), parse_error)
            pcs = _load_price_checks()
            pcs[pc_id] = {
                "id": pc_id,
                "pc_number": os.path.basename(pdf_path).replace(".pdf", "").replace("pc_upload_", "")[:40],
                "institution": "",
                "due_date": "",
                "requestor": "",
                "ship_to": "",
                "items": [],
                "source_pdf": pc_file,
                "status": "parse_error",
                "status_history": [{"from": "", "to": "parse_error", "timestamp": now, "actor": "system"}],
                "created_at": now,
                "source": source,
                "parsed": {"error": parse_error},
                "parse_error": parse_error,
                "reytech_quote_number": "",
                "linked_quote_number": "",
            }
            _save_price_checks(pcs)
            return {"ok": True, "pc_id": pc_id, "parse_error": parse_error, "items": 0}
        flash(f"Price Check parse error: {parse_error}", "error")
        return redirect("/")

    items = parsed.get("line_items", [])
    header = parsed.get("header", {})
    pc_num = header.get("price_check_number", "unknown")
    institution = header.get("institution", "")
    due_date = header.get("due_date", "")

    # ── DEDUP CHECK: same PC number + institution + due date = true duplicate ──
    pcs = _load_price_checks()
    for existing_id, existing_pc in pcs.items():
        if (existing_pc.get("pc_number", "").strip() == pc_num.strip()
                and existing_pc.get("institution", "").strip().lower() == institution.strip().lower()
                and existing_pc.get("due_date", "").strip() == due_date.strip()
                and pc_num != "unknown"):
            log.info("Dedup: PC #%s from %s (due %s) already exists as %s — skipping",
                     pc_num, institution, due_date, existing_id)
            if from_email:
                return {"dedup": True, "existing_id": existing_id}
            return redirect(f"/pricecheck/{existing_id}")

    # ── CATALOG MATCHING: enrich items with known costs, MFG#, UOM ──
    try:
        from src.agents.product_catalog import match_item as _cat_match, init_catalog_db as _cat_init
        _cat_init()
        for item in items:
            desc = (item.get("description") or "").strip()
            mfg = (item.get("mfg_number") or "").strip()
            if not desc and not mfg:
                continue
            matches = _cat_match(desc, mfg, top_n=1)
            if matches and matches[0].get("match_confidence", 0) >= 0.50:
                best = matches[0]
                # Initialize pricing dict
                pricing = item.get("pricing", {})
                if not pricing:
                    item["pricing"] = pricing
                # Pull catalog data into the item
                if best.get("cost") and not pricing.get("unit_cost"):
                    pricing["unit_cost"] = best["cost"]
                    pricing["price_source"] = "catalog"
                if best.get("sell_price") and not pricing.get("recommended_price"):
                    pricing["recommended_price"] = best["sell_price"]
                if best.get("mfg_number") and not item.get("mfg_number"):
                    item["mfg_number"] = best["mfg_number"]
                if best.get("uom"):
                    item["uom"] = best["uom"]
                if best.get("manufacturer"):
                    pricing["manufacturer"] = best["manufacturer"]
                pricing["catalog_match"] = best.get("name", "")[:50]
                pricing["catalog_id"] = best.get("id")
                pricing["catalog_confidence"] = best.get("match_confidence", 0)
                log.info("  catalog match for '%s': %s (%.0f%%) cost=$%.2f",
                         desc[:30], best.get("name", "")[:30],
                         best.get("match_confidence", 0) * 100,
                         best.get("cost", 0))
    except Exception as e:
        log.debug("Catalog matching on upload failed (non-fatal): %s", e)

    # ── Save PC Record ──
    pcs = _load_price_checks()
    pcs[pc_id] = {
        "id": pc_id,
        "pc_number": pc_num,
        "institution": institution,
        "due_date": due_date,
        "requestor": header.get("requestor", ""),
        "ship_to": parsed.get("ship_to", ""),
        "phone": header.get("phone", ""),
        "agency": institution,
        "items": items,
        "source_pdf": pc_file,
        "status": "new",
        "status_history": [
            {"from": "", "to": "parsed", "timestamp": now, "actor": "system", "notes": f"Parsed {len(items)} items"},
            {"from": "parsed", "to": "new", "timestamp": now, "actor": "system", "notes": f"Source: {source}"},
        ],
        "created_at": now,
        "source": source,
        "parsed": parsed,
        "reytech_quote_number": "",
        "linked_quote_number": "",
    }
    _save_price_checks(pcs)

    log.info("PC #%s created (%s) — %d items from %s, due %s, status=new",
             pc_num, source, len(items), institution, due_date)
    
    if from_email:
        return {"ok": True, "pc_id": pc_id, "pc_number": pc_num, "items": len(items)}
    
    flash(f"Price Check #{pc_num} — {len(items)} items from {institution}. Due {due_date}", "success")
    return redirect(f"/pricecheck/{pc_id}")


def _load_price_checks():
    """Delegate to dashboard's DB-primary implementation."""
    from src.api.dashboard import _load_price_checks as _db_load
    return _db_load()


def _save_price_checks(pcs):
    """Delegate to dashboard's DB-primary implementation."""
    from src.api.dashboard import _save_price_checks as _db_save
    _db_save(pcs)


@bp.route("/rfq/<rid>")
@auth_required
@safe_route
def detail(rid):
    _bad = _validate_rid(rid)
    if _bad: return _bad
    # WARNING: GET handler — must NEVER call save_rfqs() or modify data.
    # Data loss incident 2026-03-16: save_rfqs in GET handler corrupted items.
    # Check if this is actually a price check
    pcs = _load_price_checks()
    if rid in pcs:
        return redirect(f"/pricecheck/{rid}")
    rfqs = load_rfqs()
    _r_orig = rfqs.get(rid)
    if not _r_orig: flash("Not found", "error"); return redirect("/")

    # CRITICAL: deep copy for rendering — never mutate cached objects.
    # load_rfqs() has in-memory cache. Mutating r here (item mapping, intelligence
    # trimming) would persist in cache and corrupt data on next save.
    import copy as _copy
    r = _copy.deepcopy(_r_orig)

    # Ensure r is a plain dict (not a Jinja2-aware object)
    if not isinstance(r, dict):
        r = dict(r) if hasattr(r, 'items') else {}

    # ── Restore template paths from DB if files missing from disk (post-redeploy) ──
    tmpl = r.get("templates", {})
    db_files = list_rfq_files(rid, category="template")
    restored = False
    for db_f in db_files:
        ft = db_f.get("file_type", "").lower().replace("template_", "")
        fname = db_f.get("filename", "").lower()
        ttype = None
        if "703b" in ft or "703b" in fname:
            ttype = "703b"
        elif "704b" in ft or "704b" in fname:
            ttype = "704b"
        elif "bid" in ft or "bid" in fname:
            ttype = "bidpkg"
        if ttype and (ttype not in tmpl or not os.path.exists(tmpl.get(ttype, ""))):
            full_f = get_rfq_file(db_f["id"])
            if full_f and full_f.get("data"):
                restore_dir = os.path.join(DATA_DIR, "rfq_templates", rid)
                os.makedirs(restore_dir, exist_ok=True)
                restore_path = os.path.join(restore_dir, db_f["filename"])
                with open(restore_path, "wb") as _fw:
                    _fw.write(full_f["data"])
                tmpl[ttype] = restore_path
                restored = True
    if restored:
        r["templates"] = tmpl
        rfqs[rid] = r
        r["_needs_save"] = True  # Deferred to POST /rfq/{rid}/save-restore
    
    # ── Restore output_files from DB if empty (post-redeploy) ──
    if not r.get("output_files") and r.get("status") in ("generated", "sent", "won", "lost"):
        db_gen_files = list_rfq_files(rid, category="generated")
        if db_gen_files:
            r["output_files"] = [f["filename"] for f in db_gen_files]
            # Also restore files to disk for download
            for db_f in db_gen_files:
                fname = db_f.get("filename", "")
                sol = r.get("solicitation_number", rid)
                restore_dir = os.path.join(OUTPUT_DIR, sol)
                restore_path = os.path.join(restore_dir, fname)
                if not os.path.exists(restore_path):
                    full_f = get_rfq_file(db_f["id"])
                    if full_f and full_f.get("data"):
                        os.makedirs(restore_dir, exist_ok=True)
                        with open(restore_path, "wb") as _fw:
                            _fw.write(full_f["data"])
            rfqs[rid] = r
            r["_needs_save"] = True  # Deferred to POST /rfq/{rid}/save-restore

    # ── Enrichment DISABLED — was crashing page with Undefined serialization ──
    # TODO: re-enable after fixing stored Undefined values in rfqs.json
    pass  # enrichment disabled

    # Map items → line_items (SQLite column is "items", template expects "line_items")
    # Also handle: items might be a JSON string, a list, or missing
    if "line_items" not in r or not r["line_items"]:
        items_data = r.get("items", [])
        if isinstance(items_data, str):
            try:
                import json as _json
                items_data = _json.loads(items_data)
            except Exception:
                items_data = []
        if isinstance(items_data, list) and items_data:
            r["line_items"] = items_data

    if not isinstance(r.get("line_items"), list):
        r["line_items"] = []

    # Also map solicitation_number from rfq_number (SQLite vs JSON field names)
    if not r.get("solicitation_number") and r.get("rfq_number"):
        r["solicitation_number"] = r["rfq_number"]

    # Show link suggestion if unlinked (read-only — no save_rfqs here)
    if not r.get("linked_pc_id"):
        try:
            from src.core.pc_rfq_linker import find_matching_pc
            from src.api.dashboard import _load_price_checks as _dash_load_pcs
            pcs = _dash_load_pcs()
            pc_id, pc_data, reason = find_matching_pc(r, pcs)
            if pc_id:
                r["_suggested_pc"] = pc_id
                r["_suggested_pc_reason"] = reason
                pc_inner = pc_data.get("pc_data", pc_data)
                if isinstance(pc_inner, str):
                    try:
                        import json as _json
                        pc_inner = _json.loads(pc_inner)
                    except Exception:
                        pc_inner = {}
                r["_suggested_pc_number"] = pc_inner.get("pc_number", pc_data.get("pc_number", ""))
                r["_suggested_pc_items"] = len(pc_inner.get("items", pc_data.get("items", [])))
        except Exception:
            pass

    # Trim intelligence blobs to prevent page crash / slow render
    import json as _json_trim
    _items_list = r.get("line_items", r.get("items", []))
    try:
        _items_size = len(_json_trim.dumps(_items_list, default=str))
        if _items_size > 100000:  # >100KB of item JSON
            for _itm in _items_list:
                _itm.pop("intelligence", None)
                _itm.pop("oracle", None)
            log.warning("DETAIL %s: stripped intelligence blobs (%.0fKB)", rid, _items_size / 1024)
        else:
            for _itm in _items_list:
                _intel = _itm.get("intelligence", {})
                if isinstance(_intel, dict):
                    _cms = _intel.get("catalog_matches", [])
                    if len(_cms) > 3:
                        _intel["catalog_matches"] = _cms[:3]
                    for _cm in _intel.get("catalog_matches", []):
                        for _fld in ("description", "enriched_description"):
                            _v = _cm.get(_fld, "")
                            if len(_v) > 100:
                                _cm[_fld] = _v[:100] + "..."
    except Exception:
        pass

    log.info("RFQ detail render: rid=%s, line_items=%d", rid, len(_items_list))

    # Pass agency required_forms so checkboxes default correctly
    _agency_req = set()
    try:
        from src.core.agency_config import match_agency
        _ak, _ac = match_agency(r)
        _agency_req = set(_ac.get("required_forms", []))
    except Exception:
        pass

    return render_page("rfq_detail.html", active_page="Home", r=r, rid=rid,
                       agency_required_forms=_agency_req)


@bp.route("/rfq/<rid>/review-package")
@auth_required
@safe_route
def review_package(rid):
    """Package review screen — guided form-by-form verification before sending."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error")
        return redirect("/")

    from src.core.dal import get_latest_manifest, get_buyer_preferences, get_lifecycle_events
    manifest = get_latest_manifest(rid)
    if not manifest:
        # Try to create manifest from existing output_files
        output_files = r.get("output_files", [])
        if output_files:
            try:
                from src.core.dal import create_package_manifest
                from src.core.agency_config import match_agency
                _ak, _ac = match_agency(r)
                _gen_forms = []
                for _of in output_files:
                    _fid = "unknown"
                    _of_lower = _of.lower()
                    if "quote" in _of_lower and "704" not in _of_lower: _fid = "quote"
                    elif "703b" in _of_lower or "703c" in _of_lower: _fid = "703b"
                    elif "704b" in _of_lower: _fid = "704b"
                    elif "calrecycle" in _of_lower: _fid = "calrecycle74"
                    elif "bidderdecl" in _of_lower or "bidder" in _of_lower: _fid = "bidder_decl"
                    elif "dvbe" in _of_lower or "843" in _of_lower: _fid = "dvbe843"
                    elif "darfur" in _of_lower: _fid = "darfur_act"
                    elif "cuf" in _of_lower or "cv012" in _of_lower: _fid = "cv012_cuf"
                    elif "std204" in _of_lower or "payee" in _of_lower: _fid = "std204"
                    elif "std1000" in _of_lower: _fid = "std1000"
                    elif "seller" in _of_lower or "permit" in _of_lower: _fid = "sellers_permit"
                    _gen_forms.append({"form_id": _fid, "filename": _of})

                _mid = create_package_manifest(
                    rfq_id=rid, agency_key=_ak, agency_name=_ac.get("name", ""),
                    required_forms=_ac.get("required_forms", []),
                    generated_forms=_gen_forms,
                    quote_number=r.get("reytech_quote_number", ""),
                    quote_total=sum(float(i.get("price_per_unit",0))*int(float(i.get("qty",1))) for i in r.get("line_items", r.get("items", [])) if i.get("price_per_unit")),
                    item_count=len(r.get("line_items", r.get("items", []))),
                    created_by="recovery"
                )
                if _mid:
                    manifest = get_latest_manifest(rid)
                    log.info("Created recovery manifest %s for RFQ %s from %d output_files", _mid, rid, len(output_files))
            except Exception as _rm_e:
                log.error("Recovery manifest failed: %s", _rm_e)

    if not manifest:
        flash("No package generated yet — generate first, then review.", "error")
        return redirect(f"/rfq/{rid}")

    buyer_email = r.get("requestor_email", "")
    buyer_prefs = get_buyer_preferences(buyer_email) if buyer_email else []
    timeline = get_lifecycle_events("rfq", rid, limit=20)
    sol = r.get("solicitation_number", "") or r.get("rfq_number", "") or "unknown"

    # Get previous manifest for version diff
    prev_manifest = None
    if manifest and manifest.get("version", 1) > 1:
        try:
            from src.core.db import get_db
            with get_db() as conn:
                prev_row = conn.execute(
                    "SELECT id FROM package_manifest WHERE rfq_id = ? AND version = ?",
                    (rid, manifest["version"] - 1)).fetchone()
                if prev_row:
                    prev_manifest = get_package_manifest(prev_row[0])
        except Exception:
            pass

    return render_page("rfq_review.html",
        r=r, rid=rid, sol=sol,
        manifest=manifest,
        prev_manifest=prev_manifest,
        buyer_prefs=buyer_prefs,
        timeline=timeline,
        active_page="Home")


@bp.route("/rfq/<rid>/support")
@auth_required
@safe_route
def rfq_support_view(rid):
    """Support timeline — full RFQ lifecycle for customer support."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error"); return redirect("/")

    from src.core.dal import get_lifecycle_events, get_latest_manifest, get_buyer_preferences
    timeline = get_lifecycle_events("rfq", rid, limit=200)
    manifest = get_latest_manifest(rid)
    buyer_email = r.get("requestor_email", "")
    buyer_prefs = get_buyer_preferences(buyer_email) if buyer_email else []

    all_manifests = []
    deliveries = []
    emails = []
    try:
        from src.core.db import get_db
        with get_db() as conn:
            rows = conn.execute(
                "SELECT id, version, created_at, overall_status, total_forms, quote_number, quote_total, package_filename "
                "FROM package_manifest WHERE rfq_id = ? ORDER BY version DESC", (rid,)).fetchall()
            all_manifests = [dict(row) for row in rows]
            rows = conn.execute(
                "SELECT * FROM package_delivery WHERE rfq_id = ? ORDER BY delivered_at DESC", (rid,)).fetchall()
            deliveries = [dict(row) for row in rows]
    except Exception:
        pass
    try:
        from src.core.db import get_db
        with get_db() as conn:
            rows = conn.execute(
                "SELECT id, logged_at, direction, sender, recipient, subject, status "
                "FROM email_log WHERE rfq_id = ? ORDER BY logged_at DESC LIMIT 20", (rid,)).fetchall()
            emails = [dict(row) for row in rows]
    except Exception:
        pass

    sol = r.get("solicitation_number", "") or r.get("rfq_number", "") or "unknown"
    return render_page("rfq_support.html", r=r, rid=rid, sol=sol,
        timeline=timeline, manifest=manifest, all_manifests=all_manifests,
        deliveries=deliveries, emails=emails, buyer_prefs=buyer_prefs, active_page="Home")


@bp.route("/api/rfq/<rid>/add-buyer-pref", methods=["POST"])
@auth_required
def api_add_buyer_pref(rid):
    """Add a buyer preference from the support view."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    buyer_email = r.get("requestor_email", "")
    if not buyer_email:
        return jsonify({"ok": False, "error": "No buyer email on this RFQ"})
    data = request.get_json(force=True, silent=True) or {}
    key = data.get("preference_key", "")
    value = data.get("preference_value", "")
    notes = data.get("notes", "")
    if not key or not (value or notes):
        return jsonify({"ok": False, "error": "preference_key and notes required"})
    from src.core.dal import save_buyer_preference, log_lifecycle_event
    ok = save_buyer_preference(buyer_email, key, value or notes[:200],
        buyer_name=r.get("requestor_name", ""), agency_key=r.get("agency", ""),
        source=data.get("source", "manual"), notes=notes)
    if ok:
        log_lifecycle_event("rfq", rid, "buyer_preference_added",
            f"Preference added: {key} for {buyer_email}", actor="user",
            detail={"key": key, "notes": notes[:200]})
    return jsonify({"ok": ok})


@bp.route("/api/rfq/<rid>/lookup-tax-rate", methods=["POST"])
@auth_required
def api_lookup_tax_rate(rid):
    """Look up CA sales tax rate from delivery address via CDTFA API."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    data = request.get_json(force=True, silent=True) or {}
    address = data.get("address") or r.get("delivery_location") or r.get("ship_to") or ""
    if not address or len(address.strip()) < 5:
        return jsonify({"ok": False, "error": "No delivery address to look up"})
    try:
        from src.agents.tax_agent import get_tax_rate
        # Parse address directly — parse_ship_to expects multi-line format
        # "190 California Dr, Yountville, CA 94599" → street, city, zip
        import re as _re_tax
        _zip_match = _re_tax.search(r'\b(\d{5})\b', address)
        _city_match = _re_tax.search(r',\s*([A-Za-z\s]+),\s*[A-Z]{2}', address)
        _street_match = _re_tax.search(r'^(\d+\s+.+?)(?:,|$)', address)
        _d_street = _street_match.group(1).strip() if _street_match else ""
        _d_city = _city_match.group(1).strip() if _city_match else ""
        _d_zip = _zip_match.group(1) if _zip_match else ""
        log.info("Tax lookup: raw='%s' → street='%s' city='%s' zip='%s'",
                 address[:60], _d_street, _d_city, _d_zip)
        if _d_street and _d_city and _d_zip:
            result = get_tax_rate(street=_d_street, city=_d_city, zip_code=_d_zip)
        else:
            from src.agents.tax_agent import parse_ship_to
            _parts = [p.strip() for p in address.split(",")]
            parsed = parse_ship_to("", _parts)
            result = get_tax_rate(
                street=parsed.get("street", ""),
                city=parsed.get("city", ""),
                zip_code=parsed.get("zip", "")
            )
        if result and result.get("rate"):
            rate_pct = round(result["rate"] * 100, 3)
            r["tax_rate"] = rate_pct
            r["tax_validated"] = True
            r["tax_source"] = result.get("source", "cdtfa_api")
            r["tax_jurisdiction"] = result.get("jurisdiction", "")
            save_rfqs(rfqs)
            return jsonify({"ok": True, "rate": rate_pct,
                "jurisdiction": result.get("jurisdiction", ""),
                "city": result.get("city", ""),
                "county": result.get("county", ""),
                "confidence": result.get("confidence", ""),
                "source": result.get("source", "")})
        else:
            return jsonify({"ok": False, "error": result.get("error", "Lookup failed")})
    except Exception as e:
        log.error("Tax rate lookup for RFQ %s: %s", rid, e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/rfq/<rid>/resend-package", methods=["POST"])
@auth_required
def api_resend_package(rid):
    """Resend the latest approved package to a recipient."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    data = request.get_json(force=True, silent=True) or {}
    to_email = data.get("to", "")
    subject = data.get("subject", "")
    body_text = data.get("body", "")
    if not to_email:
        return jsonify({"ok": False, "error": "Recipient email required"})
    sol = r.get("solicitation_number", "") or "unknown"

    from src.core.dal import get_latest_manifest, log_lifecycle_event, record_package_delivery
    manifest = get_latest_manifest(rid)
    if not manifest:
        return jsonify({"ok": False, "error": "No package generated"})
    pkg_filename = manifest.get("package_filename") or f"RFQ_Package_{sol}_ReytechInc.pdf"

    # Find package data (disk or DB)
    pkg_data = None
    pkg_path = os.path.join(OUTPUT_DIR, sol, pkg_filename)
    if os.path.exists(pkg_path):
        with open(pkg_path, "rb") as _f:
            pkg_data = _f.read()
    else:
        try:
            files = list_rfq_files(rid, category="generated")
            for dbf in files:
                if "Package" in (dbf.get("filename") or "") or dbf.get("filename") == pkg_filename:
                    full = get_rfq_file(dbf["id"])
                    if full and full.get("data"):
                        pkg_data = full["data"]; break
        except Exception:
            pass
    if not pkg_data:
        return jsonify({"ok": False, "error": f"Package file not found: {pkg_filename}"})

    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.application import MIMEApplication
        smtp_user = os.environ.get("GMAIL_ADDRESS", "")
        smtp_pass = os.environ.get("GMAIL_PASSWORD", "")
        if not smtp_user or not smtp_pass:
            return jsonify({"ok": False, "error": "Email not configured"})
        msg = MIMEMultipart()
        msg["From"] = smtp_user
        msg["To"] = to_email
        msg["Subject"] = subject or f"Reytech Inc. — RFQ Response #{sol}"
        msg.attach(MIMEText(body_text or "Please find attached our RFQ response package.", "plain"))
        att = MIMEApplication(pkg_data, _subtype="pdf")
        att.add_header("Content-Disposition", "attachment", filename=pkg_filename)
        msg.attach(att)
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, [to_email], msg.as_string())
        server.quit()

        log_lifecycle_event("rfq", rid, "package_sent",
            f"Resent to {to_email} (support view)", actor="user",
            detail={"recipient": to_email, "subject": subject, "resend": True})
        if manifest.get("id"):
            import hashlib
            record_package_delivery(manifest["id"], rid, to_email,
                recipient_name=r.get("requestor_name", ""), email_subject=subject,
                package_hash=hashlib.sha256(pkg_data).hexdigest())
        return jsonify({"ok": True, "sent_to": to_email, "size": len(pkg_data)})
    except Exception as e:
        log.error("Resend RFQ %s: %s", rid, e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/support/search-packages")
@auth_required
def api_search_packages():
    """Search package manifests by form_id, agency, status, solicitation, or buyer."""
    form_id = request.args.get("form_id", "")
    agency = request.args.get("agency", "")
    status = request.args.get("status", "")
    sol = request.args.get("sol", "")
    buyer = request.args.get("buyer", "")
    limit = min(int(request.args.get("limit", 50)), 200)
    try:
        from src.core.db import get_db
        import json as _jsearch
        with get_db() as conn:
            q = "SELECT pm.* FROM package_manifest pm WHERE 1=1"
            p = []
            if form_id:
                q += " AND pm.generated_forms LIKE ?"; p.append(f"%{form_id}%")
            if agency:
                q += " AND pm.agency_key LIKE ?"; p.append(f"%{agency}%")
            if status:
                q += " AND pm.overall_status = ?"; p.append(status)
            if sol:
                q += " AND pm.rfq_id LIKE ?"; p.append(f"%{sol}%")
            q += " ORDER BY pm.created_at DESC LIMIT ?"; p.append(limit)
            rows = conn.execute(q, p).fetchall()
            results = []
            for row in rows:
                d = dict(row)
                for f in ("generated_forms", "missing_forms", "required_forms"):
                    if d.get(f):
                        try: d[f] = _jsearch.loads(d[f])
                        except Exception: pass
                results.append(d)
            return jsonify({"ok": True, "results": results, "count": len(results)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/rfq/<rid>/save-restore", methods=["POST"])
@auth_required
@safe_route
def rfq_save_restore(rid):
    """Save template/file restorations. Called via POST from rfq_detail.html, not GET."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False})
    if r.pop("_needs_save", False):
        save_rfqs(rfqs)
        return jsonify({"ok": True, "saved": True})
    return jsonify({"ok": True, "saved": False})


@bp.route("/rfq/<rid>/update", methods=["POST"])
@auth_required
@safe_route
def update(rid):
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r: return redirect("/")
    
    from src.core.validation import validate_price, validate_cost, validate_markup, validate_qty, validate_text, validate_short_text, validate_url
    for i, item in enumerate(r["line_items"]):
        for field, key, vfn in [("cost", "supplier_cost", validate_cost), ("scprs", "scprs_last_price", validate_price), ("price", "price_per_unit", validate_price), ("markup", "markup_pct", validate_markup)]:
            v = request.form.get(f"{field}_{i}")
            if v:
                val, err = vfn(v)
                if err: log.warning("RFQ update item[%d] %s: %s", i, key, err)
                item[key] = val
        # Save qty and uom from separate inputs
        qty_val = request.form.get(f"qty_{i}")
        if qty_val:
            val, err = validate_qty(qty_val)
            if err: log.warning("RFQ update item[%d] qty: %s", i, err)
            item["qty"] = val
        uom_val = request.form.get(f"uom_{i}")
        if uom_val is not None:
            val, _ = validate_short_text(uom_val, max_len=20, default="EA")
            item["uom"] = val.upper()
        # Save edited description
        desc_val = request.form.get(f"desc_{i}")
        if desc_val is not None:
            val, _ = validate_text(desc_val, max_len=5000)
            item["description"] = val
        # Save part number
        part_val = request.form.get(f"part_{i}")
        if part_val is not None:
            val, _ = validate_short_text(part_val, max_len=100)
            item["item_number"] = val
        # Save item link and auto-detect supplier
        link_raw = request.form.get(f"link_{i}", "")
        link_val, _ = validate_url(link_raw)
        item["item_link"] = link_val
        if link_val:
            try:
                from src.agents.item_link_lookup import detect_supplier
                item["item_supplier"] = detect_supplier(link_val)
            except Exception as _e:
                log.debug("Suppressed: %s", _e)
    
    # Save quote-level notes
    quote_notes_val, _ = validate_text(request.form.get("quote_notes", ""), max_len=2000)
    r["quote_notes"] = quote_notes_val

    _transition_status(r, "ready", actor="user", notes="Pricing updated")
    save_rfqs(rfqs)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, "ready")
    except Exception:
        pass

    # Save SCPRS prices for future lookups
    save_prices_from_rfq(r)
    
    # Record ALL prices to history + auto-ingest to catalog
    try:
        _record_rfq_prices(r, source="rfq_finalize")
    except Exception as _e:
        log.debug("Price recording: %s", _e)
    
    # Sync all priced items to product catalog
    cat_added, cat_updated = 0, 0
    try:
        from src.agents.product_catalog import match_item, add_to_catalog, add_supplier_price, init_catalog_db
        init_catalog_db()
        for item in r.get("line_items", []):
            desc = item.get("description", "")
            pn = item.get("item_number", "") or ""
            cost = item.get("supplier_cost") or 0
            bid = item.get("price_per_unit") or 0
            supplier = item.get("item_supplier", "")
            uom = item.get("uom", "EA")
            url = item.get("item_link", "")
            if not desc or (not cost and not bid):
                continue
            cat_matches = match_item(desc, pn, top_n=1)
            if cat_matches and cat_matches[0].get("match_confidence", 0) >= 0.5:
                pid = cat_matches[0]["id"]
                if cost > 0 and supplier:
                    add_supplier_price(pid, supplier, cost, url=url)
                if url:
                    try:
                        from src.agents.product_catalog import _get_conn
                        conn = _get_conn()
                        conn.execute(
                            "UPDATE product_catalog SET photo_url=COALESCE(NULLIF(photo_url,''),?) WHERE id=?",
                            (url, pid))
                        conn.commit(); conn.close()
                    except Exception:
                        pass
                cat_updated += 1
            else:
                pid = add_to_catalog(
                    description=desc, part_number=pn,
                    cost=cost if cost > 0 else 0,
                    sell_price=bid if bid > 0 else 0,
                    supplier_name=supplier, uom=uom,
                    supplier_url=url,
                    source=f"rfq_finalize_{r.get('solicitation_number', '')}",
                )
                if pid and cost > 0 and supplier:
                    add_supplier_price(pid, supplier, cost, url=url)
                    cat_added += 1
        if cat_added or cat_updated:
            log.info("Finalize catalog sync: +%d new, ~%d updated", cat_added, cat_updated)
    except Exception as _ce:
        log.debug("Finalize catalog sync: %s", _ce)
    
    # Auto-learn item mappings + lock costs from user pricing
    try:
        from src.core.pricing_oracle_v2 import auto_learn_mapping, lock_cost
        for _item in r.get("line_items", []):
            _desc = _item.get("description", "")
            _cost = _item.get("supplier_cost") or _item.get("unit_price")
            if _desc and _cost:
                try:
                    _cv = float(str(_cost).replace("$", "").replace(",", ""))
                except (ValueError, TypeError):
                    _cv = 0
                if _cv > 0:
                    auto_learn_mapping(_desc, _item.get("catalog_match", {}).get("name", _desc),
                                       item_number=_item.get("item_number", ""), confidence=0.7)
                    lock_cost(_desc, _cv, supplier=_item.get("item_supplier", ""),
                              source="user_pricing", expires_days=30,
                              item_number=_item.get("item_number", ""))
    except Exception:
        pass

    _log_rfq_activity(rid, "pricing_finalized",
        f"Pricing finalized for #{r.get('solicitation_number','?')} ({len(r.get('line_items',[]))} items, catalog +{cat_added}/~{cat_updated})",
        actor="user")
    
    flash("Pricing finalized — saved to catalog", "success")
    return redirect(f"/rfq/{rid}")


@bp.route("/api/rfq/<rid>/update-field", methods=["POST"])
@auth_required
def rfq_update_field(rid):
    """Update individual header fields (solicitation, requestor, due date, etc.)."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "Not found"})
    data = request.get_json(force=True, silent=True) or {}
    changed = []
    allowed = ["solicitation_number", "requestor_name", "requestor_email",
               "due_date", "ship_to", "delivery_location", "institution",
               "agency_name", "notes"]
    from src.core.validation import validate_header_field
    for field in allowed:
        if field in data:
            old = r.get(field, "")
            val, err = validate_header_field(field, data[field])
            if err:
                log.warning("RFQ %s update-field %s: %s", rid, field, err)
            r[field] = val
            if old != data[field]:
                changed.append(f"{field}: '{old}' -> '{data[field]}'")
                # Log parse gap when user fills an empty field
                if not old and data[field]:
                    try:
                        from src.core.db import get_db
                        with get_db() as _conn:
                            _conn.execute("""
                                INSERT INTO parse_gaps
                                (rfq_id, field_name, user_filled_value,
                                 source_type, email_subject, requestor_email, agency)
                                VALUES (?,?,?,?,?,?,?)
                            """, (rid, field, data[field],
                                  r.get("source", ""),
                                  r.get("email_subject", ""),
                                  r.get("requestor_email", ""),
                                  r.get("agency", "")))
                    except Exception:
                        pass
    if changed:
        save_rfqs(rfqs)
        _log_rfq_activity(rid, "field_updated",
            "; ".join(changed), actor="user")

    # Re-attempt PC linking on any field update if not already linked
    link_result = None
    if not r.get("linked_pc_id") and any(f in data for f in ["solicitation_number", "requestor_email", "requestor_name"]):
        try:
            from src.api.dashboard import _link_rfq_to_pc
            _link_trace = []
            if _link_rfq_to_pc(r, _link_trace):
                save_rfqs(rfqs)
                link_result = {"linked": True, "trace": _link_trace,
                               "pc_id": r.get("linked_pc_id", ""),
                               "pc_number": r.get("linked_pc_number", "")}
                log.info("Re-linked RFQ %s: %s", rid, _link_trace)
            else:
                link_result = {"linked": False, "trace": _link_trace}
        except Exception as _le:
            link_result = {"linked": False, "error": str(_le)}
            log.warning("Re-link: %s", _le)

    # Smart validation against buyer history
    suggestions = {}
    if "delivery_location" in data or "ship_to" in data or "institution" in data:
        try:
            from src.core.db import get_db
            with get_db() as _vconn:
                email = r.get("requestor_email", "")
                if email:
                    history = _vconn.execute("""
                        SELECT ship_to_address, dept_name, COUNT(*) as cnt
                        FROM scprs_po_master
                        WHERE buyer_email = ?
                        AND ship_to_address != ''
                        GROUP BY ship_to_address
                        ORDER BY cnt DESC LIMIT 5
                    """, (email,)).fetchall()

                    if history:
                        new_val = data.get("delivery_location", data.get("ship_to", data.get("institution", "")))
                        top_location = history[0][0]
                        top_count = history[0][2]
                        total = sum(h[2] for h in history)

                        from difflib import SequenceMatcher
                        best_match = max(
                            [(h[0], h[2], SequenceMatcher(None, new_val.lower(), h[0].lower()).ratio())
                             for h in history],
                            key=lambda x: x[2]
                        )

                        suggestions["buyer_history"] = {
                            "most_common": top_location,
                            "frequency": f"{top_count}/{total} POs",
                            "confidence": round(top_count / total * 100),
                            "all_locations": [
                                {"location": h[0], "department": h[1], "count": h[2]}
                                for h in history
                            ],
                        }

                        if best_match[2] < 0.5 and top_count >= 3:
                            suggestions["warning"] = (
                                f"This buyer usually ships to '{top_location}' "
                                f"({top_count} of {total} POs). "
                                f"You entered '{new_val}'. Confirm?"
                            )
                            suggestions["needs_confirm"] = True
        except Exception:
            pass

    return jsonify({"ok": True, "updated": changed, "suggestions": suggestions,
                    "link_result": link_result})


@bp.route("/api/rfq/<rid>/bulk-scrape-urls", methods=["POST"])
@auth_required
def api_rfq_bulk_scrape_urls(rid):
    """Bulk paste URLs → scrape each → apply cost + supplier to items by index."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404
    data = request.get_json(force=True, silent=True) or {}
    urls = data.get("urls", [])
    if not urls:
        return jsonify({"ok": False, "error": "No URLs provided"})
    items = r.get("line_items", [])
    results = []
    applied = 0
    for i, url in enumerate(urls):
        url = (url or "").strip()
        if not url:
            results.append({"line": i + 1, "url": "", "status": "skipped"})
            continue
        if i >= len(items):
            results.append({"line": i + 1, "url": url[:60], "status": "skipped"})
            continue
        try:
            from src.agents.item_link_lookup import lookup_from_url, detect_supplier
            res = lookup_from_url(url)
            price = res.get("price") or res.get("list_price") or res.get("cost")
            if price and float(price) > 0:
                price = float(price)
                item = items[i]
                item["item_link"] = url
                item["item_supplier"] = detect_supplier(url)
                item["supplier_cost"] = price
                item["cost_source"] = "Supplier URL"
                item["cost_supplier_name"] = item.get("item_supplier", "")
                markup = item.get("markup_pct") or r.get("default_markup") or 25
                try:
                    markup = float(markup)
                except (ValueError, TypeError):
                    markup = 25
                item["markup_pct"] = markup
                item["price_per_unit"] = round(price * (1 + markup / 100), 2)
                _pn = res.get("mfg_number") or res.get("part_number") or ""
                if _pn:
                    item["item_number"] = _pn
                _desc = res.get("title") or res.get("description") or ""
                if _desc and (not item.get("description") or len(item.get("description", "")) < 10):
                    item["description"] = _desc
                results.append({"line": i + 1, "url": url[:60], "status": "ok",
                               "price": price, "supplier": item["item_supplier"]})
                applied += 1
            else:
                results.append({"line": i + 1, "url": url[:60], "status": "no_price"})
        except Exception as e:
            log.error("Bulk scrape URL error line %d: %s", i + 1, e, exc_info=True)
            results.append({"line": i + 1, "url": url[:60], "status": "error", "error": str(e)[:80]})
    if applied > 0:
        save_rfqs(rfqs)
    return jsonify({"ok": True, "results": results, "applied": applied, "total": len(urls)})


@bp.route("/api/rfq/<rid>/bulk-paste-data", methods=["POST"])
@auth_required
def api_rfq_bulk_paste_data(rid):
    """Bulk paste multi-column data (description, MFG#, URL, cost, markup) into line items."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404
    data = request.get_json(force=True, silent=True) or {}
    rows = data.get("rows", [])
    if not rows:
        return jsonify({"ok": False, "error": "No data provided"})
    items = r.get("line_items", [])
    results = []
    applied = 0
    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            results.append({"line": i + 1, "status": "skipped"})
            continue
        if i >= len(items):
            results.append({"line": i + 1, "status": "skipped"})
            continue
        # Check if row has any non-empty values
        has_data = any((row.get(k) or "").strip() for k in
                       ("description", "item_number", "item_link", "supplier_cost", "markup_pct"))
        if not has_data:
            results.append({"line": i + 1, "status": "skipped"})
            continue
        try:
            item = items[i]
            fields_set = 0
            # Description
            desc = (row.get("description") or "").strip()
            if desc:
                item["description"] = desc
                fields_set += 1
            # MFG# / Item Number
            mfg = (row.get("item_number") or "").strip()
            if mfg:
                item["item_number"] = mfg
                fields_set += 1
            # Item Link / URL
            link = (row.get("item_link") or "").strip()
            if link:
                if not link.startswith("http") and ("." in link):
                    link = "https://" + link
                item["item_link"] = link
                try:
                    from src.agents.item_link_lookup import detect_supplier
                    item["item_supplier"] = detect_supplier(link)
                except Exception:
                    pass
                fields_set += 1
            # Cost
            cost_str = (row.get("supplier_cost") or "").strip().replace("$", "").replace(",", "")
            if cost_str:
                try:
                    cost = float(cost_str)
                    if cost > 0:
                        item["supplier_cost"] = cost
                        item["cost_source"] = "Bulk Paste"
                        fields_set += 1
                        # Recalculate bid price with markup
                        markup_str = (row.get("markup_pct") or "").strip().replace("%", "")
                        if markup_str:
                            try:
                                markup = float(markup_str)
                                item["markup_pct"] = markup
                            except (ValueError, TypeError):
                                markup = item.get("markup_pct") or r.get("default_markup") or 25
                        else:
                            markup = item.get("markup_pct") or r.get("default_markup") or 25
                        try:
                            markup = float(markup)
                        except (ValueError, TypeError):
                            markup = 25
                        item["markup_pct"] = markup
                        item["price_per_unit"] = round(cost * (1 + markup / 100), 2)
                except (ValueError, TypeError):
                    pass
            elif (row.get("markup_pct") or "").strip():
                # Markup without cost — update markup only if item already has cost
                markup_str = (row.get("markup_pct") or "").strip().replace("%", "")
                try:
                    markup = float(markup_str)
                    item["markup_pct"] = markup
                    if item.get("supplier_cost") and float(item["supplier_cost"]) > 0:
                        item["price_per_unit"] = round(float(item["supplier_cost"]) * (1 + markup / 100), 2)
                    fields_set += 1
                except (ValueError, TypeError):
                    pass
            if fields_set > 0:
                res_obj = {"line": i + 1, "status": "ok", "fields": fields_set}
                if item.get("supplier_cost"):
                    res_obj["price"] = item["supplier_cost"]
                if item.get("item_supplier"):
                    res_obj["supplier"] = item["item_supplier"]
                results.append(res_obj)
                applied += 1
            else:
                results.append({"line": i + 1, "status": "skipped"})
        except Exception as e:
            log.error("Bulk paste data error line %d: %s", i + 1, e, exc_info=True)
            results.append({"line": i + 1, "status": "error", "error": str(e)[:80]})
    if applied > 0:
        save_rfqs(rfqs)
    return jsonify({"ok": True, "results": results, "applied": applied, "total": len(rows)})


@bp.route("/api/rfq/<rid>/autosave", methods=["POST"])
@auth_required
def api_rfq_autosave(rid):
    """AJAX auto-save: persist line item edits without page reload."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "not found"}), 404

    data = request.get_json(force=True, silent=True) or {}
    items_data = data.get("items", [])

    from src.core.validation import validate_rfq_item
    for update in items_data:
        idx = update.get("idx")
        if idx is None or idx >= len(r["line_items"]):
            continue
        item = r["line_items"][idx]
        errs = validate_rfq_item(update, item)
        if errs:
            log.warning("RFQ autosave %s item[%s]: %s", rid, idx, "; ".join(errs))
        # Auto-detect supplier from link
        if "item_link" in update and item.get("item_link"):
            try:
                from src.agents.item_link_lookup import detect_supplier
                item["item_supplier"] = detect_supplier(item["item_link"])
            except Exception:
                pass

    # Save package form checklist if provided
    pkg_forms = data.get("package_forms")
    if pkg_forms is not None and isinstance(pkg_forms, dict):
        r["package_forms"] = pkg_forms

    # Save tax rate if provided
    tax_rate = data.get("tax_rate")
    if tax_rate is not None:
        from src.core.validation import validate_header_field
        v, _ = validate_header_field("tax_rate", tax_rate)
        r["tax_rate"] = v

    # Save shipping if provided
    if data.get("shipping_option") is not None:
        r["shipping_option"] = str(data["shipping_option"])[:20]
    if data.get("shipping_amount") is not None:
        try:
            r["shipping_amount"] = max(0, min(99999, float(data["shipping_amount"])))
        except (ValueError, TypeError):
            pass

    # Save delivery location if provided (belt-and-suspenders with saveField)
    if data.get("delivery_location"):
        r["delivery_location"] = str(data["delivery_location"])[:500]

    # Save quote notes if provided
    if "quote_notes" in data:
        from src.core.validation import validate_text
        _qn_val, _ = validate_text(data["quote_notes"], max_len=2000)
        r["quote_notes"] = _qn_val

    save_rfqs(rfqs)

    try:
        from src.core.dal import log_lifecycle_event
        _has_markup = any(u.get("markup_pct") for u in items_data)
        log_lifecycle_event("rfq", rid, "items_edited",
            f"Autosaved {len(r.get('line_items', []))} items" + (" (markup changed)" if _has_markup else ""),
            actor="user")
    except Exception:
        pass

    # F11: Check guardrails on saved data
    guardrail_warnings = _check_guardrails(r.get("line_items", []))

    # F6: Record price audits for changed items
    try:
        from src.core.db import record_audit
        sol = r.get("solicitation_number", "")
        for update in items_data:
            idx = update.get("idx")
            if idx is None or idx >= len(r["line_items"]):
                continue
            item = r["line_items"][idx]
            desc = (item.get("description", "") or "")[:60]
            # Record cost changes
            if "supplier_cost" in update and update["supplier_cost"]:
                record_audit(
                    item_description=desc, field_changed="supplier_cost",
                    old_value=0, new_value=float(update["supplier_cost"]),
                    source="manual_edit", rfq_id=rid, actor="user"
                )
            # Record bid changes
            if "price_per_unit" in update and update["price_per_unit"]:
                record_audit(
                    item_description=desc, field_changed="price_per_unit",
                    old_value=0, new_value=float(update["price_per_unit"]),
                    source="manual_edit", rfq_id=rid, actor="user"
                )
    except Exception:
        pass

    # Write priced items to catalog (same as full save does)
    try:
        sol = r.get("solicitation_number", "")
        for update in items_data:
            idx = update.get("idx")
            if idx is None or idx >= len(r["line_items"]):
                continue
            item = r["line_items"][idx]
            cost = item.get("supplier_cost") or 0
            bid = item.get("price_per_unit") or 0
            desc = item.get("description", "")
            if desc and (cost > 0 or bid > 0):
                from src.agents.product_catalog import add_to_catalog, init_catalog_db
                init_catalog_db()
                add_to_catalog(
                    description=desc,
                    part_number=item.get("item_number", ""),
                    cost=float(cost) if cost else 0,
                    sell_price=float(bid) if bid else 0,
                    source=f"rfq_autosave_{sol}",
                    supplier_name=item.get("item_supplier", ""),
                    supplier_url=item.get("item_link", ""),
                )
    except Exception:
        pass

    return jsonify({
        "ok": True, "saved": len(items_data),
        "guardrails": guardrail_warnings if guardrail_warnings else None,
    })


@bp.route("/rfq/<rid>/add-item", methods=["POST"])
@auth_required
@safe_route
def rfq_add_item(rid):
    """Add a line item to an RFQ (for generic/Cal Vet RFQs or manual entry)."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error"); return redirect("/")

    if "line_items" not in r:
        r["line_items"] = []

    next_num = len(r["line_items"]) + 1

    from src.core.validation import validate_qty, validate_short_text, validate_text
    _qty, _ = validate_qty(request.form.get("qty", 1))
    _uom, _ = validate_short_text(request.form.get("uom", "EA"), max_len=20, default="EA")
    _desc, _ = validate_text(request.form.get("description", ""), max_len=5000)
    _itemno, _ = validate_short_text(request.form.get("item_number", ""), max_len=100)
    new_item = {
        "line_number": next_num,
        "qty": _qty,
        "uom": _uom.upper(),
        "description": _desc.strip(),
        "item_number": _itemno,
        "supplier_cost": 0,
        "scprs_last_price": None,
        "source_type": "manual",
        "price_per_unit": 0,
    }

    r["line_items"].append(new_item)
    save_rfqs(rfqs)
    _log_rfq_activity(rid, "item_added",
        f"Line item #{next_num} added: {new_item['description'][:60]}",
        actor="user")
    flash(f"Item #{next_num} added", "success")
    return redirect(f"/rfq/{rid}#add-item-section")


@bp.route("/rfq/<rid>/remove-item/<int:idx>", methods=["POST"])
@auth_required
@safe_route
def rfq_remove_item(rid, idx):
    """Remove a line item from an RFQ by index."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return _item_response(rid, False, "RFQ not found")

    items = r.get("line_items", [])
    if 0 <= idx < len(items):
        removed = items.pop(idx)
        _renumber_items(items)
        save_rfqs(rfqs)
        _log_rfq_activity(rid, "item_removed",
            f"Line item removed: {removed.get('description','')[:60]}",
            actor="user")
        return _item_response(rid, True, "Item removed")
    return _item_response(rid, False, "Invalid item index")


@bp.route("/rfq/<rid>/duplicate-item/<int:idx>", methods=["POST"])
@auth_required
@safe_route
def rfq_duplicate_item(rid, idx):
    """Duplicate a line item (insert copy right after the original)."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return _item_response(rid, False, "RFQ not found")

    items = r.get("line_items", [])
    if 0 <= idx < len(items):
        import copy
        dupe = copy.deepcopy(items[idx])
        dupe.pop("_catalog_product_id", None)
        items.insert(idx + 1, dupe)
        _renumber_items(items)
        save_rfqs(rfqs)
        return _item_response(rid, True, f"Item duplicated at #{idx + 2}")
    return _item_response(rid, False, "Invalid item index")


@bp.route("/rfq/<rid>/move-item/<int:idx>/<direction>", methods=["POST"])
@auth_required
@safe_route
def rfq_move_item(rid, idx, direction):
    """Move a line item up or down."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return _item_response(rid, False, "RFQ not found")

    items = r.get("line_items", [])
    if direction == "up" and idx > 0:
        items[idx], items[idx - 1] = items[idx - 1], items[idx]
    elif direction == "down" and idx < len(items) - 1:
        items[idx], items[idx + 1] = items[idx + 1], items[idx]
    else:
        return _item_response(rid, False, "Cannot move")

    _renumber_items(items)
    save_rfqs(rfqs)
    return _item_response(rid, True, f"Item moved {direction}")


@bp.route("/rfq/<rid>/reset-items", methods=["POST"])
@auth_required
@safe_route
def rfq_reset_items(rid):
    """Clear all line items from an RFQ so it can be re-imported."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return _item_response(rid, False, "RFQ not found")

    old_count = len(r.get("line_items", []))
    r["line_items"] = []
    r.pop("linked_pc_id", None)
    r.pop("linked_pc_number", None)
    r.pop("linked_pc_match_reason", None)
    r.pop("uploaded_pc_pdf", None)
    save_rfqs(rfqs)
    _log_rfq_activity(rid, "items_reset",
        f"All {old_count} line items cleared for re-import",
        actor="user")
    return _item_response(rid, True, f"Cleared {old_count} items")


@bp.route("/rfq/<rid>/lookup-item/<int:idx>", methods=["POST"])
@auth_required
def rfq_lookup_single_item(rid, idx):
    """Run SCPRS + Catalog + Amazon lookup on a single line item by index."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    items = r.get("line_items", [])
    if idx < 0 or idx >= len(items):
        return jsonify({"ok": False, "error": "Invalid item index"})

    item = items[idx]
    desc = item.get("description", "")
    pn = item.get("item_number", "")
    results = {"scprs": None, "catalog": None, "amazon": None}

    # 1. SCPRS lookup
    try:
        from src.agents.scprs_lookup import bulk_lookup
        wrapped = bulk_lookup([item])
        if wrapped and wrapped[0].get("scprs_last_price"):
            items[idx] = wrapped[0]
            results["scprs"] = {
                "price": wrapped[0]["scprs_last_price"],
                "vendor": wrapped[0].get("scprs_vendor", ""),
                "source": wrapped[0].get("scprs_source", ""),
            }
    except Exception as e:
        results["scprs"] = {"error": str(e)[:80]}
        log.debug("Single item SCPRS error: %s", e)

    # 2. Catalog lookup
    try:
        from src.agents.product_catalog import match_item, init_catalog_db
        init_catalog_db()
        matches = match_item(desc, pn, top_n=3)
        if matches and matches[0].get("match_confidence", 0) >= 0.3:
            best = matches[0]
            items[idx]["catalog_match"] = best
            cost = best.get("best_cost") or best.get("cost") or 0
            results["catalog"] = {
                "name": best.get("name", "")[:60],
                "cost": cost,
                "supplier": best.get("best_supplier", ""),
                "confidence": round(best.get("match_confidence", 0), 2),
                "sell_price": best.get("sell_price") or best.get("recommended_price") or 0,
            }
            # Fill cost if empty
            if cost and not items[idx].get("supplier_cost"):
                items[idx]["supplier_cost"] = round(float(cost), 2)
                items[idx]["item_supplier"] = best.get("best_supplier", "")
    except Exception as e:
        results["catalog"] = {"error": str(e)[:80]}
        log.debug("Single item catalog error: %s", e)

    # 3. Amazon/web lookup
    try:
        from src.agents.web_price_research import research_items
        wrapped = research_items([items[idx]])
        if wrapped and wrapped[0].get("amazon_price"):
            items[idx] = wrapped[0]
            results["amazon"] = {
                "price": wrapped[0].get("amazon_price"),
                "url": wrapped[0].get("item_link", ""),
                "source": wrapped[0].get("item_supplier", ""),
            }
    except Exception as e:
        results["amazon"] = {"error": str(e)[:80]}
        log.debug("Single item Amazon error: %s", e)

    # 4. Unified Pricing Oracle
    try:
        from src.core.pricing_oracle_v2 import get_pricing
        oracle = get_pricing(
            description=desc,
            quantity=item.get("quantity", item.get("qty", 1)),
            cost=item.get("supplier_cost") or item.get("unit_price"),
            item_number=pn,
        )
        items[idx]["oracle"] = oracle
        results["oracle"] = {
            "quote_price": oracle.get("recommendation", {}).get("quote_price"),
            "confidence": oracle.get("recommendation", {}).get("confidence"),
            "market_avg": oracle.get("market", {}).get("weighted_avg"),
            "competitors": len(oracle.get("competitors", [])),
            "cross_sell": len(oracle.get("cross_sell", [])),
            "sources": oracle.get("sources_used", []),
        }
    except Exception as e:
        results["oracle"] = {"error": str(e)[:80]}

    save_rfqs(rfqs)

    # Build summary
    found = []
    if results["scprs"] and not results["scprs"].get("error"):
        found.append(f"SCPRS: ${results['scprs']['price']:.2f}")
    if results["catalog"] and not results["catalog"].get("error") and results["catalog"].get("cost"):
        found.append(f"Catalog: ${results['catalog']['cost']:.2f} ({results['catalog']['supplier']})")
    if results["amazon"] and not results["amazon"].get("error"):
        found.append(f"Amazon: ${results['amazon']['price']:.2f}")
    if results.get("oracle") and not results["oracle"].get("error") and results["oracle"].get("quote_price"):
        found.append(f"Oracle: ${results['oracle']['quote_price']:.2f} ({results['oracle'].get('confidence','?')})")

    return jsonify({
        "ok": True,
        "idx": idx,
        "description": desc[:60],
        "results": results,
        "summary": " | ".join(found) if found else "No prices found",
    })


@bp.route("/rfq/<rid>/upload-supplier-quote", methods=["POST"])
@auth_required
def rfq_upload_supplier_quote(rid):
    """Upload a supplier quote PDF → parse → match to RFQ items → fill costs + update catalog."""
    _bad = _validate_rid(rid)
    if _bad: return _bad
    import os
    from src.core.paths import DATA_DIR

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    f = request.files.get("file")
    if not f or not f.filename.lower().endswith(".pdf"):
        return jsonify({"ok": False, "error": "Upload a PDF file"})

    # Save uploaded file
    upload_dir = os.path.join(DATA_DIR, "uploads", "supplier_quotes")
    os.makedirs(upload_dir, exist_ok=True)
    pdf_path = os.path.join(upload_dir, f"sq_{rid}_{f.filename}")
    f.save(pdf_path)

    # Parse the quote
    try:
        from src.forms.supplier_quote_parser import parse_supplier_quote, match_quote_to_rfq
    except ImportError as e:
        return jsonify({"ok": False, "error": f"Parser not available: {e}"})

    parsed = parse_supplier_quote(pdf_path)
    if not parsed.get("ok"):
        return jsonify({"ok": False, "error": parsed.get("error", "Parse failed"),
                        "raw_text": parsed.get("raw_text", "")[:500]})

    quote_items = parsed.get("items", [])
    if not quote_items:
        return jsonify({"ok": False, "error": "No priced items found in PDF",
                        "raw_text": parsed.get("raw_text", "")[:500]})

    supplier = parsed.get("supplier", "Unknown")
    quote_num = parsed.get("quote_number", "")

    # Match to RFQ line items
    rfq_items = r.get("line_items", [])
    
    # ── If RFQ has NO items, create them from supplier quote directly ──
    if not rfq_items and quote_items:
        log.info("RFQ has 0 items — creating %d items from supplier quote", len(quote_items))
        for qi, q in enumerate(quote_items):
            new_item = {
                "line_number": qi + 1,
                "qty": q.get("qty", 1),
                "uom": (q.get("uom") or "EA").upper(),
                "description": q.get("description", ""),
                "item_number": q.get("item_number", ""),
                "supplier_cost": round(q.get("unit_price", 0), 2),
                "cost_source": "Supplier Quote",
                "cost_supplier_name": supplier,
                "item_supplier": supplier,
                "price_per_unit": 0,
                "scprs_last_price": 0,
                "amazon_price": 0,
                "_desc_source": "supplier",
            }
            rfq_items.append(new_item)
        r["line_items"] = rfq_items
        applied = len(quote_items)
        desc_upgraded = len(quote_items)
        unmatched = []
        catalog_added = 0
        catalog_updated = 0
        
        # Catalog sync for all items
        try:
            from src.agents.product_catalog import match_item, add_to_catalog, add_supplier_price, init_catalog_db
            init_catalog_db()
            for q in quote_items:
                _desc = q.get("description", "")
                _pn = q.get("item_number", "")
                _cost = q.get("unit_price", 0)
                if not _desc or _cost <= 0:
                    continue
                cat_matches = match_item(_desc, _pn, top_n=1)
                if cat_matches and cat_matches[0].get("match_confidence", 0) >= 0.5:
                    pid = cat_matches[0]["id"]
                    add_supplier_price(pid, supplier, _cost)
                    catalog_updated += 1
                else:
                    pid = add_to_catalog(description=_desc, part_number=_pn, cost=_cost,
                                         supplier_name=supplier, uom=q.get("uom", "EA"),
                                         source=f"supplier_quote_{quote_num or 'upload'}")
                    if pid:
                        add_supplier_price(pid, supplier, _cost)
                        catalog_added += 1
        except Exception as _ce:
            log.debug("Catalog sync (new items): %s", _ce)
        
        # Skip the matching loop — go straight to save
        r["_last_supplier_quote"] = {
            "supplier": supplier, "quote_number": quote_num, "pdf": pdf_path,
            "items_parsed": len(quote_items), "items_matched": applied,
            "uploaded_at": __import__("datetime").datetime.now().isoformat(),
        }
        save_rfqs(rfqs)
        
        try:
            from src.api.dashboard import save_rfq_file
            with open(pdf_path, "rb") as _qf:
                pdf_data = _qf.read()
            save_rfq_file(rid, f.filename, "application/pdf", pdf_data,
                         category="supplier_quote", uploaded_by="user")
        except Exception as _fe:
            log.warning("Failed to save supplier quote to DB: %s", _fe)
        
        _log_rfq_activity(rid, "supplier_quote_uploaded",
            f"Supplier quote from {supplier} ({quote_num or f.filename}): "
            f"CREATED {len(quote_items)} items (RFQ was empty), catalog +{catalog_added}/~{catalog_updated}",
            actor="user")
        
        try:
            from src.agents.drive_triggers import on_supplier_quote_uploaded
            on_supplier_quote_uploaded(r, pdf_path, supplier, quote_num)
        except Exception:
            pass
        
        return jsonify({
            "ok": True, "supplier": supplier, "quote_number": quote_num,
            "items_parsed": len(quote_items), "items_matched": applied,
            "desc_upgraded": desc_upgraded,
            "unmatched": [], "catalog_added": catalog_added, "catalog_updated": catalog_updated,
            "reconciliation": [{
                "line": i+1, "supplier_desc": q.get("description","")[:60],
                "supplier_pn": q.get("item_number",""), "supplier_qty": q.get("qty",0),
                "supplier_uom": q.get("uom",""), "supplier_price": q.get("unit_price",0),
                "matched": True, "confidence": 1.0, "rfq_line": i+1,
                "qty_match": True, "uom_match": True,
            } for i, q in enumerate(quote_items)],
            "warnings": {"qty_mismatches": 0, "uom_mismatches": 0},
            "created_items": True,
        })

    # ── Diagnostic logging for remote debugging ──
    log.info("SUPPLIER QUOTE MATCH: %d quote items vs %d RFQ items", len(quote_items), len(rfq_items))
    for qi, q in enumerate(quote_items[:5]):
        log.info("  Q[%d] pn='%s' $%.2f desc='%s'", qi, q.get("item_number",""), q.get("unit_price",0), q.get("description","")[:60])
    for ri, r_item in enumerate(rfq_items[:5]):
        log.info("  R[%d] pn='%s' desc='%s'", ri, r_item.get("item_number",""), r_item.get("description","")[:60])
    
    matches = match_quote_to_rfq(quote_items, rfq_items)
    
    matched_count = sum(1 for m in matches if m.get("matched"))
    log.info("MATCH RESULT: %d/%d matched", matched_count, len(matches))
    for m in matches:
        if not m.get("matched"):
            log.info("  UNMATCHED: Q[%d] pn='%s' $%.2f conf=%.2f", m["quote_idx"], m["quote_pn"], m["unit_price"], m.get("confidence",0))

    # Apply matches: fill cost, pick best description, update catalog
    applied = 0
    unmatched = []
    catalog_added = 0
    catalog_updated = 0
    desc_upgraded = 0

    for m in matches:
        cost = m.get("unit_price", 0)
        q_desc = m.get("quote_desc", "")
        q_pn = m.get("quote_pn", "")
        q_qty = m.get("qty", 1)
        q_uom = m.get("uom", "EA")

        if m["matched"] and m["rfq_idx"] is not None and cost > 0:
            idx = m["rfq_idx"]
            if idx < len(rfq_items):
                item = rfq_items[idx]
                item["supplier_cost"] = round(cost, 2)
                item["cost_source"] = "Supplier Quote"
                item["cost_supplier_name"] = supplier
                item["item_supplier"] = supplier

                # ── Description: ALWAYS use supplier's when matched ──
                # Supplier is the source of truth for what they're selling.
                # Their descriptions include McKesson #, MFG #, pack sizes.
                if q_desc and len(q_desc) > 10:
                    item["description"] = q_desc
                    item["_desc_source"] = "supplier"
                    desc_upgraded += 1

                # Fill part number if empty or from a richer source
                if q_pn and not item.get("item_number"):
                    item["item_number"] = q_pn

                applied += 1

                # ── Catalog update: always save/enrich ──
                try:
                    from src.agents.product_catalog import (
                        match_item, add_to_catalog, add_supplier_price,
                        init_catalog_db
                    )
                    init_catalog_db()
                    pn_search = q_pn or item.get("item_number", "")
                    cat_matches = match_item(q_desc, pn_search, top_n=1)
                    
                    if cat_matches and cat_matches[0].get("match_confidence", 0) >= 0.5:
                        pid = cat_matches[0]["id"]
                        # Always update catalog description with supplier's (richer)
                        if q_desc and len(q_desc) > len(cat_matches[0].get("description", "") or ""):
                            try:
                                from src.agents.product_catalog import _get_conn, _tokenize
                                conn = _get_conn()
                                conn.execute(
                                    "UPDATE product_catalog SET description=?, search_tokens=?, updated_at=? WHERE id=?",
                                    (q_desc, _tokenize(q_desc),
                                     __import__('datetime').datetime.now().isoformat(), pid)
                                )
                                conn.commit()
                                conn.close()
                            except Exception:
                                pass
                        add_supplier_price(pid, supplier, cost)
                        catalog_updated += 1
                    else:
                        pid = add_to_catalog(
                            description=q_desc,
                            part_number=pn_search,
                            cost=cost,
                            supplier_name=supplier,
                            uom=q_uom,
                            source=f"supplier_quote_{quote_num or 'upload'}",
                        )
                        if pid:
                            add_supplier_price(pid, supplier, cost)
                            catalog_added += 1
                except Exception as _ce:
                    log.debug("Catalog update from supplier quote: %s", _ce)

        else:
            unmatched.append({
                "description": q_desc,
                "part_number": q_pn,
                "unit_price": cost,
                "qty": q_qty,
            })
            # Still add unmatched items to catalog
            if cost > 0 and q_desc:
                try:
                    from src.agents.product_catalog import (
                        match_item, add_to_catalog, add_supplier_price,
                        init_catalog_db
                    )
                    init_catalog_db()
                    cat_matches = match_item(q_desc, q_pn, top_n=1)
                    if cat_matches and cat_matches[0].get("match_confidence", 0) >= 0.5:
                        pid = cat_matches[0]["id"]
                        add_supplier_price(pid, supplier, cost)
                        catalog_updated += 1
                    else:
                        pid = add_to_catalog(
                            description=q_desc,
                            part_number=q_pn,
                            cost=cost,
                            supplier_name=supplier,
                            source=f"supplier_quote_{quote_num or 'upload'}",
                        )
                        if pid:
                            add_supplier_price(pid, supplier, cost)
                            catalog_added += 1
                except Exception as _ce:
                    log.debug("Catalog update (unmatched): %s", _ce)

    # Save RFQ
    r["_last_supplier_quote"] = {
        "supplier": supplier,
        "quote_number": quote_num,
        "pdf": pdf_path,
        "items_parsed": len(quote_items),
        "items_matched": applied,
        "uploaded_at": __import__("datetime").datetime.now().isoformat(),
    }
    save_rfqs(rfqs)

    # Save supplier quote PDF to rfq_files DB (persists across deploys)
    try:
        from src.api.dashboard import save_rfq_file
        with open(pdf_path, "rb") as _qf:
            pdf_data = _qf.read()
        save_rfq_file(
            rid,
            f.filename,
            "application/pdf",
            pdf_data,
            category="supplier_quote",
            uploaded_by="user",
        )
        log.info("Supplier quote saved to DB: %s (%d bytes) for RFQ %s", f.filename, len(pdf_data), rid)
    except Exception as _fe:
        log.warning("Failed to save supplier quote to DB: %s", _fe)

    # Log activity
    sol = r.get("solicitation_number", rid)
    _log_rfq_activity(rid, "supplier_quote_uploaded",
        f"Supplier quote from {supplier} ({quote_num or f.filename}): "
        f"{len(quote_items)} parsed, {applied} matched, {desc_upgraded} desc upgraded, "
        f"catalog +{catalog_added}/~{catalog_updated}",
        actor="user")

    # ── Google Drive: archive supplier quote ──
    try:
        from src.agents.drive_triggers import on_supplier_quote_uploaded
        on_supplier_quote_uploaded(r, pdf_path, supplier, quote_num)
    except Exception as _gde:
        log.debug("Drive trigger (supplier_quote): %s", _gde)

    # ── Build reconciliation table for validation ──
    reconciliation = []
    for m in matches:
        recon = {
            "line": m.get("quote_idx", 0) + 1,
            "supplier_desc": (m.get("quote_desc") or "")[:60],
            "supplier_pn": m.get("quote_pn", ""),
            "supplier_qty": m.get("qty", 0),
            "supplier_uom": m.get("uom", ""),
            "supplier_price": m.get("unit_price", 0),
            "matched": m.get("matched", False),
            "confidence": m.get("confidence", 0),
        }
        if m["matched"] and m["rfq_idx"] is not None and m["rfq_idx"] < len(rfq_items):
            ri = rfq_items[m["rfq_idx"]]
            recon["rfq_line"] = m["rfq_idx"] + 1
            recon["rfq_qty"] = ri.get("qty", 0)
            recon["rfq_uom"] = (ri.get("uom") or "").upper()
            recon["qty_match"] = recon["supplier_qty"] == recon["rfq_qty"]
            recon["uom_match"] = recon["supplier_uom"].upper() == recon["rfq_uom"]
        reconciliation.append(recon)

    # Validation flags
    qty_mismatches = [r for r in reconciliation if r.get("matched") and not r.get("qty_match", True)]
    uom_mismatches = [r for r in reconciliation if r.get("matched") and not r.get("uom_match", True)]

    return jsonify({
        "ok": True,
        "supplier": supplier,
        "quote_number": quote_num,
        "items_parsed": len(quote_items),
        "items_matched": applied,
        "desc_upgraded": desc_upgraded,
        "unmatched": unmatched,
        "catalog_added": catalog_added,
        "catalog_updated": catalog_updated,
        "reconciliation": reconciliation,
        "warnings": {
            "qty_mismatches": len(qty_mismatches),
            "uom_mismatches": len(uom_mismatches),
        },
    })


def _renumber_items(items):
    """Re-number line items sequentially."""
    for i, it in enumerate(items):
        it["line_number"] = i + 1


def _item_response(rid, ok, msg):
    """Return JSON for AJAX or redirect for form POST."""
    if request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": ok, "message": msg})
    from flask import flash as _flash
    _flash(msg, "success" if ok else "error")
    return redirect(f"/rfq/{rid}")


@bp.route("/rfq/<rid>/upload-templates", methods=["POST"])
@auth_required
@safe_route
def upload_templates(rid):
    """Upload 703B/704B/Bid Package template PDFs for an RFQ."""
    _bad = _validate_rid(rid)
    if _bad: return _bad
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error"); return redirect("/")

    files = request.files.getlist("templates")
    if not files:
        flash("No files uploaded", "error"); return redirect(f"/rfq/{rid}")

    rfq_dir = os.path.join(UPLOAD_DIR, rid)
    os.makedirs(rfq_dir, exist_ok=True)

    saved = []
    for f in files:
        safe_fn = _safe_filename(f.filename)
        if safe_fn and safe_fn.lower().endswith(".pdf"):
            p = os.path.join(rfq_dir, safe_fn)
            f.save(p)
            saved.append(p)
            # Store in DB
            try:
                f.seek(0)
                file_data = f.read()
                if not file_data:
                    with open(p, "rb") as _rb:
                        file_data = _rb.read()
                save_rfq_file(rid, safe_fn, "template_upload", file_data, category="template", uploaded_by="user")
            except Exception:
                try:
                    with open(p, "rb") as _rb:
                        save_rfq_file(rid, safe_fn, "template_upload", _rb.read(), category="template", uploaded_by="user")
                except Exception as _e:
                    log.debug("Suppressed: %s", _e)

    if not saved:
        flash("No PDFs found in upload", "error"); return redirect(f"/rfq/{rid}")

    # Identify which forms were uploaded
    new_templates = identify_attachments(saved)

    # Merge with existing templates (don't overwrite)
    existing = r.get("templates", {})
    for key, path in new_templates.items():
        existing[key] = path
    r["templates"] = existing

    # If we now have a 704B and didn't have line items, re-parse
    if "704b" in new_templates and not r.get("line_items"):
        try:
            parsed = parse_rfq_attachments(existing)
            r["line_items"] = parsed.get("line_items", r.get("line_items", []))
            r["solicitation_number"] = parsed.get("solicitation_number", r.get("solicitation_number", ""))
            r["delivery_location"] = parsed.get("delivery_location", r.get("delivery_location", ""))
            # Auto SCPRS lookup on new items
            r["line_items"] = bulk_lookup(r.get("line_items", []))
        except Exception as e:
            log.error(f"Re-parse error: {e}")

    save_rfqs(rfqs)

    found = [k for k in new_templates.keys()]
    _log_rfq_activity(rid, "templates_uploaded",
        f"Templates uploaded: {', '.join(found).upper()} for #{r.get('solicitation_number','?')}",
        actor="user", metadata={"templates": found})
    flash(f"Templates uploaded: {', '.join(found).upper()}", "success")
    return redirect(f"/rfq/{rid}")


@bp.route("/rfq/<rid>/generate-package", methods=["POST"])
@auth_required
@safe_route
def generate_rfq_package(rid):
    """ONE BUTTON — generates complete RFQ package:
    1. Filled 703B (RFQ form)
    2. Filled 704B (Quote Worksheet)
    3. Filled Bid Package
    4. Reytech Quote on letterhead
    5. Draft email with all attachments
    """
    _bad = _validate_rid(rid)
    if _bad: return _bad
    from src.api.trace import Trace
    t = Trace("rfq_package", rfq_id=rid)
    
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        t.fail("RFQ not found")
        flash("RFQ not found", "error")
        return redirect("/")
    
    sol = r.get("solicitation_number", "unknown")
    t.step("Starting", sol=sol, items=len(r.get("line_items", [])))
    
    # ── Step 1: Save ALL fields from form (not just pricing) ──
    for i, item in enumerate(r.get("line_items", [])):
        for field, key in [("cost", "supplier_cost"), ("scprs", "scprs_last_price"), ("price", "price_per_unit"), ("markup", "markup_pct")]:
            v = request.form.get(f"{field}_{i}")
            if v:
                try:
                    item[key] = float(v)
                except Exception as _e:
                    log.debug("Suppressed: %s", _e)
        # Save description, qty, uom, part# from form too
        desc_val = request.form.get(f"desc_{i}")
        if desc_val is not None:
            item["description"] = desc_val.strip()
        qty_val = request.form.get(f"qty_{i}")
        if qty_val:
            try: item["qty"] = int(float(qty_val))
            except Exception: pass
        uom_val = request.form.get(f"uom_{i}")
        if uom_val is not None:
            item["uom"] = uom_val.strip().upper()
        part_val = request.form.get(f"part_{i}")
        if part_val is not None:
            item["item_number"] = part_val.strip()
        link_val = request.form.get(f"link_{i}", "").strip()
        if link_val:
            item["item_link"] = link_val
    
    r["sign_date"] = get_pst_date()
    safe_sol = re.sub(r'[^a-zA-Z0-9_-]', '_', sol.strip())
    out_dir = os.path.join(OUTPUT_DIR, sol)
    
    # ── Step 1.5: Archive old files (don't delete until new generation succeeds) ──
    import shutil as _sh_clean
    _old_dir = None
    if os.path.exists(out_dir):
        _old_dir = out_dir + "_prev"
        try:
            if os.path.exists(_old_dir):
                _sh_clean.rmtree(_old_dir)
            os.rename(out_dir, _old_dir)
            t.step(f"Archived old files to {sol}_prev/")
        except Exception as _ce:
            _old_dir = None
            t.warn("Archive failed, removing old files", error=str(_ce))
            try:
                _sh_clean.rmtree(out_dir)
            except Exception:
                pass
    os.makedirs(out_dir, exist_ok=True)

    r["output_files"] = []
    r.pop("draft_email", None)
    
    output_files = []
    errors = []
    
    # ── Step 2: Fill State Forms (703B, 704B, Bid Package) ──
    try:
        tmpl = r.get("templates", {})
        
        # ── DB Fallback: if template files don't exist on disk (post-redeploy),
        # reconstruct them from rfq_files DB table ──
        db_files = list_rfq_files(rid, category="template")
        type_map = {"703b": "703b", "704b": "704b", "bidpkg": "bidpkg", "bid_package": "bidpkg"}
        for db_f in db_files:
            # Determine template type from filename or file_type
            ft = db_f.get("file_type", "").lower().replace("template_", "")
            fname = db_f.get("filename", "").lower()
            ttype = None
            if "703b" in ft or "703b" in fname:
                ttype = "703b"
            elif "704b" in ft or "704b" in fname:
                ttype = "704b"
            elif "bid" in ft or "bid" in fname:
                ttype = "bidpkg"
            
            if ttype and (ttype not in tmpl or not os.path.exists(tmpl.get(ttype, ""))):
                # Restore from DB to temp location
                full_f = get_rfq_file(db_f["id"])
                if full_f and full_f.get("data"):
                    restore_dir = os.path.join(DATA_DIR, "rfq_templates", rid)
                    os.makedirs(restore_dir, exist_ok=True)
                    restore_path = os.path.join(restore_dir, db_f["filename"])
                    with open(restore_path, "wb") as _fw:
                        _fw.write(full_f["data"])
                    tmpl[ttype] = restore_path
                    t.step(f"Restored {ttype} from DB: {db_f['filename']}")
        
        # ── Auto-fallback: Use saved CDCR bid package template ONLY for CDCR/CCHCS ──
        _agency = (r.get("agency", "") or "").upper()
        _is_cdcr = any(x in _agency for x in ["CDCR", "CCHCS", "CORRECTIONS"])
        if _is_cdcr and ("bidpkg" not in tmpl or not os.path.exists(tmpl.get("bidpkg", ""))):
            default_bidpkg = os.path.join(DATA_DIR, "templates", "cdcr_bid_package_template.pdf")
            if os.path.exists(default_bidpkg):
                tmpl["bidpkg"] = default_bidpkg
                t.step("Using default CDCR bid package template")
        
        # Update templates in RFQ data
        r["templates"] = tmpl
        
        # ── Match agency FIRST — determines which forms to generate ──
        try:
            from src.core.agency_config import match_agency
            _agency_key, _agency_cfg = match_agency(r)
            _req_forms = set(_agency_cfg.get("required_forms", []))
            _opt_forms = set(_agency_cfg.get("optional_forms", []))
            t.step(f"Agency matched: {_agency_key} ({_agency_cfg.get('name','')}), {len(_req_forms)} required forms: {', '.join(sorted(_req_forms))}")
        except Exception as _ae:
            t.warn(f"Agency config load failed, using CCHCS default: {_ae}")
            _req_forms = {"703b", "703c", "704b", "bidpkg", "quote", "sellers_permit"}
            _opt_forms = set()
            _agency_key = "cchcs"
        
        # Helper: should this form be included?
        _user_forms = r.get("package_forms", {})
        def _include(form_id):
            # User checklist overrides if set
            if form_id in _user_forms:
                return bool(_user_forms[form_id])
            return form_id in _req_forms

        # ── Check buyer preferences before generating ──
        try:
            from src.core.dal import get_buyer_preferences as _gbp
            _buyer_email = r.get("requestor_email", "")
            if _buyer_email:
                for _bp in _gbp(_buyer_email):
                    _bk = _bp.get("preference_key", "")
                    _bv = _bp.get("preference_value", "")
                    if _bk == "ship_to_override" and _bv and _bv != r.get("delivery_location", ""):
                        r["delivery_location"] = _bv
                        log.info("GENERATE %s: ship_to overridden by buyer pref: %s", rid, _bv[:40])
                        t.step(f"Buyer pref: ship_to → {_bv[:40]}")
                    elif _bk == "required_forms" and _bv:
                        for _ef in [f.strip() for f in _bv.split(",") if f.strip()]:
                            if _ef not in _req_forms:
                                _req_forms.add(_ef)
                                t.step(f"Buyer pref: added form {_ef}")
                    elif _bk == "no_modify_buyer_fields":
                        t.step("Buyer pref: no_modify_buyer_fields active")
        except Exception as _bp_e:
            log.debug("Buyer pref check: %s", _bp_e)

        # ── Auto-validate tax rate if not yet validated ──
        if not r.get("tax_validated") and r.get("delivery_location"):
            try:
                from src.agents.tax_agent import get_tax_rate as _gtr
                _dl = r["delivery_location"]
                _zip_m = __import__('re').search(r'\b(\d{5})\b', _dl)
                _city_m = __import__('re').search(r',\s*([A-Za-z\s]+),\s*[A-Z]{2}', _dl)
                _street_m = __import__('re').search(r'^(\d+\s+.+?)(?:,|$)', _dl)
                _tax_r = _gtr(
                    street=_street_m.group(1).strip() if _street_m else "",
                    city=_city_m.group(1).strip() if _city_m else "",
                    zip_code=_zip_m.group(1) if _zip_m else ""
                )
                if _tax_r and _tax_r.get("rate"):
                    r["tax_rate"] = round(_tax_r["rate"] * 100, 3)
                    r["tax_validated"] = True
                    r["tax_source"] = _tax_r.get("source", "cdtfa_api")
                    r["tax_jurisdiction"] = _tax_r.get("jurisdiction", "")
                    t.step(f"Tax auto-validated: {r['tax_rate']}% ({r['tax_jurisdiction']})")
            except Exception as _te:
                log.debug("Tax auto-validate: %s", _te)

        try:
            from src.core.dal import log_lifecycle_event as _lle_start
            _lle_start("rfq", rid, "package_generate_started",
                f"Generating package for {_agency_key} ({len(_req_forms)} required forms)",
                actor="user", detail={"agency": _agency_key, "required_forms": list(_req_forms)})
        except Exception:
            pass

        # ── Template-based forms (only if agency requires them) ──
        if _include("703b") or _include("703c"):
            # Handle both 703B and 703C (Fair & Reasonable) — same fill logic
            _703_key = "703c" if "703c" in tmpl else "703b"
            _703_label = "703C" if _703_key == "703c" else "703B"
            if _703_key in tmpl and os.path.exists(tmpl[_703_key]):
                try:
                    fill_703b(tmpl[_703_key], r, CONFIG, f"{out_dir}/{sol}_{_703_label}_Reytech.pdf")
                    output_files.append(f"{sol}_{_703_label}_Reytech.pdf")
                    t.step(f"{_703_label} filled")
                except Exception as e:
                    errors.append(f"{_703_label}: {e}")
                    t.warn(f"{_703_label} fill failed", error=str(e))
            else:
                t.step(f"{_703_label} skipped — no template")
                errors.append(f"{_703_label}: no template uploaded — upload {_703_label} PDF on this RFQ page")
        
        if _include("704b"):
            if "704b" in tmpl and os.path.exists(tmpl["704b"]):
                try:
                    fill_704b(tmpl["704b"], r, CONFIG, f"{out_dir}/{sol}_704B_Reytech.pdf")
                    output_files.append(f"{sol}_704B_Reytech.pdf")
                    t.step("704B filled")
                except Exception as e:
                    errors.append(f"704B: {e}")
                    t.warn("704B fill failed", error=str(e))
            else:
                t.step("704B skipped — no template")
                errors.append("704B: no template uploaded — upload 704B PDF on this RFQ page")
        
        if _include("bidpkg"):
            if "bidpkg" in tmpl and os.path.exists(tmpl["bidpkg"]):
                try:
                    _bidpkg_path = f"{out_dir}/{sol}_BidPackage_Reytech.pdf"
                    fill_bid_package(tmpl["bidpkg"], r, CONFIG, _bidpkg_path)
                    output_files.append(f"{sol}_BidPackage_Reytech.pdf")
                    t.step("Bid Package filled")
                    # Replace 843 pages in bid package with master template version
                    try:
                        import tempfile as _tmpf
                        _843_tmp = _tmpf.mktemp(suffix=".pdf")
                        from src.forms.reytech_filler_v4 import generate_dvbe_843
                        generate_dvbe_843(r, CONFIG, _843_tmp)
                        from pypdf import PdfReader as _PR843, PdfWriter as _PW843
                        _bp_reader = _PR843(_bidpkg_path)
                        _843_reader = _PR843(_843_tmp)
                        _bp_writer = _PW843()
                        _replaced = False
                        for _bp_page in _bp_reader.pages:
                            _ptxt = (_bp_page.extract_text() or "").upper()
                            if ("DVBE DECLARATIONS" in _ptxt or "DGS PD 843" in _ptxt) and not _replaced:
                                for _843p in _843_reader.pages:
                                    _bp_writer.add_page(_843p)
                                _replaced = True
                                continue
                            _bp_writer.add_page(_bp_page)
                        if _replaced:
                            with open(_bidpkg_path, "wb") as _bpf:
                                _bp_writer.write(_bpf)
                            t.step("843 replaced with master template in bid package")
                        os.remove(_843_tmp)
                    except Exception as _843e:
                        log.debug("843 replacement in bid package: %s", _843e)
                except Exception as e:
                    errors.append(f"Bid Package: {e}")
                    t.warn("Bid Package fill failed", error=str(e))
            else:
                t.step("Bid Package skipped — no template")
                errors.append("Bid Package: no template uploaded — upload Bid Package PDF on this RFQ page")
        
        # ── AGENCY-GATED FORMS ─────
        
        # STD 204 Payee Data Record
        if _include("std204"):
            try:
                from src.forms.reytech_filler_v4 import fill_std204
                std204_tmpl = os.path.join(DATA_DIR, "templates", "std204_blank.pdf")
                if os.path.exists(std204_tmpl):
                    fill_std204(std204_tmpl, r, CONFIG, f"{out_dir}/{sol}_STD204_Reytech.pdf")
                    output_files.append(f"{sol}_STD204_Reytech.pdf")
                    t.step("STD 204 filled")
            except Exception as e:
                errors.append(f"STD 204: {e}")
        
        # Seller's Permit
        if _include("sellers_permit"):
            try:
                sellers_permit = os.path.join(DATA_DIR, "templates", "sellers_permit_reytech.pdf")
                if os.path.exists(sellers_permit):
                    import shutil
                    shutil.copy2(sellers_permit, f"{out_dir}/{sol}_SellersPermit_Reytech.pdf")
                    output_files.append(f"{sol}_SellersPermit_Reytech.pdf")
                    t.step("Seller's Permit included")
            except Exception as e:
                t.warn("Seller's Permit copy failed", error=str(e))
        
        # DVBE 843
        if _include("dvbe843"):
            try:
                from src.forms.reytech_filler_v4 import generate_dvbe_843
                generate_dvbe_843(r, CONFIG, f"{out_dir}/{sol}_DVBE843_Reytech.pdf")
                output_files.append(f"{sol}_DVBE843_Reytech.pdf")
                t.step("DVBE 843 generated")
            except Exception as e:
                errors.append(f"DVBE 843: {e}")
        
        # CV 012 CUF (Cal Vet)
        if _include("cv012_cuf"):
            try:
                from src.forms.reytech_filler_v4 import fill_cv012_cuf
                cuf_tmpl = os.path.join(DATA_DIR, "templates", "cv012_cuf_blank.pdf")
                if os.path.exists(cuf_tmpl):
                    fill_cv012_cuf(cuf_tmpl, r, CONFIG, f"{out_dir}/{sol}_CV012_CUF_Reytech.pdf")
                    output_files.append(f"{sol}_CV012_CUF_Reytech.pdf")
                    t.step("CV 012 CUF filled")
            except Exception as e:
                t.warn("CV 012 CUF failed", error=str(e))
        
        # Barstow CUF (facility-specific)
        if _include("barstow_cuf") or "barstow_cuf" in _opt_forms:
            try:
                from src.forms.reytech_filler_v4 import generate_barstow_cuf
                _rfq_text = " ".join([
                    str(r.get("ship_to", "")), str(r.get("delivery_location", "")),
                    str(r.get("institution", "")),
                ]).lower()
                if "barstow" in _rfq_text or "92311" in _rfq_text:
                    generate_barstow_cuf(r, CONFIG, f"{out_dir}/{sol}_BarstowCUF_Reytech.pdf")
                    output_files.append(f"{sol}_BarstowCUF_Reytech.pdf")
                    t.step("Barstow CUF generated")
            except Exception as e:
                t.warn("Barstow CUF failed", error=str(e))
        
        # Bidder Declaration
        if _include("bidder_decl"):
            try:
                _bd_tmpl = os.path.join(DATA_DIR, "templates", "bidder_declaration_blank.pdf")
                if os.path.exists(_bd_tmpl):
                    from src.forms.reytech_filler_v4 import fill_bidder_declaration
                    fill_bidder_declaration(_bd_tmpl, r, CONFIG, f"{out_dir}/{sol}_BidderDecl_Reytech.pdf")
                    t.step("Bidder Declaration filled from template")
                else:
                    from src.forms.reytech_filler_v4 import generate_bidder_declaration
                    generate_bidder_declaration(r, CONFIG, f"{out_dir}/{sol}_BidderDecl_Reytech.pdf")
                    t.step("Bidder Declaration generated via ReportLab (no template)")
                output_files.append(f"{sol}_BidderDecl_Reytech.pdf")
            except Exception as e:
                errors.append(f"Bidder Declaration: {e}")
                t.warn("Bidder Declaration failed", error=str(e))

        # Darfur Act — template-first, ReportLab fallback
        if _include("darfur_act"):
            try:
                _da_tmpl = os.path.join(DATA_DIR, "templates", "darfur_act_blank.pdf")
                if os.path.exists(_da_tmpl):
                    from src.forms.reytech_filler_v4 import fill_darfur_standalone
                    fill_darfur_standalone(_da_tmpl, r, CONFIG, f"{out_dir}/{sol}_DarfurAct_Reytech.pdf")
                    t.step("Darfur Act filled from template")
                else:
                    from src.forms.reytech_filler_v4 import generate_darfur_act
                    generate_darfur_act(r, CONFIG, f"{out_dir}/{sol}_DarfurAct_Reytech.pdf")
                    t.step("Darfur Act generated via ReportLab (no template)")
                output_files.append(f"{sol}_DarfurAct_Reytech.pdf")
            except Exception as e:
                errors.append(f"Darfur Act: {e}")
                t.warn("Darfur Act failed", error=str(e))
        
        # CalRecycle 74
        if _include("calrecycle74") or ("calrecycle74" in _opt_forms and len(r.get("line_items", [])) > 6):
            try:
                from src.forms.reytech_filler_v4 import fill_calrecycle_standalone
                cr_tmpl = os.path.join(DATA_DIR, "templates", "calrecycle_74_blank.pdf")
                if os.path.exists(cr_tmpl):
                    fill_calrecycle_standalone(cr_tmpl, r, CONFIG, f"{out_dir}/{sol}_CalRecycle74_Reytech.pdf")
                    output_files.append(f"{sol}_CalRecycle74_Reytech.pdf")
                    t.step("CalRecycle 74 filled")
            except Exception as e:
                t.warn("CalRecycle 74 failed", error=str(e))
        
        # STD 1000 GenAI
        if _include("std1000"):
            try:
                from src.forms.reytech_filler_v4 import fill_std1000
                std1000_tmpl = os.path.join(DATA_DIR, "templates", "std1000_blank.pdf")
                if os.path.exists(std1000_tmpl):
                    fill_std1000(std1000_tmpl, r, CONFIG, f"{out_dir}/{sol}_STD1000_Reytech.pdf")
                    output_files.append(f"{sol}_STD1000_Reytech.pdf")
                    t.step("STD 1000 filled")
            except Exception as e:
                t.warn("STD 1000 failed", error=str(e))
        
        # STD 205
        if _include("std205"):
            try:
                from src.forms.reytech_filler_v4 import generate_std205
                generate_std205(r, CONFIG, f"{out_dir}/{sol}_STD205_Reytech.pdf")
                output_files.append(f"{sol}_STD205_Reytech.pdf")
                t.step("STD 205 generated")
            except Exception as e:
                t.warn("STD 205 failed", error=str(e))
        
        # Drug-Free Workplace
        if _include("drug_free"):
            try:
                from src.forms.reytech_filler_v4 import generate_drug_free
                generate_drug_free(r, CONFIG, f"{out_dir}/{sol}_DrugFree_Reytech.pdf")
                output_files.append(f"{sol}_DrugFree_Reytech.pdf")
                t.step("Drug-Free STD 21 generated")
            except Exception as e:
                t.warn("Drug-Free failed", error=str(e))
        # GenAI 708
        if _include("genai_708"):
            _genai_tmpl = os.path.join(DATA_DIR, "templates", "genai_708_blank.pdf")
            if os.path.exists(_genai_tmpl):
                try:
                    from src.forms.reytech_filler_v4 import fill_genai_708
                    fill_genai_708(_genai_tmpl, r, CONFIG, f"{out_dir}/{sol}_GenAI708_Reytech.pdf")
                    output_files.append(f"{sol}_GenAI708_Reytech.pdf")
                    t.step("GenAI 708 filled")
                except Exception as e:
                    errors.append(f"GenAI 708: {e}")

        # STD 205 from template (prefer over ReportLab version)
        if _include("std205"):
            _std205_tmpl = os.path.join(DATA_DIR, "templates", "STD205_Payee_Data_Record_Supplement.pdf")
            if os.path.exists(_std205_tmpl) and f"{sol}_STD205_Reytech.pdf" not in output_files:
                try:
                    from src.forms.reytech_filler_v4 import fill_std205
                    fill_std205(_std205_tmpl, r, CONFIG, f"{out_dir}/{sol}_STD205_Reytech.pdf")
                    output_files.append(f"{sol}_STD205_Reytech.pdf")
                    t.step("STD 205 filled from template")
                except Exception as e:
                    errors.append(f"STD 205 template: {e}")

        # W-9 (static copy) — only if agency requires it (CalVet doesn't)
        if _include("w9"):
            _w9_path = os.path.join(DATA_DIR, "templates", "w9_reytech.pdf")
            if os.path.exists(_w9_path):
                import shutil as _sh_w9
                try:
                    _sh_w9.copy2(_w9_path, f"{out_dir}/{sol}_W9_Reytech.pdf")
                    output_files.append(f"{sol}_W9_Reytech.pdf")
                except Exception:
                    pass

        # Seller's Permit (static copy if not already added)
        if _include("sellers_permit"):
            _sp_path = os.path.join(DATA_DIR, "templates", "sellers_permit_reytech.pdf")
            if os.path.exists(_sp_path) and f"{sol}_SellersPermit_Reytech.pdf" not in output_files:
                import shutil as _sh_sp
                try:
                    _sh_sp.copy2(_sp_path, f"{out_dir}/{sol}_SellersPermit_Reytech.pdf")
                    output_files.append(f"{sol}_SellersPermit_Reytech.pdf")
                except Exception:
                    pass

    except Exception as e:
        errors.append(f"State forms: {e}")
        t.warn("State forms exception", error=str(e))

    # ── Step 2.5: 703C master template fallback (ONLY if agency requires it) ──
    if (_include("703b") or _include("703c")) and not any(f.endswith("_703B_Reytech.pdf") or f.endswith("_703C_Reytech.pdf") for f in output_files):
        _master_703c = os.path.join(DATA_DIR, "templates", "AMS 703C - RFQ.pdf")
        if os.path.exists(_master_703c):
            try:
                fill_703b(_master_703c, r, CONFIG, f"{out_dir}/{sol}_703C_Reytech.pdf")
                output_files.append(f"{sol}_703C_Reytech.pdf")
                t.step("703C filled from master template")
            except Exception as e:
                t.warn("703C master fill failed", error=str(e))

    # ── Step 3: Generate Reytech Quote on letterhead ──
    if QUOTE_GEN_AVAILABLE:
        try:
            quote_path = os.path.join(out_dir, f"{safe_sol}_Quote_Reytech.pdf")
            locked_qn = r.get("reytech_quote_number", "")

            # SQLite doesn't store reytech_quote_number — check JSON directly
            if not locked_qn:
                try:
                    import json as _jqn
                    _rfq_json_path = os.path.join(DATA_DIR, "rfqs.json")
                    if os.path.exists(_rfq_json_path):
                        with open(_rfq_json_path) as _jf:
                            _jrfqs = _jqn.load(_jf)
                        locked_qn = _jrfqs.get(rid, {}).get("reytech_quote_number", "")
                        if locked_qn:
                            r["reytech_quote_number"] = locked_qn
                            t.step(f"Recovered quote number from JSON: {locked_qn}")
                except Exception:
                    pass

            # GUARDRAIL: if this RFQ already has a quote number locked,
            # ALWAYS reuse it — never burn a new counter on regenerate.
            _was_existing_qn = bool(locked_qn)
            if locked_qn:
                t.step(f"Reusing locked quote number: {locked_qn}")
            else:
                # Allocate number BEFORE generating and save immediately
                from src.forms.quote_generator import _next_quote_number
                locked_qn = _next_quote_number()
                r["reytech_quote_number"] = locked_qn
                save_rfqs(rfqs)  # persist NOW so next generate sees it
                t.step(f"Allocated new quote number: {locked_qn}")

            result = generate_quote_from_rfq(
                r, quote_path,
                include_tax=True,
                quote_number=locked_qn,
            )

            if result.get("ok"):
                qn = result.get("quote_number", locked_qn)
                r["reytech_quote_number"] = qn
                output_files.append(f"{safe_sol}_Quote_Reytech.pdf")
                t.step("Reytech Quote generated", quote_number=qn, total=result.get("total", 0))
                # CRM log
                _log_crm_activity(qn, "quote_generated",
                                  f"Quote {qn} generated for RFQ {sol} — ${result.get('total',0):,.2f}",
                                  actor="user", metadata={"rfq_id": rid})
            else:
                errors.append(f"Quote: {result.get('error', 'unknown')}")
                t.warn("Quote generation failed", error=result.get("error", "unknown"))
                # Rollback: if we just allocated a new number and generation failed, release it
                if locked_qn and not _was_existing_qn:
                    r["reytech_quote_number"] = ""
                    log.warning("Rolled back quote number %s after generation failure", locked_qn)
        except Exception as e:
            errors.append(f"Quote: {e}")
            t.warn("Quote exception", error=str(e))
    else:
        t.step("Quote generator not available — skipped")
    
    if not output_files and not r.get("form_type") == "generic_rfq":
        t.fail("No files generated", errors=errors)
        flash(f"No files generated — {'; '.join(errors) if errors else 'No templates found'}", "error")
        # Restore old files on failure
        if _old_dir and os.path.exists(_old_dir) and not os.listdir(out_dir):
            try:
                _sh_clean.rmtree(out_dir)
                os.rename(_old_dir, out_dir)
                log.info("Restored old files after generation failure for %s", rid)
            except Exception:
                pass
        return redirect(f"/rfq/{rid}")
    
    # ── Step 3.5: Collect ALL package PDFs (state forms + original RFQ attachments) ──
    # These will be merged into one single package PDF
    package_pdfs = []  # (filepath, label) — order matters
    
    # ── Split output files into 4 separate attachments ──
    # Attachment 1: 703B (standalone filled bidder form)
    # Attachment 2: 704B (standalone filled pricing worksheet)
    # Attachment 3: Formal quote on Reytech letterhead
    # Attachment 4: RFQ Package — all supporting compliance docs merged in this order:
    #   BidPackage (CDCR Terms + CUF MC-345 + GenAI 708 + Voluntary Stats)
    #   CalRecycle 74, Bidder Declaration, Darfur Act, DVBE 843,
    #   Drug-Free STD 21, Seller's Permit, STD 204
    #
    # RULE: To add a form to the package, add its form_id to the agency config
    # required_forms AND add its filename pattern to _FORM_ORDER below.
    # NEVER remove BidPackage from the package — it contains CUF/708/Stats.

    file_703b = None
    file_704b = None
    quote_file = None

    for f in output_files:
        fpath = os.path.join(out_dir, f)
        if not os.path.exists(fpath):
            continue
        fu = f.upper()
        if "703B" in fu:
            file_703b = f
        elif "704B" in fu:
            file_704b = f
        elif "QUOTE" in fu:
            quote_file = f
        else:
            package_pdfs.append((fpath, f))

    # ── Canonical package order — matches model R25Q120_Reytech_CCHCS_BidPackage ──
    # RULE: This order is locked to the model. Do not reorder without a new sample.
    # BidPackage (pg1 CDCR Terms, pg2 CalRecycle) come first, then standalone forms
    # fill the gap (BidderDecl, Darfur), then rest of BidPackage (CUF, DVBE, GenAI, DrugFree, PD802).
    # BUT since we can't interleave pages from BidPackage with standalone files mid-merge,
    # the actual runtime order is determined by _FORM_ORDER key matching filename patterns.
    # BidPackage sorts first (position 0), then BidderDecl (2), DarfurAct (3), etc.
    # The BidPackage page filter already removed BidDecl from the template output.
    _FORM_ORDER = [
        "BidPackage",    # 0. CDCR Terms + CalRecycle + CUF + DVBE + GenAI + DrugFree + PD802
        "CalRecycle74",  # 1. Standalone CalRecycle (if present — normally inside BidPackage)
        "BidderDecl",    # 2. Standalone Bidder Declaration (ReportLab — cleaner signature)
        "DarfurAct",     # 3. Standalone Darfur Act
        "DVBE843",       # 4. Standalone DVBE (if present — normally inside BidPackage)
        "DrugFree",      # 5. Standalone Drug-Free (if present — normally inside BidPackage)
        "SellersPermit", # 6. CA Seller's Permit
        "STD204",        # 7. STD 204 Payee Data Record (if included)
        "CV012_CUF",     # 8. CUF CV 012 (CalVet only)
        "BarstowCUF",    # 9. Barstow CUF (conditional)
        "STD1000",       # 10. STD 1000 (non-CCHCS agencies)
        "STD205",        # 11. STD 205 Payee Supplement
    ]

    def _form_sort_key(filename):
        fn_upper = filename.upper().replace("-", "").replace("_", "").replace(" ", "")
        for idx, pattern in enumerate(_FORM_ORDER):
            if pattern.upper().replace("_", "") in fn_upper:
                return idx
        return len(_FORM_ORDER)

    package_pdfs.sort(key=lambda pair: _form_sort_key(pair[1]))
    
    # ── Step 4: Merge all package PDFs into ONE file ──
    final_output_files = []
    try:
        _safe_agency = (_agency_cfg.get("name", "") or "").replace(" ", "").replace("/", "")[:20]
    except NameError:
        _safe_agency = ""
    package_filename = f"RFQ_Package_{_safe_agency}_{safe_sol}_ReytechInc.pdf" if _safe_agency else f"RFQ_Package_{safe_sol}_ReytechInc.pdf"

    final_output_files = []

    # ── Attachment 1: 703B ──
    if file_703b:
        final_output_files.append(file_703b)
        t.step(f"Attachment 1 — 703B: {file_703b}")

    # ── Attachment 2: 704B ──
    if file_704b:
        final_output_files.append(file_704b)
        t.step(f"Attachment 2 — 704B: {file_704b}")

    # ── Attachment 3: Formal Quote ──
    if quote_file:
        final_output_files.append(quote_file)
        t.step(f"Attachment 3 — Quote: {quote_file}")

    # ── Attachment 4: RFQ Package (supporting compliance docs merged) ──
    if package_pdfs:
        try:
            from pypdf import PdfReader, PdfWriter
            writer = PdfWriter()
            merge_count = 0

            try:
                from src.forms.reytech_filler_v4 import _bidpkg_page_skip_reason as _skip_reason
            except Exception:
                _skip_reason = None

            for pdf_path, label in package_pdfs:
                try:
                    reader = PdfReader(pdf_path)
                    pages_added = 0

                    # ONLY apply bid-package page filter to the actual BidPackage PDF
                    # Standalone forms (BidderDecl, Darfur, etc.) must NOT be filtered
                    _is_bidpkg = "BidPackage" in label or "bidpkg" in label.lower() or "bid_package" in label.lower()

                    for page_idx, page in enumerate(reader.pages):
                        try:
                            text = page.extract_text() or ""
                        except Exception:
                            text = ""

                        # Skip XFA placeholder pages
                        if text.strip().startswith("Please wait") or (
                                "Please wait" in text and len(text.strip()) < 300):
                            continue

                        # Skip pages flagged by the bid-package page filter
                        # ONLY for the actual bid package — NOT standalone forms
                        if _is_bidpkg and _skip_reason is not None:
                            try:
                                skip = _skip_reason(page)
                            except Exception:
                                skip = None
                            if skip:
                                continue

                        # Do NOT touch /Rotate — form field appearance streams are
                        # positioned in the rotated coordinate system. Changing rotation
                        # here misplaces all field values. PDF viewers handle /Rotate natively.

                        writer.add_page(page)
                        pages_added += 1

                    if pages_added > 0:
                        merge_count += 1
                        t.step(f"  Package includes: {label} ({pages_added} pg)")
                    else:
                        t.step(f"  Package skipped: {label} (all pages filtered)")
                except Exception as _me:
                    t.warn(f"Could not merge {label}", error=str(_me))

            if merge_count > 0:
                merged_path = os.path.join(out_dir, package_filename)
                with open(merged_path, "wb") as _mf:
                    writer.write(_mf)
                final_output_files.append(package_filename)
                t.step(f"Attachment 4 — RFQ Package: {merge_count} docs → {package_filename}")
            else:
                t.warn("RFQ Package: no docs to merge")
        except Exception as _merge_err:
            t.warn("Package merge failed", error=str(_merge_err))
            final_output_files.extend([os.path.basename(p) for p, _ in package_pdfs])
    
    if not final_output_files:
        t.fail("No files generated", errors=errors)
        flash(f"No files generated — {'; '.join(errors) if errors else 'No templates found'}", "error")
        return redirect(f"/rfq/{rid}")
    
    # ── Create package manifest for audit trail ──
    _manifest_id = None
    try:
        from src.core.dal import create_package_manifest, log_lifecycle_event as _lle

        _gen_forms = []
        for _of in output_files:
            _fid = "unknown"
            _of_lower = _of.lower()
            if "quote" in _of_lower and "704" not in _of_lower: _fid = "quote"
            elif "703b" in _of_lower or "703c" in _of_lower: _fid = "703b"
            elif "704b" in _of_lower: _fid = "704b"
            elif "calrecycle" in _of_lower: _fid = "calrecycle74"
            elif "bidderdecl" in _of_lower or "bidder" in _of_lower: _fid = "bidder_decl"
            elif "dvbe" in _of_lower or "843" in _of_lower: _fid = "dvbe843"
            elif "darfur" in _of_lower: _fid = "darfur_act"
            elif "cuf" in _of_lower or "cv012" in _of_lower: _fid = "cv012_cuf"
            elif "std204" in _of_lower or "payee" in _of_lower: _fid = "std204"
            elif "std1000" in _of_lower: _fid = "std1000"
            elif "seller" in _of_lower or "permit" in _of_lower: _fid = "sellers_permit"
            elif "bidpkg" in _of_lower or "bidpackage" in _of_lower: _fid = "bidpkg"
            elif "obs" in _of_lower or "1600" in _of_lower: _fid = "obs_1600"
            elif "drug" in _of_lower: _fid = "drug_free"
            _gen_forms.append({"form_id": _fid, "filename": _of})

        _gen_ids = {f["form_id"] for f in _gen_forms}
        _missing = [f for f in _req_forms if f not in _gen_ids]

        # ── Source validation: cross-reference email ──
        _source_val = None
        try:
            from src.forms.price_check import validate_source_email
            _source_val = validate_source_email(r)
            if _source_val and not _source_val.get("ok"):
                _lle("rfq", rid, "source_validation_warning",
                     _source_val.get("summary", "Source validation issues"),
                     actor="system", detail=_source_val)
        except Exception as _sv_e:
            log.debug("Source validation: %s", _sv_e)

        # ── Field audit: verify each generated form ──
        _field_audits = {}
        try:
            from src.forms.price_check import audit_generated_form
            _expected = {"company_name": CONFIG.get("company", {}).get("name", "Reytech"), "solicitation": sol}
            for _gf in _gen_forms:
                _gf_path = os.path.join(out_dir, _gf.get("filename", ""))
                if os.path.exists(_gf_path):
                    _audit = audit_generated_form(_gf_path, _gf["form_id"], _expected)
                    _field_audits[_gf["form_id"]] = _audit
                    if not _audit.get("ok"):
                        log.warning("Field audit FAIL %s: %s", _gf["form_id"], _audit.get("errors", []))
        except Exception as _fa_e:
            log.debug("Field audit: %s", _fa_e)

        _qtotal = 0
        _qnum = r.get("reytech_quote_number", "")
        _icount = len(r.get("line_items", []))
        for _it in r.get("line_items", []):
            try:
                _p = float(_it.get("price_per_unit") or _it.get("unit_price") or 0)
                _q = int(float(_it.get("qty", 1)))
                _qtotal += _p * _q
            except (ValueError, TypeError):
                pass

        _items_snap = [{"desc": (_it.get("description") or "")[:60],
            "qty": _it.get("qty", 1),
            "price": _it.get("price_per_unit") or _it.get("unit_price") or 0,
            "cost": _it.get("vendor_cost") or _it.get("supplier_cost") or 0,
            } for _it in r.get("line_items", [])]

        _manifest_id = create_package_manifest(
            rfq_id=rid, agency_key=_agency_key,
            agency_name=_agency_cfg.get("name", ""),
            required_forms=list(_req_forms), generated_forms=_gen_forms,
            missing_forms=_missing, quote_number=_qnum,
            quote_total=round(_qtotal, 2), item_count=_icount, created_by="user",
            items_snapshot=_items_snap)

        if _missing:
            _lle("rfq", rid, "package_missing_forms",
                 f"Package missing {len(_missing)} required forms: {', '.join(_missing)}",
                 actor="system", detail={"missing": _missing, "generated": [f["form_id"] for f in _gen_forms]})

        _lle("rfq", rid, "package_generated",
             f"Generated {len(output_files)} forms, {len(_missing)} missing",
             actor="user", detail={
                 "manifest_id": _manifest_id, "forms": [f["form_id"] for f in _gen_forms],
                 "missing": _missing, "quote_number": _qnum,
                 "quote_total": round(_qtotal, 2), "errors": errors[:5] if errors else []})

        r["_manifest_id"] = _manifest_id
        r["_package_filename"] = package_filename
        if _manifest_id:
            from src.core.dal import update_manifest_status as _ums
            _ums(_manifest_id, "draft", package_filename=package_filename)
            # Store validation + audit on manifest
            if _source_val or _field_audits:
                try:
                    import json as _jm
                    from src.core.db import get_db as _gdb
                    with _gdb() as _conn:
                        _conn.execute("""UPDATE package_manifest SET source_validation = ?, field_audit = ? WHERE id = ?""",
                            (_jm.dumps(_source_val, default=str) if _source_val else None,
                             _jm.dumps(_field_audits, default=str) if _field_audits else None,
                             _manifest_id))
                except Exception:
                    pass
    except Exception as _me:
        log.error("MANIFEST CREATION FAILED: %s", _me, exc_info=True)
        errors.append(f"Manifest: {_me}")

    # ── Step 5: Store final files in DB (survive redeploys) ──
    for f in final_output_files:
        fpath = os.path.join(out_dir, f)
        try:
            if os.path.exists(fpath):
                with open(fpath, "rb") as _fb:
                    ftype = "generated_quote" if "Quote" in f else "generated_package"
                    save_rfq_file(rid, f, ftype, _fb.read(), category="generated", uploaded_by="user")
                    t.step(f"DB stored: {f}")
        except Exception as _de:
            t.warn(f"DB store failed: {f}", error=str(_de))
    
    # ── Step 6: Save, transition, create draft email ──
    _transition_status(r, "generated", actor="user", notes=f"Package: {len(final_output_files)} files")
    r["output_files"] = final_output_files

    # Learn which forms were used for this agency/buyer (improves future matching)
    try:
        from src.core.agency_config import learn_agency_forms
        learn_agency_forms(
            rid, _agency_key if '_agency_key' in dir() else r.get("agency", "unknown"),
            output_files,
            buyer_email=r.get("requestor_email", ""))
    except Exception:
        pass
    r["generated_at"] = datetime.now().isoformat()
    
    # ── Google Drive: upload package to Pending ──
    try:
        from src.agents.drive_triggers import on_package_generated
        on_package_generated(r, out_dir, final_output_files)
    except Exception as _gde:
        log.debug("Drive trigger (package_generated): %s", _gde)
    
    # Draft email with final files attached (quote + merged package)
    try:
        sender = EmailSender(CONFIG.get("email", {}))
        all_paths = [os.path.join(out_dir, f) for f in final_output_files]
        r["draft_email"] = sender.create_draft_email(r, all_paths)
        t.step("Draft email created", attachments=len(all_paths))
    except Exception as e:
        t.warn("Draft email failed", error=str(e))
    
    # Save SCPRS prices for history
    try:
        save_prices_from_rfq(r)
    except Exception as _e:
        log.debug("Suppressed: %s", _e)
    
    save_rfqs(rfqs)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, "generated")
    except Exception:
        pass

    # Build success message
    parts = []
    for f in final_output_files:
        if "Quote" in f: parts.append(f"Quote #{r.get('reytech_quote_number', '?')}")
        elif "Package" in f: parts.append(f"RFQ Package ({len(package_pdfs)} docs merged)")
        else: parts.append(os.path.basename(f))
    
    msg = f"✅ RFP Package ready: {', '.join(parts)}"
    if errors:
        msg += f" | ⚠️ {'; '.join(errors)}"
    
    # Log activity
    _log_rfq_activity(rid, "package_generated", msg, actor="user",
        metadata={"files": output_files, "quote_number": r.get("reytech_quote_number",""), "errors": errors})
    
    t.ok("Package complete", files=len(output_files), errors=len(errors))
    flash(msg, "success" if not errors else "info")

    # Clean up archived old files ONLY after successful generation
    if _old_dir and os.path.exists(_old_dir):
        try:
            _sh_clean.rmtree(_old_dir)
        except Exception:
            pass
    # Clean old DB files not in new output
    try:
        from src.core.db import get_db as _gdb_clean
        with _gdb_clean() as _conn_clean:
            if output_files:
                _ph = ",".join("?" for _ in output_files)
                _conn_clean.execute(f"DELETE FROM rfq_files WHERE rfq_id = ? AND category = 'generated' AND filename NOT IN ({_ph})", [rid] + list(output_files))
            else:
                _conn_clean.execute("DELETE FROM rfq_files WHERE rfq_id = ? AND category = 'generated'", (rid,))
    except Exception:
        pass

    return redirect(f"/rfq/{rid}/review-package")


@bp.route("/rfq/<rid>/generate", methods=["POST"])
@auth_required
def generate(rid):
    _bad = _validate_rid(rid)
    if _bad: return _bad
    log.info("Generate bid package for RFQ %s", rid)
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r: return redirect("/")
    
    # Update pricing from form
    for i, item in enumerate(r["line_items"]):
        for field, key in [("cost", "supplier_cost"), ("scprs", "scprs_last_price"), ("price", "price_per_unit"), ("markup", "markup_pct")]:
            v = request.form.get(f"{field}_{i}")
            if v:
                try: item[key] = float(v)
                except Exception as e:

                    log.debug("Suppressed: %s", e)
    
    r["sign_date"] = get_pst_date()
    sol = r["solicitation_number"]
    out = os.path.join(OUTPUT_DIR, sol)
    os.makedirs(out, exist_ok=True)
    
    try:
        t = r.get("templates", {})
        output_files = []
        
        if "703b" in t and os.path.exists(t["703b"]):
            fill_703b(t["703b"], r, CONFIG, f"{out}/{sol}_703B_Reytech.pdf")
            output_files.append(f"{sol}_703B_Reytech.pdf")
        
        if "704b" in t and os.path.exists(t["704b"]):
            fill_704b(t["704b"], r, CONFIG, f"{out}/{sol}_704B_Reytech.pdf")
            output_files.append(f"{sol}_704B_Reytech.pdf")
        
        if "bidpkg" in t and os.path.exists(t["bidpkg"]):
            fill_bid_package(t["bidpkg"], r, CONFIG, f"{out}/{sol}_BidPackage_Reytech.pdf")
            output_files.append(f"{sol}_BidPackage_Reytech.pdf")
        
        if not output_files:
            flash("No template PDFs found — upload the original RFQ PDFs first", "error")
            return redirect(f"/rfq/{rid}")
        
        _transition_status(r, "generated", actor="system", notes="Bid package filled")
        r["output_files"] = output_files
        r["generated_at"] = datetime.now().isoformat()
        
        # Note which forms are missing
        missing = []
        if "703b" not in t: missing.append("703B")
        if "704b" not in t: missing.append("704B")
        if "bidpkg" not in t: missing.append("Bid Package")
        
        # Create draft email
        sender = EmailSender(CONFIG.get("email", {}))
        output_paths = [f"{out}/{f}" for f in r["output_files"]]
        r["draft_email"] = sender.create_draft_email(r, output_paths)
        
        # Save SCPRS prices
        save_prices_from_rfq(r)
        
        save_rfqs(rfqs)
        try:
            from src.core.dal import update_rfq_status as _dal_ur
            _dal_ur(rid, "generated")
        except Exception:
            pass
        msg = f"Generated {len(output_files)} form(s) for #{sol}"
        if missing:
            msg += f" — missing: {', '.join(missing)}"
        else:
            msg += " — draft email ready"
        flash(msg, "success" if not missing else "info")
    except Exception as e:
        flash(f"Error: {e}", "error")
    
    return redirect(f"/rfq/{rid}")


@bp.route("/rfq/<rid>/generate-quote")
@auth_required
@safe_route
def rfq_generate_quote(rid):
    """Generate a standalone Reytech-branded quote PDF from an RFQ."""
    _bad = _validate_rid(rid)
    if _bad: return _bad
    from src.api.trace import Trace
    t = Trace("quote_generation", rfq_id=rid)
    
    if not QUOTE_GEN_AVAILABLE:
        t.fail("Quote generator not available")
        flash("Quote generator not available", "error")
        return redirect(f"/rfq/{rid}")
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        t.fail("RFQ not found")
        flash("RFQ not found", "error"); return redirect("/")

    # Validate before generating
    from src.core.quote_validator import validate_ready_to_generate
    validation = validate_ready_to_generate(r)
    if not validation["ok"]:
        t.fail("Validation failed", errors=validation["errors"])
        flash(f"Cannot generate: {'; '.join(validation['errors'])}", "error")
        return redirect(f"/rfq/{rid}")

    sol = r.get("solicitation_number", "unknown")
    t.step("Starting", sol=sol, items=len(r.get("line_items",[])))
    safe_sol = re.sub(r'[^a-zA-Z0-9_-]', '_', sol.strip())
    out_dir = os.path.join(OUTPUT_DIR, sol)
    os.makedirs(out_dir, exist_ok=True)
    output_path = os.path.join(out_dir, f"{safe_sol}_Quote_Reytech.pdf")

    locked_qn = r.get("reytech_quote_number", "")
    # SQLite doesn't store reytech_quote_number — check JSON directly
    if not locked_qn:
        try:
            import json as _jqn2
            _rfq_json_path2 = os.path.join(DATA_DIR, "rfqs.json")
            if os.path.exists(_rfq_json_path2):
                with open(_rfq_json_path2) as _jf2:
                    _jrfqs2 = _jqn2.load(_jf2)
                locked_qn = _jrfqs2.get(rid, {}).get("reytech_quote_number", "")
                if locked_qn:
                    r["reytech_quote_number"] = locked_qn
        except Exception:
            pass
    if not locked_qn:
        from src.forms.quote_generator import _next_quote_number
        locked_qn = _next_quote_number()
        r["reytech_quote_number"] = locked_qn
        save_rfqs(rfqs)  # persist NOW to prevent duplicates

    result = generate_quote_from_rfq(r, output_path,
                                      quote_number=locked_qn)

    if result.get("ok"):
        fname = os.path.basename(output_path)
        if "output_files" not in r:
            r["output_files"] = []
        if fname not in r["output_files"]:
            r["output_files"].append(fname)
        r["reytech_quote_number"] = result.get("quote_number", locked_qn)
        save_rfqs(rfqs)
        t.ok("Quote generated", quote_number=result.get("quote_number",""), total=result.get("total",0))
        log.info("Quote #%s generated for RFQ %s — $%s", result.get("quote_number"), rid, f"{result['total']:,.2f}")
        flash(f"Reytech Quote #{result['quote_number']} generated — ${result['total']:,.2f}", "success")
        _log_crm_activity(result.get("quote_number", ""), "quote_generated",
                          f"Quote {result.get('quote_number','')} generated from RFQ {sol} — ${result.get('total',0):,.2f}",
                          actor="user", metadata={"rfq_id": rid, "agency": result.get("agency","")})
    else:
        t.fail("Quote generation failed", error=result.get("error","unknown"))
        log.error("Quote generation failed for RFQ %s: %s", rid, result.get("error", "unknown"))
        flash(f"Quote generation failed: {result.get('error', 'unknown')}", "error")

    return redirect(f"/rfq/{rid}")

@bp.route("/rfq/<rid>/send", methods=["POST"])
@auth_required
def send_email(rid):
    _bad = _validate_rid(rid)
    if _bad: return _bad
    from src.api.trace import Trace
    t = Trace("email_send", rfq_id=rid)
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r or not r.get("draft_email"):
        t.fail("No draft to send")
        flash("No draft to send", "error"); return redirect(f"/rfq/{rid}")

    # Validate before sending
    from src.core.quote_validator import validate_ready_to_send
    validation = validate_ready_to_send(r)
    if not validation["ok"]:
        t.fail("Send validation failed", errors=validation["errors"])
        flash(f"Cannot send: {'; '.join(validation['errors'])}", "error")
        return redirect(f"/rfq/{rid}")

    try:
        sender = EmailSender(CONFIG.get("email", {}))
        sender.send(r["draft_email"])
        _transition_status(r, "sent", actor="user", notes="Email sent to buyer")
        r["sent_at"] = datetime.now().isoformat()
        save_rfqs(rfqs)
        try:
            from src.core.dal import update_rfq_status as _dal_ur
            _dal_ur(rid, "sent")
        except Exception:
            pass
        t.ok("Email sent", to=r["draft_email"].get("to",""), sol=r.get("solicitation_number","?"))
        flash(f"Bid response sent to {r['draft_email']['to']}", "success")
        _log_rfq_activity(rid, "email_sent",
            f"Bid response emailed to {r['draft_email'].get('to','')} for #{r.get('solicitation_number','?')}",
            actor="user", metadata={"to": r["draft_email"].get("to",""), "quote": r.get("reytech_quote_number","")})
        qn = r.get("reytech_quote_number", "")
        if qn and QUOTE_GEN_AVAILABLE:
            update_quote_status(qn, "sent", actor="system")
            _log_crm_activity(qn, "email_sent",
                              f"Quote {qn} emailed to {r['draft_email'].get('to','')}",
                              actor="user", metadata={"to": r['draft_email'].get('to','')})
        # Lifecycle event + package delivery record
        try:
            from src.core.dal import log_lifecycle_event as _lle_send, record_package_delivery, get_latest_manifest
            _to = r["draft_email"].get("to", "")
            _subj = r["draft_email"].get("subject", "")
            _lle_send("rfq", rid, "package_sent", f"Sent to {_to}: {_subj[:60]}", actor="user",
                detail={"recipient": _to, "subject": _subj})
            _mf = get_latest_manifest(rid)
            if _mf:
                record_package_delivery(_mf["id"], rid, _to,
                    recipient_name=r.get("requestor_name", ""), email_subject=_subj)
        except Exception:
            pass
    except Exception as e:
        t.fail("Send failed", error=str(e))
        flash(f"Send failed: {e}. Use 'Open in Mail App' instead.", "error")
    
    return redirect(f"/rfq/{rid}")


@bp.route("/api/quote/<qn>/regenerate", methods=["POST"])
@auth_required
def api_quote_regenerate(qn):
    """Regenerate the formal quote PDF for a given quote number.
    Finds the quote in quotes_log.json, regenerates PDF, and updates pdf_path.
    """
    from src.api.trace import Trace
    t = Trace("quote_regenerate", quote_number=qn)
    
    if not QUOTE_GEN_AVAILABLE:
        return jsonify({"ok": False, "error": "Quote generator not available"}), 503
    
    try:
        quotes = get_all_quotes()
        qt = None
        for q in quotes:
            if q.get("quote_number") == qn:
                qt = q
                break
        if not qt:
            return jsonify({"ok": False, "error": f"Quote {qn} not found"}), 404
        
        # Build output path
        rfq_num = qt.get("rfq_number", "") or qn
        safe_rfq = re.sub(r'[^a-zA-Z0-9_-]', '_', str(rfq_num).strip()) or qn
        out_dir = os.path.join(OUTPUT_DIR, safe_rfq)
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(out_dir, f"{safe_rfq}_Quote_Reytech.pdf")
        
        # Build quote_data from existing quote — map to generate_quote_from_rfq's expected fields
        quote_data = {
            "agency_name": qt.get("institution", ""),
            "institution": qt.get("institution", ""),
            "delivery_location": "",  # Not stored in quote, use ship_to_name instead
            "ship_to": "",
            "ship_to_name": qt.get("ship_to_name", qt.get("institution", "")),
            "ship_to_address": qt.get("ship_to_address", []),
            "agency": qt.get("agency", ""),
            "solicitation_number": rfq_num,
            "rfq_number": rfq_num,
            "requestor_email": qt.get("requestor_email", ""),
            "line_items": qt.get("items_detail", []),
            "source_pc_id": qt.get("source_pc_id", ""),
            "source_rfq_id": qt.get("source_rfq_id", ""),
        }
        
        result = generate_quote_from_rfq(
            quote_data, output_path,
            include_tax=True,
            quote_number=qn,
        )
        
        if result.get("ok"):
            # Update pdf_path in quotes_log.json
            quotes_path = os.path.join(DATA_DIR, "quotes_log.json")
            try:
                import json as _json
                all_quotes = _json.load(open(quotes_path))
                for i, q in enumerate(all_quotes):
                    if q.get("quote_number") == qn:
                        all_quotes[i]["pdf_path"] = output_path
                        break
                with open(quotes_path, "w") as f:
                    _json.dump(all_quotes, f, indent=2)
            except Exception as _e:
                t.warn(f"Could not update quotes_log: {_e}")
            
            # Store in DB for redeploy survival
            fname = os.path.basename(output_path)
            try:
                with open(output_path, "rb") as _fb:
                    save_rfq_file(qn, fname, "generated_quote", _fb.read(),
                                  category="generated", uploaded_by="user")
            except Exception:
                pass
            
            t.ok("Regenerated", path=output_path, total=result.get("total", 0))
            return jsonify({
                "ok": True,
                "quote_number": qn,
                "total": result.get("total", 0),
                "pdf_path": output_path,
                "download_url": f"/api/pricecheck/download/{fname}",
                "view_url": f"/api/pricecheck/view-pdf/{fname}",
            })
        else:
            t.fail(result.get("error", "unknown"))
            return jsonify({"ok": False, "error": result.get("error", "Generation failed")}), 500
    
    except Exception as e:
        import traceback
        t.fail(str(e))
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@bp.route("/api/rfq/<rid>/dismiss", methods=["POST"])
@auth_required
def api_rfq_dismiss(rid):
    """Dismiss an RFQ from the active queue with a reason.
    Keeps data for SCPRS intelligence. reason=delete does hard delete."""
    from datetime import datetime

    data = request.get_json(force=True) if request.data else {}
    reason = data.get("reason", "other")

    rfqs = load_rfqs()

    if rid not in rfqs:
        return jsonify({"ok": False, "error": "RFQ not found"})

    # Hard delete path
    if reason == "delete":
        sol = rfqs[rid].get("solicitation_number", "?")
        del rfqs[rid]
        save_rfqs(rfqs)
        # Direct SQLite delete — save_rfqs only does INSERT OR REPLACE
        try:
            from src.core.db import get_db
            with get_db() as conn:
                conn.execute("DELETE FROM rfqs WHERE id=?", (rid,))
        except Exception as e:
            log.debug("DAL RFQ delete: %s", e)
        log.info("Hard deleted RFQ #%s (id=%s)", sol, rid)
        return jsonify({"ok": True, "deleted": rid})

    r = rfqs[rid]
    r["status"] = "dismissed"
    r["dismiss_reason"] = reason
    r["dismissed_at"] = datetime.now().isoformat()
    rfqs[rid] = r
    save_rfqs(rfqs)
    
    # Also update SQLite status via DAL
    try:
        from src.core.dal import update_rfq_status as _dal_update_rfq
        _dal_update_rfq(rid, "dismissed")
    except Exception as e:
        log.debug("DAL RFQ dismiss update: %s", e)
    
    sol = r.get("solicitation_number", "?")
    log.info("RFQ #%s dismissed: reason=%s", sol, reason)
    
    # Queue SCPRS price intelligence on line items (async)
    scprs_queued = False
    items = r.get("line_items", [])
    if items:
        try:
            from src.agents.scprs_lookup import queue_background_lookup
            for item in items[:20]:
                desc = item.get("description", "")
                if desc and len(desc) > 3:
                    queue_background_lookup(desc, source=f"dismissed_rfq_{rid}")
            scprs_queued = True
        except Exception as _e:
            log.debug("Suppressed: %s", _e)
    
    return jsonify({
        "ok": True,
        "dismissed": rid,
        "solicitation": sol,
        "reason": reason,
        "scprs_queued": scprs_queued,
    })


@bp.route("/rfq/<rid>/delete", methods=["POST"])
@auth_required
@safe_route
def delete_rfq(rid):
    """Delete an RFQ from the queue and remove its UID from processed list."""
    _bad = _validate_rid(rid)
    if _bad: return _bad
    rfqs = load_rfqs()
    if rid in rfqs:
        sol = rfqs[rid].get("solicitation_number", "?")
        # Remove this email's UID from processed list so it can be re-imported
        email_uid = rfqs[rid].get("email_uid")
        if email_uid:
            _remove_processed_uid(email_uid)
        del rfqs[rid]
        save_rfqs(rfqs)
        # Direct SQLite delete — save_rfqs only does INSERT OR REPLACE, never DELETE
        try:
            from src.core.db import get_db
            with get_db() as conn:
                conn.execute("DELETE FROM rfqs WHERE id = ?", (rid,))
            log.info("Deleted RFQ %s from SQLite", rid)
        except Exception as _db_e:
            log.warning("SQLite delete for RFQ %s failed: %s", rid, _db_e)
        log.info("Deleted RFQ #%s (id=%s)", sol, rid)
        _log_rfq_activity(rid, "deleted", f"RFQ #{sol} deleted", actor="user")
        flash(f"Deleted RFQ #{sol}", "success")
    return redirect("/")


# ═══════════════════════════════════════════════════════════════════════
# RFQ File Management — download from DB
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/rfq/<rid>/file/<file_id>")
@auth_required
def rfq_download_file(rid, file_id):
    """Download an RFQ file from the database."""
    _bad = _validate_rid(rid)
    if _bad: return _bad
    f = get_rfq_file(file_id)
    if not f or f.get("rfq_id") != rid:
        flash("File not found", "error")
        return redirect(f"/rfq/{rid}")
    from flask import Response
    return Response(
        f["data"],
        mimetype=f.get("mime_type", "application/pdf"),
        headers={"Content-Disposition": f"inline; filename=\"{f['filename']}\""}
    )


@bp.route("/api/rfq/<rid>/files")
@auth_required
def api_rfq_files(rid):
    """List all files for an RFQ."""
    category = request.args.get("category")
    files = list_rfq_files(rid, category=category)
    return jsonify({"ok": True, "files": files, "count": len(files)})


# ═══════════════════════════════════════════════════════════════════════
# RFQ Status Management — reopen, edit, resubmit
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/rfq/<rid>/reopen", methods=["POST"])
@auth_required
@safe_route
def rfq_reopen(rid):
    """Reopen an RFQ for editing. Changes status back to 'ready'."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error")
        return redirect("/")
    
    old_status = r.get("status", "?")
    _transition_status(r, "ready", actor="user", notes=f"Reopened from '{old_status}'")
    save_rfqs(rfqs)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, "ready")
    except Exception:
        pass

    _log_rfq_activity(rid, "reopened",
        f"RFQ #{r.get('solicitation_number','?')} reopened for editing (was: {old_status})",
        actor="user", metadata={"old_status": old_status})
    
    flash(f"RFQ reopened for editing (was: {old_status})", "info")
    return redirect(f"/rfq/{rid}")


@bp.route("/api/rfq/<rid>/update-status", methods=["POST"])
@auth_required
def api_rfq_update_status_json(rid):
    """Update RFQ status via JSON (AJAX)."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    data = request.get_json(force=True, silent=True) or {}
    new_status = data.get("status", "").strip()
    notes = data.get("notes", "").strip()

    valid = {"new", "ready", "generated", "sent", "won", "lost", "no_bid", "cancelled"}
    if new_status not in valid:
        return jsonify({"ok": False, "error": f"Invalid status: {new_status}"})

    old_status = r.get("status", "?")
    r["status"] = new_status
    if notes:
        r["status_notes"] = notes
    save_rfqs(rfqs)

    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, new_status)
    except Exception:
        pass

    try:
        from src.core.dal import log_lifecycle_event
        log_lifecycle_event("rfq", rid, "status_changed",
            f"Status: {old_status} → {new_status}" + (f" ({notes})" if notes else ""),
            actor="user")
    except Exception:
        pass

    return jsonify({"ok": True, "old_status": old_status, "new_status": new_status})


@bp.route("/rfq/<rid>/update-status-form", methods=["POST"])
@auth_required
@safe_route
def rfq_update_status(rid):
    """Change RFQ status to any valid state."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error")
        return redirect("/")
    
    new_status = request.form.get("status", "").strip()
    valid = {"new", "ready", "generated", "sent", "won", "lost", "no_bid", "cancelled"}
    if new_status not in valid:
        flash(f"Invalid status: {new_status}", "error")
        return redirect(f"/rfq/{rid}")
    
    old_status = r.get("status", "?")
    notes = request.form.get("notes", "").strip()
    _transition_status(r, new_status, actor="user", notes=notes or f"Changed from {old_status}")
    save_rfqs(rfqs)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, new_status)
    except Exception:
        pass

    _log_rfq_activity(rid, "status_changed",
        f"RFQ #{r.get('solicitation_number','?')} status: {old_status} → {new_status}" + (f" ({notes})" if notes else ""),
        actor="user", metadata={"old_status": old_status, "new_status": new_status, "notes": notes})

    flash(f"Status changed: {old_status} → {new_status}", "success")
    return redirect(f"/rfq/{rid}")


# ═══════════════════════════════════════════════════════════════════════
# RFQ Activity Log
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/<rid>/activity")
@auth_required
def api_rfq_activity(rid):
    """Get activity log for an RFQ."""
    activities = _get_crm_activity(ref_id=rid, limit=50)
    return jsonify({"ok": True, "activities": activities, "count": len(activities)})


# ═══════════════════════════════════════════════════════════════════════
# Email Templates API
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/email-templates")
@auth_required
def api_list_email_templates():
    """List email templates, optionally filtered by category."""
    category = request.args.get("category")
    templates = get_email_templates_db(category)
    return jsonify({"ok": True, "templates": templates})


@bp.route("/api/email-templates/<tid>", methods=["GET"])
@auth_required
def api_get_email_template(tid):
    """Get a single email template by ID."""
    templates = get_email_templates_db()
    t = next((t for t in templates if t["id"] == tid), None)
    if not t:
        return jsonify({"ok": False, "error": "Template not found"}), 404
    return jsonify({"ok": True, "template": t})


@bp.route("/api/email-templates", methods=["POST"])
@auth_required
def api_create_email_template():
    """Create or update an email template."""
    data = request.get_json(force=True, silent=True) or request.form
    tid = save_email_template_db(
        data.get("id", ""), data.get("name", ""), data.get("category", "rfq"),
        data.get("subject", ""), data.get("body", ""), int(data.get("is_default", 0)))
    if tid:
        return jsonify({"ok": True, "id": tid})
    return jsonify({"ok": False, "error": "Save failed"}), 500


@bp.route("/api/email-templates/<tid>", methods=["DELETE"])
@auth_required
def api_delete_email_template(tid):
    """Delete an email template."""
    try:
        from src.core.db import get_db
        with get_db() as conn:
            conn.execute("DELETE FROM email_templates WHERE id = ?", (tid,))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/email-templates/render", methods=["POST"])
@auth_required
def api_render_email_template():
    """Render a template with variables. POST {template_id, variables: {...}}"""
    data = request.get_json(force=True, silent=True) or {}
    tid = data.get("template_id", "")
    variables = data.get("variables", {})
    
    templates = get_email_templates_db()
    t = next((t for t in templates if t["id"] == tid), None)
    if not t:
        return jsonify({"ok": False, "error": "Template not found"}), 404
    
    subject = t["subject"]
    body = t["body"]
    for key, val in variables.items():
        subject = subject.replace("{{" + key + "}}", str(val))
        body = body.replace("{{" + key + "}}", str(val))
    
    return jsonify({"ok": True, "subject": subject, "body": body})


# ═══════════════════════════════════════════════════════════════════════
# PDF Preview from DB
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/rfq/<rid>/preview/<file_id>")
@auth_required
def rfq_preview_pdf(rid, file_id):
    """Serve a PDF for inline preview (Content-Disposition: inline)."""
    f = get_rfq_file(file_id)
    if not f or f.get("rfq_id") != rid:
        return "File not found", 404
    from flask import Response
    return Response(
        f["data"],
        mimetype="application/pdf",
        headers={"Content-Disposition": f"inline; filename=\"{f['filename']}\""}
    )


# ═══════════════════════════════════════════════════════════════════════
# Email Signature — get/save HTML signature for outbound emails
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/email-signature")
@auth_required
def get_email_signature():
    """Get current email signature config."""
    email_cfg = CONFIG.get("email", {})
    sig_html = email_cfg.get("signature_html", "")

    # Auto-generate default signature on first load if empty
    if not sig_html:
        sig_html = _build_default_signature()
        CONFIG.setdefault("email", {})["signature_html"] = sig_html

    return jsonify({
        "ok": True,
        "signature_html": sig_html,
        "signature_enabled": email_cfg.get("signature_enabled", True),
    })

@bp.route("/api/email-signature", methods=["POST"])
@auth_required
def save_email_signature():
    """Save email signature HTML."""
    data = request.get_json(force=True)
    sig_html = data.get("signature_html", "")
    
    CONFIG.setdefault("email", {})["signature_html"] = sig_html
    CONFIG["email"]["signature_enabled"] = True
    
    # Persist to config file
    import json as _json
    for cfg_path in [
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "reytech_config.json"),
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "forms", "reytech_config.json"),
    ]:
        try:
            with open(cfg_path) as f:
                cfg = _json.load(f)
            cfg.setdefault("email", {})["signature_html"] = sig_html
            cfg["email"]["signature_enabled"] = True
            with open(cfg_path, "w") as f:
                _json.dump(cfg, f, indent=2)
        except Exception as _e:
            log.debug("Suppressed: %s", _e)
    
    return jsonify({"ok": True})


@bp.route("/api/upload-sig-logo", methods=["POST"])
@auth_required
def upload_sig_logo():
    """Upload a PNG/JPG logo for the email signature. Returns base64 data URI."""
    import base64 as _b64
    if "logo" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400
    f = request.files["logo"]
    if not f.filename:
        return jsonify({"ok": False, "error": "Empty filename"}), 400

    data = f.read()
    if len(data) > 5_000_000:
        return jsonify({"ok": False, "error": "File too large (max 5MB)"}), 400

    fname = f.filename.lower()
    if fname.endswith(".png"):
        mime = "image/png"
    elif fname.endswith((".jpg", ".jpeg")):
        mime = "image/jpeg"
    elif fname.endswith(".gif"):
        mime = "image/gif"
    else:
        return jsonify({"ok": False, "error": "PNG/JPG/GIF only"}), 400

    # Resize for email if large
    try:
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(data))
        if img.width > 200:
            ratio = 200 / img.width
            img = img.resize((200, int(img.height * ratio)), Image.LANCZOS)
            buf = _io.BytesIO()
            img.save(buf, "PNG", optimize=True)
            data = buf.getvalue()
            mime = "image/png"
    except Exception:
        pass

    b64 = _b64.b64encode(data).decode()
    data_uri = f"data:{mime};base64,{b64}"

    # Save to data/ for future use
    try:
        save_path = os.path.join(DATA_DIR, "email_logo.png")
        with open(save_path, "wb") as _fw:
            _fw.write(data)
    except Exception:
        pass

    return jsonify({"ok": True, "data_uri": data_uri, "size": len(data)})


def _build_default_signature():
    """Build the default Reytech email signature HTML for compose section."""
    return """<table cellpadding="0" cellspacing="0" style="font-family:'Segoe UI',Arial,sans-serif;margin-top:12px">
 <tr>
  <td style="padding-right:14px;vertical-align:top"><img src="cid:reytech_logo" alt="Reytech Inc." style="width:120px;height:auto;display:block"></td>
  <td style="vertical-align:top;font-size:13px;color:#444;line-height:1.5">
   <strong style="font-size:14px;color:#1a1a2e">Michael Guadan</strong><br>
   <strong>Reytech Inc.</strong><br>
   <a href="tel:9498728676" style="color:#2563eb;text-decoration:none">(949) 872-8676</a><br>
   <a href="mailto:mike@reytechinc.com" style="color:#2563eb;text-decoration:none">mike@reytechinc.com</a><br>
   DVBE / Small Business Certified<br>
   <a href="https://www.reytechinc.com" style="color:#2563eb;text-decoration:none">www.reytechinc.com</a>
  </td>
 </tr>
</table>
<div style="font-size:11px;color:#999;margin-top:8px;line-height:1.4">
CA MB/SB/SB-PW/DVBE #2002605 &middot; NY SDVOB 221449<br>
DOT DBE #44511 &middot; MBE SC6550 &middot; SBA-SDVOB (FWWSKE9113T7)
</div>"""


# ═══════════════════════════════════════════════════════════════════════
# Enhanced Email Send — DB attachments + email logging + CRM tracking
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/rfq/<rid>/save-draft", methods=["POST"])
@auth_required
def save_gmail_draft(rid):
    """Save email as Gmail draft — user reviews and sends manually from Gmail."""
    from src.api.trace import Trace
    t = Trace("email_draft", rfq_id=rid)

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error")
        return redirect("/")

    to_addr = request.form.get("to", "").strip()
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    cc = request.form.get("cc", "").strip()
    attach_ids = [x.strip() for x in request.form.get("attach_files", "").split(",") if x.strip()]

    if not to_addr or not subject:
        flash("Draft requires To and Subject", "error")
        return redirect(f"/rfq/{rid}")

    import tempfile, shutil, imaplib, time as _time
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    tmp_dir = tempfile.mkdtemp(prefix="rfq_draft_")
    try:
        # Build the MIME message
        msg = MIMEMultipart("mixed")
        email_cfg = CONFIG.get("email", {})
        from_name = email_cfg.get("from_name", "Michael Guadan - Reytech Inc.")
        from_addr = email_cfg.get("email", os.environ.get("GMAIL_ADDRESS", "sales@reytechinc.com"))
        password = email_cfg.get("email_password", os.environ.get("GMAIL_PASSWORD", ""))

        msg["From"] = f"{from_name} <{from_addr}>"
        msg["To"] = to_addr
        msg["Subject"] = subject
        if cc:
            msg["Cc"] = cc

        # HTML body with signature
        try:
            from src.core.email_signature import wrap_html_email
            body_html = wrap_html_email(body)
        except Exception:
            body_html = None

        if body_html:
            related = MIMEMultipart("related")
            related.attach(MIMEText(body_html, "html"))
            # Embed logo as CID inline attachment
            try:
                from src.core.paths import DATA_DIR as _dd2
                for _ln in ("reytech_logo_email.png", "email_logo.png", "reytech_logo.png", "logo.png"):
                    _lp = os.path.join(_dd2, _ln)
                    if os.path.exists(_lp):
                        from email.mime.image import MIMEImage
                        with open(_lp, "rb") as _lf2:
                            _lip = MIMEImage(_lf2.read(), _subtype="png")
                        _lip.add_header("Content-ID", "<reytech_logo>")
                        _lip.add_header("Content-Disposition", "inline", filename="reytech_logo.png")
                        related.attach(_lip)
                        break
            except Exception:
                pass
            alt = MIMEMultipart("alternative")
            alt.attach(MIMEText(body, "plain"))
            alt.attach(related)
            msg.attach(alt)
        else:
            msg.attach(MIMEText(body, "plain"))

        # Attach files
        attached = []
        for fid in attach_ids:
            f = get_rfq_file(fid)
            if f and f.get("data"):
                path = os.path.join(tmp_dir, f["filename"])
                with open(path, "wb") as _fw:
                    _fw.write(f["data"])
                with open(path, "rb") as _fr:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(_fr.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={f['filename']}")
                msg.attach(part)
                attached.append(f["filename"])

        # Save to Gmail Drafts via IMAP APPEND
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(from_addr, password)

        saved = False
        for folder in ['"[Gmail]/Drafts"', "[Gmail]/Drafts", "Drafts", "DRAFTS"]:
            try:
                res = imap.append(folder, "", imaplib.Time2Internaldate(_time.time()), msg.as_bytes())
                if res[0] == "OK":
                    saved = True
                    t.ok("Draft saved", folder=folder, attachments=len(attached))
                    break
            except Exception as _fe:
                log.debug("IMAP draft append %s: %s", folder, _fe)

        if not saved:
            # Auto-detect Drafts folder
            _, folders = imap.list()
            import re as _re
            for _raw in (folders or []):
                _s = _raw.decode() if isinstance(_raw, bytes) else str(_raw)
                if "draft" in _s.lower():
                    _m = _re.search(r'"([^"]+)"\s*$', _s) or _re.search(r'(\S+)$', _s)
                    if _m:
                        try:
                            res = imap.append(_m.group(1), "", imaplib.Time2Internaldate(_time.time()), msg.as_bytes())
                            if res[0] == "OK":
                                saved = True
                                t.ok("Draft saved", folder=_m.group(1))
                                break
                        except Exception:
                            pass

        imap.logout()

        if saved:
            flash(f"✅ Draft saved to Gmail — open Gmail to review and send ({len(attached)} attachments)", "success")
        else:
            flash("⚠️ Could not save to Gmail Drafts — check IMAP is enabled in Gmail settings", "error")

    except Exception as e:
        t.fail("Draft save failed", error=str(e))
        flash(f"Draft save failed: {e}", "error")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    return redirect(f"/rfq/{rid}")


@bp.route("/rfq/<rid>/send-email", methods=["POST"])
@auth_required
@safe_route
def send_email_enhanced(rid):
    """Send email with editable fields and DB-stored attachments.
    Form fields: to, subject, body, attach_files (comma-separated file IDs)
    """
    from src.api.trace import Trace
    t = Trace("email_send", rfq_id=rid)

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        t.fail("RFQ not found")
        flash("RFQ not found", "error")
        return redirect("/")

    # Validate before sending
    from src.core.quote_validator import validate_ready_to_send
    validation = validate_ready_to_send(r)
    if not validation["ok"]:
        t.fail("Send validation failed", errors=validation["errors"])
        flash(f"Cannot send: {'; '.join(validation['errors'])}", "error")
        return redirect(f"/rfq/{rid}")

    # Get editable fields from form
    to_addr = request.form.get("to", "").strip()
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    cc = request.form.get("cc", "").strip()
    bcc = request.form.get("bcc", "").strip()
    attach_ids = [x.strip() for x in request.form.get("attach_files", "").split(",") if x.strip()]
    
    if not to_addr or not subject:
        flash("Email requires To and Subject", "error")
        return redirect(f"/rfq/{rid}")
    
    t.step("Preparing email", to=to_addr, attachments=len(attach_ids))
    
    # Build attachment list from DB files
    import tempfile, shutil
    tmp_dir = tempfile.mkdtemp(prefix="rfq_send_")
    attachment_paths = []
    attachment_names = []
    
    try:
        for fid in attach_ids:
            f = get_rfq_file(fid)
            if f and f.get("data"):
                path = os.path.join(tmp_dir, f["filename"])
                with open(path, "wb") as _fw:
                    _fw.write(f["data"])
                attachment_paths.append(path)
                attachment_names.append(f["filename"])
                t.step(f"Attached: {f['filename']}")
        
        # Also check filesystem for any files not in DB yet
        if not attach_ids and r.get("output_files"):
            out_dir = os.path.join(UPLOAD_DIR, rid)
            for fname in r["output_files"]:
                fpath = os.path.join(out_dir, fname)
                if os.path.exists(fpath):
                    attachment_paths.append(fpath)
                    attachment_names.append(fname)
        
        # Send via SMTP — include HTML signature if enabled
        draft = {
            "to": to_addr,
            "subject": subject,
            "body": body,
            "cc": cc,
            "bcc": bcc,
            "attachments": attachment_paths,
        }
        
        # Threading: if RFQ came from email, reply to that thread
        msg_id = r.get("email_message_id", "")
        if msg_id:
            draft["in_reply_to"] = msg_id
            draft["references"] = msg_id
        
        include_sig = request.form.get("include_signature") == "1"
        email_cfg = CONFIG.get("email", {})
        sig_html = email_cfg.get("signature_html", "")
        
        if include_sig and sig_html:
            # Build HTML body: plain text body + signature
            import html as _html
            body_escaped = _html.escape(body).replace("\n", "<br>")
            draft["body_html"] = f"""<div style="font-family:'Segoe UI',Arial,sans-serif;font-size:14px;color:#222;line-height:1.6">
{body_escaped}
<br><br>
<div style="border-top:1px solid #ddd;padding-top:10px;margin-top:10px">
{sig_html}
</div>
</div>"""
            t.step("HTML signature included")
        
        sender = EmailSender(CONFIG.get("email", {}))
        sender.send(draft)
        
        # Transition status
        _transition_status(r, "sent", actor="user", notes=f"Email sent to {to_addr}")
        r["sent_at"] = datetime.now().isoformat()
        r["draft_email"] = {"to": to_addr, "subject": subject, "body": body, "cc": cc, "bcc": bcc}
        save_rfqs(rfqs)
        try:
            from src.core.dal import update_rfq_status as _dal_ur
            _dal_ur(rid, "sent")
        except Exception:
            pass
        
        # ── Log to email_log table ──
        sol = r.get("solicitation_number", "")
        qn = r.get("reytech_quote_number", "")
        
        # Find contact_id from recipient email
        contact_id = ""
        try:
            from src.core.db import get_db
            with get_db() as conn:
                row = conn.execute("SELECT id FROM contacts WHERE buyer_email = ?", (to_addr.lower(),)).fetchone()
                if row:
                    contact_id = row[0]
        except Exception as _e:
            log.debug("Suppressed: %s", _e)
        
        email_log_id = log_email_sent_db(
            direction="outbound", sender=sender.email_addr, recipient=to_addr,
            subject=subject, body=body, attachments=attachment_names,
            quote_number=qn, rfq_id=rid, contact_id=contact_id)
        t.step(f"Email logged (id={email_log_id})")
        
        # ── CRM activity: log against quote AND contact ──
        _log_rfq_activity(rid, "email_sent",
            f"Bid response emailed to {to_addr} for Sol #{sol} ({len(attachment_names)} attachments)",
            actor="user", metadata={"to": to_addr, "quote": qn, "files": attachment_names, "email_log_id": email_log_id})
        
        if qn:
            _log_crm_activity(qn, "email_sent",
                f"Quote {qn} emailed to {to_addr} for Sol #{sol}",
                actor="user", metadata={"to": to_addr, "rfq_id": rid})
            if QUOTE_GEN_AVAILABLE:
                update_quote_status(qn, "sent", actor="system")
        
        if contact_id:
            _log_crm_activity(contact_id, "email_sent",
                f"Bid response for Sol #{sol} (Quote {qn}) sent to {to_addr}",
                actor="user", metadata={"rfq_id": rid, "quote": qn, "solicitation": sol})
        
        t.ok("Email sent", to=to_addr, attachments=len(attachment_names))
        flash(f"✅ Email sent to {to_addr} with {len(attachment_names)} attachments", "success")
        
    except Exception as e:
        t.fail("Send failed", error=str(e))
        flash(f"Send failed: {e}. Try 'Open in Mail App' instead.", "error")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    
    return redirect(f"/rfq/{rid}")


# ═══════════════════════════════════════════════════════════════════════
# Email History API (for contact/quote level)
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/email-history")
@auth_required
def api_email_history():
    """Get email history. Filter by ?rfq_id=, ?quote_number=, ?contact_id="""
    try:
        from src.core.db import get_db
        with get_db() as conn:
            query = "SELECT id, logged_at, direction, sender, recipient, subject, body_preview, attachments_json, quote_number, rfq_id, contact_id, status FROM email_log WHERE 1=1"
            params = []
            for field in ("rfq_id", "quote_number", "contact_id"):
                val = request.args.get(field)
                if val:
                    query += f" AND {field} = ?"
                    params.append(val)
            query += " ORDER BY logged_at DESC LIMIT 50"
            rows = conn.execute(query, params).fetchall()
            return jsonify({"ok": True, "emails": [dict(r) for r in rows], "count": len(rows)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ═══════════════════════════════════════════════════════════════════════
# OBS 1600 — CA Agricultural Food Product Certification
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/food-classify", methods=["POST"])
@auth_required
def api_food_classify():
    """Classify quote/RFQ items into CDCR food category codes.
    Body: {"items": [{"description": "..."}, ...]}
    Returns classified items with food codes.
    """
    try:
        from src.forms.food_classifier import classify_quote_items
        data = request.get_json(force=True)
        items = data.get("items", [])
        classified = classify_quote_items(items)
        food_count = sum(1 for r in classified if r['is_food'])
        return jsonify({"ok": True, "items": classified, "food_count": food_count,
                        "total_count": len(classified)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/rfq/<rid>/obs1600", methods=["POST"])
@auth_required
def api_generate_obs1600(rid):
    """Generate filled OBS 1600 food certification form for an RFQ.
    Uses the bid package PDF if available, or a standalone template.
    """
    from src.api.trace import Trace
    t = Trace("obs1600_fill", rfq_id=rid)
    
    try:
        from src.forms.reytech_filler_v4 import load_config, fill_obs1600, get_pst_date
        from src.forms.food_classifier import get_food_items_for_obs1600
        
        rfqs = load_rfqs()
        r = rfqs.get(rid)
        if not r:
            t.fail("RFQ not found")
            return jsonify({"ok": False, "error": "RFQ not found"}), 404
        
        config = load_config()
        sol = r.get("solicitation_number", "unknown")
        
        # Get items — try line_items first, then items_detail from quote
        items = r.get("line_items", [])
        if not items:
            items = r.get("items_detail", r.get("items", []))
            if isinstance(items, str):
                import json as _json
                try: items = _json.loads(items)
                except Exception: items = []
        
        # Classify food items
        food_items = get_food_items_for_obs1600(items)
        
        if not food_items:
            t.step("No food items found", item_count=len(items))
            return jsonify({"ok": False, "error": "No food items found in this RFQ. Only food products need the OBS 1600 form.",
                            "items_checked": len(items)}), 400
        
        t.step("Classified food items", food_count=len(food_items),
               items=[f"{fi['description'][:40]} → Code {fi['code']}" for fi in food_items[:5]])
        
        # Output directory
        out_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "output", sol)
        os.makedirs(out_dir, exist_ok=True)
        
        # Find bid package template — use RFQ's stored template paths
        bid_pkg = None
        tmpl = r.get("templates", {})
        
        # Check bid package template from RFQ data
        if tmpl.get("bidpkg") and os.path.exists(tmpl["bidpkg"]):
            bid_pkg = tmpl["bidpkg"]
        
        # Try to restore from DB if not on disk
        if not bid_pkg:
            try:
                from src.core.db import get_db
                with get_db() as conn:
                    db_files = conn.execute(
                        "SELECT id, filename, file_type FROM rfq_files WHERE rfq_id=? AND category='template'",
                        (rid,)).fetchall()
                    for db_f in db_files:
                        fname = db_f["filename"].lower()
                        if "bid" in fname or "package" in fname or "form" in fname:
                            full_f_row = conn.execute("SELECT data FROM rfq_files WHERE id=?", (db_f["id"],)).fetchone()
                            if full_f_row and full_f_row["data"]:
                                restore_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "rfq_templates", rid)
                                os.makedirs(restore_dir, exist_ok=True)
                                restore_path = os.path.join(restore_dir, db_f["filename"])
                                with open(restore_path, "wb") as _fw:
                                    _fw.write(full_f_row["data"])
                                bid_pkg = restore_path
                                t.step(f"Restored bid package from DB: {db_f['filename']}")
                                break
            except Exception as db_err:
                t.step(f"DB restore failed: {db_err}")
        
        # Check uploaded files directory
        if not bid_pkg:
            import glob
            for pattern in [f"*{sol}*BID*PACKAGE*", f"*{sol}*bid*pack*", f"*{sol}*form*", f"*{sol}*.pdf"]:
                for search_dir in [os.path.join(DATA_DIR, "uploads"), os.path.join(DATA_DIR, "rfq_templates"), os.path.join(DATA_DIR, "output", sol)]:
                    matches = glob.glob(os.path.join(search_dir, pattern))
                    for m in matches:
                        # Verify it has OBS 1600 fields
                        try:
                            from pypdf import PdfReader
                            _r = PdfReader(m)
                            for page in _r.pages:
                                if "/Annots" in page:
                                    for annot in page["/Annots"]:
                                        obj = annot.get_object()
                                        if "OBS 1600" in str(obj.get("/T", "")):
                                            bid_pkg = m
                                            break
                                if bid_pkg: break
                        except Exception:
                            pass
                        if bid_pkg: break
                    if bid_pkg: break
                if bid_pkg: break
        
        # Fallback: use saved CDCR bid package template
        if not bid_pkg:
            default_tmpl = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "templates", "cdcr_bid_package_template.pdf")
            if os.path.exists(default_tmpl):
                bid_pkg = default_tmpl
                t.step("Using saved CDCR bid package template")
        
        # Build rfq_data for the filler
        rfq_data = {
            "solicitation_number": sol,
            "sign_date": get_pst_date(),
            "line_items": items,
        }
        
        output_path = os.path.join(out_dir, f"{sol}_OBS1600_FoodCert_Reytech.pdf")
        
        if bid_pkg and os.path.exists(bid_pkg):
            # Fill OBS 1600 fields in the existing bid package
            fill_obs1600(bid_pkg, rfq_data, config, output_path, food_items=food_items)
            t.ok(f"Filled from bid package template: {os.path.basename(bid_pkg)}")
        else:
            # Generate standalone OBS 1600 using reportlab
            _generate_standalone_obs1600(food_items, config, rfq_data, output_path)
            t.ok("Generated standalone OBS 1600")
        
        return jsonify({
            "ok": True,
            "file": output_path,
            "filename": os.path.basename(output_path),
            "food_items": food_items,
            "food_count": len(food_items),
            "download_url": f"/api/download/{sol}/{os.path.basename(output_path)}",
        })
        
    except Exception as e:
        import traceback
        t.fail(str(e))
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


def _generate_standalone_obs1600(food_items, config, rfq_data, output_path):
    """Generate a standalone OBS 1600 PDF using reportlab when no template is available."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.units import inch
    
    company = config["company"]
    sol = rfq_data.get("solicitation_number", "")
    sign_date = rfq_data.get("sign_date", "")
    
    c = rl_canvas.Canvas(output_path, pagesize=letter)
    w, h = letter
    
    # Header
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.5*inch, h - 0.5*inch, "California Department of Corrections and Rehabilitation/California Correctional Health Care Services")
    c.setFont("Helvetica", 8)
    c.drawString(0.5*inch, h - 0.65*inch, "Office of Business Services - Non-IT Goods Procurement/Acquisitions Management Section, Procurement Services")
    c.drawString(0.5*inch, h - 0.8*inch, "OBS 1600 (Rev. 1/26)")
    
    # Title
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(w/2, h - 1.2*inch, "California-Grown/Produced Agricultural Food Products Vendor Certification")
    
    # Vendor info
    y = h - 1.6*inch
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.5*inch, y, f"Vendor Name : {company['name']}")
    y -= 0.25*inch
    c.drawString(0.5*inch, y, f"Solicitation # : {sol}")
    
    # Table header
    y -= 0.45*inch
    c.setFont("Helvetica-Bold", 8)
    col_x = [0.5*inch, 1.2*inch, 4.5*inch, 5.2*inch, 6.2*inch]
    headers = ["Quoted Line\nItem #", "Food Product Description", "Code", "CA-Grown\nor Produced\n(Yes/No)", "If Yes, % of\nProduct"]
    
    # Header row background
    c.setFillColorRGB(0.9, 0.9, 0.9)
    c.rect(0.5*inch, y - 0.15*inch, 7*inch, 0.45*inch, fill=True, stroke=True)
    c.setFillColorRGB(0, 0, 0)
    
    for i, hdr in enumerate(headers):
        lines = hdr.split("\n")
        for j, line in enumerate(lines):
            c.drawString(col_x[i] + 0.05*inch, y + 0.2*inch - j*0.12*inch, line)
    
    # Data rows
    y -= 0.35*inch
    c.setFont("Helvetica", 9)
    for item in food_items[:18]:
        y -= 0.28*inch
        c.line(0.5*inch, y - 0.05*inch, 7.5*inch, y - 0.05*inch)
        c.drawString(col_x[0] + 0.1*inch, y + 0.05*inch, str(item.get("line_number", "")))
        c.drawString(col_x[1] + 0.05*inch, y + 0.05*inch, item.get("description", "")[:55])
        c.drawCentredString(col_x[2] + 0.35*inch, y + 0.05*inch, str(item.get("code", "")))
        c.drawCentredString(col_x[3] + 0.4*inch, y + 0.05*inch, item.get("ca_grown", "No"))
        c.drawCentredString(col_x[4] + 0.4*inch, y + 0.05*inch, item.get("pct", "N/A"))
    
    # Fill remaining empty rows
    for _ in range(18 - len(food_items)):
        y -= 0.28*inch
        c.line(0.5*inch, y - 0.05*inch, 7.5*inch, y - 0.05*inch)
    
    # Table border
    table_top = h - 2.1*inch
    c.rect(0.5*inch, y - 0.05*inch, 7*inch, table_top - (y - 0.05*inch))
    
    # Certification text
    y -= 0.45*inch
    c.setFont("Helvetica", 7.5)
    c.drawString(0.5*inch, y, "Pursuant to California Code, Food and Agricultural Code, Section 58595(a), I certify under the laws of the State of California")
    y -= 0.15*inch
    c.drawString(0.5*inch, y, "that the above information is true and correct.")
    
    # Signature block
    y -= 0.4*inch
    c.setFont("Helvetica", 10)
    c.drawString(0.5*inch, y, company["owner"])
    c.drawString(3.2*inch, y, company["title"])
    c.drawString(5.5*inch, y, sign_date)
    
    y -= 0.15*inch
    c.line(0.5*inch, y, 2.8*inch, y)
    c.line(3.2*inch, y, 4.8*inch, y)
    c.line(5.5*inch, y, 7.5*inch, y)
    
    y -= 0.15*inch
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.5*inch, y, "Print Name")
    c.drawString(2.2*inch, y, "Signature")
    c.drawString(3.2*inch, y, "Title")
    c.drawString(5.5*inch, y, "Date")
    
    c.save()


@bp.route("/api/download/<sol>/<filename>")
@auth_required
def api_download_file(sol, filename):
    """Download a generated file."""
    import re as _re
    # Sanitize inputs
    sol = _re.sub(r'[^a-zA-Z0-9_-]', '', sol)
    filename = os.path.basename(filename)
    filepath = os.path.join(OUTPUT_DIR, sol, filename)
    # Backwards compat: Compliance_Forms_ → RFQ_Package_ rename
    if not os.path.exists(filepath) and filename.startswith("Compliance_Forms_"):
        _old_name = filename.replace("Compliance_Forms_", "RFQ_Package_", 1)
        _old_path = os.path.join(OUTPUT_DIR, sol, _old_name)
        if os.path.exists(_old_path):
            filepath = _old_path
            filename = _old_name
    if not os.path.exists(filepath):
        # Fallback: try serving from DB
        # sol might be solicitation number OR rfq_id — try both
        try:
            found_file = None
            # Try sol as rfq_id first
            files = list_rfq_files(sol, category="generated")
            for dbf in files:
                if dbf.get("filename") == filename:
                    found_file = dbf
                    break
            # If not found, search all RFQs by solicitation number
            if not found_file:
                rfqs = load_rfqs()
                for rid, r in rfqs.items():
                    r_sol = (r.get("solicitation_number") or r.get("rfq_number") or "").replace("/", "_")
                    if r_sol == sol or rid == sol:
                        files = list_rfq_files(rid, category="generated")
                        for dbf in files:
                            if dbf.get("filename") == filename:
                                found_file = dbf
                                break
                        if found_file:
                            break
            if found_file:
                full = get_rfq_file(found_file["id"])
                if full and full.get("data"):
                    from flask import Response
                    return Response(full["data"], mimetype="application/pdf",
                                    headers={"Content-Disposition": f'inline; filename="{filename}"'})
        except Exception as _e:
            log.warning("DB download fallback failed for %s/%s: %s", sol, filename, _e)
        return jsonify({"ok": False, "error": "File not found"}), 404
    from flask import send_file
    return send_file(filepath, mimetype="application/pdf", download_name=filename)


# ═══════════════════════════════════════════════════════════════════════
# Fill ALL Bid Package Forms (CUF, Darfur, DVBE, CalRecycle, OBS 1600, etc.)
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/<rid>/fill-bid-package", methods=["POST"])
@auth_required
def api_fill_bid_package(rid):
    """Fill ALL forms in the CDCR bid package for an RFQ."""
    from src.api.trace import Trace
    t = Trace("fill_bid_package", rfq_id=rid)
    
    try:
        from src.forms.reytech_filler_v4 import load_config, fill_bid_package, get_pst_date
        
        rfqs = load_rfqs()
        r = rfqs.get(rid)
        if not r:
            t.fail("RFQ not found")
            return jsonify({"ok": False, "error": "RFQ not found"}), 404
        
        config = load_config()
        sol = r.get("solicitation_number", "unknown")
        
        # Get items
        items = r.get("line_items", [])
        if not items:
            items = r.get("items_detail", r.get("items", []))
            if isinstance(items, str):
                import json as _json
                try: items = _json.loads(items)
                except Exception: items = []
        
        # Find template
        bid_pkg = None
        tmpl = r.get("templates", {})
        if tmpl.get("bidpkg") and os.path.exists(tmpl["bidpkg"]):
            bid_pkg = tmpl["bidpkg"]
        
        # Fallback to saved template
        if not bid_pkg:
            default_tmpl = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "templates", "cdcr_bid_package_template.pdf")
            if os.path.exists(default_tmpl):
                bid_pkg = default_tmpl
        
        if not bid_pkg:
            t.fail("No bid package template found")
            return jsonify({"ok": False, "error": "No bid package template found. Upload one at /form-filler or place in data/templates/"}), 400
        
        out_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "output", sol)
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(out_dir, f"{sol}_BidPackage_Reytech.pdf")
        
        rfq_data = {
            "solicitation_number": sol,
            "sign_date": get_pst_date(),
            "line_items": items,
        }
        
        fill_bid_package(bid_pkg, rfq_data, config, output_path)
        
        # Count food items
        from src.forms.food_classifier import get_food_items_for_obs1600
        food_items = get_food_items_for_obs1600(items)
        
        t.ok(f"Filled bid package: {len(items)} items, {len(food_items)} food items")
        
        return jsonify({
            "ok": True,
            "filename": os.path.basename(output_path),
            "download_url": f"/api/download/{sol}/{os.path.basename(output_path)}",
            "total_items": len(items),
            "food_items": len(food_items),
            "forms_filled": ["CUF", "Darfur", "Bidder Declaration", "DVBE", "Drug-Free", "CalRecycle", "OBS 1600"],
        })
    except Exception as e:
        import traceback
        t.fail(str(e))
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@bp.route("/api/fill-forms", methods=["POST"])
@auth_required
def api_fill_forms_standalone():
    """Standalone form filler — fill bid package from manually entered items.
    Body: {
        "solicitation_number": "...",
        "items": [{"line_number": 1, "description": "..."}],
        "fill_type": "all" | "obs1600_only"
    }
    """
    from src.api.trace import Trace
    t = Trace("fill_forms_standalone")
    
    try:
        from src.forms.reytech_filler_v4 import load_config, fill_bid_package, fill_obs1600, get_pst_date
        from src.forms.food_classifier import get_food_items_for_obs1600
        
        data = request.get_json(force=True)
        sol = data.get("solicitation_number", "STANDALONE")
        items = data.get("items", [])
        fill_type = data.get("fill_type", "all")
        
        if not items:
            return jsonify({"ok": False, "error": "No items provided"}), 400
        
        config = load_config()
        
        # Find template
        bid_pkg = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "templates", "cdcr_bid_package_template.pdf")
        if not os.path.exists(bid_pkg):
            return jsonify({"ok": False, "error": "No bid package template found. Upload cdcr_bid_package_template.pdf to data/templates/"}), 400
        
        out_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "output", sol)
        os.makedirs(out_dir, exist_ok=True)
        
        rfq_data = {
            "solicitation_number": sol,
            "sign_date": get_pst_date(),
            "line_items": items,
        }
        
        food_items = get_food_items_for_obs1600(items)
        
        if fill_type == "obs1600_only":
            output_path = os.path.join(out_dir, f"{sol}_OBS1600_FoodCert_Reytech.pdf")
            fill_obs1600(bid_pkg, rfq_data, config, output_path, food_items=food_items)
        else:
            output_path = os.path.join(out_dir, f"{sol}_BidPackage_Reytech.pdf")
            fill_bid_package(bid_pkg, rfq_data, config, output_path)
        
        t.ok(f"Filled {fill_type}: {sol}, {len(food_items)} food items")
        
        return jsonify({
            "ok": True,
            "filename": os.path.basename(output_path),
            "download_url": f"/api/download/{sol}/{os.path.basename(output_path)}",
            "food_items": food_items,
            "food_count": len(food_items),
            "total_items": len(items),
        })
    except Exception as e:
        import traceback
        t.fail(str(e))
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@bp.route("/api/rfq/<rid>/price-intel")
@auth_required
def api_rfq_price_intel(rid):
    """Return pricing intelligence for all items in an RFQ."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False}), 404

    intel = []
    for item in r.get("line_items", []):
        desc = item.get("description", "")
        pn = item.get("item_number", "") or ""
        current_cost = item.get("supplier_cost") or 0
        current_bid = item.get("price_per_unit") or 0
        result = {"description": desc[:60], "part_number": pn}

        # Price history
        try:
            from src.core.db import get_price_history_db
            history = get_price_history_db(
                description=desc[:60] if not pn else "",
                part_number=pn, limit=10
            )
            if history:
                prices = [h["unit_price"] for h in history if h.get("unit_price")]
                result["history"] = {
                    "count": len(history),
                    "avg": round(sum(prices) / len(prices), 2) if prices else 0,
                    "min": round(min(prices), 2) if prices else 0,
                    "max": round(max(prices), 2) if prices else 0,
                    "entries": [{
                        "price": h["unit_price"],
                        "source": h.get("source", ""),
                        "date": h.get("found_at", "")[:10],
                        "quote": h.get("quote_number", ""),
                        "agency": h.get("agency", ""),
                    } for h in history[:5]]
                }

                # Freshness: compare current cost vs most recent history
                latest = history[0]
                latest_price = latest.get("unit_price", 0)
                latest_source = latest.get("source", "")
                latest_date = latest.get("found_at", "")[:10]
                try:
                    from datetime import datetime as _dt
                    days_old = (_dt.now() - _dt.fromisoformat(
                        latest["found_at"][:19])).days
                except Exception:
                    days_old = 999

                drift = None
                if current_cost > 0 and latest_price > 0 and latest_source not in ("rfq_save", "rfq_save_bid"):
                    diff = latest_price - current_cost
                    pct = diff / current_cost * 100
                    if abs(pct) > 3:  # Only flag >3% drift
                        drift = {
                            "direction": "up" if diff > 0 else "down",
                            "amount": round(abs(diff), 2),
                            "pct": round(pct, 1),
                            "new_price": latest_price,
                            "source": latest_source,
                        }

                result["freshness"] = {
                    "days_old": days_old,
                    "stale": days_old > 90,
                    "last_source": latest_source,
                    "last_date": latest_date,
                    "drift": drift,
                }
        except Exception:
            pass

        # Catalog match
        try:
            from src.core.catalog import search_catalog
            matches = search_catalog(pn or desc[:40], limit=1)
            if matches:
                m = matches[0]
                result["catalog"] = {
                    "sku": m.get("sku", ""),
                    "typical_cost": m.get("typical_cost", 0),
                    "list_price": m.get("list_price", 0),
                    "category": m.get("category", ""),
                }
        except Exception:
            pass

        # Audit trail
        try:
            from src.core.db import get_audit_trail
            audits = get_audit_trail(
                description=desc[:40],
                rfq_id=r.get("solicitation_number", ""), limit=5)
            if audits:
                result["audit"] = [{
                    "field": a["field_changed"],
                    "old": a.get("old_value"),
                    "new": a.get("new_value"),
                    "source": a.get("source", ""),
                    "ts": a.get("ts", "")[:16],
                } for a in audits]
        except Exception:
            pass

        # Pricing recommendation
        rec = _recommend_price(item)
        if rec:
            result["recommendation"] = rec

        # F6: Price conflict resolution — all known sources
        sources = {}
        if item.get("supplier_cost") and item["supplier_cost"] > 0:
            sources["Your Cost"] = round(item["supplier_cost"], 2)
        if item.get("scprs_last_price") and item["scprs_last_price"] > 0:
            sources["SCPRS"] = round(item["scprs_last_price"], 2)
        if item.get("amazon_price") and item["amazon_price"] > 0:
            sources["Amazon"] = round(item["amazon_price"], 2)
        if item.get("price_per_unit") and item["price_per_unit"] > 0:
            sources["Current Bid"] = round(item["price_per_unit"], 2)
        if result.get("catalog") and result["catalog"].get("typical_cost"):
            sources["Catalog"] = round(result["catalog"]["typical_cost"], 2)
        if result.get("catalog") and result["catalog"].get("list_price"):
            sources["Catalog List"] = round(result["catalog"]["list_price"], 2)
        if item.get("_from_pc"):
            sources["_from_pc"] = item["_from_pc"]
        if len(sources) > 1:
            result["sources"] = sources

        # F9: Duplicate item detection — same item quoted recently?
        try:
            from src.core.db import get_price_history_db
            pn = item.get("item_number", "") or ""
            recent = get_price_history_db(
                description=desc[:40] if not pn else "",
                part_number=pn, source="rfq_save_bid", limit=3
            )
            if recent:
                dupes = []
                for rh in recent:
                    dupes.append({
                        "price": rh.get("unit_price", 0),
                        "quote": rh.get("quote_number", ""),
                        "agency": rh.get("agency", ""),
                        "date": rh.get("found_at", "")[:10],
                    })
                if dupes:
                    result["recent_quotes"] = dupes
        except Exception:
            pass

        intel.append(result)

    return jsonify({"ok": True, "intel": intel})


_pricing_alerts_cache = {"data": None, "ts": 0}

@bp.route("/api/pricing-alerts")
@auth_required
def api_pricing_alerts():
    """F8: Dashboard pricing alerts — stale prices, drift, unpriced items."""
    import time as _time
    global _pricing_alerts_cache
    if _pricing_alerts_cache["data"] and (_time.time() - _pricing_alerts_cache["ts"]) < 120:
        return jsonify(_pricing_alerts_cache["data"])
    from datetime import datetime as _dt, timedelta
    rfqs = load_rfqs()
    stale_rfqs = []
    unpriced_rfqs = []
    drift_items = 0
    now = _dt.now()

    for rid, r in rfqs.items():
        if r.get("status") in ("dismissed", "sent", "won", "lost", "cancelled"):
            continue
        items = r.get("line_items", [])
        if not items:
            continue

        # Check for unpriced items
        unpriced = sum(1 for it in items if not (it.get("price_per_unit") or 0) > 0)
        if unpriced == len(items):
            unpriced_rfqs.append({"id": rid, "sol": r.get("solicitation_number", ""), "items": len(items)})
            continue

        # Check for stale pricing (created > 14 days ago, never regenerated)
        try:
            created = r.get("created_at", "")
            if created:
                age = (now - _dt.fromisoformat(created[:19])).days
                if age > 14 and r.get("status") not in ("generated",):
                    stale_rfqs.append({
                        "id": rid, "sol": r.get("solicitation_number", ""),
                        "age_days": age, "items": len(items),
                    })
        except Exception:
            pass

    # Check price_history for recent drift
    try:
        from src.core.db import get_db
        with get_db() as conn:
            # Items with multiple prices where latest differs >10% from previous
            rows = conn.execute("""
                SELECT description, COUNT(*) as cnt,
                       MAX(unit_price) as max_p, MIN(unit_price) as min_p
                FROM price_history
                WHERE found_at > ?
                GROUP BY LOWER(SUBSTR(description, 1, 40))
                HAVING cnt > 1 AND (max_p - min_p) / min_p > 0.10
            """, ((now - timedelta(days=30)).isoformat(),)).fetchall()
            drift_items = len(rows)
    except Exception:
        pass

    total_alerts = len(stale_rfqs) + len(unpriced_rfqs) + (1 if drift_items > 0 else 0)
    _pa_result = {
        "ok": True,
        "total_alerts": total_alerts,
        "stale_rfqs": stale_rfqs,
        "unpriced_rfqs": unpriced_rfqs,
        "drift_items": drift_items,
    }
    _pricing_alerts_cache["data"] = _pa_result
    _pricing_alerts_cache["ts"] = _time.time()
    return jsonify(_pa_result)


@bp.route("/api/rfq/<rid>/qa-check")
@auth_required
def api_rfq_qa_check(rid):
    """QA gate: validate all items before package generation.
    Returns per-item pass/warn/fail with reasons."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False}), 404

    items = r.get("line_items", [])
    results = []
    overall = "pass"
    warnings = 0
    failures = 0

    for i, item in enumerate(items):
        checks = []
        item_status = "pass"
        desc = (item.get("description", "") or "")[:50]
        bid = item.get("price_per_unit") or 0
        cost = item.get("supplier_cost") or 0
        scprs = item.get("scprs_last_price") or 0
        qty = item.get("qty") or 0

        # 1: Bid price > 0
        if not bid or bid <= 0:
            checks.append({"check": "bid_price", "status": "fail", "msg": "No bid price"})
            item_status = "fail"
        else:
            checks.append({"check": "bid_price", "status": "pass", "msg": f"${bid:.2f}"})

        # 2: Supplier cost > 0
        if not cost or cost <= 0:
            checks.append({"check": "cost", "status": "warn", "msg": "No supplier cost"})
            if item_status != "fail":
                item_status = "warn"
        else:
            checks.append({"check": "cost", "status": "pass", "msg": f"${cost:.2f}"})

        # 3: Margin ≥ 15%
        if bid > 0 and cost > 0:
            margin = (bid - cost) / bid * 100
            if margin < 5:
                checks.append({"check": "margin", "status": "fail",
                               "msg": f"{margin:.1f}% — dangerously low"})
                item_status = "fail"
            elif margin < 15:
                checks.append({"check": "margin", "status": "warn",
                               "msg": f"{margin:.1f}% — below 15%"})
                if item_status != "fail":
                    item_status = "warn"
            else:
                checks.append({"check": "margin", "status": "pass",
                               "msg": f"{margin:.1f}%"})

        # 4: Bid vs SCPRS
        if bid > 0 and scprs > 0:
            diff_pct = (bid - scprs) / scprs * 100
            if diff_pct > 10:
                checks.append({"check": "scprs", "status": "warn",
                               "msg": f"{diff_pct:.0f}% above SCPRS ${scprs:.2f}"})
                if item_status != "fail":
                    item_status = "warn"
            elif diff_pct < -15:
                checks.append({"check": "scprs", "status": "warn",
                               "msg": f"{abs(diff_pct):.0f}% below SCPRS — leaving margin?"})
                if item_status != "fail":
                    item_status = "warn"
            else:
                checks.append({"check": "scprs", "status": "pass",
                               "msg": f"OK vs SCPRS ${scprs:.2f}"})

        # 5: Price freshness
        try:
            from src.core.db import get_price_history_db
            pn = item.get("item_number", "") or ""
            history = get_price_history_db(
                description=desc[:40] if not pn else "",
                part_number=pn, limit=1)
            if history:
                from datetime import datetime as _dt
                days = (_dt.now() - _dt.fromisoformat(
                    history[0]["found_at"][:19])).days
                if days > 90:
                    checks.append({"check": "freshness", "status": "warn",
                                   "msg": f"Price data {days}d old"})
                    if item_status != "fail":
                        item_status = "warn"
                else:
                    checks.append({"check": "freshness", "status": "pass",
                                   "msg": f"{days}d ago"})
        except Exception:
            pass

        # 6: Qty > 0
        if not qty or qty <= 0:
            checks.append({"check": "qty", "status": "warn", "msg": "Qty is 0"})
            if item_status != "fail":
                item_status = "warn"

        if item_status == "fail":
            failures += 1
            if overall != "fail":
                overall = "fail"
        elif item_status == "warn":
            warnings += 1
            if overall == "pass":
                overall = "warn"

        results.append({"idx": i, "description": desc,
                        "status": item_status, "checks": checks})

    # PC diff warnings
    diff_notes = []
    pc_diff = r.get("pc_diff")
    if pc_diff:
        if pc_diff.get("added"):
            diff_notes.append(f"{len(pc_diff['added'])} new items not in Price Check")
        if pc_diff.get("removed"):
            diff_notes.append(f"{len(pc_diff['removed'])} PC items not in RFQ")
        if pc_diff.get("qty_changed"):
            diff_notes.append(f"{len(pc_diff['qty_changed'])} qty changes from PC")

    return jsonify({"ok": True, "overall": overall, "failures": failures,
                    "warnings": warnings, "total": len(items), "items": results,
                    "linked_pc": r.get("linked_pc_number", ""),
                    "diff_notes": diff_notes})


@bp.route("/form-filler")
@auth_required
def form_filler_page():
    """Standalone form filler page."""
    return render_page("form_filler.html", active_page="Forms")


# ═══════════════════════════════════════════════════════════════════════
# Admin: Nuke & Re-poll RFQ
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/nuke/<rid>", methods=["POST"])
@auth_required
def api_nuke_rfq(rid):
    """Nuclear delete: wipe RFQ from JSON + SQLite + processed UIDs, then re-poll.
    Usage: POST /api/rfq/nuke/<rfq_id>  or  POST /api/rfq/nuke/<solicitation_number>
    """
    import json as _json
    from src.api.dashboard import load_rfqs, save_rfqs

    rfqs = load_rfqs()
    nuked = []

    # Find by ID or by solicitation number
    targets = {}
    for k, v in rfqs.items():
        if k == rid or v.get("solicitation_number", "") == rid or v.get("rfq_number", "") == rid:
            targets[k] = v

    if not targets:
        return jsonify({"ok": False, "error": f"No RFQ found matching '{rid}'"}), 404

    for rfq_id, rfq in targets.items():
        sol = rfq.get("solicitation_number", rfq.get("rfq_number", "?"))
        email_uid = rfq.get("email_uid", "")

        # 1. Remove from JSON
        if rfq_id in rfqs:
            del rfqs[rfq_id]

        # 2. Remove from SQLite (rfqs, rfq_files, email_log, price_checks)
        try:
            with get_db() as conn:
                conn.execute("DELETE FROM rfq_files WHERE rfq_id = ?", (rfq_id,))
                conn.execute("DELETE FROM rfqs WHERE id = ?", (rfq_id,))
                # email_log by rfq_id
                conn.execute("DELETE FROM email_log WHERE rfq_id = ?", (rfq_id,))
                # price_checks by rfq_id
                conn.execute("DELETE FROM price_checks WHERE rfq_id = ?", (rfq_id,))
                conn.commit()
        except Exception as e:
            log.warning("Nuke SQLite cleanup for %s: %s", rfq_id, e)

        # 3. Remove email UID from processed list
        if email_uid:
            try:
                from src.api.modules.routes_pricecheck import _remove_processed_uid
                _remove_processed_uid(email_uid)
            except Exception:
                # Manual fallback
                proc_file = os.path.join(DATA_DIR, "processed_emails.json")
                try:
                    if os.path.exists(proc_file):
                        with open(proc_file) as f:
                            processed = _json.load(f)
                        if isinstance(processed, list) and email_uid in processed:
                            processed.remove(email_uid)
                        elif isinstance(processed, dict) and email_uid in processed:
                            del processed[email_uid]
                        with open(proc_file, "w") as f:
                            _json.dump(processed, f)
                except Exception:
                    pass

        nuked.append({"id": rfq_id, "sol": sol, "uid": email_uid})
        log.info("NUKED RFQ %s (sol=%s, uid=%s)", rfq_id, sol, email_uid)

    save_rfqs(rfqs)

    # 4. Trigger re-poll
    poll_result = None
    try:
        from src.api.modules.routes_pricecheck import do_poll_check
        imported = do_poll_check()
        poll_result = {"found": len(imported), "rfqs": [r.get("solicitation_number", "?") for r in imported]}
    except Exception as e:
        poll_result = {"error": str(e)}

    return jsonify({
        "ok": True,
        "nuked": nuked,
        "poll": poll_result,
    })


@bp.route("/api/rfq/<rid>/clear-quote", methods=["POST", "GET"])
@auth_required
def api_rfq_clear_quote(rid):
    """Clear the quote number on an RFQ so regeneration assigns a new one."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    
    old_qn = r.get("reytech_quote_number", "")
    r["reytech_quote_number"] = ""
    r["linked_quote_number"] = ""
    save_rfqs(rfqs)
    
    return jsonify({"ok": True, "cleared": old_qn, "message": f"Cleared {old_qn}. Regenerate to get a new number."})


@bp.route("/api/rfq/<rid>/set-quote-number", methods=["POST"])
@auth_required
def api_rfq_set_quote_number(rid):
    """Force-set the quote number on an RFQ. Used to fix counter drift."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    data = request.get_json(force=True, silent=True) or {}
    qn = data.get("quote_number", "").strip()
    if not qn:
        return jsonify({"ok": False, "error": "Provide quote_number"})
    old = r.get("reytech_quote_number", "")
    r["reytech_quote_number"] = qn
    save_rfqs(rfqs)
    log.info("Force-set quote number on RFQ %s: %s → %s", rid, old, qn)
    return jsonify({"ok": True, "old": old, "new": qn})


@bp.route("/api/admin/fix-quote-number/<rid>/<new_qn>/<int:counter_seq>", methods=["POST", "GET"])
@auth_required
def api_admin_fix_quote_number(rid, new_qn, counter_seq):
    """One-shot admin: set RFQ quote number + reset counter. GET-accessible for browser.

    Example: /api/admin/fix-quote-number/cab4bad5/R26Q31/31
    Sets RFQ cab4bad5 to R26Q31, counter to 31 (next = R26Q32).
    """
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    old_qn = r.get("reytech_quote_number", "")
    r["reytech_quote_number"] = new_qn
    # Also update in output files and generated package data
    if r.get("quote_number"):
        r["quote_number"] = new_qn
    save_rfqs(rfqs)
    # Reset counter
    try:
        from src.forms.quote_generator import set_quote_counter, peek_next_quote_number
        set_quote_counter(seq=counter_seq)
        nxt = peek_next_quote_number()
    except Exception as e:
        nxt = f"error: {e}"
    log.warning("ADMIN fix-quote: RFQ %s: %s → %s, counter → %d (next: %s)", rid, old_qn, new_qn, counter_seq, nxt)
    return jsonify({
        "ok": True,
        "rfq": rid,
        "old_quote": old_qn,
        "new_quote": new_qn,
        "counter_set_to": counter_seq,
        "next_quote_will_be": nxt,
    })


@bp.route("/api/rfq/<rid>/clear-generated", methods=["POST", "GET"])
@auth_required
def api_rfq_clear_generated(rid):
    """
    Force-clear all generated files for an RFQ from both DB and JSON.
    Resets status to 'ready' so the full generate-package pipeline re-runs cleanly.
    Use this when Railway redeploys cached the old output and Regenerate doesn't help.
    """
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    # Clear DB generated files
    db_deleted = 0
    try:
        from src.core.db import get_db
        with get_db() as conn:
            cur = conn.execute(
                "DELETE FROM rfq_files WHERE rfq_id = ? AND category = 'generated'",
                (rid,)
            )
            db_deleted = cur.rowcount
    except Exception as _e:
        log.warning("clear-generated DB delete failed for %s: %s", rid, _e)

    # Clear disk output files
    sol = r.get("solicitation_number", rid)
    out_dir = os.path.join(OUTPUT_DIR, sol)
    disk_deleted = 0
    if os.path.exists(out_dir):
        try:
            for fname in os.listdir(out_dir):
                fpath = os.path.join(out_dir, fname)
                if os.path.isfile(fpath):
                    os.remove(fpath)
                    disk_deleted += 1
        except Exception as _de:
            log.warning("clear-generated disk delete failed: %s", _de)

    # Reset JSON state
    old_files = r.get("output_files", [])
    r["output_files"] = []
    r.pop("draft_email", None)
    r.pop("generated_at", None)
    _transition_status(r, "ready", actor="user", notes="Cleared generated files for fresh regeneration")
    save_rfqs(rfqs)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, "ready")
    except Exception:
        pass

    msg = f"Cleared {db_deleted} DB files + {disk_deleted} disk files. Status reset to 'ready'. Click Generate Package to rebuild."
    log.info("clear-generated %s: %s", rid, msg)
    return jsonify({"ok": True, "db_deleted": db_deleted, "disk_deleted": disk_deleted,
                    "old_files": old_files, "message": msg})


@bp.route("/api/rfq/<rid>/clean-slate", methods=["POST", "GET"])
@auth_required
def api_rfq_clean_slate(rid):
    """Nuclear clean: keep ONLY line_items with pricing. Clear everything else.
    Use when package is broken — stale templates, wrong forms, old data."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    # Preserve line items with all pricing fields
    items = r.get("line_items", [])
    preserved_items = []
    for it in items:
        preserved_items.append({
            "line_number": it.get("line_number", 0),
            "qty": it.get("qty", 1),
            "uom": it.get("uom", "EA"),
            "description": it.get("description", ""),
            "item_number": it.get("item_number", ""),
            "supplier_cost": it.get("supplier_cost", 0),
            "price_per_unit": it.get("price_per_unit", 0),
            "markup_pct": it.get("markup_pct"),
            "scprs_last_price": it.get("scprs_last_price"),
            "amazon_price": it.get("amazon_price"),
            "item_link": it.get("item_link", ""),
            "item_supplier": it.get("item_supplier", ""),
            "_desc_source": it.get("_desc_source", ""),
        })

    # Preserve core RFQ identity
    sol = r.get("solicitation_number", "")
    identity = {
        "solicitation_number": sol,
        "agency": r.get("agency", ""),
        "requestor_name": r.get("requestor_name", ""),
        "requestor_email": r.get("requestor_email", ""),
        "delivery_location": r.get("delivery_location", ""),
        "due_date": r.get("due_date", ""),
        "institution": r.get("institution", ""),
        "ship_to": r.get("ship_to", ""),
        "created_at": r.get("created_at", ""),
        "source": r.get("source", ""),
        "linked_pc_id": r.get("linked_pc_id", ""),
        "reytech_quote_number": r.get("reytech_quote_number", ""),
    }

    # Clear DB files (generated + templates)
    db_deleted = 0
    try:
        from src.core.db import get_db
        with get_db() as conn:
            cur = conn.execute(
                "DELETE FROM rfq_files WHERE rfq_id = ? AND category IN ('generated', 'template')",
                (rid,)
            )
            db_deleted = cur.rowcount
    except Exception as _e:
        log.warning("clean-slate DB: %s", _e)

    # Clear disk
    disk_deleted = 0
    import shutil as _sh2
    out_dir = os.path.join(OUTPUT_DIR, sol)
    if os.path.exists(out_dir):
        try:
            _sh2.rmtree(out_dir)
            disk_deleted += 1
        except Exception:
            pass
    tmpl_dir = os.path.join(DATA_DIR, "rfq_templates", rid)
    if os.path.exists(tmpl_dir):
        try:
            _sh2.rmtree(tmpl_dir)
            disk_deleted += 1
        except Exception:
            pass

    # Rebuild RFQ with clean state
    r.clear()
    r.update(identity)
    r["line_items"] = preserved_items
    r["templates"] = {}
    r["output_files"] = []
    r["status"] = "ready"

    save_rfqs(rfqs)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, "ready")
    except Exception:
        pass

    log.info("clean-slate %s: kept %d items, cleared %d DB + %d disk",
             rid, len(preserved_items), db_deleted, disk_deleted)
    return jsonify({
        "ok": True,
        "items_preserved": len(preserved_items),
        "db_cleared": db_deleted,
        "disk_cleared": disk_deleted,
        "message": f"Clean slate: {len(preserved_items)} items preserved with pricing. All docs/templates cleared. Ready to regenerate.",
    })


@bp.route("/api/rfq/<rid>/debug-pages", methods=["GET"])
@auth_required
def api_rfq_debug_pages(rid):
    """Debug: run page-skip logic against last generated package PDF. Returns per-page decisions."""
    from src.forms.reytech_filler_v4 import _bidpkg_page_skip_reason
    from pypdf import PdfReader

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    sol = r.get("solicitation_number", rid)
    pkg_path = os.path.join(OUTPUT_DIR, sol, f"RFQ_Package_{sol}_ReytechInc.pdf")

    if not os.path.exists(pkg_path):
        return jsonify({"ok": False, "error": f"Not found: {pkg_path}"})

    reader = PdfReader(pkg_path)
    pages = []
    for i, page in enumerate(reader.pages):
        try:
            reason = _bidpkg_page_skip_reason(page)
        except Exception as e:
            reason = f"ERROR: {e}"
        text_snip = (page.extract_text() or "")[:80].replace("\n", " ")
        n_fields = len(page.get("/Annots", [])) if "/Annots" in page else 0
        pages.append({"page": i, "decision": "SKIP" if reason else "KEEP",
                      "reason": reason or "", "fields": n_fields, "text": text_snip})

    return jsonify({"ok": True, "total": len(pages),
                    "kept": sum(1 for p in pages if p["decision"] == "KEEP"),
                    "skipped": sum(1 for p in pages if p["decision"] == "SKIP"),
                    "pages": pages})


@bp.route("/api/rfq/<rid>/debug-templates", methods=["GET"])
@auth_required
def api_rfq_debug_templates(rid):
    """Dump all field names from uploaded 703B/704B/bidpkg templates. Use to diagnose fill mismatches."""
    from pypdf import PdfReader

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    result = {}
    tmpl = r.get("templates", {})

    # Also restore from DB if needed
    db_files = list_rfq_files(rid, category="template")
    for db_f in db_files:
        ft = db_f.get("file_type", "").lower().replace("template_", "")
        fname = db_f.get("filename", "").lower()
        ttype = None
        if "703b" in ft or "703b" in fname: ttype = "703b"
        elif "704b" in ft or "704b" in fname: ttype = "704b"
        elif "bid" in ft or "bid" in fname: ttype = "bidpkg"
        if ttype and (ttype not in tmpl or not os.path.exists(tmpl.get(ttype, ""))):
            full_f = get_rfq_file(db_f["id"])
            if full_f and full_f.get("data"):
                restore_dir = os.path.join(DATA_DIR, "rfq_templates", rid)
                os.makedirs(restore_dir, exist_ok=True)
                restore_path = os.path.join(restore_dir, db_f["filename"])
                with open(restore_path, "wb") as _fw:
                    _fw.write(full_f["data"])
                tmpl[ttype] = restore_path

    for tname, tpath in tmpl.items():
        if not os.path.exists(tpath):
            result[tname] = {"error": f"file missing: {tpath}"}
            continue
        try:
            rdr = PdfReader(tpath)
            fields = rdr.get_fields() or {}
            sig_fields = []
            all_pages = []
            for i, pg in enumerate(rdr.pages):
                annots = pg.get("/Annots", [])
                pg_fields = []
                for a in (annots or []):
                    obj = a.get_object() if hasattr(a, "get_object") else a
                    name = str(obj.get("/T", ""))
                    ft_val = str(obj.get("/FT", ""))
                    if ft_val == "/Sig" or "sig" in name.lower():
                        sig_fields.append({"name": name, "ft": ft_val, "page": i})
                    if name:
                        pg_fields.append(name)
                all_pages.append({"page": i, "fields": pg_fields[:10]})
            result[tname] = {
                "path": tpath,
                "pages": len(rdr.pages),
                "total_fields": len(fields),
                "sig_fields": sig_fields,
                "all_field_names": sorted(fields.keys())[:50],
                "pages_preview": all_pages,
            }
        except Exception as e:
            result[tname] = {"error": str(e)}

    return jsonify({"ok": True, "templates": result})


@bp.route("/api/rfq/<rid>/diag-package")
@auth_required
def api_diag_package(rid):
    """Diagnostic: test each form generation step and report what works/fails."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    results = {"rid": rid, "items": len(r.get("line_items", [])), "steps": []}
    
    # Check agency
    try:
        from src.core.agency_config import match_agency
        _agency_key, _agency_cfg = match_agency(r)
        _req = _agency_cfg.get("required_forms", [])
        results["agency"] = _agency_key
        results["required_forms"] = _req
        results["steps"].append({"step": "agency_match", "ok": True, "agency": _agency_key, "forms": _req})
    except Exception as e:
        results["steps"].append({"step": "agency_match", "ok": False, "error": str(e)})
        return jsonify(results)
    
    # Check templates dir
    import os
    tdir = os.path.join(os.environ.get("DATA_DIR", "data"), "templates")
    if os.path.exists(tdir):
        files = os.listdir(tdir)
        results["steps"].append({"step": "templates_dir", "ok": True, "files": files})
    else:
        results["steps"].append({"step": "templates_dir", "ok": False, "error": f"{tdir} not found"})
    
    # Check quote generator
    try:
        results["steps"].append({"step": "quote_gen_import", "ok": True})
    except Exception as e:
        results["steps"].append({"step": "quote_gen_import", "ok": False, "error": str(e)})
    
    # Check each required form's generator
    form_checks = {
        "calrecycle74": ("src.forms.reytech_filler_v4", "fill_calrecycle_standalone"),
        "std204": ("src.forms.reytech_filler_v4", "fill_std204"),
        "std1000": ("src.forms.reytech_filler_v4", "fill_std1000"),
        "dvbe843": ("src.forms.reytech_filler_v4", "generate_dvbe_843"),
        "bidder_decl": ("src.forms.reytech_filler_v4", "generate_bidder_declaration"),
        "darfur_act": ("src.forms.reytech_filler_v4", "generate_darfur_act"),
        "cv012_cuf": ("src.forms.reytech_filler_v4", "fill_cv012_cuf"),
    }
    for form_id, (mod, func) in form_checks.items():
        if form_id in _req:
            try:
                m = __import__(mod, fromlist=[func])
                fn = getattr(m, func)
                results["steps"].append({"step": f"import_{form_id}", "ok": True, "func": func})
            except Exception as e:
                results["steps"].append({"step": f"import_{form_id}", "ok": False, "error": str(e)})
    
    # Check CONFIG
    try:
        from src.api.modules.routes_rfq import CONFIG
        results["steps"].append({"step": "config", "ok": True, "company": CONFIG.get("company", {}).get("name", "?")})
    except Exception as e:
        results["steps"].append({"step": "config", "ok": False, "error": str(e)})
    
    # Check line items have pricing
    items = r.get("line_items", [])
    priced = sum(1 for i in items if i.get("price_per_unit") and i["price_per_unit"] > 0)
    results["steps"].append({"step": "pricing", "items": len(items), "priced": priced})

    return jsonify(results)


# ══ Consolidated from routes_features*.py ══════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════
# Email Draft Queue Status
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/email/queue-status")
@auth_required
def api_email_queue_status():
    """Status of email drafts: pending, approved, sent."""
    outbox_path = os.path.join(DATA_DIR, "outbox.json")
    try:
        with open(outbox_path) as f:
            outbox = json.load(f)
    except Exception:
        outbox = []

    if isinstance(outbox, dict):
        outbox = list(outbox.values())

    draft = [e for e in outbox if (e.get("status") or "").lower() in ("draft", "pending")]
    approved = [e for e in outbox if (e.get("status") or "").lower() == "approved"]
    sent = [e for e in outbox if (e.get("status") or "").lower() == "sent"]

    return jsonify({
        "ok": True,
        "drafts": len(draft),
        "approved": len(approved),
        "sent": len(sent),
        "total": len(outbox),
        "needs_review": len(draft),
        "ready_to_send": len(approved),
        "recent_drafts": [
            {"to": e.get("to", "?"), "subject": e.get("subject", "?")[:50],
             "created": e.get("created", "?"), "type": e.get("type", "?")}
            for e in sorted(draft, key=lambda x: x.get("created", ""), reverse=True)[:5]
        ]
    })


# ═══════════════════════════════════════════════════════════════════════
# RFQs Ready to Quote
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/ready-to-quote")
@auth_required
def api_rfq_ready_to_quote():
    """RFQs that need pricing/quoting — prioritized by deadline."""
    rfqs_path = os.path.join(DATA_DIR, "rfqs.json")
    if not os.path.exists(rfqs_path):
        return jsonify({"ok": True, "rfqs": [], "count": 0})

    try:
        with open(rfqs_path) as f:
            rfqs = json.load(f)
    except Exception:
        return jsonify({"ok": True, "rfqs": [], "count": 0})

    today = datetime.now().strftime("%Y-%m-%d")
    ready = []

    for rid, r in rfqs.items():
        status = (r.get("status") or "").lower()
        if status in ("new", "draft", "priced", "inbox"):
            due = r.get("due_date") or r.get("deadline") or ""
            sol = r.get("solicitation_number", rid)
            items = r.get("line_items") or r.get("items_detail") or []
            if isinstance(items, str):
                try: items = json.loads(items)
                except Exception: items = []

            overdue = due and due < today
            days_left = None
            if due:
                try:
                    dd = datetime.strptime(due[:10], "%Y-%m-%d")
                    days_left = (dd - datetime.now()).days
                except Exception: pass

            ready.append({
                "id": rid,
                "solicitation": sol[:30],
                "requestor": r.get("requestor", r.get("buyer_name", "?")),
                "institution": r.get("institution", "?"),
                "status": status.upper(),
                "items": len(items) if isinstance(items, list) else 0,
                "due": due[:10] if due else "TBD",
                "days_left": days_left,
                "overdue": overdue,
                "total": r.get("total_price", 0),
            })

    # Sort: overdue first, then by days_left
    ready.sort(key=lambda x: (not x["overdue"], x["days_left"] if x["days_left"] is not None else 999))

    return jsonify({
        "ok": True,
        "rfqs": ready[:20],
        "count": len(ready),
        "overdue": len([r for r in ready if r["overdue"]]),
        "due_this_week": len([r for r in ready if r.get("days_left") is not None and 0 <= r["days_left"] <= 7])
    })


@bp.route("/api/rfq/<rid>/clean-items", methods=["POST"])
@auth_required
def rfq_clean_items(rid):
    """Remove junk items (legal text, instructions, boilerplate) from an RFQ."""
    from src.api.dashboard import load_rfqs, save_rfqs
    rfqs = load_rfqs()
    rfq = rfqs.get(rid)
    if not rfq:
        return jsonify({"ok": False, "error": "RFQ not found"})

    items = rfq.get("line_items", [])
    original_count = len(items)

    from src.forms.price_check import _filter_junk_items
    cleaned = _filter_junk_items(items)

    rfq["line_items"] = cleaned
    if "parsed" in rfq:
        rfq["parsed"]["line_items"] = cleaned

    save_rfqs(rfqs)

    removed = original_count - len(cleaned)
    return jsonify({"ok": True, "removed": removed, "kept": len(cleaned), "original": original_count})


# ═══════════════════════════════════════════════════════════════════
# Package Manifest + Lifecycle API
# ═══════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/<rid>/manifest")
@auth_required
def api_rfq_manifest(rid):
    """Get the latest package manifest for an RFQ."""
    from src.core.dal import get_latest_manifest
    manifest = get_latest_manifest(rid)
    if not manifest:
        return jsonify({"ok": False, "error": "No package manifest found"})
    return jsonify({"ok": True, "manifest": manifest})


@bp.route("/api/rfq/<rid>/manifest/<int:manifest_id>/review", methods=["POST"])
@auth_required
def api_rfq_review_form(rid, manifest_id):
    """Record a review verdict for a form in the manifest."""
    from src.core.dal import review_form, log_lifecycle_event
    data = request.get_json(force=True, silent=True) or {}
    form_id = data.get("form_id", "")
    verdict = data.get("verdict", "approved")
    notes = data.get("notes", "")
    if not form_id:
        return jsonify({"ok": False, "error": "form_id required"})
    ok = review_form(manifest_id, form_id, verdict, reviewed_by="user", notes=notes)
    if ok:
        log_lifecycle_event("rfq", rid, "form_reviewed",
            f"Form {form_id}: {verdict}" + (f" — {notes}" if notes else ""),
            actor="user", detail={"form_id": form_id, "verdict": verdict, "manifest_id": manifest_id})
    return jsonify({"ok": ok})


@bp.route("/api/rfq/<rid>/manifest/<int:manifest_id>/approve", methods=["POST"])
@auth_required
def api_rfq_approve_package(rid, manifest_id):
    """Approve the entire package (all forms must be reviewed first)."""
    from src.core.dal import get_package_manifest, update_manifest_status, log_lifecycle_event
    manifest = get_package_manifest(manifest_id)
    if not manifest:
        return jsonify({"ok": False, "error": "Manifest not found"})
    pending = [r for r in manifest.get("reviews", []) if r.get("verdict") == "pending"]
    if pending:
        return jsonify({"ok": False, "error": f"{len(pending)} forms still pending review",
                        "pending": [r["form_id"] for r in pending]})
    rejected = [r for r in manifest.get("reviews", []) if r.get("verdict") == "rejected"]
    if rejected:
        return jsonify({"ok": False, "error": f"{len(rejected)} forms rejected",
                        "rejected": [r["form_id"] for r in rejected]})
    ok = update_manifest_status(manifest_id, "approved")
    if ok:
        log_lifecycle_event("rfq", rid, "package_approved",
            f"Package v{manifest.get('version', '?')} approved ({manifest.get('total_forms', 0)} forms)",
            actor="user", detail={"manifest_id": manifest_id, "version": manifest.get("version")})
    return jsonify({"ok": ok, "status": "approved"})


@bp.route("/api/rfq/<rid>/manifest/<int:manifest_id>/remove-form", methods=["POST"])
@auth_required
def api_rfq_remove_form(rid, manifest_id):
    """Remove a form from the package manifest and delete its file."""
    data = request.get_json(force=True, silent=True) or {}
    form_id = data.get("form_id", "")
    if not form_id:
        return jsonify({"ok": False, "error": "form_id required"})

    try:
        from src.core.db import get_db
        with get_db() as conn:
            # Verify manifest belongs to this RFQ
            _owner = conn.execute(
                "SELECT rfq_id FROM package_manifest WHERE id = ?",
                (manifest_id,)).fetchone()
            if not _owner or _owner[0] != rid:
                return jsonify({"ok": False, "error": "Manifest not found for this RFQ"})

            # Get the review record to find the filename
            row = conn.execute(
                "SELECT form_filename FROM package_review WHERE manifest_id = ? AND form_id = ?",
                (manifest_id, form_id)).fetchone()
            filename = row[0] if row else ""

            # Delete the review record
            conn.execute(
                "DELETE FROM package_review WHERE manifest_id = ? AND form_id = ?",
                (manifest_id, form_id))

            # Update the manifest's generated_forms list
            manifest_row = conn.execute(
                "SELECT generated_forms, total_forms FROM package_manifest WHERE id = ?",
                (manifest_id,)).fetchone()
            if manifest_row:
                import json as _json_rm
                gen_forms = _json_rm.loads(manifest_row[0] or "[]")
                gen_forms = [f for f in gen_forms if (f.get("form_id") if isinstance(f, dict) else f) != form_id]
                total = (manifest_row[1] or 0) - 1
                conn.execute(
                    "UPDATE package_manifest SET generated_forms = ?, total_forms = ? WHERE id = ?",
                    (_json_rm.dumps(gen_forms), max(total, 0), manifest_id))

            # Delete the actual file from disk
            if filename:
                rfqs = load_rfqs()
                r = rfqs.get(rid, {})
                sol = r.get("solicitation_number", "") or r.get("rfq_number", "") or "unknown"
                filepath = os.path.join(OUTPUT_DIR, sol, filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    log.info("Removed file %s for form %s", filename, form_id)

                # Remove from rfq output_files list
                out_files = r.get("output_files", [])
                if filename in out_files:
                    out_files.remove(filename)
                    r["output_files"] = out_files
                    save_rfqs(rfqs)

            # Log the removal
            from src.core.dal import log_lifecycle_event
            log_lifecycle_event("rfq", rid, "form_removed",
                f"Removed {form_id} from package ({filename})",
                actor="user", detail={"form_id": form_id, "filename": filename, "manifest_id": manifest_id})

        return jsonify({"ok": True, "removed": form_id, "filename": filename})
    except Exception as e:
        log.error("Remove form failed: %s", e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/rfq/<rid>/timeline")
@auth_required
def api_rfq_timeline(rid):
    """Get the full lifecycle timeline for an RFQ."""
    from src.core.dal import get_lifecycle_events
    events = get_lifecycle_events("rfq", rid, limit=200)
    return jsonify({"ok": True, "events": events, "count": len(events)})


@bp.route("/api/rfq/<rid>/buyer-prefs")
@auth_required
def api_rfq_buyer_prefs(rid):
    """Get buyer preferences for the RFQ's requestor."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    email = r.get("requestor_email", "")
    if not email:
        return jsonify({"ok": True, "preferences": [], "message": "No requestor email"})
    from src.core.dal import get_buyer_preferences
    prefs = get_buyer_preferences(email)
    return jsonify({"ok": True, "preferences": prefs, "buyer_email": email})


@bp.route("/api/rfq/<rid>/download-complete-package")
@auth_required
def api_download_complete_package(rid):
    """Download ALL forms merged into one PDF — quote + compliance."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    sol = r.get("solicitation_number", "") or r.get("rfq_number", "") or "unknown"
    out_dir = os.path.join(OUTPUT_DIR, sol)
    output_files = r.get("output_files", [])

    if not output_files:
        return jsonify({"ok": False, "error": "No files generated"})

    try:
        from pypdf import PdfReader, PdfWriter
        writer = PdfWriter()
        merged_count = 0

        # Quote first, then all other forms in order
        quote_files = [f for f in output_files if "Quote" in f and "704" not in f.upper()]
        other_files = [f for f in output_files if f not in quote_files]
        ordered = quote_files + other_files

        for f in ordered:
            fpath = os.path.join(out_dir, f)
            if not os.path.exists(fpath):
                continue
            # Skip the merged package file itself (avoid double-counting)
            if "RFQ_Package" in f or "Compliance_Forms" in f:
                continue
            try:
                reader = PdfReader(fpath)
                for page in reader.pages:
                    text = ""
                    try:
                        text = page.extract_text() or ""
                    except Exception:
                        pass
                    if text.strip().startswith("Please wait") and len(text.strip()) < 300:
                        continue
                    writer.add_page(page)
                merged_count += 1
            except Exception as _e:
                log.warning("Skip %s in complete package: %s", f, _e)

        if merged_count == 0:
            return jsonify({"ok": False, "error": "No valid PDFs to merge"})

        import io
        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)

        _safe_agency = ""
        try:
            from src.core.agency_config import match_agency
            _ak, _ac = match_agency(r)
            _safe_agency = (_ac.get("name", "") or "").replace(" ", "").replace("/", "")[:20]
        except Exception:
            pass

        filename = f"Complete_RFQ_{_safe_agency}_{sol}_ReytechInc.pdf" if _safe_agency else f"Complete_RFQ_{sol}_ReytechInc.pdf"

        from flask import send_file
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)
    except Exception as e:
        log.error("Complete package download failed: %s", e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/rfq/<rid>/export-invoice")
@auth_required
def api_export_invoice(rid):
    """Export buyer invoice as Excel for QB entry."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    items = r.get("line_items", r.get("items", []))
    sol = r.get("solicitation_number", "") or r.get("rfq_number", "")
    quote_num = r.get("reytech_quote_number", "")
    buyer = r.get("requestor_name", "")
    agency = r.get("agency", "") or r.get("agency_name", "")
    tax_rate = float(r.get("tax_rate", 0) or 0) / 100
    delivery = r.get("delivery_location", "")

    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl not installed — run: pip install openpyxl"}), 500

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        hf = Font(bold=True, size=14)
        sf = Font(size=11, color="333333")
        cf = Font(bold=True, size=10, color="FFFFFF")
        cfill = PatternFill("solid", fgColor="2E75B6")
        mf = '#,##0.00'
        bd = Border(bottom=Side(style='thin', color='CCCCCC'))

        ws.merge_cells('A1:G1')
        ws['A1'] = "REYTECH INC. — INVOICE"
        ws['A1'].font = hf
        ws['A2'] = f"Quote #: {quote_num}"
        ws['A2'].font = sf
        ws['A3'] = f"Bill To: {buyer} — {agency}"
        ws['A3'].font = sf
        ws['A4'] = f"Ship To: {delivery}"
        ws['A4'].font = sf
        ws['A5'] = f"Solicitation #: {sol}"
        ws['A5'].font = sf

        for col, h in enumerate(["#", "Description", "Part #", "QTY", "UOM", "Unit Price", "Subtotal"], 1):
            c = ws.cell(row=7, column=col, value=h)
            c.font = cf
            c.fill = cfill
            c.alignment = Alignment(horizontal='center')

        subtotal = 0
        for idx, item in enumerate(items):
            row = 8 + idx
            qty = int(float(item.get("qty", 1) or 1))
            price = float(item.get("price_per_unit", 0) or 0)
            lt = qty * price
            subtotal += lt
            ws.cell(row=row, column=1, value=idx+1).border = bd
            ws.cell(row=row, column=2, value=(item.get("description", "") or "")[:80]).border = bd
            ws.cell(row=row, column=3, value=item.get("part_number", "") or item.get("item_number", "")).border = bd
            ws.cell(row=row, column=4, value=qty).border = bd
            ws.cell(row=row, column=5, value=item.get("uom", "EA")).border = bd
            ws.cell(row=row, column=6, value=price).number_format = mf
            ws.cell(row=row, column=7, value=lt).number_format = mf

        tr = 8 + len(items) + 1
        tax_amt = subtotal * tax_rate
        ws.cell(row=tr, column=6, value="Subtotal:").font = Font(bold=True)
        ws.cell(row=tr, column=7, value=subtotal).number_format = mf
        ws.cell(row=tr+1, column=6, value=f"Tax ({r.get('tax_rate', 0)}%):").font = Font(bold=True)
        ws.cell(row=tr+1, column=7, value=tax_amt).number_format = mf
        ws.cell(row=tr+2, column=6, value="TOTAL:").font = Font(bold=True, size=12)
        ws.cell(row=tr+2, column=7, value=subtotal + tax_amt).number_format = mf
        ws.cell(row=tr+2, column=7).font = Font(bold=True, size=12)

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"Invoice_{quote_num}_{sol}_Reytech.xlsx")
    except Exception as e:
        log.error("Invoice export: %s", e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/rfq/<rid>/export-supplier-po")
@auth_required
def api_export_supplier_po(rid):
    """Export supplier PO as Excel for QB entry."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    items = r.get("line_items", r.get("items", []))
    sol = r.get("solicitation_number", "") or r.get("rfq_number", "")
    quote_num = r.get("reytech_quote_number", "")
    supplier_name = ""
    for item in items:
        sn = item.get("cost_supplier_name", "") or item.get("scprs_supplier", "")
        if sn:
            supplier_name = sn
            break

    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl not installed — run: pip install openpyxl"}), 500

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supplier PO"

        hf = Font(bold=True, size=14)
        sf = Font(size=11, color="333333")
        cf = Font(bold=True, size=10, color="FFFFFF")
        cfill = PatternFill("solid", fgColor="2D6A2E")
        mf = '#,##0.00'
        bd = Border(bottom=Side(style='thin', color='CCCCCC'))

        ws.merge_cells('A1:H1')
        ws['A1'] = "REYTECH INC. — PURCHASE ORDER"
        ws['A1'].font = hf
        ws['A2'] = f"Related Quote #: {quote_num}"
        ws['A2'].font = sf
        ws['A3'] = f"Supplier: {supplier_name}"
        ws['A3'].font = sf
        ws['A4'] = f"Solicitation #: {sol}"
        ws['A4'].font = sf

        for col, h in enumerate(["#", "Description", "Part #", "QTY", "UOM", "Supplier Cost", "Subtotal", "Source"], 1):
            c = ws.cell(row=6, column=col, value=h)
            c.font = cf
            c.fill = cfill
            c.alignment = Alignment(horizontal='center')

        subtotal = 0
        for idx, item in enumerate(items):
            row = 7 + idx
            qty = int(float(item.get("qty", 1) or 1))
            cost = float(item.get("supplier_cost", 0) or item.get("vendor_cost", 0) or 0)
            lt = qty * cost
            subtotal += lt
            source = item.get("cost_source", "")
            sname = item.get("cost_supplier_name", "")
            ws.cell(row=row, column=1, value=idx+1).border = bd
            ws.cell(row=row, column=2, value=(item.get("description", "") or "")[:80]).border = bd
            ws.cell(row=row, column=3, value=item.get("part_number", "") or item.get("item_number", "")).border = bd
            ws.cell(row=row, column=4, value=qty).border = bd
            ws.cell(row=row, column=5, value=item.get("uom", "EA")).border = bd
            ws.cell(row=row, column=6, value=cost).number_format = mf
            ws.cell(row=row, column=7, value=lt).number_format = mf
            ws.cell(row=row, column=8, value=f"{source} — {sname}" if sname else source).border = bd

        tr = 7 + len(items) + 1
        bid_total = sum(int(float(i.get("qty", 1) or 1)) * float(i.get("price_per_unit", 0) or 0) for i in items)
        ws.cell(row=tr, column=6, value="TOTAL:").font = Font(bold=True, size=12)
        ws.cell(row=tr, column=7, value=subtotal).number_format = mf
        ws.cell(row=tr, column=7).font = Font(bold=True, size=12)
        ws.cell(row=tr+2, column=5, value="Supplier Cost:").font = Font(bold=True)
        ws.cell(row=tr+2, column=7, value=subtotal).number_format = mf
        ws.cell(row=tr+3, column=5, value="Bid Total:").font = Font(bold=True)
        ws.cell(row=tr+3, column=7, value=bid_total).number_format = mf
        ws.cell(row=tr+4, column=5, value="Gross Margin:").font = Font(bold=True)
        ws.cell(row=tr+4, column=7, value=bid_total - subtotal).number_format = mf
        ws.cell(row=tr+4, column=7).font = Font(bold=True, color="2D6A2E")

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 25

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        safe_sup = supplier_name.replace(" ", "")[:20]
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"SupplierPO_{quote_num}_{sol}_{safe_sup}.xlsx")
    except Exception as e:
        log.error("Supplier PO export: %s", e)
        return jsonify({"ok": False, "error": str(e)})
