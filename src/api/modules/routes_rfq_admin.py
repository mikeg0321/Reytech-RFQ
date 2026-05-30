# routes_rfq_admin.py — Admin, Export, Settings, Diagnostics, File Management
# Split from routes_rfq.py — loaded via importlib in dashboard.py

from flask import request, jsonify, Response
from src.api.shared import bp, auth_required
import logging
log = logging.getLogger("reytech")
from src.core.error_handler import safe_route, safe_page
from src.core.security import rate_limit
from flask import redirect, flash, send_file, session
from src.core.paths import DATA_DIR, OUTPUT_DIR, UPLOAD_DIR
from src.core.db import get_db
from src.api.render import render_page
import os, json, re
from datetime import datetime, timedelta, timezone


# ═══════════════════════════════════════════════════════════════════════
# RFQ File Management — download from DB
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/rfq/<rid>/file/<file_id>")
@auth_required
@safe_page
def rfq_download_file(rid, file_id):
    """Download an RFQ file from the database."""
    _bad = _validate_rid(rid)
    if _bad: return _bad
    f = get_rfq_file(file_id)
    if not f or f.get("rfq_id") != rid:
        flash("File not found", "error")
        return redirect(f"/rfq/{rid}")
    from flask import Response
    # Validate MIME type against allowlist to prevent XSS via stored type
    _allowed_mimes = {"application/pdf", "image/png", "image/jpeg",
                      "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    _mime = f.get("mime_type", "application/pdf")
    if _mime not in _allowed_mimes:
        _mime = "application/octet-stream"
    _inline = request.args.get("inline") == "1"
    _disp = "inline" if _inline else "attachment"
    return Response(
        f["data"],
        mimetype=_mime,
        headers={"Content-Disposition": f"{_disp}; filename=\"{f['filename']}\""}
    )


@bp.route("/api/rfq/<rid>/files")
@auth_required
@safe_route
def api_rfq_files(rid):
    """List all files for an RFQ."""
    category = request.args.get("category")
    files = list_rfq_files(rid, category=category)
    return jsonify({"ok": True, "files": files, "count": len(files)})


@bp.route("/api/rfq/<rid>/file/<file_id>", methods=["DELETE"])
@auth_required
@safe_route
def api_delete_rfq_file(rid, file_id):
    """Delete a single rfq_files row.

    PR-AE 2026-05-13 (Bug 1 from PC #10846357 walkthrough): operator
    accidentally uploaded the wrong bid-package template and had no
    way to remove it without DB-side surgery. This route lets the UI
    delete a per-file chip safely.

    Safety rails:
      - file must belong to this rfq_id (defense against ID swaps)
      - category must be 'template' OR 'attachment' (NEVER 'generated' —
        those are Reytech outputs and have their own bulk-regen path)
      - returns the deleted filename so the toast can echo it back
    """
    _bad = _validate_rid(rid)
    if _bad:
        return _bad
    if not re.match(r"^rf_[a-f0-9]{8,16}$", file_id or ""):
        return jsonify({"ok": False, "error": "invalid file_id"}), 400
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, rfq_id, filename, category FROM rfq_files "
            "WHERE id=? AND rfq_id=? LIMIT 1",
            (file_id, rid),
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "file not found"}), 404
        cat = (row["category"] or "").lower()
        if cat not in ("template", "attachment"):
            return jsonify({
                "ok": False,
                "error": f"refusing to delete category '{cat}' — use "
                         "the regen path for generated outputs",
            }), 403
        conn.execute("DELETE FROM rfq_files WHERE id=? AND rfq_id=?",
                     (file_id, rid))
        conn.commit()
        log.info("Deleted rfq_file id=%s filename=%s category=%s rfq_id=%s",
                 file_id, row["filename"], cat, rid)
    return jsonify({"ok": True, "deleted": 1, "filename": row["filename"]})


# ═══════════════════════════════════════════════════════════════════════
# RFQ Status Management — reopen, edit, resubmit
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/rfq/<rid>/reopen", methods=["POST"])
@auth_required
@safe_route
def rfq_reopen(rid):
    """Reopen an RFQ for editing. Changes status back to 'ready'."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error")
        return redirect("/")
    
    old_status = r.get("status", "?")
    _transition_status(r, "ready", actor="user", notes=f"Reopened from '{old_status}'")
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, "ready")
    except Exception as _e:
        log.debug('suppressed in rfq_reopen: %s', _e)

    _log_rfq_activity(rid, "reopened",
        f"RFQ #{r.get('solicitation_number','?')} reopened for editing (was: {old_status})",
        actor="user", metadata={"old_status": old_status})
    
    flash(f"RFQ reopened for editing (was: {old_status})", "info")
    return redirect(f"/rfq/{rid}")


def _rfq_to_pc_for_qa(rfq: dict) -> dict:
    """Adapter: shape an RFQ dict into the form pc_qa_agent.run_qa expects.

    Delegates to the canonical record-field accessor so PC and RFQ shapes
    both yield correct revenue/profit readings without mutating the source.
    See src/core/record_fields.build_qa_view.
    """
    from src.core.record_fields import build_qa_view
    return build_qa_view(rfq)


@bp.route("/api/rfq/<rid>/qa", methods=["GET", "POST"])
@auth_required
@safe_route
def api_rfq_qa(rid):
    """Run the existing PC QA helper against an RFQ. Same rules — no new
    QA logic — so the RFQ page can hard-block destructive actions until
    blockers clear, mirroring the PC review page gate.
    """
    try:
        rfqs = load_rfqs()
        rfq = rfqs.get(rid)
        if not rfq:
            return jsonify({"ok": False, "error": "RFQ not found"}), 404
        pc_view = _rfq_to_pc_for_qa(rfq)
        # Default to skip-LLM on the RFQ side: the gate is meant to be fast
        # and reproducible. Caller can opt in with ?llm=1.
        use_llm = request.args.get("llm", "0") == "1"
        from src.agents.pc_qa_agent import run_qa
        import copy as _copy
        report = run_qa(_copy.deepcopy(pc_view), use_llm=use_llm)
        return jsonify(report)
    except Exception as e:
        log.error("RFQ QA error for %s: %s", rid, e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/rfq/<rid>/update-status", methods=["POST"])
@auth_required
@safe_route
def api_rfq_update_status_json(rid):
    """Update RFQ status via JSON (AJAX).

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_update_status_json_locked(rid)


def _api_rfq_update_status_json_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    data = request.get_json(force=True, silent=True) or {}
    new_status = data.get("status", "").strip()
    notes = data.get("notes", "").strip()

    from src.core.status_taxonomy import is_valid_status_for
    if not is_valid_status_for("rfq", new_status):
        return jsonify({"ok": False, "error": f"Invalid status: {new_status}"})

    old_status = r.get("status", "?")
    r["status"] = new_status
    if notes:
        r["status_notes"] = notes
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)

    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, new_status)
    except Exception as _e:
        log.debug('suppressed in api_rfq_update_status_json: %s', _e)

    try:
        from src.core.dal import log_lifecycle_event
        log_lifecycle_event("rfq", rid, "status_changed",
            f"Status: {old_status} → {new_status}" + (f" ({notes})" if notes else ""),
            actor="user")
    except Exception as _e:
        log.debug('suppressed in api_rfq_update_status_json: %s', _e)

    return jsonify({"ok": True, "old_status": old_status, "new_status": new_status})


@bp.route("/rfq/<rid>/update-status-form", methods=["POST"])
@auth_required
@safe_route
def rfq_update_status(rid):
    """Change RFQ status to any valid state."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error")
        return redirect("/")
    
    new_status = request.form.get("status", "").strip()
    from src.core.status_taxonomy import is_valid_status_for
    if not is_valid_status_for("rfq", new_status):
        flash(f"Invalid status: {new_status}", "error")
        return redirect(f"/rfq/{rid}")
    
    old_status = r.get("status", "?")
    notes = request.form.get("notes", "").strip()
    _transition_status(r, new_status, actor="user", notes=notes or f"Changed from {old_status}")
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, new_status)
    except Exception as _e:
        log.debug('suppressed in rfq_update_status: %s', _e)

    _log_rfq_activity(rid, "status_changed",
        f"RFQ #{r.get('solicitation_number','?')} status: {old_status} → {new_status}" + (f" ({notes})" if notes else ""),
        actor="user", metadata={"old_status": old_status, "new_status": new_status, "notes": notes})

    flash(f"Status changed: {old_status} → {new_status}", "success")
    return redirect(f"/rfq/{rid}")


# ═══════════════════════════════════════════════════════════════════════
# Manual mark-as-sent — escape valve for out-of-band submissions
# (Bundle-5 PR-5b, audit AA 2026-04-22).
#
# When the operator emails a quote outside the app (manual Gmail, merged
# PDF attached by hand — see scripts/merge_rfq_package_10840486.py), the
# in-app send path never fires and the record stays in status=generated
# forever. This endpoint is the explicit "I sent this myself" escape
# hatch: stamps the same metadata an in-app send would, uploads the
# attachment the operator actually sent (so the record reflects reality
# rather than the app's stale auto-generated PDF), and fires the on_sent
# hooks (Drive archive + lifecycle log + activity log).
#
# NOT called automatically — operator must click the "Mark Sent Manually"
# button on RFQ detail. Double-click is idempotent: resending to status=sent
# a record already at sent overwrites the manual metadata but never
# re-fires Drive uploads (the hook no-ops when re-run).
# ═══════════════════════════════════════════════════════════════════════


def _save_manual_attachment(rid: str, file_storage) -> dict:
    """Persist an uploaded attachment to BOTH `uploads/manual_sent/`
    (legacy filesystem path; metadata pointer for backwards compat)
    AND the `rfq_files` BLOB store (durable across redeploys, ships to
    Google Drive via the existing backup hook). Returns
    `{filename, path, size, rfq_files_id}`; empty dict on save failure.

    PR mr-wolf #4d — Mike's directive 2026-05-13 EOD: "we have google
    drive as a backup DB, should the app be saving attachments? we
    only need to save for a year, storage getting cheaper, might be
    good to capture everything for agentic use". Storage cost is
    negligible (~1-2 GB/year at current bid volume); the data
    becomes substrate for future agentic workflows (replay-bid,
    audit-corpus, alternate-prior selection in mirror-fill).

    Saves are independent — filesystem failure doesn't block BLOB
    write, BLOB failure doesn't block filesystem write. The function
    returns whichever succeeded so the caller's metadata reflects
    reality.
    """
    if not file_storage or not getattr(file_storage, "filename", ""):
        return {}
    fname = os.path.basename(file_storage.filename)
    # Strip path traversal; keep spaces and extensions.
    fname = re.sub(r"[\\\\/]", "_", fname)[:200] or "attachment.pdf"
    out: dict = {"filename": fname}

    # Read the bytes ONCE — `file_storage.save()` consumes the stream,
    # and we need the bytes for both the filesystem AND BLOB write.
    try:
        file_storage.stream.seek(0)
        blob = file_storage.stream.read()
    except Exception as _re:
        log.error("manual-sent: file_storage read failed for %s: %s", rid, _re)
        return {}
    if not blob:
        return {}

    # Filesystem write (legacy path; some operator tooling still reads
    # from `uploads/manual_sent/`). Failure here is recoverable —
    # the BLOB write below is the durable copy.
    dest_dir = os.path.join(UPLOAD_DIR, "manual_sent", rid)
    try:
        os.makedirs(dest_dir, exist_ok=True)
        dest = os.path.join(dest_dir, fname)
        with open(dest, "wb") as _f:
            _f.write(blob)
        out["path"] = dest
        out["size"] = len(blob)
    except OSError as e:
        log.warning("manual-sent FS save degraded for %s: %s", rid, e)

    # BLOB write — the durable copy. Goes to `rfq_files` with
    # category='sent_artifact'. The nightly Drive backup picks up
    # rfq_files BLOBs automatically, so this also lands the file on
    # Mike's Google Drive without an additional integration.
    try:
        from src.api.dashboard import save_rfq_file
        rfq_files_id = save_rfq_file(
            rid, fname, "manual_sent_attachment", blob,
            category="sent_artifact", uploaded_by="user",
        )
        out["rfq_files_id"] = rfq_files_id
        # If FS write failed above, fall back to a synthetic path so
        # legacy readers don't crash on `attachment["path"]` access.
        out.setdefault("path", f"rfq_files://{rfq_files_id}")
        out.setdefault("size", len(blob))
    except Exception as _be:
        log.warning("manual-sent BLOB save degraded for %s: %s", rid, _be)
        # If BOTH writes failed we have nothing to return.
        if "path" not in out:
            return {}

    return out


@bp.route("/api/rfq/<rid>/mark-sent-manually", methods=["POST"])
@auth_required
@safe_route
def api_rfq_mark_sent_manually(rid):
    """Mark RFQ as sent when the operator emailed it outside the app.

    Accepts multipart form data OR JSON:
      sent_to       — recipient email (defaults to buyer email on record)
      sent_at       — ISO datetime (defaults to now)
      notes         — freeform note saved to activity log
      attachment    — file operator actually sent (stored under
                      uploads/manual_sent/<rid>/)

    Writes: status=sent, sent_at, sent_to, sent_method="manual",
            manual_sent_metadata={source, attachment, actor, timestamp}.
    Fires: Drive archive hook, lifecycle event, CRM activity entry.

    Race-safe wrapper (PR #778 pattern).
    """
    # Read the request OUTSIDE the lock so the locked inner is pure-data
    # and reusable from non-Flask-request contexts (e.g.,
    # gmail_sent_watcher firing the same pipeline for an outbound
    # message it observed in Gmail SENT — PR #9 2026-05-26).
    is_multipart = (request.content_type or "").startswith("multipart/")
    if is_multipart:
        payload = request.form.to_dict()
        uploaded = request.files.get("attachment")
    else:
        payload = request.get_json(force=True, silent=True) or {}
        uploaded = None
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_mark_sent_manually_locked(
            rid, payload=payload, uploaded=uploaded,
        )


def _api_rfq_mark_sent_manually_locked(rid, *, payload=None, uploaded=None):
    """Inner body — always runs under `_save_rfqs_lock`.

    Pure-data entry point: callers pass `payload` (dict) and `uploaded`
    (Werkzeug FileStorage or None). The HTTP route reads the request and
    forwards; background callers (gmail_sent_watcher) construct the dict
    directly. Both go through the same lock + side-effect pipeline.
    """
    from src.api.data_layer import load_rfqs, _save_single_rfq
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404

    payload = payload or {}

    now_iso = datetime.now().isoformat()
    sent_to = (payload.get("sent_to") or r.get("requestor_email", "") or "").strip()
    sent_at = (payload.get("sent_at") or now_iso).strip()
    notes = (payload.get("notes") or "").strip()

    attachment = _save_manual_attachment(rid, uploaded)

    old_status = r.get("status", "")
    # Substrate-singleness (PR #11 / audit Item 6): single writer for
    # the 'sent' transition. Replaces the 4-step
    # `_transition_status + r["sent_at"]/sent_to/sent_method` block
    # + the propagate_sent_to_quote_row call below.
    from src.core.quote_lifecycle_shared import mark_sent_in_place
    mark_sent_in_place(
        r, sent_at=sent_at, sent_to=sent_to, sent_method="manual",
        notes=notes or "Marked sent manually (out-of-band)",
        source="user",
    )

    # Bug-2 audit (2026-05-02): the SCPRS award_tracker enrollment query
    # filters `WHERE total > 0`. If `r.total` was never persisted (some
    # paths only compute it lazily at PDF render time), the row gets
    # silently skipped from the award poll loop — sent quote never
    # triggers SCPRS won/lost detection. Compute total from items here
    # if it's missing so mark-sent always lands the row in the poll.
    try:
        if not float(r.get("total") or 0):
            _items = r.get("items") or r.get("line_items") or []
            _computed = sum(
                float(it.get("price_per_unit") or it.get("unit_price") or it.get("bid_price") or 0)
                * float(it.get("qty") or 1)
                for it in _items
                if isinstance(it, dict)
            )
            if _computed > 0:
                r["total"] = round(_computed, 2)
                log.info("RFQ %s: backfilled total=%.2f at mark-sent for award_tracker enrollment",
                         rid, r["total"])
    except (TypeError, ValueError) as _te:
        log.debug("total backfill failed for %s: %s", rid, _te)
    r["manual_sent_metadata"] = {
        "marked_at": now_iso,
        "sent_at_reported": sent_at,
        "sent_to": sent_to,
        "actor": "user",
        "notes": notes,
        "attachment": attachment or None,
        "prior_status": old_status,
    }

    # Persist — JSON + SQLite (via _save_single_rfq) + DAL update so every
    # consumer sees the flip. Mirror the defensive pattern used elsewhere in
    # this module: DAL errors are logged, not raised — the JSON write is
    # still considered authoritative.
    _save_single_rfq(rid, r)
    try:
        from src.core.dal import update_rfq_status
        update_rfq_status(rid, "sent", actor="user")
    except Exception as _e:
        log.debug("DAL update_rfq_status(sent) suppressed: %s", _e)

    # propagate to quotes table is now handled INSIDE mark_sent_in_place
    # above (PR #11 substrate-singleness — single writer for 'sent').

    # Plan §4.1: KPI telemetry — measure time-to-send for the <90s KPI.
    # Best-effort; never blocks the mark-sent flip.
    try:
        from src.core.operator_kpi import log_quote_sent
        _items = r.get("items") or r.get("line_items") or []
        log_quote_sent(
            quote_id=rid, quote_type="rfq",
            started_at=r.get("created_at") or r.get("opened_at"),
            item_count=len(_items),
            agency_key=(r.get("agency_key") or r.get("agency") or ""),
            quote_total=float(r.get("total") or 0),
        )
    except Exception as _kpi_e:
        log.debug("KPI logging suppressed: %s", _kpi_e)

    # PR-I (2026-05-13): operator-drift per-line capture. Higher-signal
    # than WR at low monthly volume — every Mark-Sent emits up to ~10
    # rows of (sent_price, rec_price, caps_applied) into the
    # operator_drift_line table. The shadow-mode cap evaluator (PR-J)
    # reads from this surface.
    # PR-K1: capture quote_number too so award_monitor (knows pc.id)
    # AND quote_lifecycle (knows quote_number) can both backfill the
    # outcome later via resolve_drift_outcome.
    _qn = (r.get("quote_number") or r.get("reytech_quote_number") or "")
    try:
        from src.core.operator_kpi import log_operator_drift
        log_operator_drift(
            quote_id=rid, quote_type="rfq",
            items=r.get("items") or r.get("line_items") or [],
            agency_key=(r.get("agency_key") or r.get("agency") or ""),
            quote_number=_qn,
        )
    except Exception as _drift_e:
        log.debug("operator_drift logging suppressed: %s", _drift_e)

    # PR-J (2026-05-13): shadow-mode cap evaluator. For every line
    # with scprs_rollup data, log what the SCPRS p75 cap WOULD have
    # done if enabled — even if the live recommendation didn't bind.
    # Answers "should I flip ORACLE_USE_SCPRS_ROLLUP on?" with real
    # data instead of a coin flip.
    try:
        from src.core.operator_kpi import log_operator_drift_shadow
        log_operator_drift_shadow(
            quote_id=rid, quote_type="rfq",
            items=r.get("items") or r.get("line_items") or [],
            agency_key=(r.get("agency_key") or r.get("agency") or ""),
            quote_number=_qn,
        )
    except Exception as _shadow_e:
        log.debug("shadow drift logging suppressed: %s", _shadow_e)

    # Fire on_sent hooks. Each wrapped so one failure cannot block the
    # mark-sent flip — the status write is the source of truth; hooks are
    # best-effort archive/log side-effects.
    try:
        from src.agents.drive_triggers import on_quote_sent
        on_quote_sent(r, email_body=notes, to_email=sent_to)
    except Exception as _e:
        log.debug("on_quote_sent suppressed: %s", _e)
    try:
        from src.core.dal import log_lifecycle_event
        log_lifecycle_event("rfq", rid, "package_sent_manual",
                            f"Marked sent manually to {sent_to or '—'}"
                            + (f" · {notes}" if notes else ""),
                            actor="user")
    except Exception as _e:
        log.debug("log_lifecycle_event(rfq sent manual) suppressed: %s", _e)
    try:
        _log_rfq_activity(rid, "sent_manually",
                          f"RFQ #{r.get('solicitation_number','?')} marked sent manually"
                          + (f" (attachment: {attachment.get('filename')})" if attachment else ""),
                          actor="user",
                          metadata={"sent_to": sent_to, "sent_at": sent_at,
                                    "attachment": bool(attachment),
                                    "prior_status": old_status})
    except Exception as _e:
        log.debug("_log_rfq_activity(sent_manually) suppressed: %s", _e)

    # PR mr-wolf #4b — auto-capture every generated form in this RFQ
    # into the prior_submissions table. Mark Sent is the operator-
    # blessed "this packet is canonical" signal; the next 703A/703C/
    # AMS 708 inbound that hits `fill_703a` / `fill_703c` / etc. will
    # find these priors and mirror-fill automatically. Fire-and-forget
    # — capture failures NEVER block the mark-sent flip.
    try:
        from src.forms.prior_submissions import (
            capture_from_rfq_generated_files,
            capture as _ps_capture,
            _form_id_from_filename as _ps_form_id_from_filename,
        )
        from src.forms.form_registry import all_form_ids as _ps_all_form_ids
        _captured = capture_from_rfq_generated_files(
            rid,
            agency_key=(r.get("agency_key") or r.get("agency") or ""),
            source_quote_number=r.get("reytech_quote_number", ""),
        )

        # PR mr-wolf #4d — ALSO capture the operator-uploaded
        # attachment (what was ACTUALLY emailed) as a BLESSED prior.
        # This is strictly more canonical than the auto-generated
        # PDFs from rfq_files (which might have been hand-edited
        # before send). The `_save_manual_attachment` upgrade earlier
        # in this PR persisted the file to rfq_files with
        # category='sent_artifact'; we read it back from the same
        # blob and stamp it into prior_submissions with blessed=True
        # so `latest_for` ranks it over any auto-generated peer.
        try:
            _manual_attachment = attachment if isinstance(attachment, dict) else {}
            _manual_path = _manual_attachment.get("path") or ""
            _manual_filename = _manual_attachment.get("filename") or ""
            _known_forms = {f.lower(): f for f in _ps_all_form_ids()}
            _manual_form_id = _ps_form_id_from_filename(_manual_filename, _known_forms)
            # Resolve bytes — prefer the BLOB store (durable), fall back
            # to filesystem path if BLOB write failed in _save_manual_attachment.
            _manual_bytes = None
            _manual_rfq_files_id = _manual_attachment.get("rfq_files_id")
            if _manual_rfq_files_id:
                try:
                    from src.core.db import get_db as _ps_get_db
                    with _ps_get_db() as _conn:
                        _row = _conn.execute(
                            "SELECT data FROM rfq_files WHERE id = ?",
                            (_manual_rfq_files_id,),
                        ).fetchone()
                        if _row and _row["data"]:
                            _manual_bytes = bytes(_row["data"])
                except Exception as _ge:
                    log.debug("prior_submissions: BLOB read for manual attachment failed: %s", _ge)
            if _manual_bytes is None and _manual_path and os.path.exists(_manual_path):
                with open(_manual_path, "rb") as _mf:
                    _manual_bytes = _mf.read()

            if _manual_bytes and _manual_form_id:
                _blessed_id = _ps_capture(
                    _manual_form_id, _manual_bytes,
                    agency_key=(r.get("agency_key") or r.get("agency") or ""),
                    source_rfq_id=rid,
                    source_quote_number=r.get("reytech_quote_number", ""),
                    filename=_manual_filename,
                    blessed=True,
                )
                if _blessed_id:
                    _captured += 1
                    log.info(
                        "prior_submissions BLESSED capture: rfq=%s form=%s "
                        "filename=%s (operator-uploaded — supersedes auto-generated peer)",
                        rid, _manual_form_id, _manual_filename,
                    )
            elif _manual_bytes and not _manual_form_id:
                log.debug(
                    "prior_submissions: manual attachment filename=%r "
                    "did not match a known form_id — skipping blessed capture",
                    _manual_filename,
                )
        except Exception as _bc_e:
            log.debug("prior_submissions blessed-attachment capture suppressed: %s", _bc_e)

        if _captured:
            log.info(
                "prior_submissions auto-capture: rfq=%s captured %d form(s) total",
                rid, _captured,
            )
    except Exception as _cap_e:
        log.debug("prior_submissions auto-capture suppressed: %s", _cap_e)

    log.info("RFQ %s marked SENT manually: sent_to=%s attachment=%s prior=%s",
             rid, sent_to, bool(attachment), old_status)
    return jsonify({
        "ok": True,
        "status": "sent",
        "sent_at": sent_at,
        "sent_to": sent_to,
        "attachment": attachment or None,
        "prior_status": old_status,
    })


# ═══════════════════════════════════════════════════════════════════════
# RFQ Activity Log
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/<rid>/activity")
@auth_required
@safe_route
def api_rfq_activity(rid):
    """Get activity log for an RFQ."""
    activities = _get_crm_activity(ref_id=rid, limit=50)
    return jsonify({"ok": True, "activities": activities, "count": len(activities)})


# ═══════════════════════════════════════════════════════════════════════
# Email Templates API
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/email-templates")
@auth_required
@safe_route
def api_list_email_templates():
    """List email templates, optionally filtered by category."""
    category = request.args.get("category")
    templates = get_email_templates_db(category)
    return jsonify({"ok": True, "templates": templates})


@bp.route("/api/email-templates/<tid>", methods=["GET"])
@auth_required
@safe_route
def api_get_email_template(tid):
    """Get a single email template by ID."""
    templates = get_email_templates_db()
    t = next((t for t in templates if t["id"] == tid), None)
    if not t:
        return jsonify({"ok": False, "error": "Template not found"}), 404
    return jsonify({"ok": True, "template": t})


@bp.route("/api/email-templates", methods=["POST"])
@auth_required
@safe_route
def api_create_email_template():
    """Create or update an email template."""
    data = request.get_json(force=True, silent=True) or request.form
    tid = save_email_template_db(
        data.get("id", ""), data.get("name", ""), data.get("category", "rfq"),
        data.get("subject", ""), data.get("body", ""), int(data.get("is_default", 0)))
    if tid:
        return jsonify({"ok": True, "id": tid})
    return jsonify({"ok": False, "error": "Save failed"}), 500


@bp.route("/api/email-templates/<tid>", methods=["DELETE"])
@auth_required
@safe_route
def api_delete_email_template(tid):
    """Delete an email template."""
    try:
        from src.core.db import get_db
        with get_db() as conn:
            conn.execute("DELETE FROM email_templates WHERE id = ?", (tid,))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/email-templates/render", methods=["POST"])
@auth_required
@safe_route
def api_render_email_template():
    """Render a template with variables. POST {template_id, variables: {...}}"""
    data = request.get_json(force=True, silent=True) or {}
    tid = data.get("template_id", "")
    variables = data.get("variables", {})
    
    templates = get_email_templates_db()
    t = next((t for t in templates if t["id"] == tid), None)
    if not t:
        return jsonify({"ok": False, "error": "Template not found"}), 404
    
    subject = t["subject"]
    body = t["body"]
    for key, val in variables.items():
        subject = subject.replace("{{" + key + "}}", str(val))
        body = body.replace("{{" + key + "}}", str(val))
    
    return jsonify({"ok": True, "subject": subject, "body": body})


# ═══════════════════════════════════════════════════════════════════════
# PDF Preview from DB
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/rfq/<rid>/preview/<file_id>")
@auth_required
@safe_page
def rfq_preview_pdf(rid, file_id):
    """Serve a PDF for inline preview (Content-Disposition: inline)."""
    f = get_rfq_file(file_id)
    if not f or f.get("rfq_id") != rid:
        return "File not found", 404
    from flask import Response
    return Response(
        f["data"],
        mimetype="application/pdf",
        headers={"Content-Disposition": f"inline; filename=\"{f['filename']}\""}
    )


# ═══════════════════════════════════════════════════════════════════════
# Email Signature — get/save HTML signature for outbound emails
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/email-signature")
@auth_required
@safe_route
def get_email_signature():
    """Get current email signature config."""
    email_cfg = CONFIG.get("email", {})
    sig_html = email_cfg.get("signature_html", "")

    # Auto-generate default signature on first load if empty
    if not sig_html:
        sig_html = _build_default_signature()
        CONFIG.setdefault("email", {})["signature_html"] = sig_html

    return jsonify({
        "ok": True,
        "signature_html": sig_html,
        "signature_enabled": email_cfg.get("signature_enabled", True),
    })

@bp.route("/api/email-signature", methods=["POST"])
@auth_required
@safe_route
def save_email_signature():
    """Save email signature HTML."""
    data = request.get_json(force=True)
    sig_html = data.get("signature_html", "")
    
    CONFIG.setdefault("email", {})["signature_html"] = sig_html
    CONFIG["email"]["signature_enabled"] = True
    
    # Persist to config file
    import json as _json
    for cfg_path in [
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "reytech_config.json"),
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "forms", "reytech_config.json"),
    ]:
        try:
            with open(cfg_path) as f:
                cfg = _json.load(f)
            cfg.setdefault("email", {})["signature_html"] = sig_html
            cfg["email"]["signature_enabled"] = True
            with open(cfg_path, "w") as f:
                _json.dump(cfg, f, indent=2)
        except Exception as _e:
            log.debug("Suppressed: %s", _e)
    
    return jsonify({"ok": True})


@bp.route("/api/upload-sig-logo", methods=["POST"])
@auth_required
@safe_route
@rate_limit("heavy")
def upload_sig_logo():
    """Upload a PNG/JPG logo for the email signature. Returns base64 data URI."""
    import base64 as _b64
    if "logo" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400
    f = request.files["logo"]
    if not f.filename:
        return jsonify({"ok": False, "error": "Empty filename"}), 400

    data = f.read()
    if len(data) > 5_000_000:
        return jsonify({"ok": False, "error": "File too large (max 5MB)"}), 400

    fname = f.filename.lower()
    if fname.endswith(".png"):
        mime = "image/png"
    elif fname.endswith((".jpg", ".jpeg")):
        mime = "image/jpeg"
    elif fname.endswith(".gif"):
        mime = "image/gif"
    else:
        return jsonify({"ok": False, "error": "PNG/JPG/GIF only"}), 400

    # Resize for email if large
    try:
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(data))
        if img.width > 200:
            ratio = 200 / img.width
            img = img.resize((200, int(img.height * ratio)), Image.LANCZOS)
            buf = _io.BytesIO()
            img.save(buf, "PNG", optimize=True)
            data = buf.getvalue()
            mime = "image/png"
    except Exception as _e:
        log.debug('suppressed in upload_sig_logo: %s', _e)

    b64 = _b64.b64encode(data).decode()
    data_uri = f"data:{mime};base64,{b64}"

    # Save to data/ for future use
    try:
        save_path = os.path.join(DATA_DIR, "email_logo.png")
        with open(save_path, "wb") as _fw:
            _fw.write(data)
    except Exception as _e:
        log.debug('suppressed in upload_sig_logo: %s', _e)

    return jsonify({"ok": True, "data_uri": data_uri, "size": len(data)})


def _build_default_signature():
    """Return empty — Gmail auto-appends the canonical signature.

    Previously auto-generated a hardcoded Reytech HTML signature on first
    load of /api/email-signature, which stacked on top of Gmail's own
    auto-sig on every send. Per CLAUDE.md "Gmail Handles Signatures", the
    app must not inject one. Users who want an explicit compose-time sig
    can still save one via POST /api/email-signature (that's a user
    choice, not an app default).
    """
    return ""


# ═══════════════════════════════════════════════════════════════════════
# Enhanced Email Send — DB attachments + email logging + CRM tracking
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/rfq/<rid>/save-draft", methods=["POST"])
@auth_required
@safe_page
def save_gmail_draft(rid):
    """Save email as Gmail draft — user reviews and sends manually from Gmail."""
    from src.api.trace import Trace
    t = Trace("email_draft", rfq_id=rid)

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        flash("RFQ not found", "error")
        return redirect("/")

    to_addr = request.form.get("to", "").strip()
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    cc = request.form.get("cc", "").strip()
    attach_ids = [x.strip() for x in request.form.get("attach_files", "").split(",") if x.strip()]

    if not to_addr or not subject:
        flash("Draft requires To and Subject", "error")
        return redirect(f"/rfq/{rid}")

    import tempfile, shutil

    tmp_dir = tempfile.mkdtemp(prefix="rfq_draft_")
    try:
        email_cfg = CONFIG.get("email", {})
        from src.core.email_signature import NAME as _SIG_NAME, COMPANY as _SIG_COMPANY
        from_name = email_cfg.get("from_name", f"{_SIG_NAME} - {_SIG_COMPANY}")
        from_addr = email_cfg.get("email", os.environ.get("GMAIL_ADDRESS", "sales@reytechinc.com"))

        try:
            from src.core.email_signature import wrap_html_email
            body_html = wrap_html_email(body)
        except Exception:
            body_html = ""

        # Stage DB-stored attachments to temp files for the Gmail API helper
        attachment_paths = []
        attached = []
        for fid in attach_ids:
            f = get_rfq_file(fid)
            if f and f.get("data"):
                path = os.path.join(tmp_dir, f["filename"])
                with open(path, "wb") as _fw:
                    _fw.write(f["data"])
                attachment_paths.append(path)
                attached.append(f["filename"])

        from src.core import gmail_api
        service = gmail_api.get_send_service("sales")
        response = gmail_api.save_draft(
            service,
            to=to_addr,
            subject=subject,
            body_plain=body,
            body_html=body_html or "",
            cc=cc or None,
            attachments=attachment_paths or None,
            from_name=from_name,
            from_addr=from_addr,
        )
        draft_id = response.get("id", "?") if isinstance(response, dict) else "?"
        t.ok("Draft saved", draft_id=draft_id, attachments=len(attached))
        flash(
            f"✅ Draft saved to Gmail — open Gmail to review and send ({len(attached)} attachments)",
            "success",
        )

    except Exception as e:
        t.fail("Draft save failed", error=str(e))
        flash(f"Draft save failed: {e}", "error")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    return redirect(f"/rfq/{rid}")


@bp.route("/rfq/<rid>/send-email", methods=["POST"])
@auth_required
@safe_route
def send_email_enhanced(rid):
    """DEPRECATED 2026-05-01 (PR-B3) — direct-send disabled.

    Mike's directive: "do not direct send, i want to see eveyrhting first."
    All RFQ sends now route through /rfq/<id>/review-package → Create
    Gmail Draft (with double-sig pre-flight + thread binding) → operator
    reviews + sends from Gmail itself.

    Redirect rather than 410 here because this route was hit via HTML
    <form action="..."> POST — a JSON 410 would render an unhelpful raw
    page in the browser. Redirect lands the operator on the right place.
    """
    flash("Send happens on the review page now — create a Gmail draft + review there.",
          "info")
    return redirect(f"/rfq/{rid}/review-package")


def _send_email_enhanced_legacy_DISABLED(rid):
    from src.api.trace import Trace
    t = Trace("email_send", rfq_id=rid)

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        t.fail("RFQ not found")
        flash("RFQ not found", "error")
        return redirect("/")

    # Validate before sending
    from src.core.quote_validator import validate_ready_to_send
    validation = validate_ready_to_send(r)
    if not validation["ok"]:
        t.fail("Send validation failed", errors=validation["errors"])
        flash(f"Cannot send: {'; '.join(validation['errors'])}", "error")
        return redirect(f"/rfq/{rid}")

    # Get editable fields from form
    to_addr = request.form.get("to", "").strip()
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    cc = request.form.get("cc", "").strip()
    bcc = request.form.get("bcc", "").strip()
    attach_ids = [x.strip() for x in request.form.get("attach_files", "").split(",") if x.strip()]
    
    if not to_addr or not subject:
        flash("Email requires To and Subject", "error")
        return redirect(f"/rfq/{rid}")
    
    t.step("Preparing email", to=to_addr, attachments=len(attach_ids))
    
    # Build attachment list from DB files
    import tempfile, shutil
    tmp_dir = tempfile.mkdtemp(prefix="rfq_send_")
    attachment_paths = []
    attachment_names = []
    
    try:
        for fid in attach_ids:
            f = get_rfq_file(fid)
            if f and f.get("data"):
                path = os.path.join(tmp_dir, f["filename"])
                with open(path, "wb") as _fw:
                    _fw.write(f["data"])
                attachment_paths.append(path)
                attachment_names.append(f["filename"])
                t.step(f"Attached: {f['filename']}")
        
        # Also check filesystem for any files not in DB yet
        if not attach_ids and r.get("output_files"):
            out_dir = os.path.join(UPLOAD_DIR, rid)
            for fname in r["output_files"]:
                fpath = os.path.join(out_dir, fname)
                if os.path.exists(fpath):
                    attachment_paths.append(fpath)
                    attachment_names.append(fname)
        
        # Send via SMTP — include HTML signature if enabled
        draft = {
            "to": to_addr,
            "subject": subject,
            "body": body,
            "cc": cc,
            "bcc": bcc,
            "attachments": attachment_paths,
        }
        
        # Threading: if RFQ came from email, reply to that thread
        msg_id = r.get("email_message_id", "")
        if msg_id:
            draft["in_reply_to"] = msg_id
            draft["references"] = msg_id
        
        include_sig = request.form.get("include_signature") == "1"
        email_cfg = CONFIG.get("email", {})
        sig_html = email_cfg.get("signature_html", "")
        
        if include_sig and sig_html:
            # Build HTML body: plain text body + signature
            import html as _html
            body_escaped = _html.escape(body).replace("\n", "<br>")
            draft["body_html"] = f"""<div style="font-family:'Segoe UI',Arial,sans-serif;font-size:14px;color:#222;line-height:1.6">
{body_escaped}
<br><br>
<div style="border-top:1px solid #ddd;padding-top:10px;margin-top:10px">
{sig_html}
</div>
</div>"""
            t.step("HTML signature included")
        
        sender = EmailSender(CONFIG.get("email", {}))
        sender.send(draft)
        
        # PR #11 substrate-singleness: single writer for 'sent'.
        # Replaces _transition_status + r["sent_at"] block + the
        # propagate_sent_to_quote_row call below.
        from src.core.quote_lifecycle_shared import mark_sent_in_place
        mark_sent_in_place(
            r, sent_method="email", notes=f"Email sent to {to_addr}",
            source="user",
        )
        r["draft_email"] = {"to": to_addr, "subject": subject, "body": body, "cc": cc, "bcc": bcc}
        from src.api.dashboard import _save_single_rfq
        _save_single_rfq(rid, r)
        try:
            from src.core.dal import update_rfq_status as _dal_ur
            _dal_ur(rid, "sent")
        except Exception as _e:
            log.debug('suppressed in send_email_enhanced: %s', _e)
        
        # ── Log to email_log table ──
        sol = r.get("solicitation_number", "")
        qn = r.get("reytech_quote_number", "")
        
        # Find contact_id from recipient email
        contact_id = ""
        try:
            from src.core.db import get_db
            with get_db() as conn:
                row = conn.execute("SELECT id FROM contacts WHERE buyer_email = ?", (to_addr.lower(),)).fetchone()
                if row:
                    contact_id = row[0]
        except Exception as _e:
            log.debug("Suppressed: %s", _e)
        
        email_log_id = log_email_sent_db(
            direction="outbound", sender=sender.email_addr, recipient=to_addr,
            subject=subject, body=body, attachments=attachment_names,
            quote_number=qn, rfq_id=rid, contact_id=contact_id)
        t.step(f"Email logged (id={email_log_id})")
        
        # ── CRM activity: log against quote AND contact ──
        _log_rfq_activity(rid, "email_sent",
            f"Bid response emailed to {to_addr} for Sol #{sol} ({len(attachment_names)} attachments)",
            actor="user", metadata={"to": to_addr, "quote": qn, "files": attachment_names, "email_log_id": email_log_id})
        
        if qn:
            _log_crm_activity(qn, "email_sent",
                f"Quote {qn} emailed to {to_addr} for Sol #{sol}",
                actor="user", metadata={"to": to_addr, "rfq_id": rid})
            if QUOTE_GEN_AVAILABLE:
                update_quote_status(qn, "sent", actor="system")
        
        if contact_id:
            _log_crm_activity(contact_id, "email_sent",
                f"Bid response for Sol #{sol} (Quote {qn}) sent to {to_addr}",
                actor="user", metadata={"rfq_id": rid, "quote": qn, "solicitation": sol})
        
        t.ok("Email sent", to=to_addr, attachments=len(attachment_names))
        flash(f"✅ Email sent to {to_addr} with {len(attachment_names)} attachments", "success")
        
    except Exception as e:
        t.fail("Send failed", error=str(e))
        flash(f"Send failed: {e}. Try 'Open in Mail App' instead.", "error")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    
    return redirect(f"/rfq/{rid}")


# ═══════════════════════════════════════════════════════════════════════
# Email History API (for contact/quote level)
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/email-history")
@auth_required
@safe_route
def api_email_history():
    """Get email history. Filter by ?rfq_id=, ?quote_number=, ?contact_id="""
    try:
        from src.core.db import get_db
        with get_db() as conn:
            query = "SELECT id, logged_at, direction, sender, recipient, subject, body_preview, attachments_json, quote_number, rfq_id, contact_id, status FROM email_log WHERE 1=1"
            params = []
            for field in ("rfq_id", "quote_number", "contact_id"):
                val = request.args.get(field)
                if val:
                    query += f" AND {field} = ?"
                    params.append(val)
            query += " ORDER BY logged_at DESC LIMIT 50"
            rows = conn.execute(query, params).fetchall()
            return jsonify({"ok": True, "emails": [dict(r) for r in rows], "count": len(rows)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ═══════════════════════════════════════════════════════════════════════
# OBS 1600 — CA Agricultural Food Product Certification
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/food-classify", methods=["POST"])
@auth_required
@safe_route
def api_food_classify():
    """Classify quote/RFQ items into CDCR food category codes.
    Body: {"items": [{"description": "..."}, ...]}
    Returns classified items with food codes.
    """
    try:
        from src.forms.food_classifier import classify_quote_items
        data = request.get_json(force=True)
        items = data.get("items", [])
        classified = classify_quote_items(items)
        food_count = sum(1 for r in classified if r['is_food'])
        return jsonify({"ok": True, "items": classified, "food_count": food_count,
                        "total_count": len(classified)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/rfq/<rid>/obs1600", methods=["POST"])
@auth_required
@safe_route
def api_generate_obs1600(rid):
    """Generate filled OBS 1600 food certification form for an RFQ.
    Uses the bid package PDF if available, or a standalone template.
    """
    from src.api.trace import Trace
    t = Trace("obs1600_fill", rfq_id=rid)
    
    try:
        from src.forms.reytech_filler_v4 import load_config, fill_obs1600, get_pst_date
        from src.forms.food_classifier import get_food_items_for_obs1600
        
        rfqs = load_rfqs()
        r = rfqs.get(rid)
        if not r:
            t.fail("RFQ not found")
            return jsonify({"ok": False, "error": "RFQ not found"}), 404
        
        config = load_config()
        sol = r.get("solicitation_number", "") or "RFQ"
        
        # Get items — try line_items first, then items_detail from quote
        items = r.get("line_items", [])
        if not items:
            items = r.get("items_detail", r.get("items", []))
            if isinstance(items, str):
                import json as _json
                try: items = _json.loads(items)
                except Exception: items = []
        
        # Classify food items
        food_items = get_food_items_for_obs1600(items)
        
        if not food_items:
            t.step("No food items found", item_count=len(items))
            return jsonify({"ok": False, "error": "No food items found in this RFQ. Only food products need the OBS 1600 form.",
                            "items_checked": len(items)}), 400
        
        t.step("Classified food items", food_count=len(food_items),
               items=[f"{fi['description'][:40]} → Code {fi['code']}" for fi in food_items[:5]])
        
        # Output directory
        out_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "output", sol)
        os.makedirs(out_dir, exist_ok=True)
        
        # Find bid package template — use RFQ's stored template paths
        bid_pkg = None
        tmpl = r.get("templates", {})
        
        # Check bid package template from RFQ data
        if tmpl.get("bidpkg") and os.path.exists(tmpl["bidpkg"]):
            bid_pkg = tmpl["bidpkg"]
        
        # Try to restore from DB if not on disk
        if not bid_pkg:
            try:
                from src.core.db import get_db
                with get_db() as conn:
                    db_files = conn.execute(
                        "SELECT id, filename, file_type FROM rfq_files WHERE rfq_id=? AND category='template'",
                        (rid,)).fetchall()
                    for db_f in db_files:
                        fname = db_f["filename"].lower()
                        if "bid" in fname or "package" in fname or "form" in fname:
                            full_f_row = conn.execute("SELECT data FROM rfq_files WHERE id=?", (db_f["id"],)).fetchone()
                            if full_f_row and full_f_row["data"]:
                                restore_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "rfq_templates", rid)
                                os.makedirs(restore_dir, exist_ok=True)
                                restore_path = os.path.join(restore_dir, db_f["filename"])
                                with open(restore_path, "wb") as _fw:
                                    _fw.write(full_f_row["data"])
                                bid_pkg = restore_path
                                t.step(f"Restored bid package from DB: {db_f['filename']}")
                                break
            except Exception as db_err:
                t.step(f"DB restore failed: {db_err}")
        
        # Check uploaded files directory
        if not bid_pkg:
            import glob
            for pattern in [f"*{sol}*BID*PACKAGE*", f"*{sol}*bid*pack*", f"*{sol}*form*", f"*{sol}*.pdf"]:
                for search_dir in [os.path.join(DATA_DIR, "uploads"), os.path.join(DATA_DIR, "rfq_templates"), os.path.join(DATA_DIR, "output", sol)]:
                    matches = glob.glob(os.path.join(search_dir, pattern))
                    for m in matches:
                        # Verify it has OBS 1600 fields
                        try:
                            from pypdf import PdfReader
                            _r = PdfReader(m)
                            for page in _r.pages:
                                if "/Annots" in page:
                                    for annot in page["/Annots"]:
                                        obj = annot.get_object()
                                        if "OBS 1600" in str(obj.get("/T", "")):
                                            bid_pkg = m
                                            break
                                if bid_pkg: break
                        except Exception as _e:
                            log.debug('suppressed in api_generate_obs1600: %s', _e)
                        if bid_pkg: break
                    if bid_pkg: break
                if bid_pkg: break
        
        # Fallback: use saved CDCR bid package template
        if not bid_pkg:
            default_tmpl = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "templates", "cdcr_bid_package_template.pdf")
            if os.path.exists(default_tmpl):
                bid_pkg = default_tmpl
                t.step("Using saved CDCR bid package template")
        
        # Build rfq_data for the filler
        rfq_data = {
            "solicitation_number": sol,
            "sign_date": get_pst_date(),
            "line_items": items,
        }
        
        output_path = os.path.join(out_dir, f"{sol}_OBS1600_FoodCert_Reytech.pdf")
        
        if bid_pkg and os.path.exists(bid_pkg):
            # Fill OBS 1600 fields in the existing bid package
            fill_obs1600(bid_pkg, rfq_data, config, output_path, food_items=food_items)
            t.ok(f"Filled from bid package template: {os.path.basename(bid_pkg)}")
        else:
            # Generate standalone OBS 1600 using reportlab
            _generate_standalone_obs1600(food_items, config, rfq_data, output_path)
            t.ok("Generated standalone OBS 1600")
        
        return jsonify({
            "ok": True,
            "file": output_path,
            "filename": os.path.basename(output_path),
            "food_items": food_items,
            "food_count": len(food_items),
            "download_url": f"/api/download/{sol}/{os.path.basename(output_path)}",
        })
        
    except Exception as e:
        import traceback
        t.fail(str(e))
        log.error("Route error: %s", e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


def _generate_standalone_obs1600(food_items, config, rfq_data, output_path):
    """Generate a standalone OBS 1600 PDF using reportlab when no template is available."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.units import inch
    
    company = config["company"]
    sol = rfq_data.get("solicitation_number", "")
    sign_date = rfq_data.get("sign_date", "")
    
    c = rl_canvas.Canvas(output_path, pagesize=letter)
    w, h = letter
    
    # Header
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.5*inch, h - 0.5*inch, "California Department of Corrections and Rehabilitation/California Correctional Health Care Services")
    c.setFont("Helvetica", 8)
    c.drawString(0.5*inch, h - 0.65*inch, "Office of Business Services - Non-IT Goods Procurement/Acquisitions Management Section, Procurement Services")
    c.drawString(0.5*inch, h - 0.8*inch, "OBS 1600 (Rev. 1/26)")
    
    # Title
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(w/2, h - 1.2*inch, "California-Grown/Produced Agricultural Food Products Vendor Certification")
    
    # Vendor info
    y = h - 1.6*inch
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.5*inch, y, f"Vendor Name : {company['name']}")
    y -= 0.25*inch
    c.drawString(0.5*inch, y, f"Solicitation # : {sol}")
    
    # Table header
    y -= 0.45*inch
    c.setFont("Helvetica-Bold", 8)
    col_x = [0.5*inch, 1.2*inch, 4.5*inch, 5.2*inch, 6.2*inch]
    headers = ["Quoted Line\nItem #", "Food Product Description", "Code", "CA-Grown\nor Produced\n(Yes/No)", "If Yes, % of\nProduct"]
    
    # Header row background
    c.setFillColorRGB(0.9, 0.9, 0.9)
    c.rect(0.5*inch, y - 0.15*inch, 7*inch, 0.45*inch, fill=True, stroke=True)
    c.setFillColorRGB(0, 0, 0)
    
    for i, hdr in enumerate(headers):
        lines = hdr.split("\n")
        for j, line in enumerate(lines):
            c.drawString(col_x[i] + 0.05*inch, y + 0.2*inch - j*0.12*inch, line)
    
    # Data rows
    y -= 0.35*inch
    c.setFont("Helvetica", 9)
    for item in food_items[:18]:
        y -= 0.28*inch
        c.line(0.5*inch, y - 0.05*inch, 7.5*inch, y - 0.05*inch)
        c.drawString(col_x[0] + 0.1*inch, y + 0.05*inch, str(item.get("line_number", "")))
        c.drawString(col_x[1] + 0.05*inch, y + 0.05*inch, item.get("description", "")[:55])
        c.drawCentredString(col_x[2] + 0.35*inch, y + 0.05*inch, str(item.get("code", "")))
        c.drawCentredString(col_x[3] + 0.4*inch, y + 0.05*inch, item.get("ca_grown", "No"))
        c.drawCentredString(col_x[4] + 0.4*inch, y + 0.05*inch, item.get("pct", "N/A"))
    
    # Fill remaining empty rows
    for _ in range(18 - len(food_items)):
        y -= 0.28*inch
        c.line(0.5*inch, y - 0.05*inch, 7.5*inch, y - 0.05*inch)
    
    # Table border
    table_top = h - 2.1*inch
    c.rect(0.5*inch, y - 0.05*inch, 7*inch, table_top - (y - 0.05*inch))
    
    # Certification text
    y -= 0.45*inch
    c.setFont("Helvetica", 7.5)
    c.drawString(0.5*inch, y, "Pursuant to California Code, Food and Agricultural Code, Section 58595(a), I certify under the laws of the State of California")
    y -= 0.15*inch
    c.drawString(0.5*inch, y, "that the above information is true and correct.")
    
    # Signature block
    y -= 0.4*inch
    c.setFont("Helvetica", 10)
    c.drawString(0.5*inch, y, company["owner"])
    c.drawString(3.2*inch, y, company["title"])
    c.drawString(5.5*inch, y, sign_date)
    
    y -= 0.15*inch
    c.line(0.5*inch, y, 2.8*inch, y)
    c.line(3.2*inch, y, 4.8*inch, y)
    c.line(5.5*inch, y, 7.5*inch, y)
    
    y -= 0.15*inch
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.5*inch, y, "Print Name")
    c.drawString(2.2*inch, y, "Signature")
    c.drawString(3.2*inch, y, "Title")
    c.drawString(5.5*inch, y, "Date")
    
    c.save()


@bp.route("/api/download/<path:sol>/<filename>")
@auth_required
@safe_route
def api_download_file(sol, filename):
    """Download a generated file."""
    import re as _re
    # Sanitize: block path traversal but preserve spaces
    sol = sol.replace("..", "").replace("/", "").replace("\\", "")
    filename = os.path.basename(filename)
    filepath = os.path.join(OUTPUT_DIR, sol, filename)
    # Backwards compat: Compliance_Forms_ → RFQ_Package_ rename
    if not os.path.exists(filepath) and filename.startswith("Compliance_Forms_"):
        _old_name = filename.replace("Compliance_Forms_", "RFQ_Package_", 1)
        _old_path = os.path.join(OUTPUT_DIR, sol, _old_name)
        if os.path.exists(_old_path):
            filepath = _old_path
            filename = _old_name
    if not os.path.exists(filepath):
        # Fallback 1: search every output subdir for the filename (P0 incident
        # 2026-05-04, RFQ 7d3c0fee Auralis). The sol slug in the URL came from
        # r.solicitation_number = "GOOD" (a known parser-junk value listed in
        # dashboard.py:1648), but generation correctly used rfq_number=
        # "RFQ-Auralis" for the output dir — so the file lives at
        # /data/output/RFQ-Auralis/... while the URL points at
        # /data/output/GOOD/... and 404s. Filenames within OUTPUT_DIR are
        # uniquely prefixed by sol, so a directory scan for an exact filename
        # match is safe — at most one hit. Path traversal is blocked at line
        # 1186 above, so the os.path.basename() filename can't escape.
        try:
            if os.path.isdir(OUTPUT_DIR):
                for _subdir in os.listdir(OUTPUT_DIR):
                    _candidate = os.path.join(OUTPUT_DIR, _subdir, filename)
                    if os.path.isfile(_candidate):
                        log.info("Download fallback: %s/%s found at %s/%s "
                                 "(sol slug mismatch)", sol, filename, _subdir,
                                 filename)
                        from flask import send_file as _sf
                        return _sf(_candidate, mimetype="application/pdf",
                                   download_name=filename)
        except Exception as _fs_e:
            log.warning("FS subdir fallback failed for %s/%s: %s",
                        sol, filename, _fs_e)

        # Fallback 2: try serving from DB
        # sol might be solicitation number OR rfq_id — try both
        try:
            found_file = None
            # Try sol as rfq_id first
            files = list_rfq_files(sol, category="generated")
            for dbf in files:
                if dbf.get("filename") == filename:
                    found_file = dbf
                    break
            # If not found, search all RFQs by solicitation number
            if not found_file:
                rfqs = load_rfqs()
                for rid, r in rfqs.items():
                    r_sol = (r.get("solicitation_number") or r.get("rfq_number") or "").replace("/", "_")
                    if r_sol == sol or rid == sol:
                        files = list_rfq_files(rid, category="generated")
                        for dbf in files:
                            if dbf.get("filename") == filename:
                                found_file = dbf
                                break
                        if found_file:
                            break
            # Last-resort fallback: filename-only DB search (any rfq, any
            # category). Filenames embed sol+agency so collisions are extremely
            # unlikely; if more than one row matches we just serve the first —
            # bytes are equivalent for the same generated artifact.
            if not found_file:
                try:
                    from src.core.db import get_db as _db
                    with _db() as _conn:
                        _row = _conn.execute(
                            "SELECT id FROM rfq_files WHERE filename = ? "
                            "ORDER BY id DESC LIMIT 1", (filename,)
                        ).fetchone()
                        if _row:
                            found_file = {"id": _row["id"], "filename": filename}
                            log.info("Download fallback: %s served via "
                                     "filename-only DB lookup (id=%s)",
                                     filename, _row["id"])
                except Exception as _fn_e:
                    log.debug("filename-only DB fallback: %s", _fn_e)
            if found_file:
                full = get_rfq_file(found_file["id"])
                if full and full.get("data"):
                    from flask import Response
                    return Response(full["data"], mimetype="application/pdf",
                                    headers={"Content-Disposition": f'inline; filename="{filename}"'})
        except Exception as _e:
            log.warning("DB download fallback failed for %s/%s: %s", sol, filename, _e)
        return jsonify({"ok": False, "error": "File not found"}), 404
    from flask import send_file
    return send_file(filepath, mimetype="application/pdf", download_name=filename)


# ═══════════════════════════════════════════════════════════════════════
# Fill ALL Bid Package Forms (CUF, Darfur, DVBE, CalRecycle, OBS 1600, etc.)
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/<rid>/fill-bid-package", methods=["POST"])
@auth_required
@safe_route
def api_fill_bid_package(rid):
    """Fill ALL forms in the CDCR bid package for an RFQ."""
    from src.api.trace import Trace
    t = Trace("fill_bid_package", rfq_id=rid)
    
    try:
        from src.forms.reytech_filler_v4 import load_config, fill_bid_package, get_pst_date
        
        rfqs = load_rfqs()
        r = rfqs.get(rid)
        if not r:
            t.fail("RFQ not found")
            return jsonify({"ok": False, "error": "RFQ not found"}), 404
        
        config = load_config()
        sol = r.get("solicitation_number", "") or "RFQ"
        
        # Get items
        items = r.get("line_items", [])
        if not items:
            items = r.get("items_detail", r.get("items", []))
            if isinstance(items, str):
                import json as _json
                try: items = _json.loads(items)
                except Exception: items = []
        
        # Find template
        bid_pkg = None
        tmpl = r.get("templates", {})
        if tmpl.get("bidpkg") and os.path.exists(tmpl["bidpkg"]):
            bid_pkg = tmpl["bidpkg"]
        
        # Fallback to saved template
        if not bid_pkg:
            default_tmpl = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "templates", "cdcr_bid_package_template.pdf")
            if os.path.exists(default_tmpl):
                bid_pkg = default_tmpl
        
        if not bid_pkg:
            t.fail("No bid package template found")
            return jsonify({"ok": False, "error": "No bid package template found. Upload one at /form-filler or place in data/templates/"}), 400
        
        out_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "output", sol)
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(out_dir, f"{sol}_BidPackage_Reytech.pdf")
        
        rfq_data = {
            "solicitation_number": sol,
            "sign_date": get_pst_date(),
            "line_items": items,
        }
        
        fill_bid_package(bid_pkg, rfq_data, config, output_path)
        
        # Count food items
        from src.forms.food_classifier import get_food_items_for_obs1600
        food_items = get_food_items_for_obs1600(items)
        
        t.ok(f"Filled bid package: {len(items)} items, {len(food_items)} food items")
        
        return jsonify({
            "ok": True,
            "filename": os.path.basename(output_path),
            "download_url": f"/api/download/{sol}/{os.path.basename(output_path)}",
            "total_items": len(items),
            "food_items": len(food_items),
            "forms_filled": ["CUF", "Darfur", "Bidder Declaration", "DVBE", "Drug-Free", "CalRecycle", "OBS 1600"],
        })
    except Exception as e:
        import traceback
        t.fail(str(e))
        log.error("Route error: %s", e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/fill-forms", methods=["POST"])
@auth_required
@safe_route
def api_fill_forms_standalone():
    """Standalone form filler — fill bid package from manually entered items.
    Body: {
        "solicitation_number": "...",
        "items": [{"line_number": 1, "description": "..."}],
        "fill_type": "all" | "obs1600_only"
    }
    """
    from src.api.trace import Trace
    t = Trace("fill_forms_standalone")
    
    try:
        from src.forms.reytech_filler_v4 import load_config, fill_bid_package, fill_obs1600, get_pst_date
        from src.forms.food_classifier import get_food_items_for_obs1600
        
        data = request.get_json(force=True)
        sol = data.get("solicitation_number", "STANDALONE")
        items = data.get("items", [])
        fill_type = data.get("fill_type", "all")
        
        if not items:
            return jsonify({"ok": False, "error": "No items provided"}), 400
        
        config = load_config()
        
        # Find template
        bid_pkg = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "templates", "cdcr_bid_package_template.pdf")
        if not os.path.exists(bid_pkg):
            return jsonify({"ok": False, "error": "No bid package template found. Upload cdcr_bid_package_template.pdf to data/templates/"}), 400
        
        out_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "output", sol)
        os.makedirs(out_dir, exist_ok=True)
        
        rfq_data = {
            "solicitation_number": sol,
            "sign_date": get_pst_date(),
            "line_items": items,
        }
        
        food_items = get_food_items_for_obs1600(items)
        
        if fill_type == "obs1600_only":
            output_path = os.path.join(out_dir, f"{sol}_OBS1600_FoodCert_Reytech.pdf")
            fill_obs1600(bid_pkg, rfq_data, config, output_path, food_items=food_items)
        else:
            output_path = os.path.join(out_dir, f"{sol}_BidPackage_Reytech.pdf")
            fill_bid_package(bid_pkg, rfq_data, config, output_path)
        
        t.ok(f"Filled {fill_type}: {sol}, {len(food_items)} food items")
        
        return jsonify({
            "ok": True,
            "filename": os.path.basename(output_path),
            "download_url": f"/api/download/{sol}/{os.path.basename(output_path)}",
            "food_items": food_items,
            "food_count": len(food_items),
            "total_items": len(items),
        })
    except Exception as e:
        import traceback
        t.fail(str(e))
        log.error("Route error: %s", e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/rfq/<rid>/price-intel")
@auth_required
@safe_route
def api_rfq_price_intel(rid):
    """Return pricing intelligence for all items in an RFQ."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False}), 404

    intel = []
    for item in r.get("line_items", []):
        desc = item.get("description", "")
        pn = item.get("item_number", "") or ""
        current_cost = item.get("supplier_cost") or 0
        current_bid = item.get("price_per_unit") or 0
        result = {"description": desc[:60], "part_number": pn}

        # Price history
        try:
            from src.core.db import get_price_history_db
            history = get_price_history_db(
                description=desc[:60] if not pn else "",
                part_number=pn, limit=10
            )
            if history:
                prices = [h["unit_price"] for h in history if h.get("unit_price")]
                result["history"] = {
                    "count": len(history),
                    "avg": round(sum(prices) / len(prices), 2) if prices else 0,
                    "min": round(min(prices), 2) if prices else 0,
                    "max": round(max(prices), 2) if prices else 0,
                    "entries": [{
                        "price": h["unit_price"],
                        "source": h.get("source", ""),
                        "date": h.get("found_at", "")[:10],
                        "quote": h.get("quote_number", ""),
                        "agency": h.get("agency", ""),
                    } for h in history[:5]]
                }

                # Freshness: compare current cost vs most recent history
                latest = history[0]
                latest_price = latest.get("unit_price", 0)
                latest_source = latest.get("source", "")
                latest_date = latest.get("found_at", "")[:10]
                try:
                    from datetime import datetime as _dt
                    days_old = (_dt.now() - _dt.fromisoformat(
                        latest["found_at"][:19])).days
                except Exception:
                    days_old = 999

                drift = None
                if current_cost > 0 and latest_price > 0 and latest_source not in ("rfq_save", "rfq_save_bid"):
                    diff = latest_price - current_cost
                    pct = diff / current_cost * 100
                    if abs(pct) > 3:  # Only flag >3% drift
                        drift = {
                            "direction": "up" if diff > 0 else "down",
                            "amount": round(abs(diff), 2),
                            "pct": round(pct, 1),
                            "new_price": latest_price,
                            "source": latest_source,
                        }

                result["freshness"] = {
                    "days_old": days_old,
                    "stale": days_old > 90,
                    "last_source": latest_source,
                    "last_date": latest_date,
                    "drift": drift,
                }
        except Exception as _e:
            log.debug('suppressed in api_rfq_price_intel: %s', _e)

        # Catalog match — switched from search_catalog to match_item on
        # 2026-05-05 (Mike P0): same popularity-sort fallthrough bug as
        # routes_rfq._enrich_items_with_intel. Confidence ≥ 0.50 gates
        # weak matches out of the price-intel surface.
        try:
            from src.agents.product_catalog import match_item
            matches = match_item(desc, pn, top_n=1)
            if matches and matches[0].get("match_confidence", 0) >= 0.50:
                m = matches[0]
                result["catalog"] = {
                    "sku": m.get("sku") or m.get("name", ""),
                    "typical_cost": m.get("cost") or m.get("best_cost") or 0,
                    "list_price": m.get("sell_price", 0),
                    "category": m.get("category", ""),
                    "match_confidence": m.get("match_confidence", 0),
                }
        except Exception as _e:
            log.debug('suppressed in api_rfq_price_intel: %s', _e)

        # Audit trail
        try:
            from src.core.db import get_audit_trail
            audits = get_audit_trail(
                description=desc[:40],
                rfq_id=r.get("solicitation_number", ""), limit=5)
            if audits:
                result["audit"] = [{
                    "field": a["field_changed"],
                    "old": a.get("old_value"),
                    "new": a.get("new_value"),
                    "source": a.get("source", ""),
                    "ts": a.get("ts", "")[:16],
                } for a in audits]
        except Exception as _e:
            log.debug('suppressed in api_rfq_price_intel: %s', _e)

        # Pricing recommendation
        rec = _recommend_price(item)
        if rec:
            result["recommendation"] = rec

        # Loss intelligence — competitive insights from past losses
        try:
            from src.agents.pricing_feedback import get_pricing_recommendation
            agency = r.get("agency", "")
            loss_rec = get_pricing_recommendation(desc, agency, float(current_cost) if current_cost else 0)
            if loss_rec.get("sources_used", 0) > 0:
                result["loss_intelligence"] = {
                    "competitor_floor": loss_rec.get("competitor_floor"),
                    "suggested_range": loss_rec.get("suggested_range"),
                    "margin_warning": loss_rec.get("margin_warning"),
                    "confidence": loss_rec.get("confidence", 0),
                    "loss_count": loss_rec.get("sources_used", 0),
                }
        except Exception as _e:
            log.debug('suppressed in api_rfq_price_intel: %s', _e)

        # F6: Price conflict resolution — all known sources
        sources = {}
        if item.get("supplier_cost") and item["supplier_cost"] > 0:
            sources["Your Cost"] = round(item["supplier_cost"], 2)
        if item.get("scprs_last_price") and item["scprs_last_price"] > 0:
            sources["SCPRS"] = round(item["scprs_last_price"], 2)
        if item.get("amazon_price") and item["amazon_price"] > 0:
            sources["Amazon"] = round(item["amazon_price"], 2)
        if item.get("price_per_unit") and item["price_per_unit"] > 0:
            sources["Current Bid"] = round(item["price_per_unit"], 2)
        if result.get("catalog") and result["catalog"].get("typical_cost"):
            sources["Catalog"] = round(result["catalog"]["typical_cost"], 2)
        if result.get("catalog") and result["catalog"].get("list_price"):
            sources["Catalog List"] = round(result["catalog"]["list_price"], 2)
        if item.get("_from_pc"):
            sources["_from_pc"] = item["_from_pc"]
        if len(sources) > 1:
            result["sources"] = sources

        # F9: Duplicate item detection — same item quoted recently?
        try:
            from src.core.db import get_price_history_db
            pn = item.get("item_number", "") or ""
            recent = get_price_history_db(
                description=desc[:40] if not pn else "",
                part_number=pn, source="rfq_save_bid", limit=3
            )
            if recent:
                dupes = []
                for rh in recent:
                    dupes.append({
                        "price": rh.get("unit_price", 0),
                        "quote": rh.get("quote_number", ""),
                        "agency": rh.get("agency", ""),
                        "date": rh.get("found_at", "")[:10],
                    })
                if dupes:
                    result["recent_quotes"] = dupes
        except Exception as _e:
            log.debug('suppressed in api_rfq_price_intel: %s', _e)

        intel.append(result)

    return jsonify({"ok": True, "intel": intel})


_pricing_alerts_cache = {"data": None, "ts": 0}

@bp.route("/api/pricing-alerts")
@auth_required
@safe_route
def api_pricing_alerts():
    """F8: Dashboard pricing alerts — stale prices, drift, unpriced items."""
    import time as _time
    global _pricing_alerts_cache
    if _pricing_alerts_cache["data"] and (_time.time() - _pricing_alerts_cache["ts"]) < 120:
        return jsonify(_pricing_alerts_cache["data"])
    from datetime import datetime as _dt, timedelta
    rfqs = load_rfqs()
    stale_rfqs = []
    unpriced_rfqs = []
    drift_items = 0
    now = _dt.now()

    # 2026-05-26 audit (Mr. Wolf 13-badge audit): exclusion set was
    # missing canonical terminal/inactive statuses. Operator-
    # dispositioned RFQs (Argarin/Ragadio no_bid per the 2026-05-26
    # substrate-wave handoff) inflated the ⚠ home-badge count.
    # Same substrate-singleness shape as PRs #1076/#1086/#1088: a KPI
    # query reads against a non-canonical filter.
    # Union of:
    #   - rfq_detail.html:1648 terminal: ('won','lost','no_bid','cancelled')
    #   - routes_intel_ops.py:853 inactive:
    #     {'dismissed','archived','deleted','duplicate','no_response'}
    #   - sent/expired (already-out-the-door states)
    _TERMINAL_OR_INACTIVE = (
        "dismissed", "sent", "won", "lost", "cancelled",
        "no_bid", "no_response", "expired",
        "archived", "deleted", "duplicate",
    )
    for rid, r in rfqs.items():
        if r.get("status") in _TERMINAL_OR_INACTIVE:
            continue
        items = r.get("line_items", [])
        if not items:
            continue

        # Check for unpriced items
        unpriced = sum(1 for it in items if not (it.get("price_per_unit") or 0) > 0)
        if unpriced == len(items):
            unpriced_rfqs.append({"id": rid, "sol": r.get("solicitation_number", ""), "items": len(items)})
            continue

        # Check for stale pricing (created > 14 days ago, never regenerated)
        try:
            created = r.get("created_at", "")
            if created:
                age = (now - _dt.fromisoformat(created[:19])).days
                if age > 14 and r.get("status") not in ("generated",):
                    stale_rfqs.append({
                        "id": rid, "sol": r.get("solicitation_number", ""),
                        "age_days": age, "items": len(items),
                    })
        except Exception as _e:
            log.debug('suppressed in api_pricing_alerts: %s', _e)

    # Check price_history for recent drift
    try:
        from src.core.db import get_db
        with get_db() as conn:
            # Items with multiple prices where latest differs >10% from previous
            rows = conn.execute("""
                SELECT description, COUNT(*) as cnt,
                       MAX(unit_price) as max_p, MIN(unit_price) as min_p
                FROM price_history
                WHERE found_at > ?
                GROUP BY LOWER(SUBSTR(description, 1, 40))
                HAVING cnt > 1 AND (max_p - min_p) / min_p > 0.10
            """, ((now - timedelta(days=30)).isoformat(),)).fetchall()
            drift_items = len(rows)
    except Exception as _e:
        log.debug('suppressed in api_pricing_alerts: %s', _e)

    total_alerts = len(stale_rfqs) + len(unpriced_rfqs) + (1 if drift_items > 0 else 0)
    _pa_result = {
        "ok": True,
        "total_alerts": total_alerts,
        "stale_rfqs": stale_rfqs,
        "unpriced_rfqs": unpriced_rfqs,
        "drift_items": drift_items,
    }
    _pricing_alerts_cache["data"] = _pa_result
    _pricing_alerts_cache["ts"] = _time.time()
    return jsonify(_pa_result)


@bp.route("/api/rfq/<rid>/qa-check")
@auth_required
@safe_route
def api_rfq_qa_check(rid):
    """QA gate: validate all items before package generation.
    Returns per-item pass/warn/fail with reasons."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False}), 404

    items = r.get("line_items", [])
    results = []
    overall = "pass"
    warnings = 0
    failures = 0

    for i, item in enumerate(items):
        checks = []
        item_status = "pass"
        desc = (item.get("description", "") or "")[:50]
        bid = item.get("price_per_unit") or 0
        cost = item.get("supplier_cost") or 0
        scprs = item.get("scprs_last_price") or 0
        qty = item.get("qty") or 0

        # 1: Bid price > 0
        if not bid or bid <= 0:
            checks.append({"check": "bid_price", "status": "fail", "msg": "No bid price"})
            item_status = "fail"
        else:
            checks.append({"check": "bid_price", "status": "pass", "msg": f"${bid:.2f}"})

        # 2: Supplier cost > 0
        if not cost or cost <= 0:
            checks.append({"check": "cost", "status": "warn", "msg": "No supplier cost"})
            if item_status != "fail":
                item_status = "warn"
        else:
            checks.append({"check": "cost", "status": "pass", "msg": f"${cost:.2f}"})

        # 3: Margin ≥ 15%
        if bid > 0 and cost > 0:
            margin = (bid - cost) / bid * 100
            if margin < 5:
                checks.append({"check": "margin", "status": "fail",
                               "msg": f"{margin:.1f}% — dangerously low"})
                item_status = "fail"
            elif margin < 15:
                checks.append({"check": "margin", "status": "warn",
                               "msg": f"{margin:.1f}% — below 15%"})
                if item_status != "fail":
                    item_status = "warn"
            else:
                checks.append({"check": "margin", "status": "pass",
                               "msg": f"{margin:.1f}%"})

        # 4: Bid vs SCPRS
        if bid > 0 and scprs > 0:
            diff_pct = (bid - scprs) / scprs * 100
            if diff_pct > 10:
                checks.append({"check": "scprs", "status": "warn",
                               "msg": f"{diff_pct:.0f}% above SCPRS ${scprs:.2f}"})
                if item_status != "fail":
                    item_status = "warn"
            elif diff_pct < -15:
                checks.append({"check": "scprs", "status": "warn",
                               "msg": f"{abs(diff_pct):.0f}% below SCPRS — leaving margin?"})
                if item_status != "fail":
                    item_status = "warn"
            else:
                checks.append({"check": "scprs", "status": "pass",
                               "msg": f"OK vs SCPRS ${scprs:.2f}"})

        # 5: Price freshness
        try:
            from src.core.db import get_price_history_db
            pn = item.get("item_number", "") or ""
            history = get_price_history_db(
                description=desc[:40] if not pn else "",
                part_number=pn, limit=1)
            if history:
                from datetime import datetime as _dt
                days = (_dt.now() - _dt.fromisoformat(
                    history[0]["found_at"][:19])).days
                if days > 90:
                    checks.append({"check": "freshness", "status": "warn",
                                   "msg": f"Price data {days}d old"})
                    if item_status != "fail":
                        item_status = "warn"
                else:
                    checks.append({"check": "freshness", "status": "pass",
                                   "msg": f"{days}d ago"})
        except Exception as _e:
            log.debug('suppressed in api_rfq_qa_check: %s', _e)

        # 6: Qty > 0
        if not qty or qty <= 0:
            checks.append({"check": "qty", "status": "warn", "msg": "Qty is 0"})
            if item_status != "fail":
                item_status = "warn"

        if item_status == "fail":
            failures += 1
            if overall != "fail":
                overall = "fail"
        elif item_status == "warn":
            warnings += 1
            if overall == "pass":
                overall = "warn"

        results.append({"idx": i, "description": desc,
                        "status": item_status, "checks": checks})

    # PC diff warnings
    diff_notes = []
    pc_diff = r.get("pc_diff")
    if pc_diff:
        if pc_diff.get("added"):
            diff_notes.append(f"{len(pc_diff['added'])} new items not in Price Check")
        if pc_diff.get("removed"):
            diff_notes.append(f"{len(pc_diff['removed'])} PC items not in RFQ")
        if pc_diff.get("qty_changed"):
            diff_notes.append(f"{len(pc_diff['qty_changed'])} qty changes from PC")

    # Requirements validation (email = contract)
    req_gaps = []
    try:
        _req_json = r.get("requirements_json", "{}")
        if _req_json and _req_json != "{}":
            from src.forms.form_qa import validate_against_requirements
            _gen_files = r.get("output_files", [])
            _vr = validate_against_requirements(_gen_files, _req_json, r)
            req_gaps = _vr.get("gaps", [])
            # Count requirement warnings
            for _gap in req_gaps:
                warnings += 1
                if overall == "pass":
                    overall = "warn"
    except Exception as _e:
        log.debug('suppressed in api_rfq_qa_check: %s', _e)

    return jsonify({"ok": True, "overall": overall, "failures": failures,
                    "warnings": warnings, "total": len(items), "items": results,
                    "linked_pc": r.get("linked_pc_number", ""),
                    "diff_notes": diff_notes,
                    "requirement_gaps": req_gaps})


@bp.route("/form-filler")
@auth_required
@safe_page
def form_filler_page():
    """Standalone form filler page."""
    return render_page("form_filler.html", active_page="Forms")


# ═══════════════════════════════════════════════════════════════════════
# Admin: Nuke & Re-poll RFQ
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/nuke/<rid>", methods=["POST"])
@auth_required
@safe_route
def api_nuke_rfq(rid):
    """Nuclear delete: wipe RFQ from JSON + SQLite + processed UIDs, then re-poll.
    Usage: POST /api/rfq/nuke/<rfq_id>  or  POST /api/rfq/nuke/<solicitation_number>
    """
    import json as _json
    from src.api.dashboard import load_rfqs, save_rfqs

    rfqs = load_rfqs()
    nuked = []

    # Find by ID or by solicitation number
    targets = {}
    for k, v in rfqs.items():
        if k == rid or v.get("solicitation_number", "") == rid or v.get("rfq_number", "") == rid:
            targets[k] = v

    if not targets:
        return jsonify({"ok": False, "error": f"No RFQ found matching '{rid}'"}), 404

    for rfq_id, rfq in targets.items():
        sol = rfq.get("solicitation_number", rfq.get("rfq_number", "?"))
        email_uid = rfq.get("email_uid", "")

        # 1. Remove from JSON
        if rfq_id in rfqs:
            del rfqs[rfq_id]

        # 2. Remove from SQLite (rfqs, rfq_files, email_log, price_checks)
        try:
            with get_db() as conn:
                conn.execute("DELETE FROM rfq_files WHERE rfq_id = ?", (rfq_id,))
                conn.execute("DELETE FROM rfqs WHERE id = ?", (rfq_id,))
                # email_log by rfq_id
                conn.execute("DELETE FROM email_log WHERE rfq_id = ?", (rfq_id,))
                # price_checks by rfq_id
                conn.execute("DELETE FROM price_checks WHERE rfq_id = ?", (rfq_id,))
                conn.commit()
        except Exception as e:
            log.warning("Nuke SQLite cleanup for %s: %s", rfq_id, e)

        # 3. Remove email UID from processed list.
        # _remove_processed_uid never existed in routes_pricecheck (phantom
        # import); the JSON-file path below is the real implementation.
        if email_uid:
            proc_file = os.path.join(DATA_DIR, "processed_emails.json")
            try:
                if os.path.exists(proc_file):
                    with open(proc_file) as f:
                        processed = _json.load(f)
                    if isinstance(processed, list) and email_uid in processed:
                        processed.remove(email_uid)
                    elif isinstance(processed, dict) and email_uid in processed:
                        del processed[email_uid]
                    with open(proc_file, "w") as f:
                        _json.dump(processed, f)
            except Exception as _e:
                log.warning("api_nuke_rfq processed_emails cleanup failed: %s", _e)

        nuked.append({"id": rfq_id, "sol": sol, "uid": email_uid})
        log.info("NUKED RFQ %s (sol=%s, uid=%s)", rfq_id, sol, email_uid)

    save_rfqs(rfqs)

    # 4. Trigger re-poll
    poll_result = None
    try:
        from src.api.modules.routes_pricecheck import do_poll_check
        imported = do_poll_check()
        poll_result = {"found": len(imported), "rfqs": [r.get("solicitation_number", "?") for r in imported]}
    except Exception as e:
        poll_result = {"error": str(e)}

    return jsonify({
        "ok": True,
        "nuked": nuked,
        "poll": poll_result,
    })


@bp.route("/api/rfq/<rid>/clear-quote", methods=["POST", "GET"])
@auth_required
@safe_route
def api_rfq_clear_quote(rid):
    """Clear the quote number on an RFQ so regeneration assigns a new one.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_clear_quote_locked(rid)


def _api_rfq_clear_quote_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    
    old_qn = r.get("reytech_quote_number", "")
    r["reytech_quote_number"] = ""
    r["linked_quote_number"] = ""
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)

    return jsonify({"ok": True, "cleared": old_qn, "message": f"Cleared {old_qn}. Regenerate to get a new number."})


@bp.route("/api/rfq/<rid>/set-quote-number", methods=["POST"])
@auth_required
@safe_route
def api_rfq_set_quote_number(rid):
    """Force-set the quote number on an RFQ. Used to fix counter drift.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_set_quote_number_locked(rid)


def _api_rfq_set_quote_number_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    data = request.get_json(force=True, silent=True) or {}
    qn = data.get("quote_number", "").strip()
    if not qn:
        return jsonify({"ok": False, "error": "Provide quote_number"})
    old = r.get("reytech_quote_number", "")
    r["reytech_quote_number"] = qn
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)
    log.info("Force-set quote number on RFQ %s: %s → %s", rid, old, qn)
    return jsonify({"ok": True, "old": old, "new": qn})


@bp.route("/api/rfq/<rid>/revise-quote", methods=["POST"])
@auth_required
@safe_route
def api_rfq_revise_quote(rid):
    """Regenerate ONLY the quote PDF with current pricing — keep all other
    package docs unchanged. Saves revision to quote_revisions table.
    Preserves quote number — never burns a new one.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_revise_quote_locked(rid)


def _api_rfq_revise_quote_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    from src.api.trace import Trace
    t = Trace("quote_revision", rfq_id=rid)

    if not QUOTE_GEN_AVAILABLE:
        return jsonify({"ok": False, "error": "Quote generator not available"})

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    locked_qn = r.get("reytech_quote_number", "")
    if not locked_qn:
        return jsonify({"ok": False, "error": "No quote number found — generate the full package first"})

    data = request.get_json(force=True, silent=True) or {}
    reason = str(data.get("reason", "Pricing updated")).strip()[:200] or "Pricing updated"

    sol = r.get("solicitation_number", "") or "RFQ"
    safe_sol = re.sub(r'[^a-zA-Z0-9_-]', '_', sol.strip())
    out_dir = os.path.join(OUTPUT_DIR, sol)
    os.makedirs(out_dir, exist_ok=True)
    output_path = os.path.join(out_dir, f"{safe_sol}_Quote_Reytech.pdf")

    # Determine revision number
    try:
        with get_db() as conn:
            row = conn.execute(
                "SELECT MAX(revision_num) as maxrev FROM quote_revisions WHERE quote_number=?",
                (locked_qn,)).fetchone()
            rev_num = (row["maxrev"] or 0) + 1 if row else 1
    except Exception:
        rev_num = (r.get("quote_revision", 0) or 0) + 1

    t.step("Revising", quote_number=locked_qn, revision=rev_num, reason=reason)

    # Snapshot pricing for diff/history
    snapshot_items = [
        {
            "line_number": it.get("line_number", i+1),
            "description": (it.get("description") or "")[:80],
            "qty": it.get("qty", 0),
            "uom": it.get("uom", ""),
            "supplier_cost": it.get("supplier_cost", 0),
            "price_per_unit": it.get("price_per_unit", 0),
            "markup_pct": it.get("markup_pct", 0),
        }
        for i, it in enumerate(r.get("line_items", []))
    ]

    # Generate quote PDF — same quote number, no new allocation
    result = generate_quote_from_rfq(r, output_path, quote_number=locked_qn)

    if not result.get("ok"):
        t.fail("Quote revision failed", error=result.get("error"))
        return jsonify({"ok": False, "error": result.get("error", "Quote generation failed")})

    new_total = result.get("total", 0)

    # Save revision to DB
    import json as _json_rev
    try:
        with get_db() as conn:
            conn.execute(
                """INSERT INTO quote_revisions
                   (quote_number, revision_num, revised_at, reason, snapshot_json, changed_by)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (locked_qn, rev_num, datetime.now().isoformat(),
                 reason, _json_rev.dumps(snapshot_items), "user"))
            conn.commit()
    except Exception as _dbe:
        log.warning("quote_revisions insert failed: %s", _dbe)

    # Update RFQ record
    r["quote_revision"] = rev_num
    r["quote_revised_at"] = datetime.now().isoformat()
    r["quote_revision_reason"] = reason
    r["pricing_snapshot"] = {
        "snapshot_at": datetime.now().isoformat(),
        "quote_number": locked_qn,
        "revision": rev_num,
        "total": new_total,
        "tax_rate": r.get("tax_rate", 0),
        "items": snapshot_items,
    }
    fname = os.path.basename(output_path)
    if "output_files" not in r:
        r["output_files"] = []
    if fname not in r["output_files"]:
        r["output_files"].append(fname)
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)

    try:
        from src.core.dal import log_lifecycle_event
        log_lifecycle_event("rfq", rid, "quote_revised",
            f"Quote {locked_qn} Rev {rev_num} — ${new_total:,.2f} — {reason}",
            actor="user",
            detail={"quote_number": locked_qn, "revision": rev_num, "total": new_total, "reason": reason})
    except Exception as _e:
        log.debug('suppressed in api_rfq_revise_quote: %s', _e)

    t.ok("Quote revised", revision=rev_num, total=new_total)
    log.info("Quote %s Rev %d generated for RFQ %s — $%.2f", locked_qn, rev_num, rid, new_total)
    return jsonify({
        "ok": True,
        "quote_number": locked_qn,
        "revision": rev_num,
        "total": new_total,
        "download": f"/api/download/{sol}/{fname}",
    })


@bp.route("/api/rfq/<rid>/revision-history", methods=["GET"])
@auth_required
@safe_route
def api_rfq_revision_history(rid):
    """Return quote revision history for an RFQ."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "not found"})
    locked_qn = r.get("reytech_quote_number", "")
    if not locked_qn:
        return jsonify({"ok": True, "revisions": [], "quote_number": ""})
    try:
        with get_db() as conn:
            rows = conn.execute(
                """SELECT revision_num, revised_at, reason, changed_by, snapshot_json
                   FROM quote_revisions WHERE quote_number=?
                   ORDER BY revision_num DESC LIMIT 20""",
                (locked_qn,)).fetchall()
        import json as _jh
        revisions = []
        for row in rows:
            snap = []
            try:
                snap = _jh.loads(row["snapshot_json"] or "[]")
            except Exception as _e:
                log.debug('suppressed in api_rfq_revision_history: %s', _e)
            total = sum((it.get("price_per_unit", 0) or 0) * (it.get("qty", 0) or 0) for it in snap)
            revisions.append({
                "revision_num": row["revision_num"],
                "revised_at": row["revised_at"],
                "reason": row["reason"],
                "changed_by": row["changed_by"],
                "total": round(total, 2),
                "item_count": len(snap),
            })
        return jsonify({"ok": True, "quote_number": locked_qn, "revisions": revisions})
    except Exception as e:
        log.warning("revision-history for %s: %s", rid, e)
        return jsonify({"ok": True, "revisions": [], "note": str(e)})


@bp.route("/api/rfq/<rid>/revert-pricing", methods=["POST"])
@auth_required
@safe_route
def api_rfq_revert_pricing(rid):
    """Revert line item prices to the last generated quote snapshot.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_revert_pricing_locked(rid)


def _api_rfq_revert_pricing_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    snapshot = r.get("pricing_snapshot")
    if not snapshot:
        return jsonify({"ok": False, "error": "No pricing snapshot found. Generate a quote first."})
    snap_items = snapshot.get("items", [])
    if len(snap_items) != len(r.get("line_items", [])):
        return jsonify({"ok": False, "error": "Item count changed since snapshot — cannot auto-revert"})
    for i, snap in enumerate(snap_items):
        item = r["line_items"][i]
        item["supplier_cost"] = snap.get("supplier_cost", 0)
        item["price_per_unit"] = snap.get("price_per_unit", 0)
        item["markup_pct"] = snap.get("markup_pct", 0)
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)
    log.info("Reverted pricing on RFQ %s to snapshot %s", rid, snapshot.get("quote_number", ""))
    return jsonify({
        "ok": True,
        "reverted_to": snapshot.get("quote_number", ""),
        "snapshot_at": snapshot.get("snapshot_at", ""),
        "items": len(snap_items),
    })


@bp.route("/api/admin/relink-rfq/<rid>", methods=["POST", "GET"])
@auth_required
@safe_route
def api_admin_relink_rfq(rid):
    """Re-run auto-linking on an existing RFQ to find its matching PC.
    GET-accessible for browser use. Example: /api/admin/relink-rfq/abc123

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_admin_relink_rfq_locked(rid)


def _api_admin_relink_rfq_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    _trace = []
    try:
        from src.api.dashboard import _link_rfq_to_pc
        linked = _link_rfq_to_pc(r, _trace)
        if linked:
            from src.api.dashboard import _save_single_rfq
            _save_single_rfq(rid, r)
            return jsonify({
                "ok": True,
                "linked": True,
                "linked_pc_id": r.get("linked_pc_id", ""),
                "linked_pc_number": r.get("linked_pc_number", ""),
                "match_reason": r.get("linked_pc_match_reason", ""),
                "ported_items": r.get("pc_diff", {}).get("ported", 0),
                "trace": _trace,
            })
        return jsonify({"ok": True, "linked": False, "trace": _trace,
                        "message": "No matching PC found for this RFQ"})
    except Exception as e:
        log.error("relink-rfq %s: %s", rid, e, exc_info=True)
        return jsonify({"ok": False, "error": str(e), "trace": _trace})


@bp.route("/api/admin/fix-quote-number/<rid>/<new_qn>/<int:counter_seq>", methods=["POST", "GET"])
@auth_required
@safe_route
def api_admin_fix_quote_number(rid, new_qn, counter_seq):
    """One-shot admin: set RFQ quote number + reset counter. GET-accessible for browser.

    Example: /api/admin/fix-quote-number/cab4bad5/R26Q31/31
    Sets RFQ cab4bad5 to R26Q31, counter to 31 (next = R26Q32).

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_admin_fix_quote_number_locked(rid, new_qn, counter_seq)


def _api_admin_fix_quote_number_locked(rid, new_qn, counter_seq):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    old_qn = r.get("reytech_quote_number", "")
    r["reytech_quote_number"] = new_qn
    # Also update in output files and generated package data
    if r.get("quote_number"):
        r["quote_number"] = new_qn
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)
    # Reset counter
    try:
        from src.forms.quote_generator import set_quote_counter, peek_next_quote_number
        set_quote_counter(seq=counter_seq)
        nxt = peek_next_quote_number()
    except Exception as e:
        nxt = f"error: {e}"
    log.warning("ADMIN fix-quote: RFQ %s: %s → %s, counter → %d (next: %s)", rid, old_qn, new_qn, counter_seq, nxt)
    return jsonify({
        "ok": True,
        "rfq": rid,
        "old_quote": old_qn,
        "new_quote": new_qn,
        "counter_set_to": counter_seq,
        "next_quote_will_be": nxt,
    })


@bp.route("/api/rfq/<rid>/clear-generated", methods=["POST", "GET"])
@auth_required
@safe_route
def api_rfq_clear_generated(rid):
    """
    Force-clear all generated files for an RFQ from both DB and JSON.
    Resets status to 'ready' so the full generate-package pipeline re-runs cleanly.
    Use this when Railway redeploys cached the old output and Regenerate doesn't help.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_clear_generated_locked(rid)


def _api_rfq_clear_generated_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    # Clear DB generated files
    db_deleted = 0
    try:
        from src.core.db import get_db
        with get_db() as conn:
            cur = conn.execute(
                "DELETE FROM rfq_files WHERE rfq_id = ? AND category = 'generated'",
                (rid,)
            )
            db_deleted = cur.rowcount
    except Exception as _e:
        log.warning("clear-generated DB delete failed for %s: %s", rid, _e)

    # Clear disk output files
    sol = r.get("solicitation_number", rid)
    out_dir = os.path.join(OUTPUT_DIR, sol)
    disk_deleted = 0
    if os.path.exists(out_dir):
        try:
            for fname in os.listdir(out_dir):
                fpath = os.path.join(out_dir, fname)
                if os.path.isfile(fpath):
                    os.remove(fpath)
                    disk_deleted += 1
        except Exception as _de:
            log.warning("clear-generated disk delete failed: %s", _de)

    # Reset JSON state
    old_files = r.get("output_files", [])
    r["output_files"] = []
    r.pop("draft_email", None)
    r.pop("generated_at", None)
    _transition_status(r, "ready", actor="user", notes="Cleared generated files for fresh regeneration")
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, "ready")
    except Exception as _e:
        log.debug('suppressed in api_rfq_clear_generated: %s', _e)

    msg = f"Cleared {db_deleted} DB files + {disk_deleted} disk files. Status reset to 'ready'. Click Generate Package to rebuild."
    log.info("clear-generated %s: %s", rid, msg)
    return jsonify({"ok": True, "db_deleted": db_deleted, "disk_deleted": disk_deleted,
                    "old_files": old_files, "message": msg})


@bp.route("/api/rfq/<rid>/clean-slate", methods=["POST", "GET"])
@auth_required
@safe_route
def api_rfq_clean_slate(rid):
    """Nuclear clean: keep ONLY line_items with pricing. Clear everything else.
    Use when package is broken — stale templates, wrong forms, old data.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_clean_slate_locked(rid)


def _api_rfq_clean_slate_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    # Preserve line items with all pricing fields
    items = r.get("line_items", [])
    preserved_items = []
    for it in items:
        preserved_items.append({
            "line_number": it.get("line_number", 0),
            "qty": it.get("qty", 1),
            "uom": it.get("uom", "EA"),
            "description": it.get("description", ""),
            "item_number": it.get("item_number", ""),
            "supplier_cost": it.get("supplier_cost", 0),
            "price_per_unit": it.get("price_per_unit", 0),
            "markup_pct": it.get("markup_pct"),
            "scprs_last_price": it.get("scprs_last_price"),
            "amazon_price": it.get("amazon_price"),
            "item_link": it.get("item_link", ""),
            "item_supplier": it.get("item_supplier", ""),
            "_desc_source": it.get("_desc_source", ""),
        })

    # Preserve core RFQ identity
    sol = r.get("solicitation_number", "")
    identity = {
        "solicitation_number": sol,
        "agency": r.get("agency", ""),
        "requestor_name": r.get("requestor_name", ""),
        "requestor_email": r.get("requestor_email", ""),
        "delivery_location": r.get("delivery_location", ""),
        "due_date": r.get("due_date", ""),
        "institution": r.get("institution", ""),
        "ship_to": r.get("ship_to", ""),
        "created_at": r.get("created_at", ""),
        "source": r.get("source", ""),
        "linked_pc_id": r.get("linked_pc_id", ""),
        "reytech_quote_number": r.get("reytech_quote_number", ""),
    }

    # Clear DB files (generated + templates)
    db_deleted = 0
    try:
        from src.core.db import get_db
        with get_db() as conn:
            cur = conn.execute(
                "DELETE FROM rfq_files WHERE rfq_id = ? AND category IN ('generated', 'template')",
                (rid,)
            )
            db_deleted = cur.rowcount
    except Exception as _e:
        log.warning("clean-slate DB: %s", _e)

    # Clear disk
    disk_deleted = 0
    import shutil as _sh2
    out_dir = os.path.join(OUTPUT_DIR, sol)
    if os.path.exists(out_dir):
        try:
            _sh2.rmtree(out_dir)
            disk_deleted += 1
        except Exception as _e:
            log.debug('suppressed in api_rfq_clean_slate: %s', _e)
    tmpl_dir = os.path.join(DATA_DIR, "rfq_templates", rid)
    if os.path.exists(tmpl_dir):
        try:
            _sh2.rmtree(tmpl_dir)
            disk_deleted += 1
        except Exception as _e:
            log.debug('suppressed in api_rfq_clean_slate: %s', _e)

    # Rebuild RFQ with clean state (sync all aliases — alias-drift substrate)
    r.clear()
    r.update(identity)
    _sync_rfq_items(r, preserved_items)
    r["templates"] = {}
    r["output_files"] = []
    r["status"] = "ready"

    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)
    try:
        from src.core.dal import update_rfq_status as _dal_ur
        _dal_ur(rid, "ready")
    except Exception as _e:
        log.debug('suppressed in api_rfq_clean_slate: %s', _e)

    log.info("clean-slate %s: kept %d items, cleared %d DB + %d disk",
             rid, len(preserved_items), db_deleted, disk_deleted)
    return jsonify({
        "ok": True,
        "items_preserved": len(preserved_items),
        "db_cleared": db_deleted,
        "disk_cleared": disk_deleted,
        "message": f"Clean slate: {len(preserved_items)} items preserved with pricing. All docs/templates cleared. Ready to regenerate.",
    })


@bp.route("/api/rfq/<rid>/debug-pages", methods=["GET"])
@auth_required
@safe_route
def api_rfq_debug_pages(rid):
    """Debug: run page-skip logic against last generated package PDF. Returns per-page decisions."""
    from src.forms.reytech_filler_v4 import _bidpkg_page_skip_reason
    from pypdf import PdfReader

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    sol = r.get("solicitation_number", rid)
    pkg_path = os.path.join(OUTPUT_DIR, sol, f"RFQ_Package_{sol}_ReytechInc.pdf")

    if not os.path.exists(pkg_path):
        return jsonify({"ok": False, "error": f"Not found: {pkg_path}"})

    reader = PdfReader(pkg_path)
    pages = []
    for i, page in enumerate(reader.pages):
        try:
            reason = _bidpkg_page_skip_reason(page)
        except Exception as e:
            reason = f"ERROR: {e}"
        text_snip = (page.extract_text() or "")[:80].replace("\n", " ")
        n_fields = len(page.get("/Annots", [])) if "/Annots" in page else 0
        pages.append({"page": i, "decision": "SKIP" if reason else "KEEP",
                      "reason": reason or "", "fields": n_fields, "text": text_snip})

    return jsonify({"ok": True, "total": len(pages),
                    "kept": sum(1 for p in pages if p["decision"] == "KEEP"),
                    "skipped": sum(1 for p in pages if p["decision"] == "SKIP"),
                    "pages": pages})


@bp.route("/api/rfq/<rid>/debug-templates", methods=["GET"])
@auth_required
@safe_route
def api_rfq_debug_templates(rid):
    """Dump all field names from uploaded 703B/704B/bidpkg templates. Use to diagnose fill mismatches."""
    from pypdf import PdfReader

    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    result = {}
    tmpl = r.get("templates", {})

    # Also restore from DB if needed
    db_files = list_rfq_files(rid, category="template")
    for db_f in db_files:
        ft = db_f.get("file_type", "").lower().replace("template_", "")
        fname = db_f.get("filename", "").lower()
        ttype = None
        if "703b" in ft or "703b" in fname: ttype = "703b"
        elif "704b" in ft or "704b" in fname: ttype = "704b"
        elif "bid" in ft or "bid" in fname: ttype = "bidpkg"
        if ttype and (ttype not in tmpl or not os.path.exists(tmpl.get(ttype, ""))):
            full_f = get_rfq_file(db_f["id"])
            if full_f and full_f.get("data"):
                restore_dir = os.path.join(DATA_DIR, "rfq_templates", rid)
                os.makedirs(restore_dir, exist_ok=True)
                restore_path = os.path.join(restore_dir, db_f["filename"])
                with open(restore_path, "wb") as _fw:
                    _fw.write(full_f["data"])
                tmpl[ttype] = restore_path

    for tname, tpath in tmpl.items():
        if not os.path.exists(tpath):
            result[tname] = {"error": f"file missing: {tpath}"}
            continue
        try:
            rdr = PdfReader(tpath)
            fields = rdr.get_fields() or {}
            sig_fields = []
            all_pages = []
            for i, pg in enumerate(rdr.pages):
                annots = pg.get("/Annots", [])
                pg_fields = []
                for a in (annots or []):
                    obj = a.get_object() if hasattr(a, "get_object") else a
                    name = str(obj.get("/T", ""))
                    ft_val = str(obj.get("/FT", ""))
                    if ft_val == "/Sig" or "sig" in name.lower():
                        sig_fields.append({"name": name, "ft": ft_val, "page": i})
                    if name:
                        pg_fields.append(name)
                all_pages.append({"page": i, "fields": pg_fields[:10]})
            result[tname] = {
                "path": tpath,
                "pages": len(rdr.pages),
                "total_fields": len(fields),
                "sig_fields": sig_fields,
                "all_field_names": sorted(fields.keys())[:50],
                "pages_preview": all_pages,
            }
        except Exception as e:
            result[tname] = {"error": str(e)}

    return jsonify({"ok": True, "templates": result})


@bp.route("/api/rfq/<rid>/diag-package")
@auth_required
@safe_route
def api_diag_package(rid):
    """Diagnostic: test each form generation step and report what works/fails."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    results = {"rid": rid, "items": len(r.get("line_items", [])), "steps": []}
    
    # Check agency
    try:
        from src.core.agency_config import match_agency
        _agency_key, _agency_cfg = match_agency(r)
        _req = _agency_cfg.get("required_forms", [])
        results["agency"] = _agency_key
        results["required_forms"] = _req
        results["steps"].append({"step": "agency_match", "ok": True, "agency": _agency_key, "forms": _req})
    except Exception as e:
        results["steps"].append({"step": "agency_match", "ok": False, "error": str(e)})
        return jsonify(results)
    
    # Check templates dir
    import os
    tdir = os.path.join(DATA_DIR, "templates")
    if os.path.exists(tdir):
        files = os.listdir(tdir)
        results["steps"].append({"step": "templates_dir", "ok": True, "files": files})
    else:
        results["steps"].append({"step": "templates_dir", "ok": False, "error": f"{tdir} not found"})
    
    # Check quote generator
    try:
        results["steps"].append({"step": "quote_gen_import", "ok": True})
    except Exception as e:
        results["steps"].append({"step": "quote_gen_import", "ok": False, "error": str(e)})
    
    # Check each required form's generator
    form_checks = {
        "calrecycle74": ("src.forms.reytech_filler_v4", "fill_calrecycle_standalone"),
        "std204": ("src.forms.reytech_filler_v4", "fill_std204"),
        "std1000": ("src.forms.reytech_filler_v4", "fill_std1000"),
        "dvbe843": ("src.forms.reytech_filler_v4", "generate_dvbe_843"),
        "bidder_decl": ("src.forms.reytech_filler_v4", "generate_bidder_declaration"),
        "darfur_act": ("src.forms.reytech_filler_v4", "generate_darfur_act"),
        "cv012_cuf": ("src.forms.reytech_filler_v4", "fill_cv012_cuf"),
    }
    for form_id, (mod, func) in form_checks.items():
        if form_id in _req:
            try:
                m = __import__(mod, fromlist=[func])
                fn = getattr(m, func)
                results["steps"].append({"step": f"import_{form_id}", "ok": True, "func": func})
            except Exception as e:
                results["steps"].append({"step": f"import_{form_id}", "ok": False, "error": str(e)})
    
    # Check CONFIG — canonical source is src.api.config (imported into
    # dashboard + route modules). The old import from src.api.modules.routes_rfq
    # was broken (routes_rfq never defines CONFIG at module scope), so this
    # diagnostic always reported `config: ok=False` even on healthy deploys.
    try:
        from src.api.config import CONFIG
        results["steps"].append({"step": "config", "ok": True, "company": CONFIG.get("company", {}).get("name", "?")})
    except Exception as e:
        results["steps"].append({"step": "config", "ok": False, "error": str(e)})
    
    # Check line items have pricing
    items = r.get("line_items", [])
    priced = sum(1 for i in items if i.get("price_per_unit") and i["price_per_unit"] > 0)
    results["steps"].append({"step": "pricing", "items": len(items), "priced": priced})

    return jsonify(results)


# ══ Consolidated from routes_features*.py ══════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════
# Email Draft Queue Status
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/email/queue-status")
@auth_required
@safe_route
def api_email_queue_status():
    """Status of email drafts: pending, approved, sent."""
    outbox_path = os.path.join(DATA_DIR, "outbox.json")
    try:
        with open(outbox_path) as f:
            outbox = json.load(f)
    except Exception:
        outbox = []

    if isinstance(outbox, dict):
        outbox = list(outbox.values())

    draft = [e for e in outbox if (e.get("status") or "").lower() in ("draft", "pending")]
    approved = [e for e in outbox if (e.get("status") or "").lower() == "approved"]
    sent = [e for e in outbox if (e.get("status") or "").lower() == "sent"]

    return jsonify({
        "ok": True,
        "drafts": len(draft),
        "approved": len(approved),
        "sent": len(sent),
        "total": len(outbox),
        "needs_review": len(draft),
        "ready_to_send": len(approved),
        "recent_drafts": [
            {"to": e.get("to", "?"), "subject": e.get("subject", "?")[:50],
             "created": e.get("created", "?"), "type": e.get("type", "?")}
            for e in sorted(draft, key=lambda x: x.get("created", ""), reverse=True)[:5]
        ]
    })


# ═══════════════════════════════════════════════════════════════════════
# RFQs Ready to Quote
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/ready-to-quote")
@auth_required
@safe_route
def api_rfq_ready_to_quote():
    """RFQs that need pricing/quoting — prioritized by deadline."""
    rfqs_path = os.path.join(DATA_DIR, "rfqs.json")
    if not os.path.exists(rfqs_path):
        return jsonify({"ok": True, "rfqs": [], "count": 0})

    try:
        with open(rfqs_path) as f:
            rfqs = json.load(f)
    except Exception:
        return jsonify({"ok": True, "rfqs": [], "count": 0})

    today = datetime.now().strftime("%Y-%m-%d")
    ready = []

    for rid, r in rfqs.items():
        status = (r.get("status") or "").lower()
        if status in ("new", "draft", "priced", "inbox"):
            due = r.get("due_date") or r.get("deadline") or ""
            sol = r.get("solicitation_number", rid)
            items = r.get("line_items") or r.get("items_detail") or []
            if isinstance(items, str):
                try: items = json.loads(items)
                except Exception: items = []

            overdue = due and due < today
            days_left = None
            if due:
                try:
                    dd = datetime.strptime(due[:10], "%Y-%m-%d")
                    days_left = (dd - datetime.now()).days
                except (ValueError, TypeError) as e:
                    log.debug("due_date parse %r: %s", due, e)

            ready.append({
                "id": rid,
                "solicitation": sol[:30],
                "requestor": r.get("requestor", r.get("buyer_name", "?")),
                "institution": r.get("institution", "?"),
                "status": status.upper(),
                "items": len(items) if isinstance(items, list) else 0,
                "due": due[:10] if due else "TBD",
                "days_left": days_left,
                "overdue": overdue,
                "total": r.get("total_price", 0),
            })

    # Sort: overdue first, then by days_left
    ready.sort(key=lambda x: (not x["overdue"], x["days_left"] if x["days_left"] is not None else 999))

    return jsonify({
        "ok": True,
        "rfqs": ready[:20],
        "count": len(ready),
        "overdue": len([r for r in ready if r["overdue"]]),
        "due_this_week": len([r for r in ready if r.get("days_left") is not None and 0 <= r["days_left"] <= 7])
    })


@bp.route("/api/admin/scan-ghost-quote-bindings")
@auth_required
@safe_route
def api_scan_ghost_quote_bindings():
    """Scan all RFQs for `reytech_quote_number` bindings that fail the
    ghost-data gate. POST to /clear to release them.

    Incident 2026-05-01 (rfq_7813c4e1, R26Q45): the new ghost-data gate
    (#675) prevents future placeholder-sol# RFQs from burning the
    counter, but pre-existing bindings remained. This endpoint reuses
    `is_ready_for_quote_allocation` to find every RFQ where the locked
    quote_number is bound to ghost markers (placeholder sol#, zero
    items, or Reytech-internal buyer email).

    Returns:
        {
          "ok": True,
          "total_with_quote": N,
          "ghost_bound": [{"rid","quote_number","sol","reasons"}, ...],
          "clean_bound": [{"rid","quote_number","sol"}, ...],
          "ghost_count": N,
          "clean_count": N,
        }
    """
    from src.api.dashboard import is_ready_for_quote_allocation
    # Resolve DATA_DIR at request time so test fixtures that monkeypatch
    # `src.core.paths.DATA_DIR` (see conftest:129) reach this code.
    from src.core import paths as _paths
    rfqs_path = os.path.join(_paths.DATA_DIR, "rfqs.json")
    if not os.path.exists(rfqs_path):
        return jsonify({"ok": True, "total_with_quote": 0,
                        "ghost_bound": [], "clean_bound": [],
                        "ghost_count": 0, "clean_count": 0})
    with open(rfqs_path) as f:
        rfqs = json.load(f)

    ghost_bound, clean_bound = [], []
    for rid, r in rfqs.items():
        qn = r.get("reytech_quote_number") or ""
        if not qn:
            continue
        ok, reasons = is_ready_for_quote_allocation(r)
        entry = {
            "rid": rid,
            "quote_number": qn,
            "sol": r.get("solicitation_number") or "",
            "agency": r.get("agency") or r.get("institution") or "",
            "requestor_email": r.get("requestor_email") or "",
            "items_count": len(r.get("line_items") or r.get("items") or []),
            "status": r.get("status") or "",
        }
        if ok:
            clean_bound.append(entry)
        else:
            entry["reasons"] = reasons
            ghost_bound.append(entry)

    # Sort ghost by quote seq descending so the most recent burns are first
    def _seq(qn: str) -> int:
        try:
            return int(qn.split("Q", 1)[-1]) if "Q" in qn else 0
        except (ValueError, IndexError):
            return 0
    ghost_bound.sort(key=lambda e: _seq(e["quote_number"]), reverse=True)
    clean_bound.sort(key=lambda e: _seq(e["quote_number"]), reverse=True)

    return jsonify({
        "ok": True,
        "total_with_quote": len(ghost_bound) + len(clean_bound),
        "ghost_count": len(ghost_bound),
        "clean_count": len(clean_bound),
        "ghost_bound": ghost_bound,
        "clean_bound": clean_bound,
    })


@bp.route("/api/admin/scan-contaminated-reprices")
@auth_required
@safe_route
def api_scan_contaminated_reprices():
    """Blast-radius scan for the Oracle cross-category contamination bug.

    Incident 2026-05-29 (rfq_fca653f6 item 5): a $2.00 composition notebook
    repriced to a $74.32 bid (~3,616% markup) because `_search_product_catalog`
    flat-OR'd its tokens and let $70 cross-category items poison the market
    average. The source + backstop fixes are forward-only — bids already
    written stay frozen. This read-only scan finds every persisted line whose
    bid implies an absurd markup over its own cost, so they can be re-priced
    or hand-corrected (§5 fix-the-data).

    A line is flagged when cost > 0, bid > 0, and implied markup
    `(bid - cost) / cost * 100` exceeds `min_markup_pct` (default 300, the
    same family as the 400% engine backstop but lower so near-misses surface).
    Sorted by extended-bid dollar impact descending — biggest damage first.

    Query params:
        min_markup_pct: int — flag threshold (default 300)

    Read-only. Does NOT mutate any quote. Returns the candidate list only.
    """
    try:
        min_markup = float(request.args.get("min_markup_pct", 300))
    except (TypeError, ValueError):
        min_markup = 300.0

    from src.core import paths as _paths
    rfqs_path = os.path.join(_paths.DATA_DIR, "rfqs.json")
    if not os.path.exists(rfqs_path):
        return jsonify({"ok": True, "min_markup_pct": min_markup,
                        "flagged_count": 0, "flagged": []})
    with open(rfqs_path) as f:
        rfqs = json.load(f)

    def _f(v):
        try:
            return float(str(v if v is not None else 0).replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            return 0.0

    flagged = []
    for rid, r in rfqs.items():
        items = r.get("line_items") or r.get("items") or []
        for idx, it in enumerate(items):
            cost = _f(it.get("supplier_cost"))
            bid = _f(it.get("price_per_unit"))
            if cost <= 0 or bid <= 0:
                continue
            implied = (bid - cost) / cost * 100
            if implied <= min_markup:
                continue
            qty = _f(it.get("qty") or it.get("quantity") or 1) or 1
            flagged.append({
                "rid": rid,
                "sol": r.get("solicitation_number") or "",
                "agency": r.get("agency") or r.get("institution") or "",
                "status": r.get("status") or "",
                "line": idx + 1,
                "description": (it.get("description") or "")[:70],
                "qty": qty,
                "cost": round(cost, 2),
                "bid": round(bid, 2),
                "implied_markup_pct": round(implied, 1),
                "ext_bid": round(bid * qty, 2),
                "repriced_reason": it.get("repriced_reason") or "",
                "scprs_ref": _f(it.get("scprs_last_price")) or None,
            })

    flagged.sort(key=lambda e: e["ext_bid"], reverse=True)
    return jsonify({
        "ok": True,
        "min_markup_pct": min_markup,
        "flagged_count": len(flagged),
        "total_exposure": round(sum(e["ext_bid"] for e in flagged), 2),
        "flagged": flagged,
    })


@bp.route("/api/admin/clear-ghost-quote-bindings", methods=["POST"])
@auth_required
@safe_route
def api_clear_ghost_quote_bindings():
    """Clear `reytech_quote_number` on every RFQ flagged by the scan.

    Idempotent — only clears RFQs that currently fail the ghost gate.
    Body: `{"dry_run": true}` returns what WOULD be cleared without
    mutating; default is to clear.
    """
    from src.api.dashboard import (
        is_ready_for_quote_allocation,
        _save_single_rfq,
    )
    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get("dry_run", False))

    from src.core import paths as _paths
    rfqs_path = os.path.join(_paths.DATA_DIR, "rfqs.json")
    if not os.path.exists(rfqs_path):
        return jsonify({"ok": True, "cleared": [], "count": 0})
    with open(rfqs_path) as f:
        rfqs = json.load(f)

    cleared = []
    for rid, r in list(rfqs.items()):
        qn = r.get("reytech_quote_number") or ""
        if not qn:
            continue
        ok, reasons = is_ready_for_quote_allocation(r)
        if ok:
            continue  # not a ghost — leave alone
        cleared.append({
            "rid": rid,
            "quote_number": qn,
            "sol": r.get("solicitation_number") or "",
            "reasons": reasons,
        })
        if not dry_run:
            rfqs[rid]["reytech_quote_number"] = ""

    # Single atomic write back to the same rfqs.json the scan read from.
    # Avoids `_save_single_rfq` here because that helper resolves DATA_DIR
    # through the dashboard module which can drift from `_paths.DATA_DIR`
    # in test fixtures, leaving the file appearing unchanged.
    if not dry_run and cleared:
        with open(rfqs_path, "w", encoding="utf-8") as f:
            json.dump(rfqs, f, indent=2)

    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "count": len(cleared),
        "cleared": cleared,
    })


@bp.route("/api/rfq/<rid>/clean-items", methods=["POST"])
@auth_required
@safe_route
def rfq_clean_items(rid):
    """Remove junk items (legal text, instructions, boilerplate) from an RFQ.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _rfq_clean_items_locked(rid)


def _rfq_clean_items_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    from src.api.dashboard import load_rfqs, save_rfqs
    rfqs = load_rfqs()
    rfq = rfqs.get(rid)
    if not rfq:
        return jsonify({"ok": False, "error": "RFQ not found"})

    items = rfq.get("line_items", [])
    original_count = len(items)

    from src.forms.price_check import _filter_junk_items
    cleaned = _filter_junk_items(items)

    _sync_rfq_items(rfq, cleaned)
    if "parsed" in rfq:
        rfq["parsed"]["line_items"] = cleaned

    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, rfq)

    removed = original_count - len(cleaned)
    return jsonify({"ok": True, "removed": removed, "kept": len(cleaned), "original": original_count})


# ═══════════════════════════════════════════════════════════════════
# Package Manifest + Lifecycle API
# ═══════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/<rid>/manifest")
@auth_required
@safe_route
def api_rfq_manifest(rid):
    """Get the latest package manifest for an RFQ."""
    from src.core.dal import get_latest_manifest
    manifest = get_latest_manifest(rid)
    if not manifest:
        return jsonify({"ok": False, "error": "No package manifest found"})
    return jsonify({"ok": True, "manifest": manifest})


@bp.route("/api/rfq/<rid>/manifest/<int:manifest_id>/review", methods=["POST"])
@auth_required
@safe_route
def api_rfq_review_form(rid, manifest_id):
    """Record a review verdict for a form in the manifest."""
    from src.core.dal import review_form, log_lifecycle_event
    data = request.get_json(force=True, silent=True) or {}
    form_id = data.get("form_id", "")
    verdict = data.get("verdict", "approved")
    notes = data.get("notes", "")
    if not form_id:
        return jsonify({"ok": False, "error": "form_id required"})
    ok, err = review_form(manifest_id, form_id, verdict, reviewed_by="user", notes=notes)
    if ok:
        log_lifecycle_event("rfq", rid, "form_reviewed",
            f"Form {form_id}: {verdict}" + (f" — {notes}" if notes else ""),
            actor="user", detail={"form_id": form_id, "verdict": verdict, "manifest_id": manifest_id})
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": err or "review not recorded"})


@bp.route("/api/rfq/<rid>/manifest/<int:manifest_id>/approve", methods=["POST"])
@auth_required
@safe_route
def api_rfq_approve_package(rid, manifest_id):
    """Approve the entire package (all forms must be reviewed first).
    Pass ?force=1 to skip pending/rejected/QA checks.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_approve_package_locked(rid, manifest_id)


def _api_rfq_approve_package_locked(rid, manifest_id):
    """Inner body — always runs under `_save_rfqs_lock`."""
    from src.core.dal import get_package_manifest, update_manifest_status, log_lifecycle_event
    manifest = get_package_manifest(manifest_id)
    if not manifest:
        return jsonify({"ok": False, "error": "Manifest not found"})
    force = request.args.get("force") == "1"
    if not force:
        pending = [r for r in manifest.get("reviews", []) if r.get("verdict") == "pending"]
        if pending:
            return jsonify({"ok": False, "error": f"{len(pending)} forms still pending review",
                            "pending": [r["form_id"] for r in pending]})
        rejected = [r for r in manifest.get("reviews", []) if r.get("verdict") == "rejected"]
        if rejected:
            return jsonify({"ok": False, "error": f"{len(rejected)} forms rejected",
                            "rejected": [r["form_id"] for r in rejected]})
        # Block if Form QA failed
        field_audit = manifest.get("field_audit") or {}
        if isinstance(field_audit, str):
            try:
                field_audit = json.loads(field_audit)
            except Exception:
                field_audit = {}
        if field_audit.get("_qa_passed") is False:
            qa_issues = field_audit.get("_qa_summary", {}).get("critical_issues", [])
            return jsonify({"ok": False,
                            "error": f"Form QA failed with {len(qa_issues)} critical issue(s). Regenerate the package to fix.",
                            "qa_issues": qa_issues[:5]})
    ok = update_manifest_status(manifest_id, "approved")
    if ok:
        log_lifecycle_event("rfq", rid, "package_approved",
            f"Package v{manifest.get('version', '?')} approved ({manifest.get('total_forms', 0)} forms)",
            actor="user", detail={"manifest_id": manifest_id, "version": manifest.get("version")})
        # Update RFQ status to ready_to_send
        try:
            rfqs = load_rfqs()
            r = rfqs.get(rid, {})
            r["status"] = "ready_to_send"
            if not r.get("draft_email"):
                r["draft_email"] = {
                    "to": r.get("requestor_email", ""),
                    "subject": f"Reytech Inc. — RFQ Response #{r.get('solicitation_number', '')}",
                    "body": "Please find attached our bid response package.",
                }
            _save_single_rfq(rid, r)
        except Exception as _e:
            log.warning("approve_package: status update failed: %s", _e)
    return jsonify({"ok": ok, "status": "approved"})


@bp.route("/api/rfq/<rid>/manifest/<int:manifest_id>/remove-form", methods=["POST"])
@auth_required
@safe_route
def api_rfq_remove_form(rid, manifest_id):
    """Remove a form from the package manifest and delete its file.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_remove_form_locked(rid, manifest_id)


def _api_rfq_remove_form_locked(rid, manifest_id):
    """Inner body — always runs under `_save_rfqs_lock`."""
    data = request.get_json(force=True, silent=True) or {}
    form_id = data.get("form_id", "")
    if not form_id:
        return jsonify({"ok": False, "error": "form_id required"})

    try:
        from src.core.db import get_db
        with get_db() as conn:
            # Verify manifest belongs to this RFQ
            _owner = conn.execute(
                "SELECT rfq_id FROM package_manifest WHERE id = ?",
                (manifest_id,)).fetchone()
            if not _owner or _owner[0] != rid:
                return jsonify({"ok": False, "error": "Manifest not found for this RFQ"})

            # Get the review record to find the filename
            row = conn.execute(
                "SELECT form_filename FROM package_review WHERE manifest_id = ? AND form_id = ?",
                (manifest_id, form_id)).fetchone()
            filename = row[0] if row else ""

            # Delete the review record
            conn.execute(
                "DELETE FROM package_review WHERE manifest_id = ? AND form_id = ?",
                (manifest_id, form_id))

            # Update the manifest's generated_forms list
            manifest_row = conn.execute(
                "SELECT generated_forms, total_forms FROM package_manifest WHERE id = ?",
                (manifest_id,)).fetchone()
            if manifest_row:
                import json as _json_rm
                gen_forms = _json_rm.loads(manifest_row[0] or "[]")
                gen_forms = [f for f in gen_forms if (f.get("form_id") if isinstance(f, dict) else f) != form_id]
                total = (manifest_row[1] or 0) - 1
                conn.execute(
                    "UPDATE package_manifest SET generated_forms = ?, total_forms = ? WHERE id = ?",
                    (_json_rm.dumps(gen_forms), max(total, 0), manifest_id))

            # Delete the actual file from disk
            if filename:
                rfqs = load_rfqs()
                r = rfqs.get(rid, {})
                sol = r.get("solicitation_number", "") or r.get("rfq_number", "") or "RFQ"
                filepath = os.path.join(OUTPUT_DIR, sol, filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    log.info("Removed file %s for form %s", filename, form_id)

                # Remove from rfq output_files list
                out_files = r.get("output_files", [])
                if filename in out_files:
                    out_files.remove(filename)
                    r["output_files"] = out_files
                    from src.api.dashboard import _save_single_rfq
                    _save_single_rfq(rid, r)

            # Log the removal
            from src.core.dal import log_lifecycle_event
            log_lifecycle_event("rfq", rid, "form_removed",
                f"Removed {form_id} from package ({filename})",
                actor="user", detail={"form_id": form_id, "filename": filename, "manifest_id": manifest_id})

        return jsonify({"ok": True, "removed": form_id, "filename": filename})
    except Exception as e:
        log.error("Remove form failed: %s", e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/rfq/<rid>/refill/<form_id>", methods=["POST"])
@auth_required
@safe_route
def api_rfq_refill_form(rid, form_id):
    """Refill a single form with updated data — inline editing from review page.

    Accepts field_overrides in JSON body, merges into RFQ data, regenerates
    only the specified form, runs QA, and resets verdict to pending.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_refill_form_locked(rid, form_id)


def _api_rfq_refill_form_locked(rid, form_id):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404

    data = request.get_json(force=True, silent=True) or {}
    overrides = data.get("field_overrides", {})

    # ── Merge overrides into RFQ ──
    # Line item overrides: [{index: 0, price_per_unit: 123.45}, ...]
    items = r.get("line_items", r.get("items", []))
    for item_override in overrides.get("line_items", []):
        idx = item_override.get("index")
        if idx is not None and 0 <= idx < len(items):
            for k, v in item_override.items():
                if k == "index":
                    continue
                # 704B guard: block buyer-field changes
                if form_id == "704b" and k in ("description", "qty", "uom", "department"):
                    continue
                items[idx][k] = v
    _sync_rfq_items(r, items)

    # Top-level overrides (solicitation_number, custom_notes, etc.)
    for k, v in overrides.items():
        if k == "line_items":
            continue
        r[k] = v

    # Update sign date
    try:
        from src.forms.reytech_filler_v4 import get_pst_date
        r["sign_date"] = get_pst_date()
    except ImportError:
        from datetime import datetime as _dt
        r["sign_date"] = _dt.now().strftime("%m/%d/%Y")

    # Save updated RFQ
    from src.api.dashboard import _save_single_rfq
    _save_single_rfq(rid, r)

    # ── Find output path and template for this form ──
    sol = r.get("solicitation_number", "") or r.get("rfq_number", "") or "RFQ"
    out_dir = os.path.join(OUTPUT_DIR, sol)
    os.makedirs(out_dir, exist_ok=True)
    tmpl = r.get("templates", {})

    # Map form_id → fill function + template + output filename
    TEMPLATE_DIR = os.path.join(DATA_DIR, "templates")
    FORM_MAP = {
        "703b":          {"fn": "fill_703b",    "tmpl_key": "703b",    "filename": f"{sol}_703B_Reytech.pdf"},
        "703c":          {"fn": "fill_703c",    "tmpl_key": "703c",    "filename": f"{sol}_703C_Reytech.pdf"},
        "704b":          {"fn": "fill_704b",    "tmpl_key": "704b",    "filename": f"{sol}_704B_Reytech.pdf"},
        "bidpkg":        {"fn": "fill_bid_package", "tmpl_key": "bidpkg", "filename": f"{sol}_BidPackage_Reytech.pdf"},
        "calrecycle74":  {"fn": "fill_calrecycle_standalone", "tmpl_file": "calrecycle_74_blank.pdf", "filename": f"{sol}_CalRecycle74_Reytech.pdf"},
        "darfur_act":    {"fn": "fill_darfur_standalone", "tmpl_file": "darfur_act_blank.pdf", "filename": f"{sol}_DarfurAct_Reytech.pdf"},
        "darfur":        {"fn": "fill_darfur_standalone", "tmpl_file": "darfur_act_blank.pdf", "filename": f"{sol}_DarfurAct_Reytech.pdf"},  # alias
        "cv012_cuf":     {"fn": "fill_cv012_cuf", "tmpl_file": "cv012_cuf_blank.pdf", "filename": f"{sol}_CV012_CUF_Reytech.pdf"},
        "std204":        {"fn": "fill_std204",  "tmpl_file": "std204_blank.pdf", "filename": f"{sol}_STD204_Reytech.pdf"},
        "std205":        {"fn": "fill_std205",  "tmpl_file": "std205_blank.pdf", "filename": f"{sol}_STD205_Reytech.pdf"},
        "std1000":       {"fn": "fill_std1000", "tmpl_file": "std1000_blank.pdf", "filename": f"{sol}_STD1000_Reytech.pdf"},
        "bidder_decl":   {"fn": "fill_bidder_declaration", "tmpl_file": "bidder_declaration_blank.pdf", "filename": f"{sol}_BidderDecl_Reytech.pdf"},
        "dvbe843":       {"fn": "generate_dvbe_843", "tmpl_file": "dvbe_843_blank.pdf", "filename": f"{sol}_DVBE843_Reytech.pdf", "no_template_arg": True},
        "sellers_permit": {"fn": "_copy_static", "tmpl_file": "sellers_permit_reytech.pdf", "filename": f"{sol}_SellersPermit_Reytech.pdf", "static": True},
    }

    if form_id == "quote":
        # Quote uses a separate generation path
        try:
            from src.forms.quote_generator import generate_quote_from_rfq
            locked_qn = r.get("reytech_quote_number", "")
            _q_output = os.path.join(out_dir, f"{sol}_Quote_Reytech.pdf")
            result = generate_quote_from_rfq(r, _q_output, quote_number=locked_qn)
            if not result.get("ok"):
                return jsonify({"ok": False, "error": f"Quote refill failed: {result.get('error')}"})
        except Exception as e:
            return jsonify({"ok": False, "error": f"Quote refill error: {e}"}), 500
    elif form_id in FORM_MAP:
        fm = FORM_MAP[form_id]
        # Resolve template path
        if "tmpl_key" in fm:
            template_path = tmpl.get(fm["tmpl_key"], "")
        else:
            template_path = os.path.join(TEMPLATE_DIR, fm["tmpl_file"])
        if not template_path or not os.path.exists(template_path):
            return jsonify({"ok": False, "error": f"Template not found for {form_id}"}), 400

        # Call the fill function
        try:
            output_path = os.path.join(out_dir, fm["filename"])
            if fm.get("static"):
                # Static file — just copy the template (e.g., sellers_permit)
                import shutil
                shutil.copy2(template_path, output_path)
            elif fm.get("no_template_arg"):
                # Function generates from scratch (e.g., generate_dvbe_843)
                import src.forms.reytech_filler_v4 as filler
                fill_fn = getattr(filler, fm["fn"])
                fill_fn(r, CONFIG, output_path)
            else:
                import src.forms.reytech_filler_v4 as filler
                fill_fn = getattr(filler, fm["fn"])
                fill_fn(template_path, r, CONFIG, output_path)
        except Exception as e:
            log.error("Refill %s failed: %s", form_id, e, exc_info=True)
            return jsonify({"ok": False, "error": f"Fill failed: {e}"}), 500
    else:
        return jsonify({"ok": False, "error": f"Unknown form_id: {form_id}"}), 400

    # ── Run QA on refilled form ──
    _qa_result = {}
    try:
        from src.forms.form_qa import verify_single_form
        _out_file = os.path.join(out_dir, FORM_MAP.get(form_id, {}).get("filename", ""))
        if form_id == "quote":
            _out_file = os.path.join(out_dir, f"{sol}_Quote_Reytech.pdf")
        _qa_result = verify_single_form(_out_file, form_id, r, CONFIG)
    except Exception as _qe:
        log.debug("Refill QA skipped: %s", _qe)

    # ── Reset verdict to pending ──
    try:
        from src.core.dal import get_latest_manifest, reset_form_verdict
        manifest = get_latest_manifest(rid)
        if manifest:
            reset_form_verdict(manifest["id"], form_id)
    except Exception as _rv:
        log.warning("Verdict reset failed: %s", _rv)

    log.info("REFILL %s/%s: success (overrides: %s)", rid, form_id,
             list(overrides.keys()) if overrides else "none")

    return jsonify({
        "ok": True,
        "form_id": form_id,
        "qa": _qa_result,
    })


@bp.route("/api/rfq/<rid>/timeline")
@auth_required
@safe_route
def api_rfq_timeline(rid):
    """Get the full lifecycle timeline for an RFQ."""
    from src.core.dal import get_lifecycle_events
    events = get_lifecycle_events("rfq", rid, limit=200)
    return jsonify({"ok": True, "events": events, "count": len(events)})


@bp.route("/api/rfq/<rid>/buyer-prefs")
@auth_required
@safe_route
def api_rfq_buyer_prefs(rid):
    """Get buyer preferences for the RFQ's requestor."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})
    email = r.get("requestor_email", "")
    if not email:
        return jsonify({"ok": True, "preferences": [], "message": "No requestor email"})
    from src.core.dal import get_buyer_preferences
    prefs = get_buyer_preferences(email)
    return jsonify({"ok": True, "preferences": prefs, "buyer_email": email})


@bp.route("/api/rfq/<rid>/download-complete-package")
@auth_required
@safe_route
def api_download_complete_package(rid):
    """Download ALL forms merged into one PDF — quote + compliance."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    sol = r.get("solicitation_number", "") or r.get("rfq_number", "") or "RFQ"
    out_dir = os.path.join(OUTPUT_DIR, sol)
    output_files = r.get("output_files", [])

    if not output_files:
        return jsonify({"ok": False, "error": "No files generated"})

    try:
        from pypdf import PdfReader, PdfWriter
        writer = PdfWriter()
        merged_count = 0

        # Quote first, then all other forms in order
        quote_files = [f for f in output_files if "Quote" in f and "704" not in f.upper()]
        other_files = [f for f in output_files if f not in quote_files]
        ordered = quote_files + other_files

        for f in ordered:
            fpath = os.path.join(out_dir, f)
            if not os.path.exists(fpath):
                continue
            # Skip the merged package file itself (avoid double-counting)
            if "RFQ_Package" in f or "Compliance_Forms" in f:
                continue
            try:
                reader = PdfReader(fpath)
                for page in reader.pages:
                    text = ""
                    try:
                        text = page.extract_text() or ""
                    except Exception as _e:
                        log.debug('suppressed in api_download_complete_package: %s', _e)
                    if text.strip().startswith("Please wait") and len(text.strip()) < 300:
                        continue
                    writer.add_page(page)
                merged_count += 1
            except Exception as _e:
                log.warning("Skip %s in complete package: %s", f, _e)

        if merged_count == 0:
            return jsonify({"ok": False, "error": "No valid PDFs to merge"})

        import io
        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)

        _safe_agency = ""
        try:
            from src.core.agency_config import match_agency
            _ak, _ac = match_agency(r)
            _safe_agency = (_ac.get("name", "") or "").replace(" ", "").replace("/", "")[:20]
        except Exception as _e:
            log.debug('suppressed in api_download_complete_package: %s', _e)

        filename = f"Complete_RFQ_{_safe_agency}_{sol}_ReytechInc.pdf" if _safe_agency else f"Complete_RFQ_{sol}_ReytechInc.pdf"

        from flask import send_file
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)
    except Exception as e:
        log.error("Complete package download failed: %s", e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/rfq/<rid>/export-invoice")
@auth_required
@safe_route
def api_export_invoice(rid):
    """Export buyer invoice as Excel for QB entry."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    items = r.get("line_items", r.get("items", []))
    sol = r.get("solicitation_number", "") or r.get("rfq_number", "")
    quote_num = r.get("reytech_quote_number", "")
    buyer = r.get("requestor_name", "")
    agency = r.get("agency", "") or r.get("agency_name", "")
    tax_rate = float(r.get("tax_rate", 0) or 0) / 100
    delivery = r.get("delivery_location", "")

    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl not installed — run: pip install openpyxl"}), 500

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        hf = Font(bold=True, size=14)
        sf = Font(size=11, color="333333")
        cf = Font(bold=True, size=10, color="FFFFFF")
        cfill = PatternFill("solid", fgColor="2E75B6")
        mf = '#,##0.00'
        bd = Border(bottom=Side(style='thin', color='CCCCCC'))

        ws.merge_cells('A1:G1')
        ws['A1'] = "REYTECH INC. — INVOICE"
        ws['A1'].font = hf
        ws['A2'] = f"Quote #: {quote_num}"
        ws['A2'].font = sf
        ws['A3'] = f"Bill To: {buyer} — {agency}"
        ws['A3'].font = sf
        ws['A4'] = f"Ship To: {delivery}"
        ws['A4'].font = sf
        ws['A5'] = f"Solicitation #: {sol}"
        ws['A5'].font = sf

        for col, h in enumerate(["#", "Description", "Part #", "QTY", "UOM", "Unit Price", "Subtotal"], 1):
            c = ws.cell(row=7, column=col, value=h)
            c.font = cf
            c.fill = cfill
            c.alignment = Alignment(horizontal='center')

        subtotal = 0
        for idx, item in enumerate(items):
            row = 8 + idx
            qty = int(float(item.get("qty", 1) or 1))
            price = float(item.get("price_per_unit", 0) or 0)
            lt = qty * price
            subtotal += lt
            ws.cell(row=row, column=1, value=idx+1).border = bd
            ws.cell(row=row, column=2, value=(item.get("description", "") or "")[:80]).border = bd
            ws.cell(row=row, column=3, value=item.get("part_number", "") or item.get("item_number", "")).border = bd
            ws.cell(row=row, column=4, value=qty).border = bd
            ws.cell(row=row, column=5, value=item.get("uom", "EA")).border = bd
            ws.cell(row=row, column=6, value=price).number_format = mf
            ws.cell(row=row, column=7, value=lt).number_format = mf

        tr = 8 + len(items) + 1
        tax_amt = subtotal * tax_rate
        ws.cell(row=tr, column=6, value="Subtotal:").font = Font(bold=True)
        ws.cell(row=tr, column=7, value=subtotal).number_format = mf
        ws.cell(row=tr+1, column=6, value=f"Tax ({r.get('tax_rate', 0)}%):").font = Font(bold=True)
        ws.cell(row=tr+1, column=7, value=tax_amt).number_format = mf
        ws.cell(row=tr+2, column=6, value="TOTAL:").font = Font(bold=True, size=12)
        ws.cell(row=tr+2, column=7, value=subtotal + tax_amt).number_format = mf
        ws.cell(row=tr+2, column=7).font = Font(bold=True, size=12)

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"Invoice_{quote_num}_{sol}_Reytech.xlsx")
    except Exception as e:
        log.error("Invoice export: %s", e)
        return jsonify({"ok": False, "error": str(e)})


@bp.route("/api/rfq/<rid>/export-supplier-po")
@auth_required
@safe_route
def api_export_supplier_po(rid):
    """Export supplier PO as Excel for QB entry."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"})

    items = r.get("line_items", r.get("items", []))
    sol = r.get("solicitation_number", "") or r.get("rfq_number", "")
    quote_num = r.get("reytech_quote_number", "")
    supplier_name = ""
    for item in items:
        sn = item.get("cost_supplier_name", "") or item.get("scprs_supplier", "")
        if sn:
            supplier_name = sn
            break

    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl not installed — run: pip install openpyxl"}), 500

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supplier PO"

        hf = Font(bold=True, size=14)
        sf = Font(size=11, color="333333")
        cf = Font(bold=True, size=10, color="FFFFFF")
        cfill = PatternFill("solid", fgColor="2D6A2E")
        mf = '#,##0.00'
        bd = Border(bottom=Side(style='thin', color='CCCCCC'))

        ws.merge_cells('A1:H1')
        ws['A1'] = "REYTECH INC. — PURCHASE ORDER"
        ws['A1'].font = hf
        ws['A2'] = f"Related Quote #: {quote_num}"
        ws['A2'].font = sf
        ws['A3'] = f"Supplier: {supplier_name}"
        ws['A3'].font = sf
        ws['A4'] = f"Solicitation #: {sol}"
        ws['A4'].font = sf

        for col, h in enumerate(["#", "Description", "Part #", "QTY", "UOM", "Supplier Cost", "Subtotal", "Source"], 1):
            c = ws.cell(row=6, column=col, value=h)
            c.font = cf
            c.fill = cfill
            c.alignment = Alignment(horizontal='center')

        subtotal = 0
        for idx, item in enumerate(items):
            row = 7 + idx
            qty = int(float(item.get("qty", 1) or 1))
            cost = float(item.get("supplier_cost", 0) or item.get("vendor_cost", 0) or 0)
            lt = qty * cost
            subtotal += lt
            source = item.get("cost_source", "")
            sname = item.get("cost_supplier_name", "")
            ws.cell(row=row, column=1, value=idx+1).border = bd
            ws.cell(row=row, column=2, value=(item.get("description", "") or "")[:80]).border = bd
            ws.cell(row=row, column=3, value=item.get("part_number", "") or item.get("item_number", "")).border = bd
            ws.cell(row=row, column=4, value=qty).border = bd
            ws.cell(row=row, column=5, value=item.get("uom", "EA")).border = bd
            ws.cell(row=row, column=6, value=cost).number_format = mf
            ws.cell(row=row, column=7, value=lt).number_format = mf
            ws.cell(row=row, column=8, value=f"{source} — {sname}" if sname else source).border = bd

        tr = 7 + len(items) + 1
        bid_total = sum(int(float(i.get("qty", 1) or 1)) * float(i.get("price_per_unit", 0) or 0) for i in items)
        ws.cell(row=tr, column=6, value="TOTAL:").font = Font(bold=True, size=12)
        ws.cell(row=tr, column=7, value=subtotal).number_format = mf
        ws.cell(row=tr, column=7).font = Font(bold=True, size=12)
        ws.cell(row=tr+2, column=5, value="Supplier Cost:").font = Font(bold=True)
        ws.cell(row=tr+2, column=7, value=subtotal).number_format = mf
        ws.cell(row=tr+3, column=5, value="Bid Total:").font = Font(bold=True)
        ws.cell(row=tr+3, column=7, value=bid_total).number_format = mf
        ws.cell(row=tr+4, column=5, value="Gross Margin:").font = Font(bold=True)
        ws.cell(row=tr+4, column=7, value=bid_total - subtotal).number_format = mf
        ws.cell(row=tr+4, column=7).font = Font(bold=True, color="2D6A2E")

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 25

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        safe_sup = supplier_name.replace(" ", "")[:20]
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"SupplierPO_{quote_num}_{sol}_{safe_sup}.xlsx")
    except Exception as e:
        log.error("Supplier PO export: %s", e)
        return jsonify({"ok": False, "error": str(e)})


# ═══════════════════════════════════════════════════════════════════════
# Bookkeeper Export — QBO CSV + deal summary
# ═══════════════════════════════════════════════════════════════════════

def _build_bookkeeper_data(r, rid):
    """Build QBO CSV content and deal summary from an RFQ dict.
    Returns (csv_content, summary, filename)."""
    import csv
    import io

    items = r.get("line_items", r.get("items", []))
    sol = r.get("solicitation_number", "")
    po = r.get("po_number", "")
    agency = r.get("agency_name", r.get("agency", ""))
    institution = r.get("institution", "")
    ship_to = r.get("delivery_location", r.get("ship_to", ""))

    # Build QBO CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Product/Service Name", "Sales Description", "Purchase Description",
        "Sales Price / Rate", "Purchase Cost", "SKU", "Taxable", "Type",
        "Income Account", "Expense Account"
    ])

    subtotal = 0
    cost_total = 0
    bid_count = 0
    for item in items:
        if item.get("no_bid"):
            continue
        bid_count += 1
        desc = (item.get("description", "") or "")[:100]
        mfg = item.get("mfg_number", "") or item.get("part_number", "") or ""
        product_name = mfg if mfg else desc[:40]
        sell_price = (item.get("unit_price")
                      or item.get("price_per_unit")
                      or (item.get("pricing", {}) or {}).get("recommended_price")
                      or 0)
        # PR mr-wolf #2: cost via the canonical reader. Was a local
        # `vendor_cost → supplier_cost → pricing.unit_cost` chain with
        # the WRONG priority (vendor_cost from scrape won over
        # operator-typed supplier_cost). Migration flips priority to
        # match the rest of the stack — operator-typed wins.
        from src.core.pricing_math import cost_from_contract as _cfc_admin
        cost = _cfc_admin(item)
        qty = item.get("qty", 1) or 1
        uom = (item.get("uom") or "EA").upper()

        subtotal += float(sell_price) * float(qty)
        cost_total += float(cost) * float(qty)

        writer.writerow([
            product_name,
            f"{product_name} {desc[:60]}",
            f"{product_name} {desc[:60]}",
            f"{float(sell_price):.2f}",
            f"{float(cost):.2f}",
            uom,
            "Yes",
            "Non-inventory",
            "Sales of Product Income",
            "Cost of Goods Sold",
        ])

    csv_content = output.getvalue()

    # Deal summary
    tax_rate = 0.0725
    tax = round(subtotal * tax_rate, 2)
    total = round(subtotal + tax, 2)
    profit = round(subtotal - cost_total, 2)
    margin = round(profit / subtotal * 100, 1) if subtotal > 0 else 0

    summary_lines = [
        f"DEAL SUMMARY — PO #{po or 'TBD'}",
        f"Agency: {agency}",
        f"Institution: {institution}",
        f"Ship To: {ship_to}",
        f"Solicitation: {sol}",
        "",
        f"Items: {bid_count}",
        f"Subtotal: ${subtotal:,.2f}",
        f"Sales Tax ({tax_rate * 100:.2f}%): ${tax:,.2f}",
        f"Total: ${total:,.2f}",
        "",
        f"Cost Total: ${cost_total:,.2f}",
        f"Profit: ${profit:,.2f} ({margin}% margin)",
    ]
    summary = "\n".join(summary_lines)
    filename = f"QBO_Import_{po or sol or rid}.csv"

    return csv_content, summary, filename


@bp.route("/api/rfq/<rid>/bookkeeper-export", methods=["POST"])
@auth_required
@safe_route
def api_bookkeeper_export(rid):
    """Generate QBO CSV + deal summary for bookkeeper."""
    try:
        bad = _validate_rid(rid)
        if bad:
            return bad
        from src.api.dashboard import load_rfqs
        rfqs = load_rfqs()
        r = rfqs.get(rid)
        if not r:
            return jsonify({"ok": False, "error": "RFQ not found"}), 404

        csv_content, summary, filename = _build_bookkeeper_data(r, rid)

        return jsonify({
            "ok": True,
            "csv": csv_content,
            "summary": summary,
            "filename": filename,
        })
    except Exception as e:
        log.error("Bookkeeper export error: %s", e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/rfq/<rid>/bookkeeper-csv")
@auth_required
@safe_route
def api_bookkeeper_csv(rid):
    """Download QBO CSV directly."""
    try:
        bad = _validate_rid(rid)
        if bad:
            return bad
        from src.api.dashboard import load_rfqs
        from flask import Response
        rfqs = load_rfqs()
        r = rfqs.get(rid)
        if not r:
            return jsonify({"ok": False, "error": "RFQ not found"}), 404

        csv_content, _summary, filename = _build_bookkeeper_data(r, rid)

        return Response(
            csv_content,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log.error("Bookkeeper CSV download error: %s", e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/rfq/<rid>/re-extract-requirements", methods=["POST"])
@auth_required
@safe_route
def api_rfq_re_extract_requirements(rid):
    """Re-run email requirement extraction on demand.

    Used when: user manually uploads new email, forwards changed, or wants
    to refresh extraction after body edits.

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_re_extract_requirements_locked(rid)


def _api_rfq_re_extract_requirements_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404

    body_text = r.get("body_text", "") or r.get("body_preview", "") or ""
    subject = r.get("email_subject", "")
    attachments = [{"filename": f} for f in r.get("attachments_raw", [])]

    if not body_text and not subject:
        return jsonify({"ok": False, "error": "No email body or subject to extract from"})

    try:
        import json as _j
        from src.agents.requirement_extractor import extract_requirements
        req = extract_requirements(body_text, subject, attachments)
        if req and req.has_requirements:
            r["requirements_json"] = _j.dumps(req.to_dict(), default=str)
            # Supplement due date if missing
            if r.get("due_date") in ("TBD", "", None) and req.due_date:
                r["due_date"] = req.due_date
            # Supplement Release/Issue Date so the 703B/703C fills (#1207 follow-up)
            if req.release_date and not r.get("release_date"):
                r["release_date"] = req.release_date
            rfqs[rid] = r
            _save_single_rfq(rid, r)
            return jsonify({
                "ok": True,
                "method": req.extraction_method,
                "confidence": req.confidence,
                "forms_found": len(req.forms_required),
                "requirements": req.to_dict(),
            })
        else:
            return jsonify({"ok": True, "method": "none", "forms_found": 0,
                           "message": "No requirements detected"})
    except Exception as e:
        log.error("Re-extract requirements error: %s", e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════
# PR-D4: editable Quote PDF (download editable copy, upload edited copy)
# ═══════════════════════════════════════════════════════════════════════
#
# Mike's 10-min escape valve (per feedback_ten_minute_escape_valve):
#   1. Operator downloads /api/rfq/<rid>/quote-editable.pdf — AcroForm
#      working copy with buyer/ship-to fields editable in any PDF viewer.
#   2. Operator edits in Adobe Reader / Preview / Chrome's PDF viewer.
#   3. Operator POSTs the edited file to /api/rfq/<rid>/upload-edited-quote.
#      Server flattens via PR-D3, applies diff_to_quote_fields to the RFQ
#      row, audit-logs the diff. Future Mark Sent flow uses the flat copy
#      as the buyer attachment.
#
# Editable fields cover the buyer/ship-to block (PR-D1 scope). Line items,
# quote_number, dates remain flat (counter collision risk + scope deferral).


@bp.route("/api/rfq/<rid>/quote-editable.pdf")
@auth_required
@safe_route
def api_rfq_quote_editable(rid):
    """Generate the AcroForm editable working copy on demand. Returns PDF bytes."""
    from src.api.dashboard import load_rfqs
    from src.core.quote_model import Quote
    from src.forms.quote_generator import generate_quote_pdf

    rfq = load_rfqs().get(rid)
    if not rfq:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404
    try:
        quote = Quote.from_legacy_dict(rfq)
        pdf_bytes = generate_quote_pdf(quote, editable=True)
    except Exception as e:
        log.error("quote-editable.pdf failed for %s: %s", rid, e, exc_info=True)
        return jsonify({"ok": False, "error": f"generation failed: {e}"}), 500
    sol = re.sub(r"[^a-zA-Z0-9_-]", "_", str(rfq.get("solicitation_number") or rid))[:40]
    fname = f"{sol}_Quote_Editable.pdf"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Quote-Mode": "editable",
        },
    )


@bp.route("/api/rfq/<rid>/upload-edited-quote", methods=["POST"])
@auth_required
@safe_route
def api_rfq_upload_edited_quote(rid):
    """Operator uploaded an edited PDF. Read AcroForm fields, sync to RFQ
    row, audit-log the diff, store the flattened bytes as the buyer copy.

    Returns JSON: {ok, applied, diff, flat_pdf_path, edits}

    Race-safe wrapper (PR #778 pattern).
    """
    from src.api.data_layer import _save_rfqs_lock
    with _save_rfqs_lock:
        return _api_rfq_upload_edited_quote_locked(rid)


def _api_rfq_upload_edited_quote_locked(rid):
    """Inner body — always runs under `_save_rfqs_lock`."""
    from src.api.dashboard import load_rfqs, _save_single_rfq
    from src.forms.quote_pdf_flatten import diff_to_quote_fields, flatten_quote_pdf

    rfq = load_rfqs().get(rid)
    if not rfq:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404

    f = request.files.get("pdf")
    if not f:
        return jsonify({"ok": False, "error": "no pdf file in upload (form field 'pdf')"}), 400

    try:
        editable_bytes = f.read()
    except Exception as e:
        return jsonify({"ok": False, "error": f"upload read failed: {e}"}), 400

    if not editable_bytes or not editable_bytes.startswith(b"%PDF-"):
        return jsonify({"ok": False, "error": "upload is not a PDF"}), 400

    flat_bytes, edits = flatten_quote_pdf(editable_bytes)
    if not edits:
        return jsonify({
            "ok": False,
            "error": "no editable AcroForm fields found in upload — was the file generated from /api/rfq/<rid>/quote-editable.pdf?",
        }), 400

    canonical = diff_to_quote_fields(edits)
    if not canonical:
        return jsonify({
            "ok": True,
            "applied": False,
            "message": "AcroForm fields present but all empty — nothing to sync",
            "edits": edits,
        })

    # Capture diff vs current row values for audit
    diff = {}
    for k, new in canonical.items():
        old = rfq.get(k)
        if old != new:
            diff[k] = {"before": old, "after": new}

    # Apply edits to the row
    rfq.update(canonical)
    rfq["last_edited_via_pdf_at"] = datetime.now().isoformat()

    # Persist flat bytes alongside the row so Mark Sent can reuse them.
    flat_path = None
    try:
        out_dir = OUTPUT_DIR if OUTPUT_DIR else DATA_DIR
        os.makedirs(out_dir, exist_ok=True)
        flat_path = os.path.join(out_dir, f"rfq_{rid}_quote_flat.pdf")
        with open(flat_path, "wb") as fh:
            fh.write(flat_bytes)
        rfq["reytech_quote_pdf_flat"] = flat_path
    except Exception as fe:
        log.warning("upload-edited-quote: writing flat copy failed: %s", fe)

    _save_single_rfq(rid, rfq)

    # Audit log — best-effort, never block the apply.
    try:
        from src.core.security import _log_audit_internal
        _log_audit_internal(
            action="quote_pdf_edited_externally",
            details=f"rfq={rid} actor={session.get('user', 'operator')} fields={len(edits)}",
            metadata={
                "target_id": rid,
                "actor": session.get("user", "operator"),
                "diff": diff,
                "flat_pdf_path": flat_path,
                "field_count": len(edits),
            },
        )
    except Exception as ae:
        log.warning("audit log failed (PR-D4 upload-edited): %s", ae)

    return jsonify({
        "ok": True,
        "applied": True,
        "diff": diff,
        "flat_pdf_path": flat_path,
        "edits": edits,
    })


# ═══════════════════════════════════════════════════════════════════════
# PR-C2 — pre-send win-validation soft warnings (UI badges)
# ═══════════════════════════════════════════════════════════════════════

@bp.route("/api/rfq/<rid>/win-warnings")
@auth_required
@safe_route
def api_rfq_win_warnings(rid):
    """Return soft win-risk warnings for the operator before send.

    Pure read endpoint. The work splits into two layers:

      1. *Lookup* — for each line item that has a part_number or
         description, query `_last_won_price_for_buyer` against this RFQ's
         contact_email. The route owns the SQLite connection so we open
         it once for the whole pass instead of N times.
      2. *Compute* — pre-enriched items go to
         `compute_win_warnings()` (pure function, no I/O).

    Returns:
      {ok: true, warnings: [...], counts: {red, orange, yellow}}

    The endpoint never blocks send. Warnings are advisory only; the
    operator is the final gate (per `feedback_ten_minute_escape_valve`).
    """
    from src.api.dashboard import load_rfqs
    from src.core.win_validation import compute_win_warnings
    rfq = load_rfqs().get(rid)
    if not rfq:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404

    items = list(rfq.get("line_items") or rfq.get("items") or [])
    contact_email = (rfq.get("contact_email")
                     or rfq.get("requestor_email") or "").strip()
    quote_number = (rfq.get("reytech_quote_number") or "").strip()

    # ── Layer 1: pre-enrich items with last-won lookups ──
    if contact_email and items:
        try:
            from src.api.modules.routes_growth_intel import (
                _last_won_price_for_buyer,
            )
            with get_db() as conn:
                conn.row_factory = __import__("sqlite3").Row
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    desc = item.get("description") or ""
                    pn = (item.get("part_number")
                          or item.get("item_number") or "")
                    lw = _last_won_price_for_buyer(
                        conn, contact_email, desc, pn,
                        exclude_quote_number=quote_number,
                    )
                    if lw and lw.get("price"):
                        item["last_won_price"] = lw["price"]
                        item["last_won_quote"] = lw.get("quote_number", "")
        except Exception as e:
            # Lookup is best-effort — never fail the warnings call because
            # of a buyer-history hiccup. compute_win_warnings still runs
            # the line-level + quote-level checks without last-won data.
            log.warning("win-warnings last-won enrichment failed for %s: %s",
                        rid, e)

    rfq_for_compute = dict(rfq)
    rfq_for_compute["line_items"] = items

    warnings = compute_win_warnings(rfq_for_compute)

    counts = {"red": 0, "orange": 0, "yellow": 0}
    for w in warnings:
        lvl = w.get("level")
        if lvl in counts:
            counts[lvl] += 1

    return jsonify({
        "ok": True,
        "warnings": warnings,
        "counts": counts,
    })


# ── PR-AV2 (2026-05-14): admin reparse for existing RFQs ──────────
# Closes the AV1 retro-heal gap. The substrate fix in #1004 only
# affects records ingested AFTER the deploy — already-persisted RFQs
# like rfq_efbdef4a (25CB021, 16 items where 9 were form-code rows)
# stay broken. This route mirrors the PC `/pricecheck/<id>/reparse`
# pattern: pulls the buyer's source PDFs from `rfq_files` blobs,
# restores them to a temp dir, and re-runs `process_buyer_request`
# with `existing_record_type="rfq"`. The new ingest_pipeline applies
# the form_code_filter + form_field_extractor to the existing record
# in place. Items drop where form codes were parsed as products; sol#
# / due_date get re-pulled from AcroForm fields.
#
# Safety:
#  - Only re-parses the buyer attachments already on the record (no
#    new upload). Operator-edited pricing on existing items is
#    preserved via the underlying ingest-pipeline's record-update
#    path which is already battle-tested by the PC reparse route.
#  - Falls back to a clean "ok: false" + reason when no buyer
#    attachments exist (e.g. an RFQ created via manual upload only).
#  - 400 when rid doesn't exist or RFQ is in a terminal state
#    (sent/won/lost/dismissed) — don't re-parse shipped quotes.


@bp.route("/api/admin/rfq/<rid>/reparse", methods=["POST"])
@auth_required
@safe_route
def api_admin_rfq_reparse(rid):
    """Re-ingest an existing RFQ from its stored buyer attachments.

    Body: ignored (route takes no params — re-parses every
    buyer_attachment on the record).

    Returns: process_buyer_request().to_dict() plus a `source_files`
    list naming which DB-blob files were fed into the pipeline.

    Skips terminal records (sent/won/lost/dismissed/duplicate/archived)
    — those represent shipped quotes; re-parsing them would mutate
    history.
    """
    from src.api.dashboard import load_rfqs, list_rfq_files, get_rfq_file
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404

    TERMINAL = {"sent", "won", "lost", "dismissed",
                "duplicate", "archived", "expired", "no_response"}
    status = (r.get("status") or "").strip().lower()
    if status in TERMINAL:
        return jsonify({
            "ok": False,
            "error": f"RFQ status is '{status}' — re-parsing terminal "
                     "records is not allowed (would mutate shipped state)",
        }), 400

    # Pull buyer-attachment files. Use the existing helper that already
    # dedups by (filename, category) so re-saved bundles don't double-
    # process. `category="buyer_attachment"` is what ingest_pipeline
    # uses for inbound PDFs.
    buyer_files = list_rfq_files(rid, category="buyer_attachment")
    if not buyer_files:
        # Fallback: try generic 'source' or no-category — some legacy
        # records pre-date the buyer_attachment classification.
        all_files = list_rfq_files(rid)
        buyer_files = [
            f for f in all_files
            if (f.get("category") or "").strip().lower()
               in ("buyer_attachment", "source", "")
        ]
    if not buyer_files:
        return jsonify({
            "ok": False,
            "error": "No buyer attachments on this RFQ — nothing to re-parse",
            "files_listed": 0,
        }), 400

    # Restore each blob to a fresh tmp dir on the data volume so
    # process_buyer_request gets real on-disk paths (its parsers expect
    # file_path, not BLOB bytes).
    import tempfile
    tmp_dir = tempfile.mkdtemp(
        prefix=f"reparse_{rid}_", dir=os.path.join(DATA_DIR, "tmp")
        if os.path.isdir(os.path.join(DATA_DIR, "tmp"))
        else None,
    )
    restored_paths = []
    restored_meta = []
    for f_meta in buyer_files:
        try:
            full = get_rfq_file(f_meta["id"])
            if not full or not full.get("data"):
                log.debug("reparse rfq %s: file %s has no blob data",
                          rid, f_meta.get("id"))
                continue
            safe_name = re.sub(r"[^A-Za-z0-9._\- ]+", "_",
                               full.get("filename") or f"{f_meta['id']}.pdf")
            target = os.path.join(tmp_dir, safe_name)
            with open(target, "wb") as _w:
                _w.write(full["data"])
            restored_paths.append(target)
            restored_meta.append({
                "file_id": f_meta["id"],
                "filename": full.get("filename"),
                "size_bytes": len(full["data"]),
            })
        except Exception as _re:
            log.warning("reparse rfq %s: restore failed for %s: %s",
                        rid, f_meta.get("id"), _re)

    if not restored_paths:
        return jsonify({
            "ok": False,
            "error": "Failed to restore any buyer attachments from DB",
            "files_listed": len(buyer_files),
        }), 500

    log.info(
        "reparse rfq %s: restored %d files from DB → %s",
        rid, len(restored_paths), tmp_dir,
    )

    # Route through the new unified pipeline. Same flag pattern as
    # /pricecheck/<id>/reparse — if classifier_v2 is off (it's been on
    # in prod for months) fall through to a 422 telling the operator
    # to flip the flag first.
    try:
        from src.core.request_classifier import classify_enabled
        if not classify_enabled():
            return jsonify({
                "ok": False,
                "error": "classifier_v2 flag is off — enable "
                         "request.classifier_v2_enabled then retry",
            }), 422

        from src.core.ingest_pipeline import process_buyer_request
        result = process_buyer_request(
            files=restored_paths,
            email_body=r.get("body_text", ""),
            email_subject=r.get("email_subject", ""),
            email_sender=r.get("buyer_email", "") or r.get("requestor_email", ""),
            existing_record_id=rid,
            existing_record_type="rfq",
        )
        return jsonify({
            "ok": result.ok,
            "rfq_id": rid,
            "items_parsed": result.items_parsed,
            "classification": result.classification,
            "ingest_warnings": list(getattr(result, "ingest_warnings", []) or []),
            "needs_review": bool(getattr(result, "needs_review", False)),
            "reasons": list(result.reasons or []),
            "errors": list(result.errors or []),
            "warnings": list(result.warnings or []),
            "source_files": restored_meta,
        })
    except Exception as _e:
        log.error("reparse rfq %s: pipeline failed: %s", rid, _e, exc_info=True)
        return jsonify({
            "ok": False,
            "error": f"reparse failed: {_e}",
            "source_files": restored_meta,
        }), 500


# ── PR-AV12 (AV-12): generated-files ledger drift — reconcile rfq_files
# DB rows with what's actually on disk in the per-RFQ output directory.
# Closes the gap flagged in the 5/14 EOD handoff: package-diag returns
# `db_files` (rows in rfq_files table) and `output_dir_files` (list
# from os.listdir on data/output/<sol>) and these don't agree on
# existing records. Files generated by older code paths landed on
# disk but never got the rfq_files row inserted; consumers that
# enumerate via list_rfq_files (Gmail draft attachments, file
# download routes, audit views) miss them.


def _reconcile_rfq_files_to_disk(rid: str, sol: str = "") -> dict:
    """Walk the per-RFQ output directory, insert rfq_files rows for
    any PDF on disk that's not already in the ledger.

    Returns {ok, added, files_added, skipped, errors}. Idempotent —
    `save_rfq_file` dedups by (rfq_id, filename, category) so
    re-running this is a no-op when the ledger is in sync.

    The `sol` arg is the solicitation_number / pc_number used as the
    output-directory name. When omitted, falls back to scanning by
    rid (legacy records used the record_id as the directory name).
    """
    from src.api.dashboard import save_rfq_file, list_rfq_files
    out: dict = {
        "ok": True,
        "added": 0,
        "files_added": [],
        "skipped": 0,
        "errors": [],
    }

    # Candidate output directories — sol wins, rid is the fallback.
    candidates = []
    if sol:
        candidates.append(os.path.join(DATA_DIR, "output", sol))
    if rid:
        candidates.append(os.path.join(DATA_DIR, "output", rid))

    out_dir = next((c for c in candidates if os.path.isdir(c)), None)
    if not out_dir:
        out["ok"] = False
        out["errors"].append(
            f"No output directory found for rid={rid} sol={sol!r}"
        )
        return out
    out["output_dir"] = out_dir

    # Existing ledger entries — read once. Compare by case-insensitive
    # filename so a disk file with mixed case still matches a ledger
    # row written by an older snake-case path.
    try:
        existing = list_rfq_files(rid) or []
    except Exception as _le:
        existing = []
        out["errors"].append(f"list_rfq_files failed: {_le}")

    existing_names = {
        (e.get("filename") or "").strip().lower()
        for e in existing
        if e.get("filename")
    }

    # Walk disk. Only sync .pdf files — that's what the ledger tracks.
    # Other files (manifest.json, audit logs) live alongside but aren't
    # consumer-facing attachments.
    try:
        disk_files = sorted(os.listdir(out_dir))
    except Exception as _lse:
        out["ok"] = False
        out["errors"].append(f"listdir({out_dir}) failed: {_lse}")
        return out

    for fname in disk_files:
        if not fname.lower().endswith(".pdf"):
            continue
        if fname.lower() in existing_names:
            out["skipped"] += 1
            continue
        fpath = os.path.join(out_dir, fname)
        try:
            with open(fpath, "rb") as fh:
                data = fh.read()
            if not data:
                out["errors"].append(f"empty file: {fname}")
                continue
            # Category heuristic: generated package PDFs land in
            # "generated"; everything else falls into "template".
            # The downstream rendering doesn't enforce a particular
            # taxonomy here — we just need a row to exist.
            category = "generated"
            save_rfq_file(
                rfq_id=rid,
                filename=fname,
                file_type="pdf",
                data=data,
                category=category,
                uploaded_by="reconcile",
            )
            out["added"] += 1
            out["files_added"].append(fname)
        except Exception as _se:
            out["errors"].append(
                f"save failed for {fname}: {_se}"
            )

    return out


@bp.route("/api/admin/rfq/<rid>/reconcile-files", methods=["POST"])
@auth_required
@safe_route
def api_admin_rfq_reconcile_files(rid):
    """Heal `rfq_files` ledger drift — insert rows for any PDF on
    disk that doesn't have a ledger entry. Idempotent; safe to call
    multiple times.

    Returns: {ok, rfq_id, output_dir, added, files_added, skipped,
              errors}
    """
    from src.api.dashboard import load_rfqs
    rfqs = load_rfqs()
    r = rfqs.get(rid)
    if not r:
        return jsonify({"ok": False, "error": "RFQ not found"}), 404

    sol = (r.get("solicitation_number") or r.get("rfq_number") or "").strip()
    result = _reconcile_rfq_files_to_disk(rid, sol)
    result["rfq_id"] = rid
    status_code = 200 if result.get("ok") else 500
    return jsonify(result), status_code


# ── PR-AL (2026-05-14): backfill — retroactively apply PR-AI + PR-AJ
# enrichment to existing non-terminal PCs and RFQs ──────────────────


@bp.route("/api/admin/heal-ingest-enrichment", methods=["POST"])
@auth_required
@safe_route
def api_heal_ingest_enrichment():
    """One-shot backfill that applies the same auto-tax + auto-price
    enrichment logic from `_create_record` (PR-AI #990 + PR-AJ #991)
    to every existing non-terminal PC and RFQ. Heals the operator
    queue's ⚠ DEFAULT records WITHOUT waiting for the next inbound.

    Body: `{"dry_run": true}` (default false) returns what WOULD be
    enriched without mutating. Pass `{"dry_run": false}` to actually
    write.

    Safety: the underlying enrichment is idempotent by design —
      - PR-AI tax-resolve only stamps when current tax_rate is zero.
      - PR-AJ Oracle reference fields only fill BLANK fields, never
        overwrite. Items with unit_cost OR supplier_cost set are
        skipped entirely.
    Re-running this endpoint multiple times converges to the same
    state; no risk of clobbering operator-typed cost/markup edits.

    Returns: {records_processed, tax_resolved, items_enriched,
              dry_run, summary[rid → counts]}
    """
    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get("dry_run", True))  # default safe: dry-run

    # Terminal statuses we skip — those records are done and shouldn't
    # see enrichment writes (operator already shipped or dismissed).
    TERMINAL = {"sent", "won", "lost", "expired", "no_response",
                "dismissed", "archived", "duplicate"}

    from src.api.dashboard import (
        _load_price_checks, _save_single_pc,
        load_rfqs, _save_single_rfq,
    )
    from src.core.quote_contract import tax_for_address
    try:
        from src.core.pricing_oracle_v2 import recommend_for_item as _rec
    except Exception:
        _rec = None

    _now = datetime.now().isoformat() if 'datetime' in dir() else ""
    if not _now:
        from datetime import datetime as _dt2
        _now = _dt2.now().isoformat()

    def _heal_record(rid, r):
        """Returns (tax_resolved_bool, items_enriched_count, ship_to_filled_bool)."""
        tax_resolved = False
        items_enriched = 0
        ship_to_filled = False

        # Skip terminal records.
        _status = (r.get("status") or "").strip().lower()
        if _status in TERMINAL:
            return (False, 0, False)

        # ── PR-AM (2026-05-14): canonical ship_to fallback ──
        # When the existing ship_to is missing or too short to parse
        # (the 4 ⚠ DEFAULT records Mike saw on /home all had
        # ship_to="CA" — 2 chars, tax_for_address skips), fall back
        # to facility_registry.resolve(institution) → canonical
        # address. Same logic PR-AI uses in _create_record at
        # ingest. Skips if no institution OR registry can't resolve
        # (agency codes like "cchcs"/"calvet" are agencies, not
        # facilities — those records genuinely need operator-typed
        # ship_to before tax can resolve).
        _ship_to = (r.get("ship_to") or "").strip()
        if not _ship_to or len(_ship_to) <= 3:
            try:
                from src.core.facility_registry import resolve as _resolve_facility
                _raw_inst = (
                    r.get("institution", "")
                    or r.get("agency", "")
                    or ""
                ).strip()
                if _raw_inst:
                    _fac = _resolve_facility(_raw_inst)
                    if _fac:
                        _canonical_ship_to = f"{_fac.address_line1}, {_fac.address_line2}"
                        if _canonical_ship_to.strip(", ") and len(_canonical_ship_to) > 3:
                            if not dry_run:
                                r["ship_to"] = _canonical_ship_to
                                # Pin canonical institution code too —
                                # converts free-form "CSP-Sacramento" /
                                # variants to the registry's stable code.
                                r["institution"] = _fac.code
                            _ship_to = _canonical_ship_to
                            ship_to_filled = True
                            log.info("heal: ship_to filled from facility_registry for %s: %s → %r",
                                     rid, _raw_inst, _canonical_ship_to[:60])
            except Exception as _fe:
                log.debug("heal: facility-registry fallback %s failed: %s", rid, _fe)

        # ── Auto-tax (PR-AI logic + PR-AN re-resolve fallback) ──
        # PR-AN (2026-05-14): trigger tax-resolve not only when rate is
        # zero, but ALSO when the stored tax_source is in the fallback
        # set. The 4 ⚠ DEFAULT records on Mike's /home queue have
        # tax_rate set to the CA statewide default (7.25 / fallback
        # rendered during initial detail-page load) with
        # tax_source="default" — pre-v3 heal correctly skipped them
        # ("tax_rate already > 0") but the visible ⚠ DEFAULT badge
        # never cleared. Now: a non-empty CDTFA-grade ship_to + a
        # current tax_source in the fallback set → re-resolve. Only
        # OVERWRITE the stored values when the new resolution returns
        # validated=True (we never DOWNGRADE a confirmed result to a
        # weaker fallback). tax_source values considered "fallback":
        # default, fallback_table, fallback, "". An operator-typed
        # "manual_operator" / "cdtfa_api" tax_source is left alone.
        _cur_rate = float(r.get("tax_rate") or 0)
        _cur_source = (r.get("tax_source") or "").strip().lower()
        FALLBACK_SOURCES = {"", "default", "fallback_table", "fallback", "ca default"}
        _needs_reresolve = (_cur_rate <= 0) or (_cur_source in FALLBACK_SOURCES)
        if _needs_reresolve:
            if _ship_to and len(_ship_to) > 3:
                try:
                    _tax = tax_for_address(_ship_to) or {}
                    _rate = float(_tax.get("rate") or 0.0)
                    _validated = bool(_tax.get("validated", False))
                    # Only upgrade — if the new result isn't validated,
                    # leave the existing fallback in place (don't make
                    # the operator surface noisier than it already is).
                    if _rate > 0 and _validated:
                        if not dry_run:
                            r["tax_rate"] = round(_rate * 100, 3)
                            r["tax_source"] = str(_tax.get("source") or "")
                            r["tax_jurisdiction"] = str(_tax.get("jurisdiction") or "")
                            r["tax_validated"] = True
                        tax_resolved = True
                except Exception as _te:
                    log.debug("heal: tax-resolve %s failed: %s", rid, _te)

        # ── Auto-price (PR-AJ logic) ──
        # Items key differs: PC=items, RFQ=line_items. Both possible.
        _items = r.get("items") or r.get("line_items") or []
        if _rec is not None and isinstance(_items, list):
            for _it in _items:
                if not isinstance(_it, dict):
                    continue
                _desc = (_it.get("description") or "").strip()
                if not _desc:
                    continue
                # Skip operator-confirmed cost (sacred per PR-AC).
                if _it.get("unit_cost") or _it.get("supplier_cost"):
                    continue
                # PR-AN short-circuit: skip items already Oracle-touched.
                # Oracle returns non-deterministic sparse data per call —
                # pre-fix the counter incremented on every run as each
                # call filled different blank fields (Mike's diagnostic:
                # 22 items_enriched on 3 consecutive calls even though
                # all 22 already had auto_priced_at_ingest=True visible
                # on the detail page). Now: one heal pass per item;
                # subsequent runs converge to 0 unless a NEW item lands
                # without the flag (re-ingest or operator-pasted row).
                if _it.get("auto_priced_at_ingest"):
                    continue
                try:
                    _r = _rec(
                        description=_desc,
                        part_number=str(_it.get("part_number") or _it.get("item_number") or _it.get("mfg_number") or ""),
                        qty=float(_it.get("quantity") or _it.get("qty") or 1) or 1,
                        upc=str(_it.get("upc") or ""),
                    )
                except Exception as _ie:
                    log.debug("heal: per-item rec %s/%r failed: %s", rid, _desc[:40], _ie)
                    continue
                if not _r:
                    continue
                _stamped = False
                if not _it.get("catalog_cost") and _r.get("catalog_cost"):
                    if not dry_run:
                        _it["catalog_cost"] = _r["catalog_cost"]
                    _stamped = True
                if not _it.get("supplier") and _r.get("supplier"):
                    if not dry_run:
                        _it["supplier"] = _r["supplier"]
                    _stamped = True
                if not _it.get("source_url") and _r.get("source_url"):
                    if not dry_run:
                        _it["source_url"] = _r["source_url"]
                    _stamped = True
                if not _it.get("asin") and _r.get("asin"):
                    if not dry_run:
                        _it["asin"] = _r["asin"]
                    _stamped = True
                if _it.get("confidence") in (None, 0, 0.0) and _r.get("confidence") is not None:
                    if not dry_run:
                        _it["confidence"] = float(_r["confidence"])
                    _stamped = True
                if _stamped:
                    if not dry_run:
                        _it["auto_priced_at_ingest"] = True
                        _it["auto_price_at"] = _now
                    items_enriched += 1
        # PR-AL hotfix (2026-05-14): force alias-parity BEFORE save.
        # data_layer._save_single_rfq:315-319 runs `r["items"] =
        # list(r["line_items"])` (and vice versa) on every save to
        # keep the SQL `items` column in lockstep with the data_json
        # blob. Pre-fix, on a record where r["items"] and
        # r["line_items"] came back from json deserialization as
        # SEPARATE lists with same content (typical of round-tripped
        # records), our heal mutated `_items` (whichever the `or`
        # picked) but the save-time alias-sync then OVERWROTE r["items"]
        # with the unmutated r["line_items"]. Result: 22 items reported
        # enriched per run but ZERO persisted across runs (Mike's
        # diagnostic — 3 identical heal responses in a row + no 🔮
        # badges visible on rfq_4a723a40). The fix: explicitly pin
        # both keys to the SAME mutated list so the alias-sync
        # becomes a list(_items) shallow-copy of THE mutated list
        # rather than the unmutated peer.
        if not dry_run and items_enriched > 0 and isinstance(_items, list):
            if "items" in r:
                r["items"] = _items
            if "line_items" in r:
                r["line_items"] = _items
        return (tax_resolved, items_enriched, ship_to_filled)

    summary = {}
    total_tax = 0
    total_items = 0
    total_ship_to = 0

    # PCs
    pcs = _load_price_checks() or {}
    for pcid, pc in list(pcs.items()):
        tax_r, items_e, ship_f = _heal_record(pcid, pc)
        if tax_r or items_e or ship_f:
            summary[pcid] = {"type": "pc", "tax_resolved": tax_r,
                             "items_enriched": items_e,
                             "ship_to_filled": ship_f}
            if tax_r:
                total_tax += 1
            if ship_f:
                total_ship_to += 1
            total_items += items_e
            if not dry_run:
                _save_single_pc(pcid, pc)

    # RFQs
    rfqs = load_rfqs() or {}
    for rid, r in list(rfqs.items()):
        tax_r, items_e, ship_f = _heal_record(rid, r)
        if tax_r or items_e or ship_f:
            summary[rid] = {"type": "rfq", "tax_resolved": tax_r,
                            "items_enriched": items_e,
                            "ship_to_filled": ship_f}
            if tax_r:
                total_tax += 1
            if ship_f:
                total_ship_to += 1
            total_items += items_e
            if not dry_run:
                _save_single_rfq(rid, r)

    log.info(
        "heal-ingest-enrichment: dry_run=%s records=%d tax=%d items=%d",
        dry_run, len(summary), total_tax, total_items,
    )
    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "records_processed": len(summary),
        "ship_to_filled": total_ship_to,
        "tax_resolved": total_tax,
        "items_enriched": total_items,
        "summary": summary,
    })


@bp.route("/api/admin/operator-drift-stats", methods=["GET"])
@auth_required
@safe_route
def api_operator_drift_stats():
    """PR-AP diagnostic — `operator_drift_line` table state at a glance.

    The motivating signal: /admin/auto-recommendations (PR-S) renders
    "No operator_drift_line rows in last 7d — was Mark-Sent used?"
    even though three Mark-Sent variants (PC mark-sent, PC mark-sent-
    manually, RFQ mark-sent-manually) all wire drift logging. The
    empty state could mean:

      A. Mark-Sent isn't being clicked → operator-funnel gap, fix
         with a UI nudge.
      B. Drift logging fires but silently fails (lookup_failed, audit
         missing, etc.) → fix the logging path.
      C. Rows exist but the 7d read-window query is buggy → fix the
         aggregation.

    Without prod-state visibility, the fix is a coin flip. This
    endpoint reads the table directly and reports:
      - total rows
      - rows in the last 7 / 30 / 90 day windows
      - most recent N sent_at timestamps + agency_key + quote_id
      - per-quote_type counts (pc vs rfq)
      - distinct agency_keys present
      - distribution of drift_pct (median + p25/p75)

    GET only — read-only. Audit lives on the /admin/auto-recommendations
    page anyway; this is the structured backend it exposes.

    Returns: {ok, total, by_window[7d,30d,90d], by_quote_type,
              recent[ {quote_id,quote_type,sent_at,agency_key} ],
              agencies, drift_pct_stats}
    """
    from src.core.operator_kpi import get_drift_diagnostic
    result = get_drift_diagnostic()
    if not result.get("ok"):
        return jsonify(result), 500
    return jsonify(result)


@bp.route("/api/admin/heal-drift-backfill", methods=["POST"])
@auth_required
@safe_route
def api_heal_drift_backfill():
    """PR-AR — retroactive drift logging for already-sent records.

    PR-AQ confirmed 0/235 items in 35 sent records carry oracle_audit.
    PR-AR's synthesizer makes log_operator_drift produce rows from the
    Oracle-suggested prices already on each item, but those past
    Mark-Sent events fired before the synthesizer existed — so prod
    operator_drift_line is still empty.

    This route walks every active PC/RFQ with status="sent" and fires
    `log_operator_drift` (which now includes the synthesizer) for each.
    Already-logged rows from earlier passes are NOT deduped — the
    drift table allows multiple rows per (quote_id, line_idx) and the
    aggregator reads windowed snapshots. If you re-run this route, it
    will insert another set of rows; in practice this should run
    exactly once.

    Body: {"dry_run": true} (default true) reports what WOULD log
    without inserting. {"dry_run": false} actually writes.

    Returns: {records_scanned, records_logged, rows_inserted,
              synthesized_audits_total, dry_run,
              summary[rid → {rows, synthesized}]}
    """
    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get("dry_run", True))

    TERMINAL_NON_SENT = {"won", "lost", "expired", "no_response",
                         "dismissed", "archived", "duplicate"}

    from src.api.dashboard import _load_price_checks, load_rfqs
    from src.core.operator_kpi import log_operator_drift

    summary = {}
    rows_inserted_total = 0
    synthesized_total = 0
    records_scanned = 0
    records_logged = 0

    def _process(rid: str, r: dict, kind: str):
        nonlocal rows_inserted_total, synthesized_total
        nonlocal records_scanned, records_logged
        if not isinstance(r, dict):
            return
        if (r.get("status") or "").strip().lower() != "sent":
            return
        records_scanned += 1
        items = r.get("items") or r.get("line_items") or []
        if not items:
            return
        agency_key = (
            r.get("agency_key") or r.get("agency")
            or r.get("institution") or ""
        )
        if dry_run:
            # Simulate by walking the items with the synthesizer to
            # count what WOULD log — without inserting.
            from src.core.operator_kpi import _synthesize_oracle_audit
            from datetime import datetime as _dt
            now_iso = _dt.now().isoformat()
            simulated = 0
            synth = 0
            for it in items:
                if not isinstance(it, dict):
                    continue
                audit = it.get("oracle_audit") or {}
                if not (isinstance(audit, dict) and audit):
                    audit = _synthesize_oracle_audit(it, now_iso) or {}
                    if audit:
                        synth += 1
                    else:
                        continue
                # Has audit → check price gate too
                for key in ("unit_price", "bid_price", "price_per_unit"):
                    v = it.get(key)
                    try:
                        if v is not None and float(v) > 0:
                            simulated += 1
                            break
                    except (TypeError, ValueError):
                        continue
            if simulated > 0:
                records_logged += 1
                rows_inserted_total += simulated
                synthesized_total += synth
                summary[rid] = {
                    "type": kind, "rows": simulated, "synthesized": synth,
                }
            return
        # Real run: fire log_operator_drift.
        qn = (r.get("quote_number") or r.get("reytech_quote_number") or "")
        try:
            result = log_operator_drift(
                quote_id=rid, quote_type=kind,
                items=items, agency_key=agency_key,
                quote_number=qn,
            )
        except Exception as e:
            log.debug("heal-drift-backfill %s: %s", rid, e)
            return
        rows = int(result.get("rows_logged") or 0)
        synth = int(result.get("synthesized_audits") or 0)
        if rows > 0:
            records_logged += 1
            rows_inserted_total += rows
            synthesized_total += synth
            summary[rid] = {
                "type": kind, "rows": rows, "synthesized": synth,
            }

    pcs = _load_price_checks() or {}
    for pcid, pc in list(pcs.items()):
        _process(pcid, pc, "pc")
    rfqs = load_rfqs() or {}
    for rid, r in list(rfqs.items()):
        _process(rid, r, "rfq")

    log.info(
        "heal-drift-backfill: dry_run=%s scanned=%d logged=%d "
        "rows=%d synthesized=%d",
        dry_run, records_scanned, records_logged,
        rows_inserted_total, synthesized_total,
    )
    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "records_scanned": records_scanned,
        "records_logged": records_logged,
        "rows_inserted": rows_inserted_total,
        "synthesized_audits_total": synthesized_total,
        "summary": summary,
    })


@bp.route("/api/admin/operator-drift-audit-coverage", methods=["GET"])
@auth_required
@safe_route
def api_operator_drift_audit_coverage():
    """PR-AQ — explain WHY operator_drift_line is empty.

    PR-AP showed 0 rows table-wide. /admin/funnel shows sent PCs in the
    30-day window. Drift logging fires on every Mark-Sent but produces
    zero rows. `log_operator_drift` skips lines lacking oracle_audit OR
    lacking a positive unit_price/bid_price/price_per_unit.

    This endpoint walks every active PC/RFQ with status="sent" and
    reports per-record + aggregate coverage so the substrate fix
    targets the right gate.

    GET only — read-only.
    """
    from src.core.operator_kpi import get_drift_audit_coverage
    result = get_drift_audit_coverage()
    if not result.get("ok"):
        return jsonify(result), 500
    return jsonify(result)


@bp.route("/api/admin/heal-due-dates", methods=["POST"])
@auth_required
@safe_route
def api_heal_due_dates():
    """PR-AO retroactive backfill — upgrade `due_date_source == "default"`
    records by scanning their stored buyer-attachment PDFs for a parsable
    deadline.

    For every active (non-terminal) PC and RFQ whose due_date_source is
    `default`, this:
      1. Lists rfq_files for category="buyer_attachment" (PR-A persistence)
      2. Writes each PDF BLOB to a temp file
      3. Runs `extract_deadline_from_pdf` (pdfplumber text + regex)
      4. On first hit: stamps due_date, due_time, due_date_source=
         "attachment", due_date_attachment=filename + saves

    Body: {"dry_run": true} (default true) returns what WOULD upgrade
    without mutating. {"dry_run": false} actually writes.

    Safety: idempotent — re-running on an already-upgraded record is a
    no-op because `apply_attachment_if_default` short-circuits when
    source != "default". Records whose attachments don't yield a
    deadline are left untouched (anchor never drifts).

    Returns: {records_upgraded, records_scanned, dry_run, summary[rid → {date, attachment}]}
    """
    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get("dry_run", True))

    TERMINAL = {"sent", "won", "lost", "expired", "no_response",
                "dismissed", "archived", "duplicate"}

    from src.api.dashboard import (
        _load_price_checks, _save_single_pc,
        load_rfqs, _save_single_rfq,
        list_rfq_files,
    )
    from src.core.attachment_deadline import extract_deadline_from_pdf

    import tempfile
    import os as _os

    def _scan_one(rid: str, r: dict) -> dict:
        """Returns {'upgraded': bool, 'date': str|None, 'attachment': str|None}."""
        _status = (r.get("status") or "").strip().lower()
        if _status in TERMINAL:
            return {"upgraded": False, "date": None, "attachment": None}
        if (r.get("due_date_source") or "").lower() != "default":
            return {"upgraded": False, "date": None, "attachment": None}

        # PR-A buyer attachments live in rfq_files under
        # category="buyer_attachment". The legacy category="source"
        # is also valid for older records.
        files = list_rfq_files(rid, category="buyer_attachment") or []
        if not files:
            files = list_rfq_files(rid, category="source") or []
        if not files:
            return {"upgraded": False, "date": None, "attachment": None}

        # Iterate files; first PDF that yields a deadline wins.
        for f in files:
            fname = f.get("filename") or ""
            if not fname.lower().endswith(".pdf"):
                continue
            file_id = f.get("id")
            if not file_id:
                continue
            # Pull BLOB
            try:
                from src.core.db import get_db
                with get_db() as conn:
                    row = conn.execute(
                        "SELECT data FROM rfq_files WHERE id = ?", (file_id,)
                    ).fetchone()
            except Exception as _de:
                log.debug("heal-due-dates: BLOB read %s/%s failed: %s",
                          rid, file_id, _de)
                continue
            if not row or not row["data"]:
                continue
            # Write to temp file for pdfplumber
            try:
                with tempfile.NamedTemporaryFile(
                    suffix=".pdf", delete=False
                ) as tf:
                    tf.write(row["data"])
                    tmp_path = tf.name
            except Exception as _we:
                log.debug("heal-due-dates: tempfile %s failed: %s", rid, _we)
                continue
            try:
                date_iso, time_str = extract_deadline_from_pdf(tmp_path)
            except Exception as _ee:
                log.debug("heal-due-dates: extract %s failed: %s", rid, _ee)
                date_iso, time_str = None, None
            finally:
                try:
                    _os.unlink(tmp_path)
                except Exception:
                    pass
            if date_iso:
                if not dry_run:
                    r["due_date"] = date_iso
                    if time_str:
                        r["due_time"] = time_str
                    r["due_date_source"] = "attachment"
                    r["due_date_attachment"] = fname
                return {"upgraded": True, "date": date_iso, "attachment": fname}

        return {"upgraded": False, "date": None, "attachment": None}

    summary = {}
    scanned = 0
    upgraded_count = 0

    # PCs
    pcs = _load_price_checks() or {}
    for pcid, pc in list(pcs.items()):
        _status = (pc.get("status") or "").strip().lower()
        if _status in TERMINAL:
            continue
        if (pc.get("due_date_source") or "").lower() != "default":
            continue
        scanned += 1
        res = _scan_one(pcid, pc)
        if res["upgraded"]:
            summary[pcid] = {
                "type": "pc",
                "date": res["date"],
                "attachment": res["attachment"],
            }
            upgraded_count += 1
            if not dry_run:
                _save_single_pc(pcid, pc)

    # RFQs
    rfqs = load_rfqs() or {}
    for rid, r in list(rfqs.items()):
        _status = (r.get("status") or "").strip().lower()
        if _status in TERMINAL:
            continue
        if (r.get("due_date_source") or "").lower() != "default":
            continue
        scanned += 1
        res = _scan_one(rid, r)
        if res["upgraded"]:
            summary[rid] = {
                "type": "rfq",
                "date": res["date"],
                "attachment": res["attachment"],
            }
            upgraded_count += 1
            if not dry_run:
                _save_single_rfq(rid, r)

    log.info(
        "heal-due-dates: dry_run=%s scanned=%d upgraded=%d",
        dry_run, scanned, upgraded_count,
    )
    return jsonify({
        "ok": True,
        "dry_run": dry_run,
        "records_scanned": scanned,
        "records_upgraded": upgraded_count,
        "summary": summary,
    })
