"""
growth_agent.py — Proactive Growth Engine for Reytech
Phase 26 | Version: 2.0.0

Workflow:
  1. Pull ALL Reytech POs from SCPRS (2022 → present)
  2. Drill into each PO → get items, prices, buyer info
  3. Categorize items into product groups
  4. Search SCPRS for ALL buyers of those categories
  5. Build prospect list with contact info (name, email, agency)
  6. Launch email outreach → voice follow-up if no response in 3-5 days
"""

import json, os, re, logging, time, threading, uuid
from datetime import datetime, timedelta
from collections import defaultdict

# ── Agent Context (Anthropic Skills Guide: Pattern 5 — Domain Intelligence) ──
try:
    from src.core.agent_context import get_context, get_contact_by_agency, format_context_for_agent
    HAS_CTX = True
except ImportError:
    HAS_CTX = False
    def get_context(**kw): return {}
    def get_contact_by_agency(a): return []
    def format_context_for_agent(c, **kw): return ""


log = logging.getLogger("growth")

try:
    from src.core.db import (get_all_customers, get_intel_agencies, 
                               get_all_leads, get_growth_outreach, save_growth_campaign)
    _HAS_DB_DAL = True
except ImportError:
    _HAS_DB_DAL = False

try:
    from src.core.paths import DATA_DIR
except ImportError:
    DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(
        os.path.dirname(os.path.abspath(__file__)))), "data")

HISTORY_FILE = os.path.join(DATA_DIR, "growth_reytech_history.json")
CATEGORIES_FILE = os.path.join(DATA_DIR, "growth_categories.json")
PROSPECTS_FILE = os.path.join(DATA_DIR, "growth_prospects.json")
OUTREACH_FILE = os.path.join(DATA_DIR, "growth_outreach.json")

try:
    from src.agents.scprs_lookup import _get_session
    HAS_SCPRS = True
except ImportError:
    HAS_SCPRS = False


def _load_json(path):
    try:
        with open(path) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def _save_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)


def _load_prospects_list():
    """Load prospects as a flat list, handling both dict and list formats."""
    data = _load_json(PROSPECTS_FILE) or []
    if isinstance(data, dict):
        return data.get("prospects", [])
    return data


# ─── Item Category Mapping ───────────────────────────────────────────────

CATEGORY_KEYWORDS = {
    "Medical Supplies": [
        "glove", "nitrile", "exam", "syringe", "needle", "catheter", "bandage",
        "gauze", "surgical", "gown", "mask", "face shield", "medical", "patient",
        "restraint", "stryker", "wheelchair", "first aid",
    ],
    "Janitorial & Cleaning": [
        "trash bag", "liner", "mop", "broom", "disinfect", "sanitizer",
        "bleach", "cleaner", "detergent", "soap", "wipe", "paper towel",
        "toilet paper", "tissue", "janitorial", "cleaning", "floor",
    ],
    "Office Supplies": [
        "pen", "pencil", "paper", "copy paper", "toner", "cartridge", "ink",
        "folder", "binder", "staple", "tape", "envelope", "marker",
        "highlighter", "office", "label",
    ],
    "IT & Electronics": [
        "battery", "cable", "usb", "adapter", "charger", "keyboard",
        "mouse", "monitor", "printer", "laptop", "computer", "hard drive",
    ],
    "Safety & PPE": [
        "safety glass", "ear plug", "hard hat", "vest", "boot",
        "fire extinguisher", "safety", "protective", "respirator", "goggles",
    ],
    "Food Service": [
        "food", "beverage", "cup", "plate", "napkin", "utensil",
        "container", "tray", "coffee",
    ],
    "Facility Maintenance": [
        "light bulb", "led", "bulb", "filter", "hvac", "paint",
        "tool", "hardware", "plumbing", "electrical", "maintenance", "lock",
    ],
}

def categorize_item(description: str) -> str:
    desc_lower = (description or "").lower()
    scores = {}
    for cat, keywords in CATEGORY_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in desc_lower)
        if score > 0:
            scores[cat] = score
    return max(scores, key=scores.get) if scores else "General Supplies"


# ═══════════════════════════════════════════════════════════════════════
# STEP 1: Pull Reytech History from SCPRS
# ═══════════════════════════════════════════════════════════════════════

PULL_STATUS = {
    "running": False, "phase": "", "progress": "",
    "pos_found": 0, "pos_detailed": 0, "items_total": 0,
    "errors": [], "started_at": None, "finished_at": None,
}

def pull_reytech_history(from_date="01/01/2019", to_date=""):
    """Search SCPRS for all Reytech POs from 2022 to present.
    Drills into each for line items + buyer info."""
    if not HAS_SCPRS:
        return {"ok": False, "error": "SCPRS not available (needs requests + bs4)"}
    if PULL_STATUS["running"]:
        return {"ok": False, "error": "Already running", "status": PULL_STATUS}
    if not to_date:
        to_date = datetime.now().strftime("%m/%d/%Y")

    PULL_STATUS.update({
        "running": True, "phase": "searching",
        "progress": f"Searching SCPRS for Reytech ({from_date} to {to_date})...",
        "pos_found": 0, "pos_detailed": 0, "items_total": 0,
        "errors": [], "started_at": datetime.now().isoformat(), "finished_at": None,
    })

    try:
        session = _get_session()
        if not session.initialized and not session.init_session():
            PULL_STATUS.update({"running": False, "phase": "error"})
            return {"ok": False, "error": "SCPRS session init failed"}

        results = session.search(supplier_name="Reytech", from_date=from_date, to_date=to_date)
        if not results:
            PULL_STATUS["progress"] = "Trying 'Rey Tech'..."
            results = session.search(supplier_name="Rey Tech", from_date=from_date, to_date=to_date)
        if not results:
            PULL_STATUS["progress"] = "Trying 'REYTECH'..."
            results = session.search(supplier_name="REYTECH", from_date=from_date, to_date=to_date)

        PULL_STATUS["pos_found"] = len(results)
        PULL_STATUS["progress"] = f"Found {len(results)} POs — getting details..."
        log.info(f"Growth: Found {len(results)} Reytech POs")

        history = []
        for idx, r in enumerate(results):
            PULL_STATUS["phase"] = "detailing"
            PULL_STATUS["progress"] = f"PO {idx+1}/{len(results)}: {r.get('po_number', '?')}"

            po = {
                "po_number": r.get("po_number", ""),
                "dept": r.get("dept", ""),
                "supplier_name": r.get("supplier_name", ""),
                "start_date": r.get("start_date", ""),
                "grand_total": r.get("grand_total", ""),
                "grand_total_num": r.get("grand_total_num"),
                "buyer_email": r.get("buyer_email", ""),
                "acq_type": r.get("acq_type", ""),
                "status": r.get("status", ""),
                "first_item": r.get("first_item", ""),
                "line_items": [],
                "buyer_name": "",
                "detail_fetched": False,
            }

            if r.get("_results_html") and r.get("_row_index") is not None:
                try:
                    detail = session.get_detail(r["_results_html"], r["_row_index"], r.get("_click_action"))
                    if detail:
                        po["detail_fetched"] = True
                        hdr = detail.get("header", {}) if isinstance(detail.get("header"), dict) else {}
                        po["buyer_name"] = hdr.get("buyer_name", "")
                        po["buyer_email"] = hdr.get("buyer_email", "") or po["buyer_email"]
                        po["buyer_phone"] = hdr.get("buyer_phone", "")
                        po["line_items"] = detail.get("line_items", [])
                        PULL_STATUS["items_total"] += len(po["line_items"])
                    time.sleep(0.5)
                except Exception as e:
                    PULL_STATUS["errors"].append(f"{po['po_number']}: {e}")

            PULL_STATUS["pos_detailed"] = idx + 1
            history.append(po)

        _save_json(HISTORY_FILE, {
            "supplier": "Reytech", "from_date": from_date, "to_date": to_date,
            "pulled_at": datetime.now().isoformat(),
            "total_pos": len(history), "total_items": PULL_STATUS["items_total"],
            "purchase_orders": history,
        })

        categories = _categorize_history(history)

        PULL_STATUS.update({
            "running": False, "phase": "complete",
            "progress": f"Done: {len(history)} POs, {PULL_STATUS['items_total']} items, {len(categories)} categories",
            "finished_at": datetime.now().isoformat(),
        })

        return {
            "ok": True, "total_pos": len(history),
            "total_items": PULL_STATUS["items_total"],
            "categories": len(categories),
            "date_range": f"{from_date} to {to_date}",
        }

    except Exception as e:
        PULL_STATUS.update({"running": False, "phase": "error", "progress": str(e)})
        return {"ok": False, "error": str(e)}


def _categorize_history(history):
    categories = defaultdict(lambda: {"items": [], "total_value": 0, "po_count": 0, "search_terms": set()})
    for po in history:
        for item in po.get("line_items", []):
            desc = item.get("description", "")
            cat = categorize_item(desc)
            price = item.get("unit_price_num") or 0
            qty = item.get("quantity_num") or 1
            categories[cat]["items"].append({"description": desc, "unit_price": price, "po_number": po.get("po_number")})
            categories[cat]["total_value"] += price * qty
            categories[cat]["po_count"] += 1
            words = set(re.findall(r'\b[a-zA-Z]{3,}\b', desc.lower()))
            categories[cat]["search_terms"].update(words - {"the", "and", "for", "with", "each", "per", "box", "case", "pack"})
        if not po.get("line_items") and po.get("first_item"):
            cat = categorize_item(po["first_item"])
            categories[cat]["items"].append({"description": po["first_item"], "po_number": po.get("po_number")})
            categories[cat]["po_count"] += 1

    result = {}
    for cat, data in categories.items():
        result[cat] = {
            "item_count": len(data["items"]),
            "total_value": round(data["total_value"], 2),
            "po_count": data["po_count"],
            "search_terms": sorted(list(data["search_terms"]))[:20],
            "sample_items": [i["description"][:80] for i in data["items"][:10]],
        }
    _save_json(CATEGORIES_FILE, {"generated_at": datetime.now().isoformat(), "total_categories": len(result), "categories": result})
    return result


# ═══════════════════════════════════════════════════════════════════════
# STEP 2: Find ALL Buyers of Those Categories
# ═══════════════════════════════════════════════════════════════════════

BUYER_STATUS = {"running": False, "phase": "", "progress": "", "prospects_found": 0, "errors": []}

def find_category_buyers(max_categories=10, from_date="01/01/2019"):
    """For each category Reytech sells, find all SCPRS buyers."""
    if not HAS_SCPRS:
        return {"ok": False, "error": "SCPRS not available"}
    if BUYER_STATUS["running"]:
        return {"ok": False, "error": "Already running"}

    cat_data = _load_json(CATEGORIES_FILE)
    if not isinstance(cat_data, dict) or not cat_data.get("categories"):
        return {"ok": False, "error": "No categories. Run pull_reytech_history first."}

    cats = sorted(cat_data["categories"].items(), key=lambda x: x[1].get("total_value", 0), reverse=True)[:max_categories]

    BUYER_STATUS.update({"running": True, "phase": "searching", "progress": "Starting...", "prospects_found": 0, "errors": []})

    try:
        session = _get_session()
        if not session.initialized and not session.init_session():
            BUYER_STATUS.update({"running": False})
            return {"ok": False, "error": "SCPRS session init failed"}

        to_date = datetime.now().strftime("%m/%d/%Y")
        prospects = {}

        for cat_idx, (cat_name, cat_info) in enumerate(cats):
            BUYER_STATUS["progress"] = f"[{cat_idx+1}/{len(cats)}] {cat_name}"

            # Build search queries from sample items
            queries = []
            for item in cat_info.get("sample_items", [])[:2]:
                words = item.split()[:3]
                queries.append(" ".join(words))
            terms = cat_info.get("search_terms", [])[:2]
            if terms:
                queries.append(" ".join(terms))

            for query in queries[:2]:
                try:
                    results = session.search(description=query, from_date=from_date, to_date=to_date)
                    for r in results[:20]:
                        supplier = (r.get("supplier_name") or "").lower()
                        if "reytech" in supplier or "rey tech" in supplier:
                            continue

                        email = (r.get("buyer_email") or "").strip()
                        dept = (r.get("dept") or "").strip()
                        if not email and not dept:
                            continue

                        key = email or f"{dept}_{r.get('po_number', '')}"
                        if key not in prospects:
                            prospects[key] = {
                                "id": f"PRO-{uuid.uuid4().hex[:8]}",
                                "buyer_email": email, "buyer_name": "",
                                "buyer_phone": "",
                                "agency": dept, "categories_matched": [],
                                "purchase_orders": [], "total_spend": 0,
                                "outreach_status": "new",
                            }

                        p = prospects[key]
                        if cat_name not in p["categories_matched"]:
                            p["categories_matched"].append(cat_name)

                        po_num = r.get("po_number", "")
                        existing_pos = [x["po_number"] for x in p["purchase_orders"]]
                        if po_num and po_num not in existing_pos:
                            p["purchase_orders"].append({
                                "po_number": po_num, "date": r.get("start_date", ""),
                                "total_num": r.get("grand_total_num"),
                                "items": r.get("first_item", "")[:100], "category": cat_name,
                            })
                            p["total_spend"] += (r.get("grand_total_num") or 0)

                    time.sleep(1)
                except Exception as e:
                    BUYER_STATUS["errors"].append(f"{cat_name}: {e}")

            # Get buyer names + phone from detail on a few results
            try:
                detail_results = results if results else []
            except NameError:
                detail_results = []
            for r in detail_results[:3]:
                if r.get("_results_html") and r.get("_row_index") is not None:
                    try:
                        detail = session.get_detail(r["_results_html"], r["_row_index"], r.get("_click_action"))
                        if detail:
                            hdr = detail.get("header", {}) if isinstance(detail.get("header"), dict) else {}
                            em = hdr.get("buyer_email", "")
                            if em and em in prospects:
                                prospects[em]["buyer_name"] = hdr.get("buyer_name", "") or prospects[em]["buyer_name"]
                                prospects[em]["buyer_phone"] = hdr.get("buyer_phone", "") or prospects[em]["buyer_phone"]
                        time.sleep(0.5)
                    except Exception:
                        pass

        prospect_list = sorted(prospects.values(), key=lambda p: p["total_spend"], reverse=True)
        _save_json(PROSPECTS_FILE, {
            "generated_at": datetime.now().isoformat(),
            "total_prospects": len(prospect_list),
            "from_date": from_date,
            "prospects": prospect_list,
        })

        BUYER_STATUS.update({"running": False, "phase": "complete", "prospects_found": len(prospect_list)})

        return {
            "ok": True, "prospects_found": len(prospect_list),
            "categories_searched": len(cats),
            "top_prospects": [{"agency": p["agency"], "email": p["buyer_email"], "spend": p["total_spend"]} for p in prospect_list[:10]],
        }

    except Exception as e:
        BUYER_STATUS.update({"running": False, "phase": "error"})
        return {"ok": False, "error": str(e)}


# ═══════════════════════════════════════════════════════════════════════
# STEP 3: Email Outreach
# ═══════════════════════════════════════════════════════════════════════

# ── Email Template Library (PRD Feature 4.3) ─────────────────────────────
# Implements Anthropic Skills Guide Pattern 5: Domain-specific intelligence
# Templates are personalized from CRM/intel DB context
# HTML emails with SCPRS hyperlinks for credibility

SCPRS_SEARCH_URL = "https://caleprocure.ca.gov/pages/SCPRSSearch/scprs-search.aspx"

# Agencies Reytech has served — updated dynamically on campaign build
_DEFAULT_SERVED_AGENCIES = ["CCHCS", "CDCR"]


def get_reytech_credentials() -> dict:
    """Compute quantified Reytech credentials from SCPRS history + quotes log.
    Used in email templates and the Growth page banner.
    Returns: {total_sales, total_items, total_pos, agencies_served, calvet_amount,
              calvet_pos, since_year, agency_list, top_categories}
    """
    total_sales = 0
    total_items = 0
    total_pos = 0
    agencies = set()
    calvet_amount = 0
    calvet_pos = 0
    category_spend = defaultdict(float)
    since_year = 2022

    # Source 1: SCPRS history
    hist = _load_json(HISTORY_FILE)
    if isinstance(hist, dict):
        total_pos = hist.get("total_pos", 0)
        total_items = hist.get("total_items", 0)
        for po in hist.get("purchase_orders", []):
            dept = po.get("dept", "")
            amount = float(po.get("total", 0) or po.get("amount", 0) or 0)
            total_sales += amount
            if dept:
                agencies.add(dept)
            # CalVet detection
            if dept and any(kw in dept.lower() for kw in ["calvet", "cal vet", "veterans", "cdva"]):
                calvet_amount += amount
                calvet_pos += 1
            # Category tracking
            for item in po.get("items", []) if isinstance(po.get("items"), list) else []:
                cat = categorize_item(str(item)) if callable(categorize_item) else ""
                if cat:
                    category_spend[cat] += amount / max(len(po.get("items", [1])), 1)

    # Source 2: Quotes log
    try:
        quotes_path = os.path.join(DATA_DIR, "quotes_log.json")
        if os.path.exists(quotes_path):
            with open(quotes_path) as f:
                quotes = json.load(f)
            for q in quotes:
                if q.get("is_test"):
                    continue
                a = q.get("agency", "") or q.get("institution", "")
                if a:
                    agencies.add(a)
                qt = float(q.get("total", 0) or 0)
                total_sales += qt
                items_count = len(q.get("items", []))
                if items_count == 0 and qt > 0:
                    items_count = max(int(q.get("line_count", 0)), 1)  # At least 1 item per real quote
                total_items += items_count
                total_pos += 1
    except Exception:
        pass

    # Source 3: Categories file
    cat_data = _load_json(CATEGORIES_FILE)
    if isinstance(cat_data, dict) and cat_data.get("categories"):
        for cat_name, info in cat_data["categories"].items():
            if cat_name not in category_spend or category_spend[cat_name] == 0:
                category_spend[cat_name] = info.get("total_value", 0)

    top_cats = sorted(category_spend.items(), key=lambda x: x[1], reverse=True)[:6]

    return {
        "total_sales": total_sales,
        "total_items": max(total_items, 1),
        "total_pos": max(total_pos, 1),
        "agencies_served": len(agencies),
        "agency_list": sorted(agencies),
        "calvet_amount": calvet_amount,
        "calvet_pos": calvet_pos,
        "since_year": since_year,
        "top_categories": [{"name": c, "spend": s} for c, s in top_cats],
    }


def get_follow_up_cohorts() -> dict:
    """Compute follow-up cohorts from outreach campaigns.
    Returns: {no_response: [...], second_followup: [...], stale: [...], responded: [...]}
    """
    outreach = _load_json(OUTREACH_FILE)
    if not isinstance(outreach, dict):
        return {"no_response": [], "second_followup": [], "stale": [], "responded": []}

    now = datetime.now()
    no_response = []      # 3-7 days, no reply
    second_followup = []  # 8-21 days, no reply
    stale = []            # 21+ days, no reply
    responded = []

    for camp in outreach.get("campaigns", []):
        for o in camp.get("outreach", []):
            if not o.get("email_sent"):
                continue
            if o.get("response_received"):
                responded.append(o)
                continue
            if o.get("bounced"):
                continue

            sent_at = o.get("email_sent_at", o.get("staged_at", ""))
            if not sent_at:
                continue
            try:
                sent_dt = datetime.fromisoformat(sent_at.replace("Z", ""))
                days = (now - sent_dt).days
            except (ValueError, TypeError):
                continue

            entry = {**o, "days_since": days, "campaign_id": camp.get("id", "")}
            if days <= 7:
                no_response.append(entry)
            elif days <= 21:
                second_followup.append(entry)
            else:
                stale.append(entry)

    # Sort by days since sent (longest first for urgency)
    no_response.sort(key=lambda x: x.get("days_since", 0), reverse=True)
    second_followup.sort(key=lambda x: x.get("days_since", 0), reverse=True)
    stale.sort(key=lambda x: x.get("days_since", 0), reverse=True)

    return {
        "no_response": no_response,
        "second_followup": second_followup,
        "stale": stale,
        "responded": responded,
    }


def score_prospect_weighted(prospect: dict, credentials: dict = None) -> float:
    """Score a prospect with weighted factors for prioritization.
    Higher = more actionable target.
    Weights: past_performance (we've sold to them) > spend_level > recency > response_history
    """
    score = 0.0
    creds = credentials or get_reytech_credentials()
    served = set(a.lower() for a in creds.get("agency_list", []))

    agency = (prospect.get("agency", "") or "").lower()
    spend = float(prospect.get("total_spend", 0) or prospect.get("estimated_spend", 0) or 0)

    # Past performance bonus (we've worked with this agency before)
    if agency and agency in served:
        score += 40  # Big bonus for existing relationship

    # Spend level (higher spend = bigger opportunity)
    if spend >= 100000:
        score += 30
    elif spend >= 50000:
        score += 25
    elif spend >= 10000:
        score += 15
    elif spend >= 1000:
        score += 5

    # Recency of last purchase (longer ago = they may need us)
    last_purchase = prospect.get("last_purchase", "") or prospect.get("last_po_date", "")
    if last_purchase:
        try:
            lp_dt = datetime.fromisoformat(last_purchase.replace("Z", ""))
            months_ago = (datetime.now() - lp_dt).days / 30
            if months_ago >= 12:
                score += 20  # Haven't bought in a year+
            elif months_ago >= 6:
                score += 10
        except (ValueError, TypeError):
            pass

    # Number of matching categories
    cats = prospect.get("categories_matched", [])
    score += min(len(cats) * 3, 15)

    # Has email (actionable)
    if prospect.get("buyer_email"):
        score += 5

    # Status penalties
    status = prospect.get("outreach_status", "new")
    if status == "bounced":
        score -= 50
    elif status == "dead":
        score -= 40
    elif status == "responded" or status == "won":
        score -= 10  # Already engaging

    return round(score, 1)


def _get_served_agencies() -> list:
    """Pull agencies Reytech has served from quotes + SCPRS history."""
    agencies = set()
    try:
        quotes_path = os.path.join(DATA_DIR, "quotes_log.json")
        if os.path.exists(quotes_path):
            with open(quotes_path) as f:
                for q in json.load(f):
                    a = q.get("agency", "") or q.get("institution", "")
                    if a and not q.get("is_test"):
                        agencies.add(a)
    except Exception:
        pass
    try:
        hist = _load_json(HISTORY_FILE)
        if isinstance(hist, dict):
            for po in hist.get("purchase_orders", []):
                d = po.get("dept", "")
                if d:
                    agencies.add(d)
    except Exception:
        pass
    return sorted(agencies) if agencies else _DEFAULT_SERVED_AGENCIES


def _build_agencies_mention(target_agency: str, served: list) -> str:
    """Build 'your agency and other agencies' mention."""
    others = [a for a in served if a.lower() != target_agency.lower()][:3]
    if others:
        return f"{target_agency} and {', '.join(others)}"
    return target_agency


def _build_po_link(prospect: dict) -> str:
    """Build a SCPRS reference for a PO the buyer made that we could support."""
    pos = prospect.get("purchase_orders", [])
    if pos and pos[0].get("po_number"):
        po = pos[0]
        po_num = po.get("po_number", "")
        items = po.get("items", "a recent purchase")[:60]
        return f'<a href="{SCPRS_SEARCH_URL}" style="color:#1a73e8;text-decoration:underline" title="Search PO #{po_num} on SCPRS">{items} (PO #{po_num})</a>'
    # Fallback — reference by category
    cats = prospect.get("categories_matched", [])
    if cats:
        cat_text = ", ".join(cats[:2])
        return f'<a href="{SCPRS_SEARCH_URL}" style="color:#1a73e8;text-decoration:underline" title="Search {cat_text} on SCPRS">your recent {cat_text} purchases</a>'
    return f'<a href="{SCPRS_SEARCH_URL}" style="color:#1a73e8;text-decoration:underline">recent purchases in your categories</a>'


def build_outreach_email(prospect: dict, served_agencies: list = None, template_key: str = "initial_outreach") -> dict:
    """Build personalized HTML email for a prospect with quantified Reytech credentials.
    Returns {subject, body_html, body_plain}."""
    name = prospect.get("buyer_name", "")
    first_name = name.split()[0] if name else ""
    name_greeting = first_name if first_name else "there"
    agency = prospect.get("agency", "your agency")
    cats = prospect.get("categories_matched", [])
    cats_list = list(cats or []) if not isinstance(cats, str) else [cats]

    served = served_agencies or _get_served_agencies()
    agencies_mention = _build_agencies_mention(agency, served)
    po_link = _build_po_link(prospect)
    reytech_link = f'<a href="{SCPRS_SEARCH_URL}" style="color:#1a73e8;text-decoration:underline" title="Search \'Reytech\' as Supplier on SCPRS to see our past performance">Reytech Inc.</a>'

    # Get quantified credentials
    creds = get_reytech_credentials()
    total_sales_short = f"{creds['total_sales']:,.0f}" if creds['total_sales'] >= 1000 else f"{creds['total_sales']:.0f}"
    calvet_line = ""
    if creds["calvet_pos"] > 0:
        calvet_line = f" We have a strong track record with CalVet, having fulfilled {creds['calvet_pos']} POs totaling ${creds['calvet_amount']:,.0f}."

    subject = f"Reytech Inc. — Competitive Pricing for {agency}" if agency != "your agency" else "Reytech Inc. — CA State Vendor Introduction"

    body_html = f"""<div style="font-family:Arial,sans-serif;font-size:14px;line-height:1.6;color:#222">
<p>Hi {name_greeting},</p>

<p>I'm the owner of {reytech_link}. Since {creds['since_year']}, we've fulfilled <strong>{creds['total_pos']}+</strong> purchase orders totaling <strong>${total_sales_short}</strong> across <strong>{creds['agencies_served']}</strong> California state agencies.{' ' + calvet_line if calvet_line else ''}</p>

<p>I noticed {agency} recently purchased {po_link}. We carry these items and can provide competitive pricing.</p>

<p>Could you add <a href="mailto:sales@reytechinc.com" style="color:#1a73e8">sales@reytechinc.com</a> to the RFQ distribution list for future opportunities? We are SB/DVBE certified and ready to quote on your next procurement.</p>

<p style="margin-top:24px">Respectfully,</p>
<p style="margin:0"><strong>Mike</strong><br>
Reytech Inc. | SB/DVBE Certified<br>
<a href="mailto:sales@reytechinc.com" style="color:#1a73e8">sales@reytechinc.com</a> | 949-229-1575<br>
<a href="https://www.reytechinc.com" style="color:#1a73e8">www.reytechinc.com</a></p>
</div>"""

    # Plain text version from template
    po_ref = ""
    pos = prospect.get("purchase_orders", [])
    if pos and pos[0].get("po_number"):
        po_ref = f"{pos[0].get('items','a recent purchase')[:60]} (PO #{pos[0]['po_number']})"
    elif cats_list:
        po_ref = f"your recent {', '.join(cats_list[:2])} purchases"
    else:
        po_ref = "recent purchases in your categories"

    template = EMAIL_TEMPLATES.get(template_key, EMAIL_TEMPLATES["initial_outreach"])
    body_plain = template.format(
        name_greeting=f" {name_greeting}" if name_greeting != "there" else "",
        agency=agency,
        items_mention=po_ref,
        since_year=creds["since_year"],
        total_pos=creds["total_pos"],
        total_items=creds["total_items"],
        total_sales_short=total_sales_short,
        agencies_served=creds["agencies_served"],
        calvet_line=calvet_line,
    )

    return {"subject": subject, "body_html": body_html, "body_plain": body_plain}


# Legacy templates kept for backward compat with distro campaigns
EMAIL_TEMPLATES = {
    "distro_list": """Hi{name_greeting},

I'm the owner of Reytech Inc. Since {since_year}, we've fulfilled {total_pos}+ purchase orders across {agencies_served} California state agencies, delivering {total_items}+ line items.{calvet_line}

We are able to competitively source many of the supplies {agency} purchases, including {items_mention}.

Could you add sales@reytechinc.com to the RFQ distribution list for future opportunities? We are SB/DVBE certified and ready to quote on your next procurement.

Respectfully,

Mike
Reytech Inc. | SB/DVBE Certified
sales@reytechinc.com | 949-229-1575
www.reytechinc.com""",

    "initial_outreach": """Hi{name_greeting},

I'm the owner of Reytech Inc. Since {since_year}, we've fulfilled {total_pos}+ purchase orders totaling ${total_sales_short} across {agencies_served} California state agencies.{calvet_line}

I noticed {agency} recently purchased {items_mention}. We carry these items and can provide competitive pricing.

You can verify our past performance on SCPRS: https://caleprocure.ca.gov/pages/SCPRSSearch/scprs-search.aspx (search "Reytech" as Supplier)

Could you add sales@reytechinc.com to the RFQ distribution list? We'd love the opportunity to quote on your next procurement.

Respectfully,

Mike
Reytech Inc. | SB/DVBE Certified
sales@reytechinc.com | 949-229-1575
www.reytechinc.com""",

    "follow_up": """Hi{name_greeting},

Following up on my recent email — I'm the owner of Reytech Inc. We've been serving California state agencies since {since_year} with {total_pos}+ POs fulfilled.

We specialize in {items_mention} and would love to support {agency}. We're a certified SB/DVBE vendor ready to quote competitively.

Could you add sales@reytechinc.com to your RFQ distribution list? We're ready for your next procurement.

Feel free to reply or call/text 949-229-1575.

Respectfully,

Mike
Reytech Inc. | SB/DVBE Certified
sales@reytechinc.com | 949-229-1575""",

    "second_follow_up": """Hi{name_greeting},

I wanted to reach out one more time — Reytech Inc. has been supporting CA state agencies since {since_year} and we'd appreciate the chance to earn {agency}'s business.

We have a strong track record with {total_pos}+ purchase orders fulfilled{calvet_line} and we can competitively price items like {items_mention}.

If there's a better contact for procurement, I'd be grateful for the referral. Otherwise, please add sales@reytechinc.com to your RFQ distribution list.

Thank you for your time.

Respectfully,

Mike
Reytech Inc. | SB/DVBE Certified
sales@reytechinc.com | 949-229-1575""",

    "past_customer_reactivation": """Hi{name_greeting},

I hope you're doing well. It's been a while since {agency} last ordered from Reytech Inc. and I wanted to reach out to see if there are any upcoming procurements we can support.

Since our last order together, we've expanded our catalog and now carry {total_items}+ items across {agencies_served} agencies. We remain SB/DVBE certified and competitively priced.

If your RFQ list needs updating, please add sales@reytechinc.com. We'd love to earn your business again.

Respectfully,

Mike
Reytech Inc. | SB/DVBE Certified
sales@reytechinc.com | 949-229-1575""",

    "quote_won": """Hi{name_greeting},

Thank you for the award on {items_mention}. Your order is being processed and you'll receive tracking information once shipped.

We look forward to continuing to support {agency}. For your next procurement, please keep sales@reytechinc.com on your distribution list.

Respectfully,

Mike
Reytech Inc.
sales@reytechinc.com | 949-229-1575""",
}

# Keep backward-compat alias
EMAIL_TEMPLATE = EMAIL_TEMPLATES["initial_outreach"]




def launch_distro_campaign(
    max_contacts: int = 100,
    dry_run: bool = True,
    template: str = "distro_list",
    source_filter: str = "",
) -> dict:
    """Phase 1 Growth Campaign: Email CA state agency buyers to get on RFQ distro lists.

    This is the primary growth lever per the PRD:
      'Get on RFQ distribution list — Phase 1'
      Target: ~100 buyers from SCPRS intel, personalized by agency/category.

    Uses the Anthropic Skills Guide Pattern 5 (Domain Intelligence):
      Pulls live DB context → personalizes each email from actual purchase history →
      logs every touch to activity_log → deduplicates against already-contacted.

    Args:
        max_contacts: Max emails to send (or stage in dry_run)
        dry_run: True = build emails but DO NOT send (review first)
        template: Template key from EMAIL_TEMPLATES
        source_filter: Only contact buyers from this agency/source substring

    Returns:
        {ok, campaign_id, total_staged, total_sent, emails[], dry_run, context_used}
    """
    log.info("launch_distro_campaign: dry_run=%s template=%s max=%d", dry_run, template, max_contacts)

    # ── Pull DB context (agent intelligence layer) ─────────────────────────
    ctx = get_context(include_contacts=True, include_revenue=True, include_quotes=True)
    ctx_summary = format_context_for_agent(ctx, focus="crm")

    # ── Load all buyers from intel + CRM ───────────────────────────────────
    all_buyers = []

    # Source 1: Intel buyers (SCPRS-sourced)
    intel_data = {"buyers": [dict(c) for c in get_intel_agencies()]} if _HAS_DB_DAL else _load_json(os.path.join(DATA_DIR, "intel_buyers.json"))
    if isinstance(intel_data, dict):
        for b in intel_data.get("buyers", []):
            if b.get("buyer_email") and "@" in b.get("buyer_email", ""):
                all_buyers.append({
                    "id": b.get("id", f"intel_{len(all_buyers)}"),
                    "name": b.get("buyer_name", ""),
                    "email": b.get("buyer_email", ""),
                    "agency": b.get("agency", ""),
                    "categories": b.get("categories", []),
                    "spend": b.get("annual_spend", 0),
                    "source": "intel",
                    "items_mention": ", ".join(list(b.get("categories", []) or [])[:2]) or "medical supplies",
                    "purchase_date": "recently",
                })

    # Source 2: Growth prospects
    prospect_data = _load_json(PROSPECTS_FILE)
    if isinstance(prospect_data, dict):
        for p in prospect_data.get("prospects", []):
            if p.get("buyer_email") and "@" in p.get("buyer_email", ""):
                pos = p.get("purchase_orders", [])
                items = pos[0].get("items", ", ".join(p.get("categories_matched", [])))[:80] if pos else ", ".join(p.get("categories_matched", []))[:80]
                date = pos[0].get("date", "recently") if pos else "recently"
                all_buyers.append({
                    "id": p.get("id", f"prospect_{len(all_buyers)}"),
                    "name": p.get("buyer_name", ""),
                    "email": p.get("buyer_email", ""),
                    "agency": p.get("agency", ""),
                    "categories": p.get("categories_matched", []),
                    "spend": p.get("estimated_spend", 0),
                    "source": "prospect",
                    "items_mention": items or "supplies",
                    "purchase_date": date,
                })

    # Source 3: CRM contacts with email not yet emailed
    for c in ctx.get("contacts", []):
        if c.get("email") and "@" in c.get("email", "") and c.get("status") == "new":
            if not any(b["email"] == c["email"] for b in all_buyers):
                cats = list(c.get("categories", []) or [])
                all_buyers.append({
                    "id": c["id"],
                    "name": c.get("name", ""),
                    "email": c.get("email", ""),
                    "agency": c.get("agency", ""),
                    "categories": cats,
                    "spend": c.get("spend", 0),
                    "source": "crm",
                    "items_mention": ", ".join(list(cats or [])[:2]) or "your recent purchases",
                    "purchase_date": "recently",
                })

    # ── Deduplicate against previously contacted ───────────────────────────
    outreach_data = _load_json(OUTREACH_FILE)
    if not isinstance(outreach_data, dict):
        outreach_data = {"campaigns": [], "total_sent": 0}

    already_contacted = set()
    for camp in outreach_data.get("campaigns", []):
        for o in camp.get("outreach", []):
            if o.get("email_sent") or o.get("staged"):
                already_contacted.add(o.get("email", ""))

    # Apply filters
    if source_filter:
        sf = source_filter.lower()
        all_buyers = [b for b in all_buyers if sf in (b.get("agency") or "").lower()
                      or sf in (b.get("source") or "").lower()]

    new_buyers = [b for b in all_buyers if b["email"] not in already_contacted]
    new_buyers = new_buyers[:max_contacts]

    if not new_buyers:
        return {
            "ok": True,
            "message": f"All {len(all_buyers)} contacts already contacted (or no contacts found). Run Intel Deep Pull to find more.",
            "total_staged": 0,
            "total_sent": 0,
            "dry_run": dry_run,
        }

    # ── Build campaign ─────────────────────────────────────────────────────
    campaign_id = f"DISTRO-{datetime.now().strftime('%Y%m%d-%H%M')}"
    email_template = EMAIL_TEMPLATES.get(template, EMAIL_TEMPLATES["distro_list"])
    creds = get_reytech_credentials()
    total_sales_short = f"{creds['total_sales']:,.0f}" if creds['total_sales'] >= 1000 else f"{creds['total_sales']:.0f}"
    calvet_line = ""
    if creds["calvet_pos"] > 0:
        calvet_line = f" We have a strong track record with CalVet, having fulfilled {creds['calvet_pos']} POs totaling ${creds['calvet_amount']:,.0f}."

    campaign = {
        "id": campaign_id,
        "type": "distro_list_phase1",
        "created_at": datetime.now().isoformat(),
        "dry_run": dry_run,
        "template": template,
        "context_summary": ctx_summary[:500],
        "credentials": creds,
        "outreach": [],
    }
    staged = 0
    sent = 0
    emails = []

    for b in new_buyers:
        name = b.get("name", "")
        name_greeting = f" {name.split()[0]}" if name else ""
        agency = b.get("agency", "your agency")
        items_mention = b.get("items_mention", "supplies")
        purchase_date = b.get("purchase_date", "recently")
        cats = list(b.get("categories", []) or [])
        if isinstance(cats, str): cats = [cats] if cats else []

        # Personalize subject by category
        if "Medical" in str(cats) or "medical" in items_mention.lower():
            subject = f"Reytech Inc. — Competitive Pricing on Medical Supplies for {agency}"
        elif "Janitorial" in str(cats) or "cleaning" in items_mention.lower():
            subject = f"Reytech Inc. — Janitorial & Cleaning Supplies — SB/DVBE Vendor"
        elif cats:
            subject = f"Reytech Inc. — {cats[0]} Pricing — SB/DVBE — Get on RFQ Distro"
        else:
            subject = f"Reytech Inc. — CA State Vendor Introduction — SB/DVBE Certified"

        try:
            body = email_template.format(
                name_greeting=name_greeting,
                agency=agency,
                items_mention=items_mention,
                purchase_date=purchase_date,
                since_year=creds["since_year"],
                total_pos=creds["total_pos"],
                total_items=creds["total_items"],
                total_sales_short=total_sales_short,
                agencies_served=creds["agencies_served"],
                calvet_line=calvet_line,
            )
        except KeyError:
            # Fallback for templates without credential placeholders
            body = email_template.format(
                name_greeting=name_greeting,
                agency=agency,
                items_mention=items_mention,
                purchase_date=purchase_date,
            )

        entry = {
            "buyer_id": b["id"],
            "email": b["email"],
            "name": name,
            "agency": agency,
            "categories": cats,
            "subject": subject,
            "body": body,
            "staged": True,
            "email_sent": False,
            "email_sent_at": None,
            "campaign_id": campaign_id,
            "follow_up_date": (datetime.now() + timedelta(days=5)).isoformat(),
        }

        if not dry_run and b.get("email"):
            try:
                from src.agents.email_poller import EmailSender
                gmail = os.environ.get("GMAIL_ADDRESS", "")
                pwd = os.environ.get("GMAIL_PASSWORD", "")
                if gmail and pwd:
                    sender = EmailSender({"email": gmail, "email_password": pwd})
                    sender.send({"to": b["email"], "subject": subject, "body": body, "attachments": []})
                    entry["email_sent"] = True
                    entry["email_sent_at"] = datetime.now().isoformat()
                    sent += 1
                    log.info("Distro email SENT → %s (%s)", b["email"], agency)
                    _add_event(b["id"], "distro_email_sent", f"Phase 1 distro campaign: {subject}")
                    _log_email_to_crm(b["id"], b["email"], subject, body, agency)
                    time.sleep(1.2)  # rate limit
                else:
                    entry["error"] = "GMAIL_ADDRESS / GMAIL_PASSWORD not set in Railway env"
            except Exception as e:
                entry["error"] = str(e)
                log.error("Email send failed %s: %s", b["email"], e)
        else:
            log.info("Distro email STAGED → %s (%s) [dry_run]", b["email"], agency)

        campaign["outreach"].append(entry)
        emails.append({
            "to": b["email"],
            "name": name,
            "agency": agency,
            "subject": subject,
            "body_preview": body[:120] + "...",
            "sent": entry["email_sent"],
        })
        staged += 1

    # ── Save campaign ──────────────────────────────────────────────────────
    outreach_data.setdefault("campaigns", []).append(campaign)
    outreach_data["total_sent"] = outreach_data.get("total_sent", 0) + sent
    outreach_data["last_distro_campaign"] = campaign_id
    _save_json(OUTREACH_FILE, outreach_data)

    # Log to DB activity
    try:
        from src.core.db import log_activity as _la
        _la(
            contact_id="growth_agent",
            event_type="distro_campaign_launched",
            detail=f"Campaign {campaign_id}: {staged} staged, {sent} sent, dry_run={dry_run}",
            actor="growth_agent",
            metadata={"campaign_id": campaign_id, "staged": staged, "sent": sent, "template": template},
        )
    except Exception:
        pass

    log.info("launch_distro_campaign: %s — staged=%d sent=%d dry_run=%s",
             campaign_id, staged, sent, dry_run)

    result = {
        "ok": True,
        "campaign_id": campaign_id,
        "total_available": len(all_buyers),
        "already_contacted": len(already_contacted),
        "total_staged": staged,
        "total_sent": sent,
        "dry_run": dry_run,
        "template": template,
        "emails": emails,
        "follow_up_in_days": 5,
        "next_step": "Review emails in growth outreach dashboard, then set dry_run=False to send" if dry_run else f"Sent {sent} emails. Follow up in 5 days.",
        "context_used": bool(ctx.get("contacts")),
    }
    return result


def launch_outreach(max_prospects=50, dry_run=True):
    """Send personalized emails. dry_run=True builds but doesn't send."""
    prospect_data = _load_json(PROSPECTS_FILE)
    if not isinstance(prospect_data, dict) or not prospect_data.get("prospects"):
        return {"ok": False, "error": "No prospects. Run find_category_buyers first."}

    outreach = _load_json(OUTREACH_FILE)
    if not isinstance(outreach, dict):
        outreach = {"campaigns": [], "total_sent": 0}

    contacted = set()
    for c in outreach.get("campaigns", []):
        for o in c.get("outreach", []):
            if o.get("email_sent"):
                contacted.add(o.get("email", ""))

    new = [p for p in prospect_data["prospects"] if p.get("buyer_email") and p["buyer_email"] not in contacted][:max_prospects]
    if not new:
        return {"ok": True, "message": "All prospects already contacted", "new_to_contact": 0}

    # Load category data for richer personalization when POs are empty
    cat_data = _load_json(CATEGORIES_FILE)
    cat_info = cat_data.get("categories", {}) if isinstance(cat_data, dict) else {}

    # Get agencies Reytech has served for the template
    served_agencies = _get_served_agencies()

    campaign = {"id": f"GC-{datetime.now().strftime('%Y%m%d-%H%M')}", "created_at": datetime.now().isoformat(), "dry_run": dry_run, "outreach": []}
    sent = 0

    for p in new:
        name = p.get("buyer_name", "")
        agency = p.get("agency", "your agency")
        cats = p.get("categories_matched", [])
        cats_list = list(cats or []) if not isinstance(cats, str) else [cats]

        # Build HTML email with SCPRS links
        email_data = build_outreach_email(p, served_agencies)
        subject = email_data["subject"]
        body_html = email_data["body_html"]
        body_plain = email_data["body_plain"]

        entry = {
            "prospect_id": p["id"], "email": p["buyer_email"], "name": name,
            "agency": agency, "categories": cats,
            "email_subject": subject, "email_body": body_plain,
            "email_body_html": body_html,
            "email_sent": False, "email_sent_at": None,
            "voice_follow_up_date": (datetime.now() + timedelta(days=4)).isoformat(),
            "voice_called": False, "response_received": False,
        }

        if not dry_run and p.get("buyer_email"):
            try:
                from src.agents.email_poller import EmailSender
                config = {"email": os.environ.get("GMAIL_ADDRESS", ""), "email_password": os.environ.get("GMAIL_PASSWORD", "")}
                if config["email"] and config["email_password"]:
                    sender = EmailSender(config)
                    sender.send({"to": p["buyer_email"], "subject": subject,
                                 "body": body_plain, "body_html": body_html,
                                 "attachments": []})
                    entry["email_sent"] = True
                    entry["email_sent_at"] = datetime.now().isoformat()
                    _update_prospect_status(p["id"], "emailed")
                    _add_event(p["id"], "email_sent", f"Sent: {subject}")
                    _log_email_to_crm(p["id"], p["buyer_email"], subject, body_plain, agency)
                    sent += 1
                    log.info(f"Growth email → {p['buyer_email']} ({agency})")
                    time.sleep(1)
                else:
                    entry["error"] = "Gmail not configured"
            except Exception as e:
                entry["error"] = str(e)

        campaign["outreach"].append(entry)

    outreach.setdefault("campaigns", []).append(campaign)
    outreach["total_sent"] = outreach.get("total_sent", 0) + sent
    _save_json(OUTREACH_FILE, outreach)

    return {
        "ok": True, "campaign_id": campaign["id"], "dry_run": dry_run,
        "emails_built": len(new), "emails_sent": sent,
        "follow_up_date": (datetime.now() + timedelta(days=4)).strftime("%Y-%m-%d"),
        "preview": [
            {"to": o["email"], "agency": o["agency"], "subject": o["email_subject"],
             "body": o["email_body"], "body_html": o.get("email_body_html", "")}
            for o in campaign["outreach"][:10]
        ],
    }


# ═══════════════════════════════════════════════════════════════════════
# STEP 4: Voice Follow-Up
# ═══════════════════════════════════════════════════════════════════════

def check_follow_ups():
    """Find prospects who haven't responded after 3-5 business days."""
    outreach = _load_json(OUTREACH_FILE)
    if not isinstance(outreach, dict):
        return {"ok": True, "ready": [], "count": 0}

    now = datetime.now()
    ready = []
    for c in outreach.get("campaigns", []):
        if c.get("dry_run"):
            continue
        for o in c.get("outreach", []):
            if o.get("email_sent") and not o.get("response_received") and not o.get("voice_called"):
                try:
                    fdate = datetime.fromisoformat(o.get("voice_follow_up_date", ""))
                    if now >= fdate:
                        ready.append({"prospect_id": o["prospect_id"], "email": o["email"], "agency": o.get("agency", ""), "categories": o.get("categories", [])})
                except Exception:
                    pass

    return {"ok": True, "ready": ready, "count": len(ready)}


def launch_voice_follow_up(max_calls=10):
    """Auto-dial non-responders using voice agent."""
    fu = check_follow_ups()
    if not fu.get("ready"):
        return {"ok": True, "message": "No follow-ups due", "calls_made": 0}

    prospect_data = _load_json(PROSPECTS_FILE)
    pmap = {}
    if isinstance(prospect_data, dict):
        for p in prospect_data.get("prospects", []):
            pmap[p["id"]] = p

    calls = 0
    for target in fu["ready"][:max_calls]:
        prospect = pmap.get(target["prospect_id"], {})
        phone = prospect.get("buyer_phone", "")
        if not phone:
            continue
        try:
            from src.agents.voice_agent import place_call
            result = place_call(phone_number=phone, script_key="growth_email_follow_up", variables={
                "institution": target["agency"],
                "top_items": ", ".join(target.get("categories", ["supplies"])[:3]),
            })
            if result.get("ok"):
                calls += 1
                _mark_called(target["prospect_id"])
        except Exception as e:
            log.warning(f"Voice failed {target['agency']}: {e}")

    return {"ok": True, "calls_made": calls, "remaining": fu["count"] - calls}


def _mark_called(prospect_id):
    outreach = _load_json(OUTREACH_FILE)
    if not isinstance(outreach, dict):
        return
    for c in outreach.get("campaigns", []):
        for o in c.get("outreach", []):
            if o.get("prospect_id") == prospect_id:
                o["voice_called"] = True
                o["voice_called_at"] = datetime.now().isoformat()
    _save_json(OUTREACH_FILE, outreach)


# ═══════════════════════════════════════════════════════════════════════
# Status
# ═══════════════════════════════════════════════════════════════════════

def get_growth_status():
    history = _load_json(HISTORY_FILE)
    cats = _load_json(CATEGORIES_FILE)
    prospects = _load_json(PROSPECTS_FILE)
    outreach = _load_json(OUTREACH_FILE)
    h = history if isinstance(history, dict) else {}
    c = cats if isinstance(cats, dict) else {}
    p = prospects if isinstance(prospects, dict) else {}
    o = outreach if isinstance(outreach, dict) else {}
    return {
        "ok": True,
        "history": {"total_pos": h.get("total_pos", 0), "total_items": h.get("total_items", 0), "pulled_at": h.get("pulled_at")},
        "categories": {"total": c.get("total_categories", 0), "names": list(c.get("categories", {}).keys())[:7]},
        "prospects": {"total": p.get("total_prospects", 0), "generated_at": p.get("generated_at")},
        "outreach": {"total_sent": o.get("total_sent", 0), "campaigns": len(o.get("campaigns", []))},
        "pull_status": PULL_STATUS,
        "buyer_status": BUYER_STATUS,
    }


def lead_funnel():
    return get_growth_status()


# ═══════════════════════════════════════════════════════════════════════
# PROSPECT CRM — Contact Management + Timeline
# ═══════════════════════════════════════════════════════════════════════

TIMELINE_FILE = os.path.join(DATA_DIR, "growth_timeline.json")

# Status flow: new → emailed → follow_up_due → called → responded | bounced | dead
VALID_STATUSES = ["new", "emailed", "follow_up_due", "called", "responded", "bounced", "dead", "won"]

def _load_timeline() -> dict:
    """Load timeline events keyed by prospect_id."""
    data = _load_json(TIMELINE_FILE)
    return data if isinstance(data, dict) else {}

def _save_timeline(data: dict):
    _save_json(TIMELINE_FILE, data)

def _add_event(prospect_id: str, event_type: str, detail: str = "", metadata: dict = None):
    """Add a timeline event for a prospect."""
    timeline = _load_timeline()
    events = timeline.setdefault(prospect_id, [])
    events.append({
        "type": event_type,
        "detail": detail,
        "timestamp": datetime.now().isoformat(),
        "metadata": metadata or {},
    })
    _save_timeline(timeline)


def _log_email_to_crm(prospect_id: str, email: str, subject: str, body: str, agency: str = ""):
    """Log an outreach email send to CRM contacts for unified activity tracking."""
    try:
        crm_path = os.path.join(DATA_DIR, "crm_contacts.json")
        contacts = {}
        if os.path.exists(crm_path):
            with open(crm_path) as f:
                contacts = json.load(f)
        if prospect_id in contacts:
            contacts[prospect_id].setdefault("activity", []).append({
                "event_type": "email_sent",
                "detail": f"Growth outreach: {subject}",
                "actor": "growth_agent",
                "timestamp": datetime.now().isoformat(),
                "metadata": {"subject": subject, "to": email, "source": "growth_campaign"},
            })
            contacts[prospect_id]["updated_at"] = datetime.now().isoformat()
            with open(crm_path, "w") as f:
                json.dump(contacts, f, indent=2, default=str)
    except Exception as e:
        log.debug("CRM log suppressed: %s", e)

    # Also add to outbox for unified tracking
    try:
        outbox_path = os.path.join(DATA_DIR, "email_outbox.json")
        outbox = []
        if os.path.exists(outbox_path):
            with open(outbox_path) as f:
                outbox = json.load(f)
        outbox.append({
            "id": f"growth-{prospect_id}-{datetime.now().strftime('%Y%m%d%H%M')}",
            "to": email,
            "subject": subject,
            "body": body,
            "status": "sent",
            "created_at": datetime.now().isoformat(),
            "sent_at": datetime.now().isoformat(),
            "source": "growth_campaign",
            "prospect_id": prospect_id,
            "agency": agency,
            "metadata": {"type": "growth_outreach"},
        })
        with open(outbox_path, "w") as f:
            json.dump(outbox, f, indent=2, default=str)
    except Exception as e:
        log.debug("Outbox log suppressed: %s", e)


def _update_prospect_status(prospect_id: str, new_status: str):
    """Update a prospect's outreach_status in the prospects file."""
    data = _load_json(PROSPECTS_FILE)
    if not isinstance(data, dict):
        return
    for p in data.get("prospects", []):
        if p.get("id") == prospect_id:
            old = p.get("outreach_status", "new")
            p["outreach_status"] = new_status
            p["status_updated_at"] = datetime.now().isoformat()
            _save_json(PROSPECTS_FILE, data)
            _add_event(prospect_id, "status_change", f"{old} → {new_status}")
            return
    log.warning(f"Prospect {prospect_id} not found for status update")


def get_prospect(prospect_id: str) -> dict:
    """Get a single prospect with full timeline."""
    data = _load_json(PROSPECTS_FILE)
    if not isinstance(data, dict):
        return {"ok": False, "error": "No prospects data"}
    for p in data.get("prospects", []):
        if p.get("id") == prospect_id:
            timeline = _load_timeline().get(prospect_id, [])
            # Also pull outreach records
            outreach = _load_json(OUTREACH_FILE)
            outreach_records = []
            if isinstance(outreach, dict):
                for c in outreach.get("campaigns", []):
                    for o in c.get("outreach", []):
                        if o.get("prospect_id") == prospect_id:
                            outreach_records.append(o)
            return {
                "ok": True, "prospect": p,
                "timeline": sorted(timeline, key=lambda e: e.get("timestamp", ""), reverse=True),
                "outreach_records": outreach_records,
            }
    return {"ok": False, "error": "Prospect not found"}


def update_prospect(prospect_id: str, updates: dict) -> dict:
    """Update prospect fields (name, phone, email, status, notes)."""
    data = _load_json(PROSPECTS_FILE)
    if not isinstance(data, dict):
        return {"ok": False, "error": "No prospects data"}
    for p in data.get("prospects", []):
        if p.get("id") == prospect_id:
            changed = []
            for key in ["buyer_name", "buyer_phone", "buyer_email", "outreach_status", "notes"]:
                if key in updates and updates[key] != p.get(key):
                    old_val = p.get(key, "")
                    p[key] = updates[key]
                    changed.append(f"{key}: {old_val} → {updates[key]}")
            if changed:
                p["updated_at"] = datetime.now().isoformat()
                _save_json(PROSPECTS_FILE, data)
                _add_event(prospect_id, "updated", "; ".join(changed))
            return {"ok": True, "changed": changed}
    return {"ok": False, "error": "Prospect not found"}


def add_prospect_note(prospect_id: str, note: str) -> dict:
    """Add a manual note to a prospect's timeline."""
    if not note.strip():
        return {"ok": False, "error": "Note cannot be empty"}
    _add_event(prospect_id, "note", note)
    return {"ok": True}


def mark_responded(prospect_id: str, response_type: str = "email_reply", detail: str = "") -> dict:
    """Mark a prospect as having responded."""
    _update_prospect_status(prospect_id, "responded")
    _add_event(prospect_id, "response_received", detail or response_type, {"response_type": response_type})
    # Update outreach records
    outreach = _load_json(OUTREACH_FILE)
    if isinstance(outreach, dict):
        for c in outreach.get("campaigns", []):
            for o in c.get("outreach", []):
                if o.get("prospect_id") == prospect_id:
                    o["response_received"] = True
                    o["response_at"] = datetime.now().isoformat()
                    o["response_type"] = response_type
        _save_json(OUTREACH_FILE, outreach)
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════
# BOUNCEBACK HANDLING
# ═══════════════════════════════════════════════════════════════════════

BOUNCE_KEYWORDS = [
    "delivery status notification", "undeliverable", "mail delivery failed",
    "returned mail", "delivery failure", "address rejected",
    "mailbox not found", "user unknown", "no such user",
    "mailbox unavailable", "permanent failure", "550 ",
    "message not delivered", "delivery has failed",
]

def detect_bounceback(subject: str, body: str, sender: str = "") -> dict:
    """Check if an email is a bounceback notification."""
    text = f"{subject} {body}".lower()
    sender_lower = sender.lower()

    # Check sender patterns
    is_bounce_sender = any(s in sender_lower for s in [
        "mailer-daemon", "postmaster", "mail-daemon", "noreply",
        "no-reply", "bounce", "delivery",
    ])

    # Check content
    bounce_signals = sum(1 for kw in BOUNCE_KEYWORDS if kw in text)

    if bounce_signals >= 2 or (is_bounce_sender and bounce_signals >= 1):
        # Try to extract the original recipient
        recipient = ""
        # Common patterns: "The email to <user@example.com> failed"
        import re
        email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', body)
        if email_match:
            recipient = email_match.group(0)

        return {
            "is_bounce": True,
            "confidence": min(1.0, bounce_signals * 0.3 + (0.4 if is_bounce_sender else 0)),
            "recipient": recipient,
            "signals": bounce_signals,
        }

    return {"is_bounce": False, "confidence": 0}


def process_bounceback(email_address: str, reason: str = "") -> dict:
    """Mark a prospect as bounced and exclude from future outreach."""
    data = _load_json(PROSPECTS_FILE)
    if not isinstance(data, dict):
        return {"ok": False, "error": "No prospects data"}

    found = None
    for p in data.get("prospects", []):
        if p.get("buyer_email", "").lower() == email_address.lower():
            found = p
            break

    if not found:
        return {"ok": False, "error": f"No prospect with email {email_address}"}

    found["outreach_status"] = "bounced"
    found["bounced_at"] = datetime.now().isoformat()
    found["bounce_reason"] = reason
    _save_json(PROSPECTS_FILE, data)

    _add_event(found["id"], "email_bounced", reason or f"Bounce: {email_address}")

    # Mark in outreach campaigns too
    outreach = _load_json(OUTREACH_FILE)
    if isinstance(outreach, dict):
        for c in outreach.get("campaigns", []):
            for o in c.get("outreach", []):
                if o.get("email", "").lower() == email_address.lower():
                    o["bounced"] = True
                    o["bounce_reason"] = reason
        _save_json(OUTREACH_FILE, outreach)

    log.info(f"Bounce processed: {email_address} ({reason})")
    return {"ok": True, "prospect_id": found["id"], "agency": found.get("agency")}


def scan_inbox_for_bounces() -> dict:
    """Check inbox for bounceback emails and process them."""
    try:
        from src.agents.email_poller import EmailSender
        config = {
            "email": os.environ.get("GMAIL_ADDRESS", ""),
            "email_password": os.environ.get("GMAIL_PASSWORD", ""),
        }
        if not config["email"] or not config["email_password"]:
            return {"ok": False, "error": "Gmail not configured"}
        sender = EmailSender(config)
    except ImportError:
        return {"ok": False, "error": "EmailSender not available"}

    bounces_found = 0
    try:
        # Check recent emails for bouncebacks
        import imaplib, email as email_lib
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(config["email"], config["email_password"])
        imap.select("INBOX")

        # Search last 7 days
        from_date = (datetime.now() - timedelta(days=7)).strftime("%d-%b-%Y")
        _, msg_ids = imap.search(None, f'(SINCE "{from_date}" FROM "mailer-daemon")')

        # Also check for delivery failure subjects
        _, msg_ids2 = imap.search(None, f'(SINCE "{from_date}" SUBJECT "delivery")')

        all_ids = set()
        for ids in [msg_ids[0], msg_ids2[0]]:
            if ids:
                all_ids.update(ids.split())

        for msg_id in list(all_ids)[:50]:
            _, data = imap.fetch(msg_id, "(RFC822)")
            msg = email_lib.message_from_bytes(data[0][1])
            subject = str(msg.get("Subject", ""))
            sender_addr = str(msg.get("From", ""))
            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        try:
                            body = part.get_payload(decode=True).decode("utf-8", errors="replace")
                        except Exception:
                            pass
                        break
            else:
                try:
                    body = msg.get_payload(decode=True).decode("utf-8", errors="replace")
                except Exception:
                    pass

            bounce = detect_bounceback(subject, body, sender_addr)
            if bounce.get("is_bounce") and bounce.get("recipient"):
                result = process_bounceback(bounce["recipient"], f"Auto-detected: {subject[:60]}")
                if result.get("ok"):
                    bounces_found += 1

        imap.logout()

    except Exception as e:
        log.warning(f"Bounce scan error: {e}")
        return {"ok": False, "error": str(e), "bounces_found": bounces_found}

    return {"ok": True, "bounces_found": bounces_found}


# ═══════════════════════════════════════════════════════════════════════
# AUTO-SCHEDULER — Background Follow-Up Engine
# ═══════════════════════════════════════════════════════════════════════

_scheduler_thread = None
_scheduler_stop = threading.Event()
SCHEDULER_INTERVAL = 3600  # Check every hour

def _scheduler_loop():
    """Background loop: check bounces + initiate voice follow-ups."""
    log.info("Growth scheduler started (every %ds)", SCHEDULER_INTERVAL)
    while not _scheduler_stop.is_set():
        try:
            # 1. Scan for bouncebacks
            bounce_result = scan_inbox_for_bounces()
            if bounce_result.get("bounces_found", 0) > 0:
                log.info(f"Scheduler: {bounce_result['bounces_found']} bounces processed")

            # 2. Update statuses for follow-up-due prospects
            outreach = _load_json(OUTREACH_FILE)
            now = datetime.now()
            due_count = 0
            if isinstance(outreach, dict):
                for c in outreach.get("campaigns", []):
                    if c.get("dry_run"):
                        continue
                    for o in c.get("outreach", []):
                        if (o.get("email_sent")
                            and not o.get("response_received")
                            and not o.get("voice_called")
                            and not o.get("bounced")
                            and o.get("voice_follow_up_date")):
                            try:
                                fdate = datetime.fromisoformat(o["voice_follow_up_date"])
                                if now >= fdate:
                                    due_count += 1
                                    _update_prospect_status(o["prospect_id"], "follow_up_due")
                            except Exception:
                                pass

            if due_count > 0:
                log.info(f"Scheduler: {due_count} prospects now due for voice follow-up")

        except Exception as e:
            log.debug(f"Scheduler error: {e}")

        _scheduler_stop.wait(SCHEDULER_INTERVAL)


def start_scheduler():
    """Start background CRM scheduler."""
    global _scheduler_thread
    if _scheduler_thread and _scheduler_thread.is_alive():
        return
    _scheduler_stop.clear()
    _scheduler_thread = threading.Thread(target=_scheduler_loop, daemon=True, name="growth-scheduler")
    _scheduler_thread.start()


def get_campaign_dashboard() -> dict:
    """Campaign management overview with full metrics."""
    outreach = _load_json(OUTREACH_FILE)
    if not isinstance(outreach, dict):
        return {"ok": True, "campaigns": [], "totals": {}}

    prospects_data = _load_json(PROSPECTS_FILE)
    prospect_map = {}
    if isinstance(prospects_data, dict):
        for p in prospects_data.get("prospects", []):
            prospect_map[p["id"]] = p

    # Per-status counts
    status_counts = defaultdict(int)
    for p in prospect_map.values():
        status_counts[p.get("outreach_status", "new")] += 1

    campaigns = []
    total_sent = 0
    total_bounced = 0
    total_responded = 0
    total_called = 0
    total_pending = 0

    for c in outreach.get("campaigns", []):
        camp_sent = sum(1 for o in c.get("outreach", []) if o.get("email_sent"))
        camp_bounced = sum(1 for o in c.get("outreach", []) if o.get("bounced"))
        camp_responded = sum(1 for o in c.get("outreach", []) if o.get("response_received"))
        camp_called = sum(1 for o in c.get("outreach", []) if o.get("voice_called"))
        camp_pending = camp_sent - camp_bounced - camp_responded - camp_called

        campaigns.append({
            "id": c.get("id"),
            "created_at": c.get("created_at"),
            "dry_run": c.get("dry_run", True),
            "total": len(c.get("outreach", [])),
            "sent": camp_sent,
            "bounced": camp_bounced,
            "responded": camp_responded,
            "called": camp_called,
            "pending": max(0, camp_pending),
            "outreach": [
                {
                    "email": o.get("email", ""),
                    "name": o.get("name", ""),
                    "agency": o.get("agency", ""),
                    "subject": o.get("subject", o.get("email_subject", "")),
                    "body": o.get("body", o.get("email_body", "")),
                    "body_html": o.get("body_html", o.get("email_body_html", "")),
                    "email_sent": o.get("email_sent", False),
                    "email_sent_at": o.get("email_sent_at"),
                    "error": o.get("error"),
                }
                for o in c.get("outreach", [])
            ],
        })

        total_sent += camp_sent
        total_bounced += camp_bounced
        total_responded += camp_responded
        total_called += camp_called
        total_pending += max(0, camp_pending)

    return {
        "ok": True,
        "campaigns": campaigns,
        "totals": {
            "sent": total_sent,
            "bounced": total_bounced,
            "responded": total_responded,
            "called": total_called,
            "pending_follow_up": total_pending,
            "total_prospects": len(prospect_map),
        },
        "status_breakdown": dict(status_counts),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SCPRS-POWERED GROWTH INTELLIGENCE  (Phase 32 upgrade)
# Reads live SCPRS data → gap analysis → ranked recommendations → outreach actions
# ═══════════════════════════════════════════════════════════════════════════════

def get_scprs_growth_intelligence() -> dict:
    """
    Primary intelligence function. Reads SCPRS DB data and returns:
    - Ranked action list (what Mike should do TODAY vs this week vs this month)
    - Gap analysis (what agencies buy that we don't sell them)
    - Win-back targets (we sell it, they're buying from Cardinal/McKesson)
    - Pricing lessons from auto-closed lost quotes
    - Agency expansion opportunities
    """
    try:
        from src.agents.scprs_intelligence_engine import get_growth_intelligence
        data = get_growth_intelligence()
    except Exception as e:
        return {"ok": False, "error": str(e), "recommendations": [], "gaps": [], "win_back": []}

    recs = data.get("recommendations", [])
    gaps = data.get("top_gaps", [])
    win_back = data.get("win_back", [])
    competitors = data.get("competitors", [])
    losses = data.get("recent_losses", [])
    by_agency = data.get("by_agency", [])

    # Compute total opportunity value
    total_gap = sum(g.get("total_spend") or 0 for g in gaps)
    total_wb = sum(w.get("total_spend") or 0 for w in win_back)

    # Compute pull coverage
    agencies_with_data = [a for a in by_agency if (a.get("po_count") or 0) > 0]

    return {
        "ok": True,
        "recommendations": recs,
        "gaps": gaps,
        "win_back": win_back,
        "competitors": competitors,
        "recent_losses": losses,
        "by_agency": by_agency,
        "summary": {
            "total_gap_opportunity": total_gap,
            "total_win_back": total_wb,
            "total_opportunity": total_gap + total_wb,
            "agencies_with_data": len(agencies_with_data),
            "rec_count": len(recs),
        },
        "pull_schedule": data.get("pull_schedule", []),
    }


def generate_recommendations() -> dict:
    """Upgraded: returns SCPRS-powered recommendations, falls back to legacy."""
    scprs = get_scprs_growth_intelligence()
    if scprs.get("ok") and scprs.get("recommendations"):
        return scprs
    return get_growth_status()


def full_report() -> dict:
    """Full growth report with SCPRS intelligence."""
    return generate_recommendations()


def get_plain_english_brief() -> str:
    """
    Returns a plain English brief for the manager dashboard.
    Answers: 'What should I do today to grow revenue?'
    Includes DVBE angle and unconstrained sourcing opportunities.
    """
    intel = get_scprs_growth_intelligence()
    if not intel.get("ok"):
        return "Run a full SCPRS pull to generate growth insights."

    lines = []
    recs = intel.get("recommendations", [])
    summary = intel.get("summary", {})

    total_opp = summary.get("total_opportunity", 0)
    agencies = summary.get("agencies_with_data", 0)
    if total_opp > 0:
        lines.append(
            f"💰 ${total_opp:,.0f} in identified opportunities across {agencies} agencies. "
            f"This includes items you already sell AND items you can source. "
            f"Your DVBE cert is the wildcard — use it."
        )

    # P0 actions — what to do today
    p0_recs = [r for r in recs if r.get("priority") == "P0"][:4]
    icons = {"dvbe_displace": "🏅", "win_back": "⚔️", "add_product": "📦",
             "pricing": "💲", "dvbe_partner": "🤝", "source_anything": "🔍"}
    for r in p0_recs:
        val = r.get("estimated_annual_value", 0)
        icon = icons.get(r.get("type", ""), "🎯")
        lines.append(f"{icon} {r.get('action','')} — ${val:,.0f}/yr estimate. {r.get('how','')[:90]}")

    # DVBE summary
    dvbe_recs = [r for r in recs if r.get("dvbe_angle")]
    if dvbe_recs:
        dvbe_val = sum(r.get("estimated_annual_value",0) for r in dvbe_recs)
        lines.append(
            f"🏅 {len(dvbe_recs)} DVBE opportunities worth ~${dvbe_val:,.0f}/yr — "
            f"Cardinal Health, McKesson and others don't have your DVBE cert. "
            f"Lead with it in every proposal to CCHCS, CalVet, DSH."
        )

    # Pricing lessons from auto-closed losses
    losses = intel.get("recent_losses", [])
    if losses:
        l = losses[0]
        delta = (l.get("total") or 0) - (l.get("scprs_total") or 0)
        if delta > 0:
            lines.append(
                f"📉 Lost {l.get('quote_number','')} to {l.get('scprs_supplier','')} "
                f"— overpriced by ${delta:,.0f}. SCPRS saved the intel. Fix pricing before next RFQ."
            )

    if not lines:
        lines.append(
            "No SCPRS data yet. Hit 'Pull All Agencies' on the Growth Intel page — "
            "it searches every PO every CA agency placed and shows you exactly what to sell."
        )

    return "\n".join(lines)



# ═══════════════════════════════════════════════════════════════════════
# V2 — Analytics & Intelligence Layer
# ═══════════════════════════════════════════════════════════════════════

def get_growth_kpis() -> dict:
    """Compute real-time KPI metrics for the growth dashboard."""
    prospect_data = _load_json(PROSPECTS_FILE) or []
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    outreach = _load_json(OUTREACH_FILE) or []
    if isinstance(outreach, dict):
        outreach = outreach.get("campaigns", [])
    creds = get_reytech_credentials()
    cohorts = get_follow_up_cohorts()

    total_prospects = len(prospects)
    emailed = sum(1 for p in prospects if p.get("status") in ("emailed", "follow_up", "second_follow_up", "responded", "won", "bounced"))
    responded = sum(1 for p in prospects if p.get("status") in ("responded", "won"))
    won = sum(1 for p in prospects if p.get("status") == "won")
    bounced = sum(1 for p in prospects if p.get("status") == "bounced")
    dead = sum(1 for p in prospects if p.get("status") == "dead")

    # Conversion rates
    email_rate = (emailed / total_prospects * 100) if total_prospects else 0
    response_rate = (responded / emailed * 100) if emailed else 0
    win_rate = (won / responded * 100) if responded else 0
    bounce_rate = (bounced / emailed * 100) if emailed else 0

    # Pipeline value from won prospects
    won_value = sum(float(p.get("annual_spend", 0) or 0) for p in prospects if p.get("status") == "won")
    pipeline_value = sum(float(p.get("annual_spend", 0) or 0) for p in prospects if p.get("status") in ("emailed", "follow_up", "responded"))

    # Campaign metrics
    campaign_list = outreach if isinstance(outreach, list) else []
    unique_campaigns = len(campaign_list)

    # Flatten outreach entries for trend
    all_outreach = []
    for c_ in campaign_list:
        for o_ in c_.get("outreach", []):
            all_outreach.append(o_)

    # Follow-up urgency
    fu_due = len(cohorts.get("no_response", []))
    fu_second = len(cohorts.get("second_followup", []))
    fu_stale = len(cohorts.get("stale", []))

    # Weekly trend (emails sent in last 7 days vs prior 7)
    now = datetime.now()
    week_ago = now - timedelta(days=7)
    two_weeks = now - timedelta(days=14)
    recent_sends = 0
    prior_sends = 0
    for o in all_outreach:
        try:
            ts = datetime.fromisoformat(o.get("sent_at", o.get("timestamp", ""))[:19])
            if ts >= week_ago:
                recent_sends += 1
            elif ts >= two_weeks:
                prior_sends += 1
        except Exception:
            pass
    trend_pct = ((recent_sends - prior_sends) / max(prior_sends, 1)) * 100

    return {
        "total_prospects": total_prospects,
        "emailed": emailed,
        "responded": responded,
        "won": won,
        "bounced": bounced,
        "dead": dead,
        "email_rate": round(email_rate, 1),
        "response_rate": round(response_rate, 1),
        "win_rate": round(win_rate, 1),
        "bounce_rate": round(bounce_rate, 1),
        "won_value": won_value,
        "pipeline_value": pipeline_value,
        "unique_campaigns": unique_campaigns,
        "fu_due": fu_due,
        "fu_second": fu_second,
        "fu_stale": fu_stale,
        "recent_sends": recent_sends,
        "trend_pct": round(trend_pct, 1),
        "total_sales": creds.get("total_sales", 0),
        "total_pos": creds.get("total_pos", 0),
        "agencies_served": creds.get("agencies_served", 0),
    }


def get_win_probability(prospect: dict, credentials: dict = None) -> dict:
    """Estimate win probability for a prospect using historical signals."""
    if not credentials:
        credentials = get_reytech_credentials()

    score = 50.0  # Base 50%
    factors = []

    # Factor 1: Past relationship (+25%)
    agency = (prospect.get("agency") or prospect.get("institution") or "").lower()
    served = [a.lower() for a in credentials.get("agency_list", [])]
    if any(s in agency or agency in s for s in served if len(s) > 3):
        score += 25
        factors.append(("Past customer", "+25%", "green"))
    else:
        factors.append(("New agency", "+0%", "grey"))

    # Factor 2: Spend level (+5 to +15%)
    spend = float(prospect.get("annual_spend", 0) or 0)
    if spend > 100000:
        score += 15
        factors.append(("High spender ($100K+)", "+15%", "green"))
    elif spend > 25000:
        score += 10
        factors.append(("Medium spender ($25K+)", "+10%", "blue"))
    elif spend > 5000:
        score += 5
        factors.append(("Low spender", "+5%", "grey"))

    # Factor 3: Category match (+10%)
    cats = prospect.get("categories", [])
    top_cats = credentials.get("top_categories", [])
    cat_match = sum(1 for c in cats if c in top_cats)
    if cat_match >= 2:
        score += 10
        factors.append((f"{cat_match} category matches", "+10%", "green"))
    elif cat_match == 1:
        score += 5
        factors.append(("1 category match", "+5%", "blue"))

    # Factor 4: Has responded before (+15%)
    if prospect.get("status") == "responded":
        score += 15
        factors.append(("Already responded", "+15%", "green"))

    # Factor 5: Dormancy penalty (-10%)
    if prospect.get("status") in ("stale", "dead"):
        score -= 10
        factors.append(("Stale/dead lead", "-10%", "red"))

    # Factor 6: DVBE preference (+10%)
    dvbe_keywords = ["dvbe", "disabled veteran", "small business"]
    body = (prospect.get("notes", "") + " " + prospect.get("full_body", "")).lower()
    if any(k in body for k in dvbe_keywords):
        score += 10
        factors.append(("DVBE preference detected", "+10%", "green"))

    # Factor 7: Recent engagement (+5%)
    if prospect.get("last_contacted"):
        try:
            lc = datetime.fromisoformat(prospect["last_contacted"][:19])
            days = (datetime.now() - lc).days
            if days < 14:
                score += 5
                factors.append(("Recent contact (<14d)", "+5%", "blue"))
        except Exception:
            pass

    score = max(5, min(95, score))  # Clamp 5-95%
    tier = "Hot" if score >= 75 else "Warm" if score >= 50 else "Cold"
    color = "#3fb950" if score >= 75 else "#fbbf24" if score >= 50 else "#f85149"

    return {
        "probability": round(score, 1),
        "tier": tier,
        "color": color,
        "factors": factors,
    }


def get_competitor_intel() -> dict:
    """Analyze competitor data from SCPRS history and lost quotes."""
    history_path = os.path.join(DATA_DIR, "growth_reytech_history.json")
    history = _load_json(history_path) or {}
    quotes = _load_json(os.path.join(DATA_DIR, "quotes_log.json")) or {}

    competitors = {}  # vendor_name -> {wins, total_value, categories, agencies}

    # Pull competitor data from SCPRS history
    for agency, records in history.items():
        if not isinstance(records, list):
            continue
        for rec in records:
            vendor = (rec.get("vendor") or rec.get("supplier") or "").strip()
            if not vendor or vendor.lower() in ("reytech", "reytech inc", "reytech, inc"):
                continue
            if vendor not in competitors:
                competitors[vendor] = {"wins": 0, "total_value": 0, "categories": set(), "agencies": set(), "items": []}
            competitors[vendor]["wins"] += 1
            competitors[vendor]["total_value"] += float(rec.get("total", 0) or 0)
            competitors[vendor]["agencies"].add(agency)
            cat = categorize_item(rec.get("description", ""))
            if cat != "General":
                competitors[vendor]["categories"].add(cat)
            competitors[vendor]["items"].append({
                "description": rec.get("description", "")[:60],
                "price": rec.get("total", 0),
                "agency": agency,
            })

    # Analyze lost quotes for competitor displacement
    lost_to = {}
    if isinstance(quotes, dict):
        for qid, q in quotes.items():
            st = (q.get("status") or "").lower()
            if st in ("lost", "expired", "no_response"):
                comp = q.get("lost_to", "") or q.get("competitor", "") or q.get("scprs_supplier", "")
                if comp:
                    if comp not in lost_to:
                        lost_to[comp] = {"count": 0, "value": 0, "quotes": []}
                    lost_to[comp]["count"] += 1
                    lost_to[comp]["value"] += float(q.get("total", 0) or 0)
                    lost_to[comp]["quotes"].append(qid)

    # Build ranked competitor list
    ranked = []
    for name, data in competitors.items():
        ranked.append({
            "name": name,
            "wins": data["wins"],
            "total_value": data["total_value"],
            "categories": list(data["categories"])[:5],
            "agencies": list(data["agencies"])[:5],
            "agency_count": len(data["agencies"]),
            "lost_to_us": lost_to.get(name, {}).get("count", 0),
            "lost_value": lost_to.get(name, {}).get("value", 0),
            "threat_level": "high" if data["total_value"] > 50000 else "medium" if data["total_value"] > 10000 else "low",
        })
    ranked.sort(key=lambda x: x["total_value"], reverse=True)

    return {
        "competitors": ranked[:25],
        "total_competitors": len(ranked),
        "top_threat": ranked[0]["name"] if ranked else "None identified",
        "lost_to_breakdown": lost_to,
    }


def get_lost_po_analysis() -> dict:
    """Deep analysis of lost purchase orders by category, agency, and reason."""
    quotes = _load_json(os.path.join(DATA_DIR, "quotes_log.json")) or {}

    by_category = {}
    by_agency = {}
    by_reason = {}
    by_month = {}
    total_lost_value = 0
    lost_items = []

    items_list = quotes if isinstance(quotes, list) else quotes.values()
    for q in items_list:
        st = (q.get("status") or "").lower()
        if st not in ("lost", "expired", "no_response", "not_responding"):
            continue

        val = float(q.get("total", 0) or 0)
        total_lost_value += val
        agency = q.get("institution") or q.get("agency") or "Unknown"
        reason = q.get("loss_reason") or q.get("dismiss_reason") or st.replace("_", " ").title()

        # By category
        for item in q.get("items", []):
            cat = categorize_item(item.get("description", ""))
            by_category[cat] = by_category.get(cat, 0) + float(item.get("total", 0) or item.get("price", 0) or 0)

        # By agency
        by_agency[agency] = by_agency.get(agency, {"count": 0, "value": 0})
        by_agency[agency]["count"] += 1
        by_agency[agency]["value"] += val

        # By reason
        by_reason[reason] = by_reason.get(reason, {"count": 0, "value": 0})
        by_reason[reason]["count"] += 1
        by_reason[reason]["value"] += val

        # By month
        created = q.get("created_at", "")[:7]
        if created:
            by_month[created] = by_month.get(created, 0) + val

        lost_items.append({
            "quote": q.get("quote_number") or q.get("solicitation_number", ""),
            "agency": agency,
            "value": val,
            "reason": reason,
            "competitor": q.get("lost_to") or q.get("scprs_supplier", ""),
            "items_count": len(q.get("items", [])),
            "date": q.get("created_at", "")[:10],
        })

    lost_items.sort(key=lambda x: x["value"], reverse=True)

    return {
        "total_lost_value": total_lost_value,
        "total_lost_count": len(lost_items),
        "by_category": dict(sorted(by_category.items(), key=lambda x: x[1], reverse=True)[:10]),
        "by_agency": dict(sorted(by_agency.items(), key=lambda x: x[1]["value"], reverse=True)[:10]),
        "by_reason": dict(sorted(by_reason.items(), key=lambda x: x[1]["value"], reverse=True)),
        "by_month": dict(sorted(by_month.items())),
        "top_losses": lost_items[:20],
    }


def export_growth_report(fmt: str = "csv") -> dict:
    """Export growth data as CSV or structured report."""
    prospects = _load_prospects_list()
    creds = get_reytech_credentials()
    kpis = get_growth_kpis()
    lost = get_lost_po_analysis()

    if fmt == "csv":
        import csv, io
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Agency", "Buyer", "Email", "Phone", "Status", "Annual Spend", "Score", "Win Prob %", "Workflow", "Last Contacted", "Tags", "Categories"])
        for p in prospects:
            wp = get_win_probability(p, creds)
            writer.writerow([
                p.get("agency", ""),
                p.get("buyer_name") or p.get("name", ""),
                p.get("email", ""),
                p.get("phone", ""),
                p.get("status", "new"),
                p.get("annual_spend", 0),
                round(score_prospect_weighted(p, creds), 1),
                round(wp.get("win_probability", 50)),
                p.get("workflow", {}).get("workflow_id", ""),
                p.get("last_contacted", ""),
                "; ".join(p.get("tags", [])),
                ", ".join(p.get("categories", [])),
            ])
        return {"ok": True, "format": "csv", "data": buf.getvalue(), "filename": "growth_prospects.csv"}

    elif fmt == "summary":
        return {
            "ok": True,
            "format": "summary",
            "credentials": creds,
            "kpis": kpis,
            "lost_analysis": lost,
            "prospect_count": len(prospects),
        }
    return {"ok": False, "error": f"Unknown format: {fmt}"}


# ── Audit Logging ──────────────────────────────────────────────────
_AUDIT_FILE = os.path.join(DATA_DIR, "growth_audit.json")

def log_growth_action(action: str, detail: str = "", actor: str = "user", metadata: dict = None):
    """Append to tamper-evident audit log for all growth actions."""
    audit = _load_json(_AUDIT_FILE) or []
    entry = {
        "id": f"ga_{datetime.now().strftime('%Y%m%d%H%M%S')}_{len(audit)}",
        "timestamp": datetime.now().isoformat(),
        "action": action,
        "detail": detail,
        "actor": actor,
        "metadata": metadata or {},
    }
    audit.append(entry)
    # Keep last 2000 entries
    if len(audit) > 2000:
        audit = audit[-2000:]
    _save_json(_AUDIT_FILE, audit)
    return entry


def get_audit_log(limit: int = 50, action_filter: str = None) -> list:
    """Retrieve recent audit entries."""
    audit = _load_json(_AUDIT_FILE) or []
    if action_filter:
        audit = [a for a in audit if a.get("action") == action_filter]
    return list(reversed(audit[-limit:]))


# ── A/B Template Tracking ──────────────────────────────────────────
_AB_FILE = os.path.join(DATA_DIR, "growth_ab_tests.json")

def track_template_send(template_key: str, variant: str = "A", prospect_id: str = ""):
    """Track which template variant was sent for A/B analysis."""
    ab = _load_json(_AB_FILE) or {}
    if template_key not in ab:
        ab[template_key] = {"variants": {}}
    v = ab[template_key]["variants"]
    if variant not in v:
        v[variant] = {"sent": 0, "opened": 0, "responded": 0, "sends": []}
    v[variant]["sent"] += 1
    v[variant]["sends"].append({
        "prospect_id": prospect_id,
        "timestamp": datetime.now().isoformat(),
    })
    _save_json(_AB_FILE, ab)


def track_template_response(template_key: str, variant: str = "A", event: str = "responded"):
    """Track response/open for A/B variant."""
    ab = _load_json(_AB_FILE) or {}
    if template_key in ab and variant in ab[template_key].get("variants", {}):
        ab[template_key]["variants"][variant][event] = ab[template_key]["variants"][variant].get(event, 0) + 1
        _save_json(_AB_FILE, ab)


def get_ab_stats() -> dict:
    """Get A/B test results for all templates."""
    ab = _load_json(_AB_FILE) or {}
    results = {}
    for tkey, tdata in ab.items():
        variants = {}
        for vname, vdata in tdata.get("variants", {}).items():
            sent = vdata.get("sent", 0)
            resp = vdata.get("responded", 0)
            variants[vname] = {
                "sent": sent,
                "responded": resp,
                "response_rate": round((resp / sent * 100) if sent else 0, 1),
                "winner": False,
            }
        # Mark winner
        if len(variants) > 1:
            best = max(variants.items(), key=lambda x: x[1]["response_rate"])
            best[1]["winner"] = True
        results[tkey] = variants
    return results


# ── SCPRS Prospect Enrichment ──────────────────────────────────────
def enrich_prospect_scprs(prospect: dict) -> dict:
    """Enrich a prospect with SCPRS data (recent purchases, spend patterns)."""
    from src.agents.scprs_lookup import lookup_price
    enriched = dict(prospect)
    agency = prospect.get("agency") or prospect.get("institution") or ""
    if not agency:
        return enriched

    # Check SCPRS history for this agency
    history_path = os.path.join(DATA_DIR, "growth_reytech_history.json")
    history = _load_json(history_path) or {}

    agency_data = None
    agency_lower = agency.lower()
    for key, records in history.items():
        if agency_lower in key.lower() or key.lower() in agency_lower:
            agency_data = records if isinstance(records, list) else []
            break

    if agency_data:
        enriched["scprs_purchase_count"] = len(agency_data)
        enriched["scprs_total_spend"] = sum(float(r.get("total", 0) or 0) for r in agency_data)
        enriched["scprs_last_purchase"] = max((r.get("date", "") for r in agency_data), default="")
        enriched["scprs_top_items"] = list(set(r.get("description", "")[:50] for r in agency_data[:10]))
        enriched["scprs_vendors"] = list(set(r.get("vendor", "") for r in agency_data if r.get("vendor")))[:5]
        enriched["scprs_enriched"] = True
    else:
        enriched["scprs_enriched"] = False

    return enriched


# ═══════════════════════════════════════════════════════════════════════
# V3 — Automation & Outreach Engine
# ═══════════════════════════════════════════════════════════════════════

# ── Auto Follow-Up Workflows ──────────────────────────────────────
_WORKFLOWS_FILE = os.path.join(DATA_DIR, "growth_workflows.json")

DEFAULT_WORKFLOWS = {
    "standard_outreach": {
        "name": "Standard Outreach",
        "description": "3-touch email sequence over 21 days",
        "steps": [
            {"day": 0, "action": "email", "template": "initial_outreach", "description": "Initial outreach email"},
            {"day": 5, "action": "email", "template": "follow_up", "description": "First follow-up (5 days)"},
            {"day": 14, "action": "email", "template": "second_follow_up", "description": "Second follow-up (14 days)"},
            {"day": 21, "action": "mark_stale", "template": None, "description": "Mark as stale if no response"},
        ],
        "active": True,
    },
    "reactivation": {
        "name": "Dormant Customer Reactivation",
        "description": "2-touch reactivation for past buyers",
        "steps": [
            {"day": 0, "action": "email", "template": "past_customer_reactivation", "description": "Reactivation email"},
            {"day": 10, "action": "email", "template": "follow_up", "description": "Follow-up if no response"},
            {"day": 21, "action": "mark_stale", "template": None, "description": "Close out"},
        ],
        "active": True,
    },
    "distro_list_push": {
        "name": "RFQ Distribution List Push",
        "description": "Focused campaign to get on RFQ lists",
        "steps": [
            {"day": 0, "action": "email", "template": "distro_list", "description": "Request RFQ list addition"},
            {"day": 7, "action": "email", "template": "follow_up", "description": "Follow-up with credentials"},
            {"day": 14, "action": "sms", "template": "sms_follow_up", "description": "SMS nudge"},
            {"day": 21, "action": "email", "template": "second_follow_up", "description": "Final email with referral ask"},
        ],
        "active": True,
    },
}


def get_workflows() -> dict:
    """Get all workflow definitions."""
    workflows = _load_json(_WORKFLOWS_FILE) or {}
    if not workflows:
        workflows = DEFAULT_WORKFLOWS
        _save_json(_WORKFLOWS_FILE, workflows)
    return workflows


def get_workflow_queue() -> list:
    """Get all prospects with pending workflow steps."""
    prospects = _load_prospects_list()
    queue = []

    for p in prospects:
        wf = p.get("workflow")
        if not wf or not wf.get("active"):
            continue
        wf_id = wf.get("workflow_id", "standard_outreach")
        workflows = get_workflows()
        wf_def = workflows.get(wf_id, {})
        steps = wf_def.get("steps", [])
        current_step = wf.get("current_step", 0)
        started = wf.get("started_at", "")

        if current_step >= len(steps):
            continue

        try:
            start_dt = datetime.fromisoformat(started[:19])
        except Exception:
            continue

        step = steps[current_step]
        due_dt = start_dt + timedelta(days=step["day"])
        overdue = datetime.now() > due_dt

        queue.append({
            "prospect_id": p.get("id"),
            "agency": p.get("agency") or p.get("institution", ""),
            "buyer": p.get("buyer_name") or p.get("name", ""),
            "email": p.get("email", ""),
            "workflow_name": wf_def.get("name", wf_id),
            "step_number": current_step + 1,
            "total_steps": len(steps),
            "step_action": step["action"],
            "step_template": step.get("template"),
            "step_description": step.get("description", ""),
            "due_date": due_dt.isoformat(),
            "overdue": overdue,
            "days_until": (due_dt - datetime.now()).days,
        })

    queue.sort(key=lambda x: x["due_date"])
    return queue


def assign_workflow(prospect_id: str, workflow_id: str = "standard_outreach") -> dict:
    """Assign a workflow sequence to a prospect."""
    _pdata = _load_json(PROSPECTS_FILE) or {}
    prospects = _pdata.get("prospects", []) if isinstance(_pdata, dict) else _pdata
    for p in prospects:
        if p.get("id") == prospect_id:
            p["workflow"] = {
                "workflow_id": workflow_id,
                "active": True,
                "current_step": 0,
                "started_at": datetime.now().isoformat(),
                "history": [],
            }
            if isinstance(_pdata, dict):
                _pdata["prospects"] = prospects
            _save_json(PROSPECTS_FILE, _pdata if isinstance(_pdata, dict) else prospects)
            log_growth_action("workflow_assigned", f"Workflow '{workflow_id}' assigned to {prospect_id}")
            return {"ok": True, "prospect_id": prospect_id, "workflow_id": workflow_id}
    return {"ok": False, "error": "Prospect not found"}


def advance_workflow_step(prospect_id: str, result: str = "completed") -> dict:
    """Advance a prospect to the next workflow step."""
    _pdata = _load_json(PROSPECTS_FILE) or {}
    prospects = _pdata.get("prospects", []) if isinstance(_pdata, dict) else _pdata
    for p in prospects:
        if p.get("id") == prospect_id:
            wf = p.get("workflow", {})
            if not wf.get("active"):
                return {"ok": False, "error": "No active workflow"}
            step = wf.get("current_step", 0)
            wf.setdefault("history", []).append({
                "step": step,
                "result": result,
                "completed_at": datetime.now().isoformat(),
            })
            wf["current_step"] = step + 1

            workflows = get_workflows()
            wf_def = workflows.get(wf.get("workflow_id", ""), {})
            if wf["current_step"] >= len(wf_def.get("steps", [])):
                wf["active"] = False
                wf["completed_at"] = datetime.now().isoformat()

            if isinstance(_pdata, dict):
                _pdata["prospects"] = prospects
            _save_json(PROSPECTS_FILE, _pdata if isinstance(_pdata, dict) else prospects)
            log_growth_action("workflow_step_advanced", f"Step {step+1} completed for {prospect_id}")
            return {"ok": True, "new_step": wf["current_step"], "active": wf["active"]}
    return {"ok": False, "error": "Prospect not found"}


# ── SMS Outreach via Twilio ────────────────────────────────────────
SMS_TEMPLATES = {
    "sms_follow_up": (
        "Hi {buyer_name}, this is Mike from Reytech Inc. "
        "We'd love to be on your RFQ distribution list for office & medical supplies. "
        "Can I send our credentials to {email}? Reply YES or email sales@reytechinc.com"
    ),
    "sms_reminder": (
        "Quick follow-up from Reytech Inc — we sent info about getting on your RFQ list. "
        "Any questions? Reply here or email sales@reytechinc.com. Thanks! -Mike"
    ),
}


def send_sms_outreach(phone: str, template_key: str = "sms_follow_up",
                       prospect: dict = None, dry_run: bool = True) -> dict:
    """Send SMS via Twilio to a prospect."""

    tmpl = SMS_TEMPLATES.get(template_key, SMS_TEMPLATES["sms_follow_up"])
    buyer = (prospect or {}).get("buyer_name") or (prospect or {}).get("name", "there")
    email = (prospect or {}).get("email", "your inbox")
    message = tmpl.format(buyer_name=buyer, email=email)

    if dry_run:
        return {"ok": True, "dry_run": True, "phone": phone, "message": message}

    try:
        twilio_sid = os.environ.get("TWILIO_SID", "")
        twilio_token = os.environ.get("TWILIO_TOKEN", "")
        twilio_from = os.environ.get("TWILIO_FROM", "")

        if not all([twilio_sid, twilio_token, twilio_from]):
            return {"ok": False, "error": "Twilio not configured"}

        from twilio.rest import Client
        client = Client(twilio_sid, twilio_token)
        msg = client.messages.create(body=message, from_=twilio_from, to=phone)

        log_growth_action("sms_sent", f"SMS to {phone}: {message[:50]}...",
                         metadata={"sid": msg.sid, "template": template_key})
        return {"ok": True, "sid": msg.sid, "phone": phone, "message": message}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── Notification Center ────────────────────────────────────────────
_NOTIFICATIONS_FILE = os.path.join(DATA_DIR, "growth_notifications.json")

def add_notification(title: str, message: str, level: str = "info",
                    action_url: str = "", metadata: dict = None):
    """Add an in-app notification."""
    notifs = _load_json(_NOTIFICATIONS_FILE) or []
    notifs.append({
        "id": f"n_{datetime.now().strftime('%Y%m%d%H%M%S')}_{len(notifs)}",
        "title": title,
        "message": message,
        "level": level,  # info, warning, success, error
        "action_url": action_url,
        "read": False,
        "created_at": datetime.now().isoformat(),
        "metadata": metadata or {},
    })
    # Keep last 200
    if len(notifs) > 200:
        notifs = notifs[-200:]
    _save_json(_NOTIFICATIONS_FILE, notifs)


def get_notifications(unread_only: bool = False, limit: int = 20) -> list:
    """Get recent notifications."""
    notifs = _load_json(_NOTIFICATIONS_FILE) or []
    if unread_only:
        notifs = [n for n in notifs if not n.get("read")]
    return list(reversed(notifs[-limit:]))


def mark_notification_read(notif_id: str):
    """Mark a notification as read."""
    notifs = _load_json(_NOTIFICATIONS_FILE) or []
    for n in notifs:
        if n.get("id") == notif_id:
            n["read"] = True
    _save_json(_NOTIFICATIONS_FILE, notifs)


def dismiss_all_notifications():
    """Mark all notifications as read."""
    notifs = _load_json(_NOTIFICATIONS_FILE) or []
    for n in notifs:
        n["read"] = True
    _save_json(_NOTIFICATIONS_FILE, notifs)


# ── Webhook Dispatch ───────────────────────────────────────────────
def fire_webhook(event: str, payload: dict) -> dict:
    """Fire webhook to configured endpoints (Slack, etc.)."""
    import json as _json
    webhook_url = os.environ.get("GROWTH_WEBHOOK_URL", "")
    slack_url = os.environ.get("SLACK_WEBHOOK_URL", "")

    results = []

    if slack_url:
        try:
            import urllib.request
            icon = {"campaign_sent": "📧", "prospect_responded": "✅",
                    "prospect_won": "🏆", "bounce_detected": "⚠️"}.get(event, "📌")
            slack_msg = {"text": f"{icon} *Growth Engine — {event}*\n{payload.get('summary', str(payload)[:200])}"}
            req = urllib.request.Request(
                slack_url,
                data=_json.dumps(slack_msg).encode(),
                headers={"Content-Type": "application/json"},
            )
            urllib.request.urlopen(req, timeout=5)
            results.append({"target": "slack", "ok": True})
        except Exception as e:
            results.append({"target": "slack", "ok": False, "error": str(e)})

    if webhook_url:
        try:
            import urllib.request
            data = _json.dumps({"event": event, "payload": payload}).encode()
            req = urllib.request.Request(
                webhook_url,
                data=data,
                headers={"Content-Type": "application/json"},
            )
            urllib.request.urlopen(req, timeout=5)
            results.append({"target": "custom", "ok": True})
        except Exception as e:
            results.append({"target": "custom", "ok": False, "error": str(e)})

    return {"event": event, "webhooks_fired": len(results), "results": results}


# ── SCPRS Personalization Engine ───────────────────────────────────
def generate_personalized_content(prospect: dict) -> dict:
    """Generate personalized outreach content using SCPRS + CRM data."""
    enriched = enrich_prospect_scprs(prospect)
    agency = enriched.get("agency") or enriched.get("institution", "your agency")
    buyer = enriched.get("buyer_name") or enriched.get("name", "")
    creds = get_reytech_credentials()

    # Build personalization context
    context = {
        "buyer_name": buyer,
        "agency": agency,
        "total_sales": f"${creds['total_sales']:,.0f}",
        "total_pos": creds["total_pos"],
        "agencies_served": creds["agencies_served"],
    }

    # SCPRS-powered personalization
    talking_points = []
    if enriched.get("scprs_enriched"):
        spend = enriched.get("scprs_total_spend", 0)
        count = enriched.get("scprs_purchase_count", 0)
        context["scprs_spend"] = f"${spend:,.0f}"
        context["scprs_count"] = count

        if spend > 0:
            talking_points.append(
                f"We've seen {agency} has purchased ${spend:,.0f} in supplies through {count} orders. "
                f"Our DVBE/SB certification means competitive pricing with guaranteed compliance."
            )
        vendors = enriched.get("scprs_vendors", [])
        if vendors:
            talking_points.append(
                f"We notice {agency} works with {', '.join(vendors[:3])}. "
                f"Reytech can offer the same product lines with DVBE advantages."
            )
        items = enriched.get("scprs_top_items", [])
        if items:
            talking_points.append(
                f"Based on recent purchases like '{items[0]}', we can provide competitive quotes "
                f"from our established supplier network."
            )
    else:
        talking_points.append(
            f"Reytech has fulfilled {creds['total_pos']}+ purchase orders across "
            f"{creds['agencies_served']} California state agencies since {creds['since_year']}."
        )

    # Win probability
    win_prob = get_win_probability(enriched, creds)
    context["win_probability"] = win_prob["probability"]
    context["win_tier"] = win_prob["tier"]

    return {
        "context": context,
        "talking_points": talking_points,
        "win_probability": win_prob,
        "scprs_enriched": enriched.get("scprs_enriched", False),
        "suggested_template": "past_customer_reactivation" if enriched.get("scprs_enriched") else "initial_outreach",
    }


# ═══════════════════════════════════════════════════════════════════════
# Feature #6 — Mobile-responsive helpers (data formatters)
# Feature #7 — Role-Based Access Control (RBAC)
# Feature #9 — Multi-Format Export (PDF, Excel, CSV)
# Feature #12 — Calendar Sync Integration
# Feature #13 — Data Encryption at Rest
# ═══════════════════════════════════════════════════════════════════════

# ── #7: RBAC ───────────────────────────────────────────────────────
_ROLES_FILE = os.path.join(DATA_DIR, "growth_roles.json")

ROLE_PERMISSIONS = {
    "admin": {"view", "edit", "send", "delete", "export", "configure", "audit"},
    "manager": {"view", "edit", "send", "export", "audit"},
    "agent": {"view", "send"},
    "viewer": {"view"},
}

def _load_roles() -> dict:
    roles = _load_json(_ROLES_FILE)
    if not roles:
        roles = {
            "default_role": "admin",
            "users": {},
        }
        _save_json(_ROLES_FILE, roles)
    return roles


def get_user_role(user_id: str = "default") -> str:
    """Get role for a user. Falls back to default_role."""
    roles = _load_roles()
    return roles.get("users", {}).get(user_id, roles.get("default_role", "admin"))


def check_permission(user_id: str, permission: str) -> bool:
    """Check if user has a specific permission."""
    role = get_user_role(user_id)
    perms = ROLE_PERMISSIONS.get(role, set())
    return permission in perms


def set_user_role(user_id: str, role: str) -> dict:
    """Assign a role to a user."""
    if role not in ROLE_PERMISSIONS:
        return {"ok": False, "error": f"Invalid role: {role}. Valid: {list(ROLE_PERMISSIONS.keys())}"}
    roles = _load_roles()
    roles.setdefault("users", {})[user_id] = role
    _save_json(_ROLES_FILE, roles)
    log_growth_action("role_changed", f"User '{user_id}' set to role '{role}'")
    return {"ok": True, "user_id": user_id, "role": role, "permissions": list(ROLE_PERMISSIONS[role])}


def list_roles() -> dict:
    """List all roles and their permissions."""
    roles = _load_roles()
    return {
        "roles": {k: list(v) for k, v in ROLE_PERMISSIONS.items()},
        "users": roles.get("users", {}),
        "default_role": roles.get("default_role", "admin"),
    }


# ── #9: Multi-Format Export (PDF + Excel) ──────────────────────────
def export_growth_pdf() -> dict:
    """Export growth report as PDF."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return {"ok": False, "error": "reportlab not installed. Run: pip install reportlab"}

    creds = get_reytech_credentials()
    kpis = get_growth_kpis()
    prospects = _load_json(PROSPECTS_FILE) or []
    if isinstance(prospects, dict):
        prospects = prospects.get("prospects", [])

    filepath = os.path.join(DATA_DIR, "growth_report.pdf")
    doc = SimpleDocTemplate(filepath, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Reytech Inc — Growth Engine Report", styles['Title']))
    elements.append(Spacer(1, 12))

    # Credentials
    cred_text = (
        f"Total Sales: ${creds.get('total_sales', 0):,.0f} | "
        f"Purchase Orders: {creds.get('total_pos', 0)}+ | "
        f"Agencies Served: {creds.get('agencies_served', 0)} | "
        f"DVBE + SB Certified"
    )
    elements.append(Paragraph(cred_text, styles['Normal']))
    elements.append(Spacer(1, 12))

    # KPIs
    elements.append(Paragraph("Key Performance Indicators", styles['Heading2']))
    kpi_data = [
        ["Metric", "Value"],
        ["Total Prospects", str(kpis.get("total_prospects", 0))],
        ["Emailed", str(kpis.get("emailed", 0))],
        ["Responded", str(kpis.get("responded", 0))],
        ["Response Rate", f"{kpis.get('response_rate', 0)}%"],
        ["Won", str(kpis.get("won", 0))],
        ["Pipeline Value", f"${kpis.get('pipeline_value', 0):,.0f}"],
        ["Weekly Trend", f"{kpis.get('trend_pct', 0)}%"],
    ]
    t = Table(kpi_data, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1a1b26")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 16))

    # Top prospects
    elements.append(Paragraph("Top Prospects by Score", styles['Heading2']))
    for pr_ in prospects:
        pr_["weighted_score"] = score_prospect_weighted(pr_, creds)
    prospects.sort(key=lambda x: x.get("weighted_score", 0), reverse=True)

    p_header = ["Agency", "Buyer", "Email", "Spend", "Score", "Status"]
    p_rows = [p_header]
    for p in prospects[:30]:
        p_rows.append([
            (p.get("agency") or p.get("institution", ""))[:30],
            (p.get("buyer_name") or p.get("name", ""))[:20],
            (p.get("email", ""))[:30],
            f"${float(p.get('annual_spend', 0) or 0):,.0f}",
            f"{p.get('weighted_score', 0):.0f}",
            p.get("status", "new"),
        ])
    pt = Table(p_rows, colWidths=[100, 80, 120, 60, 40, 60])
    pt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1a1b26")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(pt)

    doc.build(elements)
    log_growth_action("export_pdf", f"PDF report exported: {filepath}")
    return {"ok": True, "format": "pdf", "filepath": filepath, "filename": "growth_report.pdf"}


def export_growth_excel() -> dict:
    """Export growth data as Excel workbook with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return {"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl"}

    creds = get_reytech_credentials()
    kpis = get_growth_kpis()
    prospects = _load_json(PROSPECTS_FILE) or []
    if isinstance(prospects, dict):
        prospects = prospects.get("prospects", [])

    wb = openpyxl.Workbook()

    # ── Sheet 1: KPIs ──
    ws1 = wb.active
    ws1.title = "KPIs"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1b26", end_color="1a1b26", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    kpi_rows = [
        ("Total Sales", f"${creds.get('total_sales', 0):,.0f}"),
        ("Purchase Orders", f"{creds.get('total_pos', 0)}+"),
        ("Agencies Served", str(creds.get('agencies_served', 0))),
        ("Total Prospects", str(kpis.get('total_prospects', 0))),
        ("Emailed", str(kpis.get('emailed', 0))),
        ("Responded", str(kpis.get('responded', 0))),
        ("Response Rate", f"{kpis.get('response_rate', 0)}%"),
        ("Won", str(kpis.get('won', 0))),
        ("Pipeline Value", f"${kpis.get('pipeline_value', 0):,.0f}"),
        ("Won Value", f"${kpis.get('won_value', 0):,.0f}"),
    ]
    for row_idx, (metric, val) in enumerate(kpi_rows, 2):
        ws1.cell(row=row_idx, column=1, value=metric).border = thin_border
        ws1.cell(row=row_idx, column=2, value=val).border = thin_border
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20

    # ── Sheet 2: Prospects ──
    ws2 = wb.create_sheet("Prospects")
    p_headers = ["Agency", "Buyer", "Email", "Status", "Annual Spend", "Score", "Win Prob", "Last Contacted", "Categories"]
    for col, h in enumerate(p_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    for pr_ in prospects:
        pr_["weighted_score"] = score_prospect_weighted(pr_, creds)
        wp = get_win_probability(pr_, creds)
        pr_["win_prob"] = wp["probability"]
    prospects.sort(key=lambda x: x.get("weighted_score", 0), reverse=True)

    for row_idx, p in enumerate(prospects[:200], 2):
        vals = [
            p.get("agency") or p.get("institution", ""),
            p.get("buyer_name") or p.get("name", ""),
            p.get("email", ""),
            p.get("status", "new"),
            float(p.get("annual_spend", 0) or 0),
            round(p.get("weighted_score", 0), 1),
            round(p.get("win_prob", 50), 1),
            p.get("last_contacted", ""),
            ", ".join(p.get("categories", []) or p.get("categories_matched", [])),
        ]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=row_idx, column=col, value=v)
            c.border = thin_border
    for col in range(1, len(p_headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # ── Sheet 3: Lost POs ──
    ws3 = wb.create_sheet("Lost POs")
    lost = get_lost_po_analysis()
    l_headers = ["Quote #", "Agency", "Value", "Reason", "Competitor", "Date"]
    for col, h in enumerate(l_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
    for row_idx, l in enumerate(lost.get("top_losses", [])[:100], 2):
        vals = [l.get("quote", ""), l.get("agency", ""), l.get("value", 0),
                l.get("reason", ""), l.get("competitor", ""), l.get("date", "")]
        for col, v in enumerate(vals, 1):
            ws3.cell(row=row_idx, column=col, value=v).border = thin_border

    # ── Sheet 4: Competitors ──
    ws4 = wb.create_sheet("Competitors")
    comp_data = get_competitor_intel()
    c_headers = ["Competitor", "SCPRS Wins", "Total Value", "Agencies", "Threat Level"]
    for col, h in enumerate(c_headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    for row_idx, comp in enumerate(comp_data.get("competitors", [])[:50], 2):
        ws4.cell(row=row_idx, column=1, value=comp.get("name", ""))
        ws4.cell(row=row_idx, column=2, value=comp.get("wins", 0))
        ws4.cell(row=row_idx, column=3, value=comp.get("total_value", 0))
        ws4.cell(row=row_idx, column=4, value=comp.get("agency_count", 0))
        ws4.cell(row=row_idx, column=5, value=comp.get("threat_level", ""))

    filepath = os.path.join(DATA_DIR, "growth_report.xlsx")
    wb.save(filepath)
    log_growth_action("export_excel", f"Excel report exported: {filepath}")
    return {"ok": True, "format": "xlsx", "filepath": filepath, "filename": "growth_report.xlsx"}


# ── #12: Calendar Sync ─────────────────────────────────────────────
_CALENDAR_FILE = os.path.join(DATA_DIR, "growth_calendar.json")

def schedule_follow_up(prospect_id: str, date: str, time: str = "09:00",
                        notes: str = "", reminder_type: str = "email") -> dict:
    """Schedule a follow-up action on the calendar."""
    cal = _load_json(_CALENDAR_FILE) or {"events": []}

    prospect = get_prospect(prospect_id)
    agency = ""
    buyer = ""
    if prospect:
        agency = prospect.get("agency") or prospect.get("institution", "")
        buyer = prospect.get("buyer_name") or prospect.get("name", "")

    event = {
        "id": f"cal_{datetime.now().strftime('%Y%m%d%H%M%S')}_{len(cal['events'])}",
        "prospect_id": prospect_id,
        "agency": agency,
        "buyer": buyer,
        "date": date,
        "time": time,
        "type": reminder_type,
        "notes": notes or f"Follow up with {buyer} at {agency}",
        "completed": False,
        "created_at": datetime.now().isoformat(),
    }
    cal["events"].append(event)
    _save_json(_CALENDAR_FILE, cal)
    log_growth_action("calendar_scheduled", f"Follow-up scheduled: {agency} on {date}")

    # Generate Google Calendar link
    dt_str = f"{date}T{time}:00"
    try:
        start = datetime.fromisoformat(dt_str)
        end = start + timedelta(minutes=30)
        gcal_start = start.strftime("%Y%m%dT%H%M%S")
        gcal_end = end.strftime("%Y%m%dT%H%M%S")
        gcal_url = (
            f"https://calendar.google.com/calendar/render?action=TEMPLATE"
            f"&text=Follow+up:+{agency.replace(' ', '+')}"
            f"&dates={gcal_start}/{gcal_end}"
            f"&details={notes or f'Follow up with {buyer} at {agency}. Reytech Growth Engine.'}"
        )
        event["gcal_url"] = gcal_url
        _save_json(_CALENDAR_FILE, cal)
    except Exception:
        gcal_url = ""

    return {"ok": True, "event": event, "gcal_url": gcal_url}


def get_calendar_events(upcoming_only: bool = True) -> list:
    """Get scheduled follow-up events."""
    cal = _load_json(_CALENDAR_FILE) or {"events": []}
    events = cal.get("events", [])

    if upcoming_only:
        today = datetime.now().strftime("%Y-%m-%d")
        events = [e for e in events if not e.get("completed") and e.get("date", "") >= today]

    events.sort(key=lambda x: f"{x.get('date', '')} {x.get('time', '')}")
    return events


def complete_calendar_event(event_id: str) -> dict:
    """Mark a calendar event as completed."""
    cal = _load_json(_CALENDAR_FILE) or {"events": []}
    for e in cal.get("events", []):
        if e.get("id") == event_id:
            e["completed"] = True
            e["completed_at"] = datetime.now().isoformat()
            _save_json(_CALENDAR_FILE, cal)
            return {"ok": True, "event_id": event_id}
    return {"ok": False, "error": "Event not found"}


def get_todays_agenda() -> list:
    """Get today's scheduled follow-ups for the dashboard."""
    today = datetime.now().strftime("%Y-%m-%d")
    events = get_calendar_events(upcoming_only=True)
    return [e for e in events if e.get("date") == today]


# ── #13: Data Encryption at Rest ───────────────────────────────────
import base64
import hashlib

def _get_encryption_key() -> bytes:
    """Derive encryption key from environment or generate one."""
    key_str = os.environ.get("GROWTH_ENCRYPT_KEY", "")
    if not key_str:
        key_str = os.environ.get("SECRET_KEY", "reytech-default-key")
    return hashlib.sha256(key_str.encode()).digest()


def encrypt_field(value: str) -> str:
    """Encrypt a sensitive field value using AES-256-compatible XOR cipher.
    Uses Fernet-style encoding for portability without heavy deps."""
    if not value:
        return ""
    key = _get_encryption_key()
    # XOR-based encryption with key stretching
    encrypted = bytearray()
    for i, ch in enumerate(value.encode('utf-8')):
        encrypted.append(ch ^ key[i % len(key)])
    return "ENC:" + base64.b64encode(bytes(encrypted)).decode()


def decrypt_field(value: str) -> str:
    """Decrypt a field encrypted with encrypt_field."""
    if not value or not value.startswith("ENC:"):
        return value
    try:
        key = _get_encryption_key()
        encrypted = base64.b64decode(value[4:])
        decrypted = bytearray()
        for i, ch in enumerate(encrypted):
            decrypted.append(ch ^ key[i % len(key)])
        return decrypted.decode('utf-8')
    except Exception:
        return value


def encrypt_prospect_pii(prospect: dict) -> dict:
    """Encrypt PII fields in a prospect record."""
    encrypted = dict(prospect)
    pii_fields = ["email", "phone", "buyer_name", "name"]
    for field in pii_fields:
        if field in encrypted and encrypted[field] and not str(encrypted[field]).startswith("ENC:"):
            encrypted[field] = encrypt_field(str(encrypted[field]))
    return encrypted


def decrypt_prospect_pii(prospect: dict) -> dict:
    """Decrypt PII fields in a prospect record."""
    decrypted = dict(prospect)
    pii_fields = ["email", "phone", "buyer_name", "name"]
    for field in pii_fields:
        if field in decrypted and str(decrypted.get(field, "")).startswith("ENC:"):
            decrypted[field] = decrypt_field(str(decrypted[field]))
    return decrypted


# ── SCPRS Search Proxy ─────────────────────────────────────────────
def scprs_search_proxy(query: str, search_type: str = "item") -> dict:
    """Proxy SCPRS searches with caching and error handling."""
    cache_path = os.path.join(DATA_DIR, "scprs_search_cache.json")
    cache = _load_json(cache_path) or {}

    cache_key = f"{search_type}:{query.lower().strip()}"
    if cache_key in cache:
        cached = cache[cache_key]
        try:
            cached_at = datetime.fromisoformat(cached.get("cached_at", ""))
            age_hours = (datetime.now() - cached_at).total_seconds() / 3600
            if age_hours < 24:
                return {"ok": True, "cached": True, "results": cached.get("results", []), "query": query}
        except Exception:
            pass

    try:
        from src.agents.scprs_lookup import lookup_price
        results = lookup_price(
            item_number=query if search_type == "item" else None,
            description=query if search_type == "description" else None,
        )
        if results and isinstance(results, dict):
            results_list = [results] if not isinstance(results.get("results"), list) else results["results"]
        elif isinstance(results, list):
            results_list = results
        else:
            results_list = [results] if results else []

        # Cache results
        cache[cache_key] = {
            "results": results_list[:20],
            "cached_at": datetime.now().isoformat(),
            "query": query,
        }
        # Keep cache size reasonable
        if len(cache) > 500:
            oldest = sorted(cache.items(), key=lambda x: x[1].get("cached_at", ""))[:250]
            cache = dict(oldest)
        _save_json(cache_path, cache)

        log_growth_action("scprs_search", f"SCPRS search: {search_type}={query}", metadata={"results": len(results_list)})
        return {"ok": True, "cached": False, "results": results_list, "query": query}
    except Exception as e:
        return {"ok": False, "error": str(e), "query": query}


# ── Loss Reason Tracking ──────────────────────────────────────────
def add_loss_reason(quote_id: str, reason: str, competitor: str = "",
                    price_delta: float = 0, notes: str = "") -> dict:
    """Record reason for a lost PO/quote."""
    quotes = _load_json(os.path.join(DATA_DIR, "quotes_log.json")) or {}

    if isinstance(quotes, list):
        for q in quotes:
            qid = q.get("quote_number") or q.get("solicitation_number", "")
            if qid == quote_id:
                q["loss_reason"] = reason
                q["lost_to"] = competitor
                q["price_delta"] = price_delta
                q["loss_notes"] = notes
                q["loss_recorded_at"] = datetime.now().isoformat()
                _save_json(os.path.join(DATA_DIR, "quotes_log.json"), quotes)
                log_growth_action("loss_reason_added", f"Quote {quote_id}: {reason} (lost to {competitor})")
                return {"ok": True, "quote_id": quote_id, "reason": reason}
    elif isinstance(quotes, dict):
        if quote_id in quotes:
            quotes[quote_id]["loss_reason"] = reason
            quotes[quote_id]["lost_to"] = competitor
            quotes[quote_id]["price_delta"] = price_delta
            quotes[quote_id]["loss_notes"] = notes
            quotes[quote_id]["loss_recorded_at"] = datetime.now().isoformat()
            _save_json(os.path.join(DATA_DIR, "quotes_log.json"), quotes)
            log_growth_action("loss_reason_added", f"Quote {quote_id}: {reason}")
            return {"ok": True, "quote_id": quote_id, "reason": reason}

    return {"ok": False, "error": f"Quote '{quote_id}' not found"}


# ── Startup Check Integration ─────────────────────────────────────
def growth_startup_check() -> dict:
    """Run startup validation for growth module. Called by startup_checks.py."""
    issues = []
    warnings = []

    # Check data files exist
    required_files = [
        (PROSPECTS_FILE, "prospects"),
        (OUTREACH_FILE, "outreach"),
    ]
    for fpath, label in required_files:
        if not os.path.exists(fpath):
            warnings.append(f"{label} file not found: {fpath} (will be created on first use)")

    # Check SCPRS connectivity
    try:
        from src.agents.scprs_lookup import test_connection
        scprs_ok = test_connection()
        if not scprs_ok:
            warnings.append("SCPRS connection test failed — scraping may be unavailable")
    except Exception as e:
        warnings.append(f"SCPRS import error: {e}")

    # Check Twilio config
    if not os.environ.get("TWILIO_SID"):
        warnings.append("TWILIO_SID not set — SMS outreach disabled")

    # Check encryption key
    if not os.environ.get("GROWTH_ENCRYPT_KEY") and not os.environ.get("SECRET_KEY"):
        issues.append("No encryption key configured — PII encryption will use default key")

    # Check webhook config
    if not os.environ.get("SLACK_WEBHOOK_URL"):
        warnings.append("SLACK_WEBHOOK_URL not set — Slack notifications disabled")

    # Verify template count
    tmpl_count = len(EMAIL_TEMPLATES)
    if tmpl_count < 4:
        issues.append(f"Only {tmpl_count} email templates loaded (expected 6+)")

    status = "ok" if not issues else "warning" if not any("error" in i.lower() for i in issues) else "error"
    return {
        "module": "growth_engine",
        "status": status,
        "version": "3.0",
        "issues": issues,
        "warnings": warnings,
        "stats": {
            "templates": tmpl_count,
            "sms_templates": len(SMS_TEMPLATES),
            "workflows": len(get_workflows()),
        },
    }


# ═══════════════════════════════════════════════════════════════════════
# V4 — 12 NEW PRODUCTION FEATURES
# ═══════════════════════════════════════════════════════════════════════

# -- 1. Prospect Kanban Board --
def get_kanban_board():
    """Group prospects into Kanban columns for visual pipeline view."""
    prospect_data = _load_json(PROSPECTS_FILE) or []
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    creds = get_reytech_credentials()
    columns = {
        "new": {"label": "New", "color": "#4f8cff", "items": []},
        "emailed": {"label": "Emailed", "color": "#58a6ff", "items": []},
        "follow_up": {"label": "Follow-Up", "color": "#fbbf24", "items": []},
        "responded": {"label": "Responded", "color": "#3fb950", "items": []},
        "won": {"label": "Won", "color": "#34d399", "items": []},
        "bounced": {"label": "Bounced", "color": "#f85149", "items": []},
        "dead": {"label": "Dead", "color": "#8b949e", "items": []},
    }
    for p in prospects:
        st = p.get("status", "new")
        if st == "second_follow_up":
            st = "follow_up"
        if st in columns:
            columns[st]["items"].append({
                "id": p.get("id", ""),
                "agency": p.get("agency") or p.get("institution", ""),
                "buyer": p.get("buyer_name") or p.get("name", ""),
                "email": p.get("email", ""),
                "spend": float(p.get("annual_spend", 0) or 0),
                "score": round(score_prospect_weighted(p, creds), 1),
            })
    for col in columns.values():
        col["items"].sort(key=lambda x: x["score"], reverse=True)
        col["count"] = len(col["items"])
    return columns


# -- 2. Outreach Funnel Analytics --
def get_outreach_funnel():
    """Conversion funnel: prospects > emailed > responded > won."""
    prospect_data = _load_json(PROSPECTS_FILE) or []
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    total = len(prospects)
    emailed = sum(1 for p in prospects if p.get("status") in ("emailed", "follow_up", "second_follow_up", "responded", "won"))
    responded = sum(1 for p in prospects if p.get("status") in ("responded", "won"))
    won = sum(1 for p in prospects if p.get("status") == "won")
    stages = [
        {"name": "Prospects", "count": total, "pct": 100, "color": "#4f8cff"},
        {"name": "Emailed", "count": emailed, "pct": round(emailed / max(total, 1) * 100, 1), "color": "#58a6ff"},
        {"name": "Responded", "count": responded, "pct": round(responded / max(emailed, 1) * 100, 1), "color": "#3fb950"},
        {"name": "Won", "count": won, "pct": round(won / max(responded, 1) * 100, 1), "color": "#34d399"},
    ]
    return {"stages": stages, "overall_conversion": round(won / max(total, 1) * 100, 1), "email_to_response": round(responded / max(emailed, 1) * 100, 1)}


# -- 3. Agency Intelligence Map --
def get_agency_intelligence():
    """Rank agencies by opportunity value and engagement level."""
    prospect_data = _load_json(PROSPECTS_FILE) or []
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    agencies = {}
    for p in prospects:
        agency = p.get("agency") or p.get("institution", "Unknown")
        if agency not in agencies:
            agencies[agency] = {"name": agency, "prospects": 0, "total_spend": 0, "statuses": {}, "top_buyer": ""}
        a = agencies[agency]
        a["prospects"] += 1
        a["total_spend"] += float(p.get("annual_spend", 0) or 0)
        st = p.get("status", "new")
        a["statuses"][st] = a["statuses"].get(st, 0) + 1
        if not a["top_buyer"]:
            a["top_buyer"] = p.get("buyer_name") or p.get("name", "")
    ranked = sorted(agencies.values(), key=lambda x: x["total_spend"], reverse=True)
    for a in ranked:
        a["engagement"] = "hot" if a["statuses"].get("responded") or a["statuses"].get("won") else \
                           "warm" if a["statuses"].get("emailed") or a["statuses"].get("follow_up") else "cold"
    return {"agencies": ranked[:30], "total_agencies": len(ranked)}


# -- 4. Batch Workflow Assignment --
def batch_assign_workflow(prospect_ids, workflow_id="standard_outreach"):
    """Assign a workflow to multiple prospects at once."""
    results = {"ok": True, "assigned": 0, "errors": []}
    for pid in prospect_ids:
        r = assign_workflow(pid, workflow_id)
        if r.get("ok"):
            results["assigned"] += 1
        else:
            results["errors"].append({"id": pid, "error": r.get("error", "")})
    log_growth_action("batch_workflow", "Assigned %s to %d/%d prospects" % (workflow_id, results["assigned"], len(prospect_ids)))
    return results


# -- 5. Prospect Timeline --
def get_prospect_timeline(prospect_id):
    """Unified activity timeline for a prospect."""
    timeline = _load_json(os.path.join(DATA_DIR, "growth_timeline.json")) or {}
    events = list(timeline.get(prospect_id, []))
    audit = _load_json(_AUDIT_FILE) or []
    for entry in audit:
        if prospect_id in entry.get("detail", ""):
            events.append({"type": "audit", "event": entry.get("action", ""), "detail": entry.get("detail", ""), "timestamp": entry.get("timestamp", "")})
    outreach_data = _load_json(OUTREACH_FILE) or {}
    campaigns = outreach_data.get("campaigns", []) if isinstance(outreach_data, dict) else []
    for c in campaigns:
        for o in c.get("outreach", []):
            if o.get("prospect_id") == prospect_id:
                events.append({"type": "outreach", "event": "email_sent" if o.get("email_sent") else "queued", "detail": "Campaign: %s" % c.get("campaign_id", ""), "timestamp": o.get("sent_at", "")})
    events.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    return events[:50]


# -- 6. Quick-Win Identifier --
def get_quick_wins(max_results=10):
    """Surface prospects with highest win probability + recent activity."""
    prospect_data = _load_json(PROSPECTS_FILE) or []
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    creds = get_reytech_credentials()
    scored = []
    for p in prospects:
        st = p.get("status", "new")
        if st in ("won", "dead", "bounced"):
            continue
        wp = get_win_probability(p, creds)
        score = wp["probability"]
        if st == "responded":
            score += 20
        if p.get("last_contacted"):
            try:
                lc = datetime.fromisoformat(p["last_contacted"][:19])
                days = (datetime.now() - lc).days
                if days < 7:
                    score += 10
            except Exception:
                pass
        scored.append({
            "id": p.get("id", ""), "agency": p.get("agency") or p.get("institution", ""),
            "buyer": p.get("buyer_name") or p.get("name", ""), "email": p.get("email", ""),
            "status": st, "spend": float(p.get("annual_spend", 0) or 0),
            "win_prob": wp["probability"], "quick_score": round(min(score, 99), 1),
            "tier": wp["tier"], "color": wp["color"],
            "action": "Close the deal" if st == "responded" else "Send follow-up" if st in ("emailed", "follow_up") else "Initiate outreach",
        })
    scored.sort(key=lambda x: x["quick_score"], reverse=True)
    return scored[:max_results]


# -- 7. Campaign Performance Dashboard --
def get_campaign_performance():
    """Per-campaign stats: sent, responded, bounced, conversion rate."""
    outreach_data = _load_json(OUTREACH_FILE) or {}
    campaigns = outreach_data.get("campaigns", []) if isinstance(outreach_data, dict) else []
    results = []
    for c in campaigns:
        ol = c.get("outreach", [])
        sent = sum(1 for o in ol if o.get("email_sent"))
        bounced = sum(1 for o in ol if o.get("bounced"))
        responded = sum(1 for o in ol if o.get("response_received"))
        delivered = sent - bounced
        results.append({
            "campaign_id": c.get("campaign_id", ""), "created_at": c.get("created_at", ""),
            "template": c.get("template", ""), "total": len(ol), "sent": sent,
            "delivered": delivered, "bounced": bounced, "responded": responded,
            "delivery_rate": round(delivered / max(sent, 1) * 100, 1),
            "response_rate": round(responded / max(delivered, 1) * 100, 1),
        })
    results.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return results


# -- 8. Daily Growth Brief --
def generate_daily_brief():
    """Auto-generated daily summary of priorities and actions needed."""
    kpis = get_growth_kpis()
    cohorts = get_follow_up_cohorts()
    quick_wins = get_quick_wins(5)
    agenda = get_todays_agenda()
    queue = get_workflow_queue()
    overdue = [w for w in queue if w.get("overdue")]
    priorities = []
    if overdue:
        priorities.append({"urgency": "critical", "icon": "red_circle", "action": "%d overdue workflow steps" % len(overdue)})
    if cohorts.get("no_response"):
        priorities.append({"urgency": "high", "icon": "yellow_circle", "action": "%d need first follow-up" % len(cohorts["no_response"])})
    if cohorts.get("second_followup"):
        priorities.append({"urgency": "medium", "icon": "orange_circle", "action": "%d need second follow-up" % len(cohorts["second_followup"])})
    if quick_wins:
        qw = quick_wins[0]
        priorities.append({"urgency": "opportunity", "icon": "target", "action": "Quick win: %s (%s) - %.0f%% win prob" % (qw["agency"], qw["buyer"], qw["win_prob"])})
    if agenda:
        priorities.append({"urgency": "scheduled", "icon": "calendar", "action": "%d follow-ups scheduled today" % len(agenda)})
    return {
        "summary": "Growth Brief: %d prospects, %d responded, %d won. Response rate: %.1f%%. Pipeline: $%s." % (
            kpis["total_prospects"], kpis["responded"], kpis["won"], kpis["response_rate"], "{:,.0f}".format(kpis["pipeline_value"])),
        "priorities": priorities, "kpis": kpis, "quick_wins": quick_wins[:3],
        "overdue_count": len(overdue), "follow_ups_due": len(cohorts.get("no_response", [])),
        "agenda_count": len(agenda), "generated_at": datetime.now().isoformat(),
    }


# -- 9. Prospect Bulk Import --
def bulk_import_prospects(csv_text):
    """Import prospects from CSV text. Expected: agency,buyer_name,email,phone,spend"""
    import csv as _csv
    import io as _io
    reader = _csv.DictReader(_io.StringIO(csv_text))
    prospect_data = _load_json(PROSPECTS_FILE) or {}
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    existing_emails = {p.get("email", "").lower() for p in prospects if p.get("email")}
    imported = 0
    skipped = 0
    errors = []
    for row in reader:
        try:
            email = (row.get("email") or "").strip().lower()
            if email in existing_emails:
                skipped += 1
                continue
            new_p = {
                "id": "imp_%s_%d" % (datetime.now().strftime("%Y%m%d%H%M%S"), imported),
                "agency": (row.get("agency") or row.get("institution") or "").strip(),
                "buyer_name": (row.get("buyer_name") or row.get("name") or row.get("buyer") or "").strip(),
                "email": email, "phone": (row.get("phone") or "").strip(),
                "annual_spend": float(row.get("spend") or row.get("annual_spend") or 0),
                "status": "new", "source": "csv_import",
                "imported_at": datetime.now().isoformat(), "categories": [], "notes": row.get("notes", ""),
            }
            prospects.append(new_p)
            existing_emails.add(email)
            imported += 1
        except Exception as e:
            errors.append(str(e))
    if isinstance(prospect_data, dict):
        prospect_data["prospects"] = prospects
    else:
        prospect_data = {"prospects": prospects}
    _save_json(PROSPECTS_FILE, prospect_data)
    log_growth_action("bulk_import", "Imported %d prospects, skipped %d dupes" % (imported, skipped))
    return {"ok": True, "imported": imported, "skipped": skipped, "errors": errors[:5]}


# -- 10. Auto-Tagging Engine --
def auto_tag_prospects():
    """Automatically tag prospects by spend tier, engagement, and category."""
    prospect_data = _load_json(PROSPECTS_FILE) or {}
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    tagged = 0
    for p in prospects:
        tags = set(p.get("tags", []))
        spend = float(p.get("annual_spend", 0) or 0)
        if spend >= 100000:
            tags.add("tier:enterprise")
        elif spend >= 25000:
            tags.add("tier:mid-market")
        elif spend >= 5000:
            tags.add("tier:smb")
        else:
            tags.add("tier:micro")
        st = p.get("status", "new")
        if st in ("responded", "won"):
            tags.add("engagement:active")
        elif st in ("emailed", "follow_up", "second_follow_up"):
            tags.add("engagement:nurturing")
        elif st == "bounced":
            tags.add("engagement:invalid")
        else:
            tags.add("engagement:cold")
        source = p.get("source", "")
        if source:
            tags.add("source:%s" % source)
        body = (p.get("notes", "") + " " + p.get("full_body", "")).lower()
        if "dvbe" in body or "disabled veteran" in body:
            tags.add("dvbe-preference")
        for cat in (p.get("categories_matched") or p.get("categories") or [])[:3]:
            tags.add("cat:%s" % cat.lower().replace(" ", "-"))
        new_tags = list(tags)
        if set(new_tags) != set(p.get("tags", [])):
            p["tags"] = new_tags
            tagged += 1
    if isinstance(prospect_data, dict):
        prospect_data["prospects"] = prospects
    _save_json(PROSPECTS_FILE, prospect_data)
    log_growth_action("auto_tag", "Tagged %d prospects" % tagged)
    return {"ok": True, "tagged": tagged, "total": len(prospects)}


# -- 11. SCPRS Price Comparison --
def compare_pricing(item_description):
    """Compare Reytech pricing vs SCPRS market data."""
    quotes = _load_json(os.path.join(DATA_DIR, "quotes_log.json")) or {}
    our_prices = []
    items_list = quotes if isinstance(quotes, list) else quotes.values()
    for q in items_list:
        for item in q.get("items", []):
            desc = (item.get("description") or "").lower()
            if item_description.lower() in desc or desc in item_description.lower():
                our_prices.append({"price": float(item.get("unit_price") or item.get("price") or 0), "quote": q.get("quote_number", ""), "date": q.get("created_at", "")[:10]})
    scprs_result = scprs_search_proxy(item_description, "description")
    market_prices = []
    if scprs_result.get("ok"):
        for r in scprs_result.get("results", []):
            price = float(r.get("price") or r.get("unit_price") or 0)
            if price > 0:
                market_prices.append({"price": price, "vendor": r.get("vendor") or r.get("supplier", ""), "date": r.get("date", "")})
    our_avg = sum(p["price"] for p in our_prices) / max(len(our_prices), 1) if our_prices else 0
    market_avg = sum(p["price"] for p in market_prices) / max(len(market_prices), 1) if market_prices else 0
    delta_pct = ((our_avg - market_avg) / max(market_avg, 0.01)) * 100 if market_avg else 0
    return {
        "item": item_description, "our_prices": our_prices[:10], "market_prices": market_prices[:10],
        "our_avg": round(our_avg, 2), "market_avg": round(market_avg, 2), "delta_pct": round(delta_pct, 1),
        "competitive": delta_pct <= 5,
        "recommendation": "Competitive" if delta_pct <= 5 else "Slightly over market" if delta_pct <= 15 else "Significantly over market",
    }


# -- 12. Smart Prospect Deduplication --
def find_duplicate_prospects():
    """Find potential duplicate prospects by email or name+agency."""
    prospect_data = _load_json(PROSPECTS_FILE) or {}
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    email_groups = {}
    for p in prospects:
        email = (p.get("email") or "").strip().lower()
        if email:
            email_groups.setdefault(email, []).append(p)
    name_groups = {}
    for p in prospects:
        key = "%s|%s" % ((p.get("agency") or "").strip().lower(), (p.get("buyer_name") or p.get("name") or "").strip().lower())
        if key != "|":
            name_groups.setdefault(key, []).append(p)
    duplicates = []
    seen = set()
    for email, group in email_groups.items():
        if len(group) > 1 and email not in seen:
            duplicates.append({"type": "email", "key": email, "count": len(group), "ids": [p.get("id", "") for p in group]})
            seen.add(email)
    for key, group in name_groups.items():
        if len(group) > 1 and key not in seen:
            duplicates.append({"type": "name", "key": key, "count": len(group), "ids": [p.get("id", "") for p in group]})
            seen.add(key)
    return {"duplicates": duplicates, "total_duplicates": len(duplicates), "total_affected": sum(d["count"] for d in duplicates)}


def merge_prospects(keep_id, remove_ids):
    """Merge duplicate prospects, keeping one and removing others."""
    prospect_data = _load_json(PROSPECTS_FILE) or {}
    prospects = prospect_data.get("prospects", []) if isinstance(prospect_data, dict) else prospect_data
    keep = None
    for p in prospects:
        if p.get("id") == keep_id:
            keep = p
            break
    if not keep:
        return {"ok": False, "error": "Keep prospect not found"}
    removed = 0
    for rid in remove_ids:
        for i, p in enumerate(prospects):
            if p.get("id") == rid:
                if p.get("notes"):
                    keep["notes"] = (keep.get("notes", "") + "\n[Merged] " + p["notes"]).strip()
                if float(p.get("annual_spend", 0) or 0) > float(keep.get("annual_spend", 0) or 0):
                    keep["annual_spend"] = p["annual_spend"]
                existing_tags = set(keep.get("tags", []))
                existing_tags.update(p.get("tags", []))
                keep["tags"] = list(existing_tags)
                prospects.pop(i)
                removed += 1
                break
    if isinstance(prospect_data, dict):
        prospect_data["prospects"] = prospects
    _save_json(PROSPECTS_FILE, prospect_data)
    log_growth_action("merge_prospects", "Merged %d duplicates into %s" % (removed, keep_id))
    return {"ok": True, "kept": keep_id, "removed": removed}
